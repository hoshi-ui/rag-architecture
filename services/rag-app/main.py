from fastapi import FastAPI, HTTPException, Depends, Request, UploadFile, File, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any, Tuple
from dataclasses import dataclass, field
import os
import asyncio
import json
import logging
import io
import csv
import mimetypes
import zipfile
import importlib.util
import hashlib
from datetime import datetime, timedelta
from dotenv import load_dotenv
from fastapi.responses import FileResponse, JSONResponse
import re
import math
import threading
import sqlite3
import shutil
from functools import lru_cache
from xml.etree import ElementTree
import tempfile

# 加载环境变量
_dotenv_candidates = [
    os.getenv("APP_ENV_FILE"),
    os.path.join(os.path.dirname(__file__), "..", "..", "config", "app.env"),
    "/app/config/app.env",
]
for _dotenv_path in _dotenv_candidates:
    if _dotenv_path and os.path.exists(_dotenv_path):
        load_dotenv(_dotenv_path)
        break
else:
    load_dotenv()

# 配置日志
logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO"),
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger("rag-app")


_DEFAULT_RETRIEVAL_POLICY: Dict[str, Any] = {
    "query_route": {
        "order": ["existence", "visibility_probe", "version_switch", "explicit_doc_reference", "explicit_regulation_reference", "weak_title_reference", "business_topic_qa", "open_regulation_qa", "content_qa"],
        "weak_title_reference": {"require_no_filenames": True},
    },
    "source_resolution": {
        "title_candidate_limit": 5,
        "fallback_candidate_limit": 3,
        "clarification_examples_limit": 3,
    },
    "question_type_patterns": [
        {"type": "screening", "keywords": ["哪些文档", "哪些条例", "哪些包含", "哪些涉及", "涉及哪些", "全量筛查", "筛查"]},
        {"type": "single_doc_extract", "keywords": ["提取条款", "提取措施", "提取处罚", "抽取条款", "抽取措施", "抽取处罚", "条款提取"]},
        {"type": "summary", "keywords": ["概括", "总结", "核心内容", "概要", "综述"]},
        {"type": "arch", "keywords": ["架构", "系统设计", "模块", "组件", "存储", "调用", "链路", "流程图", "数据流"]},
        {"type": "compare", "keywords": ["区别", "差异", "对比", "比较", "不同", "分别", "各自", "vs", "versus", "difference between", "compare"]},
        {"type": "howto", "keywords": ["怎么做", "如何", "步骤", "流程", "实施", "配置", "安装", "how to", "steps"]},
        {"type": "definition", "keywords": ["什么是", "是什么", "定义", "meaning of", "what is"]},
    ],
    "rerank": {
        "section_lookup": {
            "intent_keywords": ["第几章", "哪一章", "哪一节", "章节", "条款", "条文", "总则", "附则", "奖惩", "法律责任", "立法程序", "监督检查"],
            "restriction_keywords": ["限制", "禁止", "约束", "限养", "禁养"],
            "restriction_subject_keywords": ["养犬", "犬只", "携犬"],
            "trigger_qtypes": ["regulation_execution"],
        },
        "clause_lookup": {
            "keywords": ["处罚", "罚款", "罚则", "限制", "禁止", "不得", "应当", "流程", "程序", "步骤", "登记", "审批", "申请", "标准", "时限", "材料", "条款"],
            "trigger_qtypes": ["single_doc_extract", "regulation_execution", "howto"],
        },
        "broad": {
            "keywords": ["相关内容", "主要内容", "核心内容", "概括", "总结", "综述", "介绍", "是什么", "有哪些"],
            "trigger_qtypes": ["summary"],
        },
    },
    "query_filters": {
        "doc_type_rules": [
            {"value": "regulation", "match_any": ["地方性法规", "管理规定", "法规"]},
            {"value": "research_report", "match_any": ["研究报告", "调研报告", "白皮书"]},
        ],
        "topic_rules": [
            {"value": "环保治理制度设计", "match_any": ["环保治理制度设计"], "match_all": ["环保", "制度"]},
            {"value": "AI成熟度研究", "match_any": ["ai成熟度研究"], "match_all": ["人工智能", "成熟度"]},
        ],
    },
    "weak_reference": {
        "generic_doc_markers": ["条例", "办法", "规定", "规则", "条款"],
        "generic_need_markers": ["核心", "处罚", "罚款", "程序", "要求", "责任", "禁止", "是否", "可否", "怎么", "如何"],
    },
}


def _deep_merge_dict(base: Dict[str, Any], override: Dict[str, Any]) -> Dict[str, Any]:
    merged = dict(base or {})
    for key, value in (override or {}).items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key] = _deep_merge_dict(merged[key], value)
        else:
            merged[key] = value
    return merged


def _load_retrieval_policy() -> Dict[str, Any]:
    candidates = [
        os.getenv("RETRIEVAL_POLICY_FILE"),
        os.path.join(os.path.dirname(__file__), "..", "..", "config", "retrieval_policy.json"),
        "/app/config/retrieval_policy.json",
    ]
    for path in candidates:
        if not path or not os.path.exists(path):
            continue
        try:
            with open(path, "r", encoding="utf-8") as f:
                payload = json.load(f)
            return _deep_merge_dict(_DEFAULT_RETRIEVAL_POLICY, payload)
        except Exception as exc:
            logger.warning(f"Failed to load retrieval policy from {path}: {exc}")
    return dict(_DEFAULT_RETRIEVAL_POLICY)


RETRIEVAL_POLICY = _load_retrieval_policy()


def _policy_get(path: str, default: Any = None) -> Any:
    node: Any = RETRIEVAL_POLICY
    for part in (path or "").split("."):
        if not part:
            continue
        if not isinstance(node, dict) or part not in node:
            return default
        node = node.get(part)
    return node if node is not None else default


def _policy_keywords(path: str) -> List[str]:
    values = _policy_get(path, default=[])
    if not isinstance(values, list):
        return []
    return [str(value) for value in values if str(value).strip()]


def _policy_match_rule(query: str, rule: Dict[str, Any]) -> bool:
    q = (query or "").lower()
    match_any = [str(value).lower() for value in (rule or {}).get("match_any") or [] if str(value).strip()]
    match_all = [str(value).lower() for value in (rule or {}).get("match_all") or [] if str(value).strip()]
    branches: List[bool] = []
    if match_any:
        branches.append(any(token in q for token in match_any))
    if match_all:
        branches.append(all(token in q for token in match_all))
    if not branches:
        return False
    return any(branches)

# 创建应用
app = FastAPI(
    title="RAG Application",
    description="企业级检索增强生成系统",
    version="1.0.0"
)

# 跨域配置
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 前端目录（静态挂载在 API 路由之后）
WEB_DIR = os.path.join(os.path.dirname(__file__), "web")
LEGACY_UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
PROJECT_ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
DEFAULT_DATABASE_DIR = os.path.join(PROJECT_ROOT_DIR, "database")


def _resolve_runtime_upload_dir() -> str:
    default_root = "/storage" if os.getenv("APP_ENV", "").lower() != "test_local" else LEGACY_UPLOAD_DIR
    runtime_root = os.getenv("RAG_RUNTIME_ROOT", default_root)
    candidate = runtime_root if os.path.basename(runtime_root.rstrip("/")) == "uploads" else os.path.join(runtime_root, "uploads")
    try:
        os.makedirs(candidate, exist_ok=True)
        return candidate
    except Exception:
        os.makedirs(LEGACY_UPLOAD_DIR, exist_ok=True)
        return LEGACY_UPLOAD_DIR


def _resolve_runtime_database_dir() -> str:
    candidate = os.getenv("RAG_DATABASE_DIR", DEFAULT_DATABASE_DIR)
    try:
        os.makedirs(candidate, exist_ok=True)
        return candidate
    except Exception:
        os.makedirs(DEFAULT_DATABASE_DIR, exist_ok=True)
        return DEFAULT_DATABASE_DIR


UPLOAD_DIR = _resolve_runtime_upload_dir()
DATABASE_DIR = _resolve_runtime_database_dir()
TASKS: Dict[str, Dict[str, Any]] = {}
TASKS_FILE = os.path.join(UPLOAD_DIR, "tasks.json")
LEXICAL_DB_FILE = os.path.join(DATABASE_DIR, "lexical_index.db")
_LEX_DB = None
_SOURCE_LOCKS: Dict[str, threading.Lock] = {}
_SOURCE_ASYNC_TASKS: Dict[str, set] = {}
_REDIS_CLIENT = None
_REDIS_DISABLED = False
_PENDING_CLARIFICATIONS: Dict[str, Dict[str, Any]] = {}
_PENDING_CLARIFICATIONS_LOCK = threading.Lock()
_CURRENT_LOCKED_DOCUMENTS: Dict[str, Dict[str, Any]] = {}
_CURRENT_LOCKED_DOCUMENTS_LOCK = threading.Lock()


def _safe_filename(name: str) -> str:
    n = (name or "").strip().replace("\\", "/")
    n = n.split("/")[-1]
    n = "".join([c for c in n if c.isalnum() or c in (".", "_", "-", " ")])
    return n or ("file_" + datetime.now().strftime("%Y%m%d_%H%M%S"))


def _clarification_user_key(user_id: str) -> str:
    uid = re.sub(r"[^a-zA-Z0-9_\-:.]", "_", (user_id or "anonymous").strip())
    return uid or "anonymous"


def _clarification_redis_key(user_id: str) -> str:
    return f"rag:clarify:pending:{_clarification_user_key(user_id)}"


def _clarification_pending_ttl_sec() -> int:
    try:
        return max(60, int(os.getenv("CLARIFICATION_PENDING_TTL_SEC", "900")))
    except Exception:
        return 900


def _current_locked_document_redis_key(user_id: str) -> str:
    return f"rag:locked_doc:current:{_clarification_user_key(user_id)}"


def _current_locked_document_ttl_sec() -> int:
    try:
        return max(120, int(os.getenv("CURRENT_LOCKED_DOCUMENT_TTL_SEC", "1800")))
    except Exception:
        return 1800


def _redis_client():
    global _REDIS_CLIENT, _REDIS_DISABLED
    if _REDIS_CLIENT is not None:
        return _REDIS_CLIENT
    if _REDIS_DISABLED:
        return None
    try:
        import redis  # type: ignore
        url = os.getenv("REDIS_URL", "redis://redis:6379/0")
        client = redis.Redis.from_url(
            url,
            decode_responses=True,
            socket_timeout=1.5,
            socket_connect_timeout=1.5,
        )
        client.ping()
        _REDIS_CLIENT = client
        return _REDIS_CLIENT
    except Exception:
        _REDIS_DISABLED = True
        return None


def _set_pending_clarification(user_id: str, query: str, candidates: List[str], reason: str = "section_anchor_ambiguous"):
    safe_candidates = [(_normalize_filename_for_match(x) or "").strip() for x in (candidates or [])]
    safe_candidates = [x for x in safe_candidates if x]
    if not safe_candidates:
        return
    payload = {
        "query": _normalize_query(query),
        "candidates": safe_candidates,
        "reason": reason,
        "created_at": datetime.now().isoformat(),
    }
    ttl = _clarification_pending_ttl_sec()
    client = _redis_client()
    if client is not None:
        try:
            client.setex(_clarification_redis_key(user_id), ttl, json.dumps(payload, ensure_ascii=False))
            return
        except Exception:
            pass
    with _PENDING_CLARIFICATIONS_LOCK:
        _PENDING_CLARIFICATIONS[_clarification_user_key(user_id)] = {
            "payload": payload,
            "expires_at": datetime.now().timestamp() + float(ttl),
        }


def _get_pending_clarification(user_id: str) -> Optional[Dict[str, Any]]:
    client = _redis_client()
    if client is not None:
        try:
            raw = client.get(_clarification_redis_key(user_id))
            if not raw:
                return None
            data = json.loads(raw)
            if isinstance(data, dict):
                return data
        except Exception:
            pass
    uid = _clarification_user_key(user_id)
    with _PENDING_CLARIFICATIONS_LOCK:
        row = _PENDING_CLARIFICATIONS.get(uid)
        if not row:
            return None
        if float(row.get("expires_at") or 0.0) < datetime.now().timestamp():
            _PENDING_CLARIFICATIONS.pop(uid, None)
            return None
        payload = row.get("payload")
        return payload if isinstance(payload, dict) else None


def _clear_pending_clarification(user_id: str):
    client = _redis_client()
    if client is not None:
        try:
            client.delete(_clarification_redis_key(user_id))
        except Exception:
            pass
    with _PENDING_CLARIFICATIONS_LOCK:
        _PENDING_CLARIFICATIONS.pop(_clarification_user_key(user_id), None)


def _set_current_locked_document(
    user_id: str,
    source: str,
    *,
    reason: str = "",
    reliable: bool = True,
    lock_mode: str = "hard_lock",
):
    safe_source = _normalize_filename_for_match(source or "")
    if not safe_source:
        return
    payload = {
        "source": safe_source,
        "reason": (reason or "").strip(),
        "reliable": bool(reliable),
        "lock_mode": (lock_mode or "hard_lock").strip(),
        "created_at": datetime.now().isoformat(),
    }
    ttl = _current_locked_document_ttl_sec()
    client = _redis_client()
    if client is not None:
        try:
            client.setex(_current_locked_document_redis_key(user_id), ttl, json.dumps(payload, ensure_ascii=False))
            return
        except Exception:
            pass
    with _CURRENT_LOCKED_DOCUMENTS_LOCK:
        _CURRENT_LOCKED_DOCUMENTS[_clarification_user_key(user_id)] = {
            "payload": payload,
            "expires_at": datetime.now().timestamp() + float(ttl),
        }


def _get_current_locked_document(user_id: str) -> Optional[Dict[str, Any]]:
    client = _redis_client()
    if client is not None:
        try:
            raw = client.get(_current_locked_document_redis_key(user_id))
            if not raw:
                return None
            data = json.loads(raw)
            if isinstance(data, dict):
                return data
        except Exception:
            pass
    uid = _clarification_user_key(user_id)
    with _CURRENT_LOCKED_DOCUMENTS_LOCK:
        row = _CURRENT_LOCKED_DOCUMENTS.get(uid)
        if not row:
            return None
        if float(row.get("expires_at") or 0.0) < datetime.now().timestamp():
            _CURRENT_LOCKED_DOCUMENTS.pop(uid, None)
            return None
        payload = row.get("payload")
        return payload if isinstance(payload, dict) else None


def _clear_current_locked_document(user_id: str):
    client = _redis_client()
    if client is not None:
        try:
            client.delete(_current_locked_document_redis_key(user_id))
        except Exception:
            pass
    with _CURRENT_LOCKED_DOCUMENTS_LOCK:
        _CURRENT_LOCKED_DOCUMENTS.pop(_clarification_user_key(user_id), None)


def _zh_number_to_int(token: str) -> Optional[int]:
    mapping = {"一": 1, "二": 2, "两": 2, "三": 3, "四": 4, "五": 5, "六": 6, "七": 7, "八": 8, "九": 9, "十": 10}
    t = (token or "").strip()
    if not t:
        return None
    if t in mapping:
        return mapping[t]
    if t.startswith("十") and len(t) == 2 and t[1] in mapping:
        return 10 + mapping[t[1]]
    if len(t) == 2 and t[0] in mapping and t[1] == "十":
        return mapping[t[0]] * 10
    if len(t) == 3 and t[0] in mapping and t[1] == "十" and t[2] in mapping:
        return mapping[t[0]] * 10 + mapping[t[2]]
    return None


def _parse_pending_candidate_selection(query: str, max_options: int) -> Optional[int]:
    q = _normalize_query(query)
    if not q or max_options <= 0 or len(q) > 20:
        return None
    number: Optional[int] = None
    if re.fullmatch(r"\d{1,2}", q):
        number = int(q)
    else:
        m = re.fullmatch(r"(?:选|选择)?第?(\d{1,2})(?:个|条|项|份)?", q)
        if m:
            number = int(m.group(1))
        else:
            m2 = re.fullmatch(r"(?:选|选择)?第?([一二两三四五六七八九十]{1,3})(?:个|条|项|份)?", q)
            if m2:
                number = _zh_number_to_int(m2.group(1))
    if number is None:
        return None
    if 1 <= number <= int(max_options):
        return number - 1
    return None


def _public_task_status(status: Optional[str]) -> str:
    s = (status or "").strip().lower()
    if not s:
        return "unknown"
    if s == "uploaded":
        return "accepted"
    if s in {"validating", "parsing", "chunking", "embedding", "embedding_partial", "indexing", "indexing_sqlite", "indexing_vector", "profile_building", "publish_pending", "reindexing", "vector_pending"}:
        return "indexing"
    if s in {"indexed", "completed"}:
        return "completed"
    if s in {"failed", "parse_failed", "parse_empty", "parse_low_quality", "unsupported_or_corrupt", "encrypted_file", "suspicious_file_type", "index_failed", "profile_failed", "publish_failed", "vector_failed", "delete_failed"}:
        return "failed"
    return s

def _new_task_id() -> str:
    import uuid
    return uuid.uuid4().hex

def _task_log(task_id: str, event: str, info: Optional[Dict[str, Any]] = None):
    t = TASKS.setdefault(task_id, {})
    logs = t.setdefault("logs", [])
    entry = {"ts": datetime.now().isoformat(), "event": event}
    if info:
        for k, v in info.items():
            entry[k] = v
    logs.append(entry)
    _save_tasks()

def _load_tasks():
    try:
        if os.path.exists(TASKS_FILE):
            import json
            with open(TASKS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict):
                    TASKS.update(data)
    except Exception:
        pass

def _save_tasks():
    try:
        import json
        tmp = TASKS.copy()
        with open(TASKS_FILE, "w", encoding="utf-8") as f:
            json.dump(tmp, f, ensure_ascii=False, indent=2)
    except Exception:
        pass
def _lex_db_connect():
    global _LEX_DB
    if _LEX_DB is None:
        real_db_path = os.getenv("LEXICAL_DB_FILE", LEXICAL_DB_FILE)
        os.makedirs(os.path.dirname(real_db_path), exist_ok=True)
        _LEX_DB = sqlite3.connect(real_db_path, check_same_thread=False)
        _LEX_DB.execute("PRAGMA journal_mode=WAL;")
        _LEX_DB.execute("PRAGMA synchronous=NORMAL;")
    return _LEX_DB
def _lex_db_init():
    conn = _lex_db_connect()
    conn.execute("CREATE TABLE IF NOT EXISTS chunks_meta (id INTEGER PRIMARY KEY AUTOINCREMENT, source TEXT, chunk_id INTEGER, section TEXT, metadata TEXT)")
    conn.execute("CREATE VIRTUAL TABLE IF NOT EXISTS chunks_fts USING fts5(text)")
    conn.execute(
        "CREATE TABLE IF NOT EXISTS document_ir_meta ("
        "source TEXT, "
        "doc_id TEXT, "
        "doc_version INTEGER, "
        "metadata TEXT, "
        "parser_name TEXT, "
        "parser_version TEXT, "
        "PRIMARY KEY(source, doc_version)"
        ")"
    )
    conn.execute(
        "CREATE TABLE IF NOT EXISTS document_ir ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT, "
        "source TEXT, "
        "doc_id TEXT, "
        "doc_version INTEGER, "
        "element_id TEXT, "
        "page_no INTEGER, "
        "section_path TEXT, "
        "element_type TEXT, "
        "bbox TEXT, "
        "reading_order INTEGER, "
        "text_raw TEXT, "
        "text_normalized TEXT, "
        "html TEXT, "
        "markdown TEXT, "
        "json_payload TEXT, "
        "ocr_used INTEGER, "
        "ocr_confidence REAL, "
        "parser_name TEXT, "
        "parser_version TEXT"
        ")"
    )
    conn.execute("CREATE TABLE IF NOT EXISTS doc_status (source TEXT PRIMARY KEY, status TEXT, updated_at TEXT)")
    conn.execute("CREATE TABLE IF NOT EXISTS documents (source TEXT PRIMARY KEY, status TEXT, active_version INTEGER, pending_version INTEGER, last_error TEXT, updated_at TEXT)")
    conn.execute(
        "CREATE TABLE IF NOT EXISTS document_profiles ("
        "source TEXT, "
        "source_id TEXT, "
        "doc_version INTEGER, "
        "original_filename TEXT, "
        "canonical_title TEXT, "
        "region TEXT, "
        "doc_type TEXT, "
        "publish_date TEXT, "
        "effective_date TEXT, "
        "doc_version_label TEXT, "
        "parse_quality_score REAL, "
        "quality_flags TEXT, "
        "source_resolution_fields TEXT, "
        "parser_route TEXT, "
        "parser_backend TEXT, "
        "mime_type TEXT, "
        "detected_ext TEXT, "
        "file_size INTEGER, "
        "page_count INTEGER, "
        "content_sha256 TEXT, "
        "created_at TEXT, "
        "updated_at TEXT, "
        "PRIMARY KEY(source, doc_version)"
        ")"
    )
    conn.execute(
        "CREATE TABLE IF NOT EXISTS document_aliases ("
        "source TEXT, "
        "doc_version INTEGER, "
        "alias TEXT, "
        "alias_type TEXT, "
        "weight REAL DEFAULT 1.0, "
        "PRIMARY KEY(source, doc_version, alias)"
        ")"
    )
    conn.execute(
        "CREATE TABLE IF NOT EXISTS document_sections ("
        "source TEXT, "
        "doc_version INTEGER, "
        "section_key TEXT, "
        "section_title TEXT, "
        "section_level INTEGER, "
        "chunk_start INTEGER, "
        "chunk_end INTEGER, "
        "section_path TEXT, "
        "PRIMARY KEY(source, doc_version, section_key)"
        ")"
    )
    conn.execute(
        "CREATE TABLE IF NOT EXISTS document_topics ("
        "source TEXT, "
        "doc_version INTEGER, "
        "topic TEXT, "
        "topic_type TEXT, "
        "weight REAL DEFAULT 1.0, "
        "PRIMARY KEY(source, doc_version, topic, topic_type)"
        ")"
    )
    conn.execute(
        "CREATE TABLE IF NOT EXISTS pending_delete_queue ("
        "source TEXT PRIMARY KEY, "
        "enqueued_at TEXT, "
        "next_retry_at TEXT, "
        "retry_count INTEGER DEFAULT 0, "
        "last_error TEXT, "
        "delete_files INTEGER DEFAULT 1"
        ")"
    )
    # migrate columns if missing
    cols = set([r[1] for r in conn.execute("PRAGMA table_info(documents)").fetchall()])
    def add_col(name: str, ddl: str):
        if name not in cols:
            try:
                conn.execute(f"ALTER TABLE documents ADD COLUMN {ddl}")
            except Exception:
                pass
    add_col("canonical_title", "canonical_title TEXT")
    add_col("title_tokens", "title_tokens TEXT")
    add_col("aliases", "aliases TEXT")
    add_col("filename_stem", "filename_stem TEXT")
    add_col("doc_type", "doc_type TEXT")
    add_col("topic", "topic TEXT")
    add_col("source_id", "source_id TEXT")
    add_col("original_filename", "original_filename TEXT")
    add_col("content_sha256", "content_sha256 TEXT")
    add_col("mime_type", "mime_type TEXT")
    add_col("detected_ext", "detected_ext TEXT")
    add_col("file_size", "file_size INTEGER")
    add_col("page_count", "page_count INTEGER")
    add_col("parser_route", "parser_route TEXT")
    add_col("parser_backend", "parser_backend TEXT")
    add_col("parse_status", "parse_status TEXT")
    add_col("parse_quality_score", "parse_quality_score REAL")
    add_col("quality_flags", "quality_flags TEXT")
    add_col("searchable", "searchable INTEGER DEFAULT 0")
    add_col("publish_gate", "publish_gate TEXT")
    add_col("duplicate_state", "duplicate_state TEXT")
    add_col("duplicate_of", "duplicate_of TEXT")
    add_col("same_title_group", "same_title_group TEXT")
    add_col("suspicious_file_type", "suspicious_file_type INTEGER DEFAULT 0")
    conn.execute("CREATE VIRTUAL TABLE IF NOT EXISTS documents_fts USING fts5(filename, title, aliases, doc_type, topic, filename_stem)")
    conn.commit()
def _lex_tx_begin():
    conn = _lex_db_connect()
    conn.execute("BEGIN IMMEDIATE")
def _lex_tx_commit():
    conn = _lex_db_connect()
    conn.execute("COMMIT")
def _lex_tx_rollback():
    conn = _lex_db_connect()
    try:
        conn.execute("ROLLBACK")
    except Exception:
        pass
def _lex_tx_savepoint(name: str):
    conn = _lex_db_connect()
    conn.execute(f"SAVEPOINT {name}")
def _lex_tx_release(name: str):
    conn = _lex_db_connect()
    conn.execute(f"RELEASE SAVEPOINT {name}")


def _lex_commit_if_needed(conn: sqlite3.Connection, had_outer_tx: bool):
    if not had_outer_tx:
        conn.commit()


def _lex_db_checkpoint(mode: str = "PASSIVE"):
    conn = _lex_db_connect()
    try:
        conn.execute(f"PRAGMA wal_checkpoint({mode})")
    except Exception:
        pass


def _fault_injection_stages() -> set[str]:
    values = []
    for env_name in ("RAG_FAULT_INJECT_STAGE", "LEX_DB_CRASH_INJECT_STAGE"):
        raw = os.getenv(env_name, "")
        if raw:
            values.extend(raw.split(","))
    return {v.strip().lower() for v in values if v and v.strip()}


def _crash_inject(stage: str):
    if stage.strip().lower() in _fault_injection_stages():
        raise RuntimeError(f"Crash injection at stage: {stage}")


def _query_match_terms(query: str) -> List[str]:
    q = (query or "").strip()
    parsed = _llm_query_parse_cache_get(q) or {}
    values: List[str] = []
    for key in ("anchors", "aspects", "section_targets"):
        items = parsed.get(key)
        if isinstance(items, list):
            for item in items:
                token = _normalize_query(str(item or ""))
                if len(token) < 2:
                    continue
                if token not in values:
                    values.append(token)
    if not values:
        for token in _tok_terms(_normalize_query(query)):
            if len(token) < 2:
                continue
            if token not in values:
                values.append(token)
    return values[:8]


def _token_overlap_score(query: str, text: str) -> float:
    q_terms = _query_match_terms(query)
    if not q_terms:
        return 0.0
    hay = (text or "")
    score = 0.0
    for term in q_terms:
        if term in hay:
            score += max(1.0, min(len(term), 8) / 2.0)
    return score


def _dedupe_keep_order(values: List[str]) -> List[str]:
    out: List[str] = []
    for value in values:
        item = (value or "").strip()
        if item and item not in out:
            out.append(item)
    return out


def _strip_leading_region_prefix(title: str) -> str:
    text = (title or "").strip()
    if not text:
        return ""
    suffixes = ["特别行政区", "自治州", "自治区", "自治县", "地区", "盟", "省", "市", "区", "县", "旗"]
    current = text
    for _ in range(2):
        changed = False
        for suffix in suffixes:
            idx = current.find(suffix)
            if idx < 2 or idx > 12:
                continue
            candidate = current[idx + len(suffix):].strip(" _-")
            if len(candidate) < 4:
                continue
            current = candidate
            changed = True
            break
        if not changed:
            break
    return current if current != text else ""


def _title_short_subject_variants(title: str) -> List[str]:
    text = (title or "").strip().replace("_", " ")
    if not text:
        return []
    suffix_pattern = r"(条例|办法|规定|规则|细则|决定|通知)"
    variants: List[str] = []
    for middle in ("安全管理",):
        match = re.match(rf"(.{{2,}}?){middle}{suffix_pattern}$", text)
        if not match:
            continue
        subject, suffix = match.groups()
        candidate = f"{subject}{suffix}".strip()
        if len(candidate) >= 4 and candidate != text and candidate not in variants:
            variants.append(candidate)
    return variants


def _strip_region_admin_tokens(text: str) -> str:
    current = (text or "").strip()
    if not current:
        return ""
    removable_tokens = [
        "黎族自治县", "回族自治县", "蒙古自治县", "苗族自治县", "土家族自治县", "藏族自治州", "自治州", "自治区", "自治县",
        "特别行政区", "地区", "盟", "省", "市", "区", "县", "旗",
        "黎族", "回族", "壮族", "苗族", "蒙古族", "土家族", "藏族", "彝族", "哈尼族", "傣族",
    ]
    changed = True
    while changed and current:
        changed = False
        for token in removable_tokens:
            if current.endswith(token):
                current = current[: -len(token)].strip()
                changed = True
                break
    return current


def _compact_region_title_variant(title: str) -> str:
    text = (title or "").strip().replace("_", " ")
    if not text:
        return ""
    suffix_pattern = r"(?:条例|办法|规定|规则|细则|决定|通知)"
    match = re.search(suffix_pattern, text)
    title_body = text[: match.start()] if match else text
    remainder = _strip_leading_region_prefix(title_body)
    if not remainder:
        return ""
    region_prefix = title_body[: title_body.find(remainder)].strip(" _-")
    region_compact = _strip_region_admin_tokens(region_prefix)
    if not region_compact or region_compact == region_prefix:
        return ""
    suffix = text[len(title_body):]
    return f"{region_compact}{remainder}{suffix}".strip()


def _extract_region_token(title: str) -> str:
    text = (title or "").strip().replace("_", " ")
    if not text:
        return ""
    remainder = _strip_leading_region_prefix(text)
    if not remainder:
        return ""
    idx = text.find(remainder)
    if idx < 1:
        return ""
    return _strip_region_admin_tokens(text[:idx].strip(" _-"))


def _region_tokens_compatible(left: str, right: str) -> bool:
    left_token = _extract_region_token(left)
    right_token = _extract_region_token(right)
    if not left_token or not right_token:
        return True
    return left_token == right_token


def _explicit_title_region_compatible(target_title: str, candidate_title: str) -> bool:
    target_token = _extract_region_token(target_title)
    if not target_token:
        return True
    candidate_token = _extract_region_token(candidate_title)
    if not candidate_token:
        return False
    return target_token == candidate_token


def _expand_title_aliases(*titles: str) -> List[str]:
    replacements = [
        ("非物质文化遗产保护", "非遗"),
        ("非物质文化遗产", "非遗"),
        ("城市管理行政执法", "城管执法"),
        ("养犬管理", "养犬"),
        ("安全管理", "安全"),
        ("城市绿化", "绿化"),
        ("人民代表大会常务委员会", "人大常委会"),
        ("人民代表大会", "人大"),
        ("常务委员会", "常委会"),
    ]
    variants: List[str] = []
    queue: List[str] = []
    for title in titles:
        base = (title or "").strip()
        if not base:
            continue
        queue.extend([
            base,
            base.replace("_", " ").strip(),
            base.replace(" ", ""),
        ])
        compact_region = _compact_region_title_variant(base)
        if compact_region:
            queue.extend([compact_region, compact_region.replace(" ", "")])
        stripped = _strip_leading_region_prefix(base.replace("_", " ").strip())
        if stripped:
            queue.extend([stripped, stripped.replace(" ", "")])
    seen: set[str] = set()
    while queue:
        current = (queue.pop(0) or "").strip()
        if not current or current in seen:
            continue
        seen.add(current)
        variants.append(current)
        normalized = current.replace(" ", "")
        if normalized and normalized not in seen:
            queue.append(normalized)
        region_free = _strip_leading_region_prefix(current)
        if region_free and region_free not in seen:
            queue.append(region_free)
        for short_variant in _title_short_subject_variants(current):
            if short_variant not in seen:
                queue.append(short_variant)
        for old, new in replacements:
            if old in current:
                replaced = current.replace(old, new).strip()
                if replaced and replaced not in seen:
                    queue.append(replaced)
    return _dedupe_keep_order(variants)


def _doc_title_alias_candidates(source: str) -> List[str]:
    info = _doc_get(source)
    section_titles = set(_doc_profile_section_titles(source))
    aliases = [
        part.strip()
        for part in (info.get("aliases") or "").split(",")
        if part and part.strip() and _normalize_query(part) not in section_titles
    ]
    return _expand_title_aliases(
        info.get("canonical_title") or "",
        info.get("filename_stem") or _filename_stem(source),
        _filename_stem(source),
        source,
        *_doc_profile_alias_candidates(source, alias_types={"title_alias", "auto"}),
        *aliases,
    )


def _doc_profile_version_key(source: str) -> Optional[int]:
    current = _doc_get(source)
    version = current.get("active_version")
    if version is None:
        version = current.get("pending_version")
    try:
        return int(version) if version is not None else None
    except Exception:
        return None


def _doc_profile_alias_candidates(source: str, alias_types: Optional[set[str]] = None) -> List[str]:
    version = _doc_profile_version_key(source)
    if version is None:
        return []
    conn = _lex_db_connect()
    rows = conn.execute(
        "SELECT alias, alias_type FROM document_aliases WHERE source = ? AND doc_version = ? ORDER BY weight DESC, alias ASC",
        (_safe_filename(source), int(version)),
    ).fetchall()
    out: List[str] = []
    for alias, alias_type in rows:
        kind = str(alias_type or "").strip()
        if alias_types and kind not in alias_types:
            continue
        value = _normalize_query(str(alias or ""))
        if value and value not in out:
            out.append(value)
    return out


def _doc_profile_section_titles(source: str) -> List[str]:
    version = _doc_profile_version_key(source)
    if version is None:
        return []
    conn = _lex_db_connect()
    rows = conn.execute(
        "SELECT section_title FROM document_sections WHERE source = ? AND doc_version = ? ORDER BY section_level ASC, section_title ASC",
        (_safe_filename(source), int(version)),
    ).fetchall()
    out: List[str] = []
    for (section_title,) in rows:
        value = _normalize_query(str(section_title or ""))
        if value and value not in out:
            out.append(value)
    return out


def _doc_profile_topic_terms(source: str) -> List[str]:
    version = _doc_profile_version_key(source)
    if version is None:
        return []
    conn = _lex_db_connect()
    rows = conn.execute(
        "SELECT topic FROM document_topics WHERE source = ? AND doc_version = ? ORDER BY weight DESC, topic ASC",
        (_safe_filename(source), int(version)),
    ).fetchall()
    out: List[str] = []
    for (topic,) in rows:
        value = _normalize_query(str(topic or ""))
        if value and value not in out:
            out.append(value)
    return out


def _doc_title_alias_score(source: str, query: str) -> float:
    q = _normalize_query(query)
    if not q:
        return 0.0
    best = 0.0
    for candidate in _doc_title_alias_candidates(source):
        cand = (candidate or "").strip()
        if not cand:
            continue
        if cand in q:
            best = max(best, 6.0)
        if len(q) >= 4 and q in cand:
            best = max(best, 4.0)
        overlap = _token_overlap_score(query, cand)
        if overlap > 0:
            best = max(best, min(5.0, overlap + 1.0))
    return best


def _query_doc_identity_terms(query: str) -> List[str]:
    q = _normalize_query(query)
    if not q:
        return []
    ordered_terms = ["管理条例", "管理办法", "实施办法", "议事规则", "条例", "办法", "规定", "规则", "细则", "决定", "通知", "法规", "文件"]
    out: List[str] = []
    for term in ordered_terms:
        if term in q and term not in out:
            out.append(term)
    return out


def _query_has_doc_identity_term(query: str) -> bool:
    return bool(_query_doc_identity_terms(query))


def _source_supports_doc_identity_term(source: str, query: str) -> bool:
    terms = _query_doc_identity_terms(query)
    if not terms:
        return False
    title_candidates = [_doc_get(source).get("canonical_title") or ""] + _doc_title_alias_candidates(source)
    haystack = "\n".join(_normalize_query(item) for item in title_candidates if _normalize_query(item))
    if not haystack:
        return False
    return any(_normalize_query(term) in haystack for term in terms)


def _profile_source_recall(query: str, limit: int, source_filter: Optional[str] = None) -> Dict[str, Dict[str, Any]]:
    q = _normalize_query(query)
    if not q:
        return {}
    section_targets = _local_validate_section_targets(_extract_section_query_targets(q), limit=max(2, int(limit) * 2))
    match_terms = _query_match_terms(q)
    conn = _lex_db_connect()
    rows = conn.execute("SELECT source FROM documents").fetchall()
    ranked: Dict[str, Dict[str, Any]] = {}
    for (raw_source,) in rows:
        src = _normalize_filename_for_match(raw_source or "")
        if not src:
            continue
        if source_filter and src != source_filter:
            continue
        if not _source_state(src).get("visible"):
            continue
        score = 0.0
        reasons: List[str] = []
        matched_terms: List[str] = []

        title_aliases = _doc_profile_alias_candidates(src, alias_types={"title_alias", "auto"})
        section_titles = _doc_profile_section_titles(src)
        topic_terms = _doc_profile_topic_terms(src)

        for target in section_targets:
            if any(target in title for title in section_titles):
                score += 3.4
                if "profile_section" not in reasons:
                    reasons.append("profile_section")
                if target not in matched_terms:
                    matched_terms.append(target)

        for term in match_terms:
            alias_hit = any(term in alias for alias in title_aliases)
            section_hit = any(term in title for title in section_titles)
            topic_hit = any(term in topic for topic in topic_terms)
            if alias_hit:
                score += 1.8
                if "profile_alias" not in reasons:
                    reasons.append("profile_alias")
            if section_hit:
                score += 1.4
                if "profile_section" not in reasons:
                    reasons.append("profile_section")
            if topic_hit:
                score += 1.1
                if "profile_topic" not in reasons:
                    reasons.append("profile_topic")
            if alias_hit or section_hit or topic_hit:
                if term not in matched_terms:
                    matched_terms.append(term)

        if score <= 0:
            continue
        coverage = float(len(matched_terms)) / float(max(len(set(match_terms + section_targets)), 1))
        ranked[src] = {
            "score": float(score + coverage),
            "matched_terms": matched_terms,
            "hit_count": len(matched_terms),
            "coverage": coverage,
            "reasons": reasons,
        }
    ordered = sorted(ranked.items(), key=lambda item: (-float(item[1].get("score", 0.0)), item[0]))[: int(limit)]
    return {src: info for src, info in ordered}


def _normalize_title_probe_text(text: str) -> str:
    core = _strip_section_question_tail(text) or _normalize_query(text)
    return _normalize_reference_text(core)


def _rank_title_source_matches(query: str, limit: int = 5, include_topic_like: bool = True) -> List[Dict[str, Any]]:
    q = _normalize_query(query)
    if not q:
        return []
    query_norm = _normalize_reference_text(q)
    query_core_norm = _normalize_title_probe_text(q)
    conn = _lex_db_connect()
    rows = conn.execute("SELECT source FROM documents").fetchall()
    ranked: List[Tuple[float, str, str, str]] = []
    for row in rows:
        source = _normalize_filename_for_match((row[0] if row else "") or "")
        if not source:
            continue
        if not _source_state(source).get("visible"):
            continue
        info = _doc_get(source)
        canonical_title = (info.get("canonical_title") or "").strip()
        alias_candidates = _doc_title_alias_candidates(source)
        canonical_norm = _normalize_reference_text(canonical_title)
        alias_norms = [
            _normalize_reference_text(candidate)
            for candidate in alias_candidates
            if _normalize_reference_text(candidate)
        ]
        match_kind = ""
        matched_text = ""
        score = 0.0
        if canonical_norm and canonical_norm in query_norm:
            match_kind = "exact_title"
            score = 10.0
            matched_text = canonical_title or source
        elif canonical_norm and query_core_norm and len(query_core_norm) >= 4 and query_core_norm in canonical_norm:
            match_kind = "exact_title"
            score = 9.0
            matched_text = canonical_title or source
        else:
            alias_hit = False
            for alias_norm, alias_raw in zip(alias_norms, alias_candidates):
                if not alias_norm:
                    continue
                if alias_norm in query_norm:
                    alias_hit = True
                    score = max(score, 8.0)
                    if len(alias_norm) >= len(_normalize_reference_text(matched_text)):
                        matched_text = alias_raw or alias_norm
                elif query_core_norm and len(query_core_norm) >= 4 and query_core_norm in alias_norm:
                    alias_hit = True
                    score = max(score, 7.0)
                    if len(alias_norm) >= len(_normalize_reference_text(matched_text)):
                        matched_text = alias_raw or alias_norm
            if alias_hit:
                match_kind = "alias_title"
        if not match_kind and include_topic_like:
            topic_like_score = _doc_title_alias_score(source, q)
            if topic_like_score >= 2.0:
                match_kind = "topic_like_title"
                score = topic_like_score
                matched_text = query
        if match_kind:
            ranked.append((score, source, match_kind, matched_text))
    ranked.sort(key=lambda item: (-item[0], item[1]))
    out: List[Dict[str, Any]] = []
    for score, source, match_kind, matched_text in ranked:
        if any(entry["source"] == source for entry in out):
            continue
        out.append({"source": source, "score": score, "match_kind": match_kind, "matched_text": matched_text})
        if len(out) >= max(1, int(limit)):
            break
    return out


def _extract_strong_title_source_matches(query: str, limit: int = 5) -> List[Dict[str, Any]]:
    return [
        entry
        for entry in _rank_title_source_matches(query, limit=max(limit * 2, 6), include_topic_like=False)
        if (entry.get("match_kind") or "") in {"exact_title", "alias_title"}
    ][: max(1, int(limit))]


def _classify_title_reference_route(query: str, fnames: Optional[List[str]] = None) -> str:
    q = _normalize_query(query)
    names = fnames if fnames is not None else _extract_filename_candidates(q)
    if not q or names:
        return ""
    strong_matches = _extract_strong_title_source_matches(q, limit=3)
    if strong_matches:
        if any((entry.get("match_kind") or "") == "exact_title" for entry in strong_matches):
            return "exact_title_reference"
        return "alias_title_reference"
    ranked_matches = _rank_title_source_matches(q, limit=3, include_topic_like=True)
    if any((entry.get("match_kind") or "") == "topic_like_title" for entry in ranked_matches):
        return "topic_like_title"
    if _is_weak_reference_query(q):
        return "weak_title_reference"
    return ""


def _extract_title_source_candidates(query: str, limit: int = 5) -> List[str]:
    return [
        entry.get("source")
        for entry in _rank_title_source_matches(query, limit=limit, include_topic_like=True)
        if entry.get("source")
    ]


def _build_document_clarification_prompt(candidate_sources: List[str]) -> str:
    suggestions: List[str] = []
    limit = max(1, int(_policy_get("source_resolution.clarification_examples_limit", 3)))
    for source in candidate_sources or []:
        info = _doc_get(source)
        title = (info.get("canonical_title") or "").strip() or _filename_stem(source)
        if title and title not in suggestions:
            suggestions.append(title)
        if len(suggestions) >= limit:
            break
    if suggestions:
        joined = "、".join(suggestions)
        return f"请先说明要查询哪一部法规文档，再继续检索；可直接回复文档名，例如：{joined}。"
    return "请先说明要查询哪一部法规文档，再继续检索；可直接回复具体法规名称或文件名。"


def _clarification_probe_terms(query: str) -> List[str]:
    q = _normalize_query(query)
    if not q:
        return []
    terms: List[str] = []

    def _add(term: str):
        value = _normalize_query(term)
        if len(value) >= 2 and value not in terms:
            terms.append(value)

    for candidate in _open_topic_anchor_terms(q) + _query_anchor_terms(q) + [_strip_query_intent_phrases(q) or q, q]:
        text = _normalize_query(candidate)
        if not text:
            continue
        text = re.sub(r"^(关于|对于|有关|请问)", "", text)
        text = re.sub(r"(是什么|有哪些|有什么|如何|怎么|怎么办|吗|呢|呀)$", "", text)
        _add(text)
        trimmed = re.sub(r"(的)?(规定|要求|内容|情形|责任|职责|条件|范围|程序|标准)$", "", text)
        trimmed = re.sub(r"(使用|管理|认定|处理)$", "", trimmed)
        _add(trimmed)
        if trimmed.startswith("禁止") and len(trimmed) >= 4:
            _add(trimmed[2:])
        if trimmed.startswith("道") and len(trimmed) >= 4:
            _add(trimmed[1:])
    return sorted(terms, key=len, reverse=True)[:8]


def _clarification_chunk_candidate_sources(query: str, limit: int = 5) -> List[str]:
    terms = _clarification_probe_terms(query)
    if not terms:
        return []
    conn = _lex_db_connect()
    scored: Dict[str, float] = {}
    for term in terms:
        rows = conn.execute(
            "SELECT m.source, COUNT(*) AS hit_count FROM chunks_fts_content c JOIN chunks_meta m ON m.id = c.id WHERE c.c0 LIKE ? GROUP BY m.source ORDER BY hit_count DESC LIMIT ?",
            (f"%{term}%", max(10, int(limit) * 6)),
        ).fetchall()
        for source, hit_count in rows:
            safe_source = _normalize_filename_for_match(source or "")
            if not safe_source or not _source_state(safe_source).get("visible"):
                continue
            scored[safe_source] = scored.get(safe_source, 0.0) + float(hit_count or 0) * max(1.0, min(float(len(term)), 6.0))
    ranked = sorted(scored.items(), key=lambda item: (-item[1], item[0]))
    return [source for source, _ in ranked[: max(1, int(limit))]]


def _retrieval_backed_clarification_candidates(query: str, seed_sources: Optional[List[str]] = None, limit: int = 3) -> List[str]:
    out: List[str] = []
    for source in seed_sources or []:
        safe_source = _normalize_filename_for_match(source or "")
        if safe_source and _source_state(safe_source).get("visible") and safe_source not in out:
            out.append(safe_source)
    for source in _clarification_chunk_candidate_sources(query, limit=max(limit * 2, 6)):
        if source not in out:
            out.append(source)
        if len(out) >= max(1, int(limit)):
            return out[: max(1, int(limit))]
    fallback = _doc_fallback_source_candidates(query, limit=max(limit * 2, 6))
    for source in list(fallback.get("sources") or []) + _doc_recall_fallback(query, limit=max(limit * 2, 6)):
        safe_source = _normalize_filename_for_match(source or "")
        if safe_source and safe_source not in out:
            out.append(safe_source)
        if len(out) >= max(1, int(limit)):
            break
    return out[: max(1, int(limit))]


def _build_retrieval_grounded_clarification_prompt(query: str, candidate_sources: List[str], reason: str = "document_target_required") -> str:
    titles: List[str] = []
    for source in candidate_sources or []:
        title = _source_display_title(source)
        if title and title not in titles:
            titles.append(title)
    joined = "、".join([f"《{title}》" for title in titles[:3]])
    return f"""你是法规知识库的澄清助手。

用户问题：{query}
系统检索到的可能相关法规：{joined or '（暂无）'}
当前原因：{reason}

请生成一段简短澄清话术，规则：
1) 只做澄清，不要假装已经回答了问题
2) 如果有候选法规，明确说“我找到几部可能相关的法规”，并优先点出最相关的 1~3 部
3) 结尾必须请用户确认是否指其中之一，或补充其他法规名称
4) 不要编造法规，不要输出条列，不超过两句话

直接输出澄清话术："""


def _build_document_not_found_prompt(target: str) -> str:
    title = (target or "").strip() or "目标法规"
    if not title.startswith("《") and not title.endswith("》") and not _extract_filename_candidates(title):
        title = f"《{title}》"
    return f"当前知识库中没有找到{title}，因此不能基于该法规回答。请确认文档是否已上传，或更换已收录的法规名称。"


def _build_compare_target_not_found_prompt(missing_targets: List[str], matched_sources: Optional[List[str]] = None) -> str:
    missing_items = [item.strip() for item in (missing_targets or []) if (item or "").strip()]
    found_titles: List[str] = []
    for source in matched_sources or []:
        info = _doc_get(source)
        title = (info.get("canonical_title") or "").strip() or _filename_stem(source)
        if title and title not in found_titles:
            found_titles.append(title)
    missing_text = "、".join(missing_items) if missing_items else "部分目标文档"
    if found_titles:
        return f"对比目标中有文档未命中：{missing_text}。当前只定位到：{'、'.join(found_titles)}，不能在缺少目标文档证据的情况下生成法规对比答案。"
    return f"对比目标文档未命中：{missing_text}。当前知识库无法定位足够的对比对象，不能生成法规对比答案。"


@dataclass
class CompareSubjectSpan:
    raw_text: str
    clean_text: str
    span_start: int = -1
    span_end: int = -1
    connector_before: str = ""
    doc_like: bool = False
    source: str = ""
    match_kind: str = ""
    prior: float = 0.0


@dataclass
class ComparePlan:
    raw_query: str
    has_intent: bool = False
    route: str = ""
    reason: str = "not_compare"
    required: bool = False
    resolved: bool = False
    subject_zone: str = ""
    tail_span: str = ""
    subjects: List[CompareSubjectSpan] = field(default_factory=list)
    matched_sources: List[str] = field(default_factory=list)
    whole_query_sources: List[str] = field(default_factory=list)
    missing_targets: List[str] = field(default_factory=list)
    doc_like_subjects: List[str] = field(default_factory=list)
    common_aspects: List[str] = field(default_factory=list)
    topic_pair: List[str] = field(default_factory=list)
    canonical_aspects: List[str] = field(default_factory=list)
    expanded_aspects: List[str] = field(default_factory=list)
    source_subqueries: Dict[str, Dict[str, str]] = field(default_factory=dict)
    compare_status: str = "not_compare"


_COMPARE_LEAD_PATTERN = re.compile(
    r"^(请|请帮我|帮我|麻烦|我想了解|想了解|我想|想)?\s*(对比一下|比较一下|对比|比较|区别一下|区别|分析一下|分析|分别比较一下|分别比较|分别说明|分别说说|分别介绍)?\s*"
)
_COMPARE_CONNECTOR_PATTERN = re.compile(r"(?:和|与|跟|及|以及|vs|VS|Vs|versus)")
_COMPARE_TAIL_MARKERS = ["里的", "中的", "中", "里", "分别", "各自", "有什么", "有何", "有哪些", "怎么", "如何", "是否", "的区别", "的差异", "的不同"]
_COMPARE_ASPECT_CANONICAL_MAP = {
    "处罚规定": "处罚",
    "处罚条款": "处罚",
    "处罚要求": "处罚",
    "处罚标准": "处罚",
    "违法处理方式": "违法处理",
    "违法行为处理": "违法处理",
    "程序要求": "程序",
    "程序规定": "程序",
    "监督机制": "监督检查",
    "监督管理": "监督检查",
    "管理责任": "管理职责",
    "安全职责": "安全责任",
}
_COMPARE_ASPECT_EXPANSIONS = {
    "处罚": ["处罚", "罚款", "罚则", "法律责任"],
    "违法处理": ["违法处理", "违法行为", "处罚", "法律责任"],
    "安全责任": ["安全责任", "责任", "职责", "安全要求"],
    "管理职责": ["管理职责", "职责", "责任", "政府职责"],
    "政府职责": ["政府职责", "职责", "责任", "管理职责"],
    "监督检查": ["监督检查", "监督", "监管", "检查"],
    "程序": ["程序", "流程", "步骤", "审议"],
    "登记要求": ["登记要求", "登记", "备案", "申请材料"],
    "保护措施": ["保护措施", "措施", "保护", "扶持措施"],
    "扶持措施": ["扶持措施", "支持措施", "保护措施"],
    "执法主体": ["执法主体", "执法人员", "主管部门"],
    "禁止行为": ["禁止行为", "禁止", "不得", "限制"],
}


def _compare_unique_texts(items: List[str], limit: Optional[int] = None) -> List[str]:
    out: List[str] = []
    for item in items or []:
        value = _normalize_query(item)
        if not value or value in out:
            continue
        out.append(value)
        if limit and len(out) >= limit:
            break
    return out


_REGION_HINT_RE = re.compile(r"([\u4e00-\u9fff]{2,12}(?:省|市|自治区|自治州|地区|盟|县|区))")


def _extract_region_hint(text: str) -> str:
    t = _normalize_query(text)
    if not t:
        return ""
    m = _REGION_HINT_RE.search(t)
    return m.group(1) if m else ""


def _strip_raw_text_mentions(query: str, terms: List[str]) -> str:
    q = _normalize_query(query)
    if not q:
        return ""
    for term in terms or []:
        t = _normalize_query(term)
        if t:
            q = q.replace(t, " ")
    q = re.sub(r"\s+", " ", q).strip()
    q = re.sub(r"[，。；;：:?？!！,.\s]+$", "", q).strip()
    return q


def _strip_compare_noise_terms(text: str) -> str:
    q = _normalize_query(text)
    if not q:
        return ""
    q = re.sub(r"\b(vs|versus|compare)\b", " ", q, flags=re.IGNORECASE)
    q = re.sub(r"(对比一下|比较一下|区别一下|对比|比较|区别|差异|不同|分别|各自|哪个|哪一个|哪种|哪类|更|更为|是否更|有没有更)", " ", q)
    q = re.sub(r"(有什么不同|有什么区别|有什么差异|有何不同|有何区别|有何差异)", " ", q)
    q = re.sub(r"\s+", " ", q).strip()
    q = re.sub(r"[，。；;：:?？!！,.\s]+$", "", q).strip()
    return q


def _compare_plan_to_dict(plan: ComparePlan) -> Dict[str, Any]:
    return {
        "is_compare": bool(plan.has_intent),
        "route": plan.route,
        "reason": plan.reason,
        "required": bool(plan.required),
        "resolved": bool(plan.resolved),
        "subjects": [subject.clean_text or subject.raw_text for subject in plan.subjects],
        "subject_spans": [
            {
                "raw_text": subject.raw_text,
                "clean_text": subject.clean_text,
                "span_start": subject.span_start,
                "span_end": subject.span_end,
                "doc_like": subject.doc_like,
                "source": subject.source,
                "match_kind": subject.match_kind,
                "prior": subject.prior,
            }
            for subject in plan.subjects
        ],
        "subject_matches": [
            {
                "subject": subject.clean_text or subject.raw_text,
                "source": subject.source,
                "match_kind": subject.match_kind,
                "doc_like": subject.doc_like,
                "prior": subject.prior,
            }
            for subject in plan.subjects
        ],
        "sources": list(plan.matched_sources),
        "doc_like_subjects": list(plan.doc_like_subjects),
        "missing_doc_targets": list(plan.missing_targets),
        "common_aspects": list(plan.common_aspects),
        "topic_pair": list(plan.topic_pair),
        "canonical_aspects": list(plan.canonical_aspects),
        "expanded_aspects": list(plan.expanded_aspects),
        "source_subqueries": dict(plan.source_subqueries),
        "target_text": "、".join(plan.missing_targets[:3] or [subject.clean_text or subject.raw_text for subject in plan.subjects[:2]]),
        "clarification": "",
        "strip_title_mentions": bool(plan.matched_sources),
        "compare_status": plan.compare_status,
        "compare_plan": {
            "raw_query": plan.raw_query,
            "subject_zone": plan.subject_zone,
            "tail_span": plan.tail_span,
            "route": plan.route,
            "reason": plan.reason,
            "required": plan.required,
            "resolved": plan.resolved,
            "subjects": [
                {
                    "raw_text": subject.raw_text,
                    "clean_text": subject.clean_text,
                    "span_start": subject.span_start,
                    "span_end": subject.span_end,
                    "doc_like": subject.doc_like,
                    "source": subject.source,
                    "match_kind": subject.match_kind,
                    "prior": subject.prior,
                }
                for subject in plan.subjects
            ],
            "matched_sources": list(plan.matched_sources),
            "whole_query_sources": list(plan.whole_query_sources),
            "missing_targets": list(plan.missing_targets),
            "doc_like_subjects": list(plan.doc_like_subjects),
            "common_aspects": list(plan.common_aspects),
            "topic_pair": list(plan.topic_pair),
            "canonical_aspects": list(plan.canonical_aspects),
            "expanded_aspects": list(plan.expanded_aspects),
            "source_subqueries": dict(plan.source_subqueries),
            "compare_status": plan.compare_status,
        },
    }


def _query_has_compare_intent(query: str) -> bool:
    q = _normalize_query(query)
    if not q:
        return False
    if _classify_question_type(q) == "compare":
        return True
    lower_q = q.lower()
    if any(marker in q for marker in ["对比", "比较", "区别", "差异", "相比", "分别", "各自"]):
        return True
    if any(marker in lower_q for marker in ["vs", "versus", "compare"]):
        return True
    return bool(re.search(r"(?:和|与|跟|及|以及|vs|VS|Vs|versus)", q)) and any(marker in q for marker in ["分别", "各自", "区别", "差异", "不同"])


def _clean_compare_subject_text(text: str) -> str:
    subject = _normalize_query(text)
    if not subject:
        return ""
    subject = re.sub(r"^(请|请帮我|帮我|麻烦|我想了解|想了解|我想|想)?\s*", "", subject)
    subject = re.sub(r"^(对比一下|比较一下|对比|比较|区别一下|区别|分析一下|分析|分别比较一下|分别比较|分别说明|分别说说|分别介绍)\s*", "", subject)
    subject = re.sub(r"(之间|各自|分别|的区别|的差异|的不同)+$", "", subject)
    subject = re.sub(r"[，。；;：:?？!！\s]+$", "", subject)
    subject = re.sub(r"(?:里的?|中的?|内的?)$", "", subject)
    subject = _strip_section_question_tail(subject) or _normalize_query(subject)
    return re.sub(r"[，。；;：:?？!！\s]+$", "", subject)


def _extract_compare_subject_spans(query: str) -> Dict[str, Any]:
    q = _normalize_query(query)
    if not q or not _query_has_compare_intent(q):
        return {"subjects": [], "subject_zone": "", "tail_span": ""}
    lead_match = _COMPARE_LEAD_PATTERN.match(q)
    lead_end = lead_match.end() if lead_match else 0
    working = q[lead_end:]
    working = re.sub(r"[。！？!?]+$", "", working)
    connector_candidates = list(_COMPARE_CONNECTOR_PATTERN.finditer(working))
    connector_match = connector_candidates[0] if connector_candidates else None
    if connector_candidates:
        suffixes = ["条例", "办法", "规定", "规则", "细则", "决定", "通知"]
        for candidate in connector_candidates:
            prefix = working[: candidate.start()].rstrip()
            if prefix.endswith("》") or any(prefix.endswith(suffix) for suffix in suffixes):
                connector_match = candidate
                break
    if not connector_match:
        return {"subjects": [], "subject_zone": working.strip(), "tail_span": ""}
    tail_start: Optional[int] = None
    for marker in _COMPARE_TAIL_MARKERS:
        idx = working.find(marker, connector_match.end())
        if idx != -1 and (tail_start is None or idx < tail_start):
            tail_start = idx
    for punct in ["，", ",", "；", ";", "：", ":"]:
        idx = working.find(punct, connector_match.end())
        if idx != -1 and (tail_start is None or idx < tail_start):
            tail_start = idx
    subject_zone = working[:tail_start] if tail_start is not None else working
    tail_span = working[tail_start:] if tail_start is not None else ""
    parts = re.split(
        r"(\s*(?:vs|VS|Vs|versus)\s*|(?:(?<=条例)|(?<=办法)|(?<=规定)|(?<=规则)|(?<=细则)|(?<=决定)|(?<=通知)|(?<=》))\s*(?:和|与|跟|及|以及)\s*)",
        subject_zone,
    )
    out: List[CompareSubjectSpan] = []
    cursor = lead_end
    connector_before = ""
    for idx, part in enumerate(parts):
        if idx % 2 == 1:
            cursor += len(part)
            connector_before = part.strip()
            continue
        raw = part or ""
        raw_value = raw.strip()
        if not raw_value:
            cursor += len(part)
            continue
        leading_ws = len(raw) - len(raw.lstrip())
        trailing_ws = len(raw.rstrip())
        subject = _clean_compare_subject_text(raw_value)
        out.append(
            CompareSubjectSpan(
                raw_text=raw_value,
                clean_text=subject,
                span_start=cursor + leading_ws,
                span_end=cursor + trailing_ws,
                connector_before=connector_before,
            )
        )
        cursor += len(part)
        connector_before = ""
        if len(out) >= 4:
            break
    return {
        "subjects": out,
        "subject_zone": subject_zone.strip(),
        "tail_span": tail_span.strip(),
    }


def _extract_compare_subjects(query: str) -> List[str]:
    spans = _extract_compare_subject_spans(query)
    return [subject.clean_text for subject in spans.get("subjects") or [] if subject.clean_text]


def _looks_like_compare_document_target(text: str) -> bool:
    q = _normalize_query(text)
    if not q:
        return False
    suffixes = ["条例", "办法", "规定", "规则", "细则", "决定", "通知"]
    matched_suffix = next((suffix for suffix in suffixes if q.endswith(suffix)), "")
    if not matched_suffix:
        return False
    core = q[: -len(matched_suffix)].strip()
    if not core:
        return False
    generic_terms = {"处罚", "要求", "责任", "程序", "措施", "条款", "职责", "标准", "条件", "管理", "规定", "规则", "办法", "条例", "通知", "决定"}
    core_terms = [term for term in _query_anchor_terms(core) if term not in generic_terms]
    if core_terms:
        return True
    return core not in generic_terms and len(core) >= 2


def _compare_clean_aspect_span(text: str) -> str:
    value = _normalize_query(text)
    if not value:
        return ""
    value = re.sub(r"^[，,、;；:：\s]+", "", value)
    value = re.sub(r"^(里的?|中的?|内的?|关于|就|在)\s*", "", value)
    value = re.sub(r"^(对比一下|比较一下|对比|比较|区别一下|区别|分析一下|分析|分别|各自)\s*", "", value)
    value = re.sub(r"(有什么|有何|有哪些|怎么|如何|是否|吗|呢|的区别|的差异|的不同|。|？|!)+$", "", value)
    return _normalize_query(value)


def _canonicalize_compare_aspect(term: str) -> str:
    value = _compare_clean_aspect_span(term)
    value = _normalize_coverage_aspect(value) or value
    value = _COMPARE_ASPECT_CANONICAL_MAP.get(value, value)
    return _normalize_query(value)


def _expand_compare_aspects(aspects: List[str], limit: int = 8) -> List[str]:
    out: List[str] = []
    for aspect in aspects or []:
        canonical = _canonicalize_compare_aspect(aspect)
        candidates = [canonical]
        candidates.extend(_coverage_aspect_variants(canonical))
        candidates.extend(_COMPARE_ASPECT_EXPANSIONS.get(canonical, []))
        if "责任" in canonical and canonical not in {"责任", "安全责任", "管理职责", "政府职责"}:
            candidates.extend([canonical.replace("责任", "职责"), canonical.replace("责任", "要求")])
        if "处罚" in canonical and canonical != "处罚":
            candidates.extend(["处罚", "法律责任", "罚则"])
        if "程序" in canonical:
            candidates.extend(["流程", "步骤", "审议"])
        for candidate in candidates:
            value = _normalize_query(candidate)
            if len(value) < 2 or value in out:
                continue
            out.append(value)
            if len(out) >= limit:
                return out
    return out


def _compare_aspects_from_span(text: str, limit: int = 4) -> List[str]:
    raw = _compare_clean_aspect_span(text)
    if not raw:
        return []
    out: List[str] = []

    def _add(term: str):
        value = _canonicalize_compare_aspect(term)
        if len(value) < 2 or value in out or _looks_like_compare_document_target(value):
            return
        out.append(value)

    for target in _extract_section_query_targets(raw):
        _add(target)
    semantic = _query_semantic_aspects(raw)
    for term in semantic.get("terms") or []:
        _add(term)
    for term in _query_match_terms(raw):
        _add(term)
    if not out:
        for piece in re.split(r"\s*(?:和|与|跟|及|以及|、|vs|VS|Vs|versus)\s*", raw):
            _add(piece)
    if not out:
        _add(raw)
    return out[:limit]


def _extract_compare_common_aspects(plan: ComparePlan) -> List[str]:
    aspects = _compare_aspects_from_span(plan.tail_span, limit=4)
    if aspects:
        return aspects
    return _compare_aspects_from_span(plan.raw_query, limit=4)


def _locate_source_mention_span(query: str, source: str) -> Tuple[int, int]:
    q = _normalize_query(query)
    for candidate in sorted(_doc_title_alias_candidates(source), key=len, reverse=True):
        value = _normalize_query(candidate)
        if len(value) < 2:
            continue
        idx = q.find(value)
        if idx != -1:
            return idx, idx + len(value)
    return -1, -1


def _extract_single_doc_compare_topic_pair(query: str, source: str) -> List[str]:
    q = _normalize_query(query)
    start, end = _locate_source_mention_span(q, source)
    tail = q[end:] if end != -1 else q
    tail = _compare_clean_aspect_span(tail)
    pair: List[str] = []

    def _add(term: str):
        value = _canonicalize_compare_aspect(term)
        if len(value) < 2 or value in pair or _looks_like_compare_document_target(value):
            return
        pair.append(value)

    for piece in re.split(r"\s*(?:和|与|跟|及|以及|、|vs|VS|Vs|versus)\s*", tail):
        _add(piece)
    if len(pair) < 2:
        for term in _compare_aspects_from_span(tail, limit=4):
            _add(term)
    return pair[:2]


def _build_compare_source_subqueries(plan: ComparePlan) -> Dict[str, Dict[str, str]]:
    out: Dict[str, Dict[str, str]] = {}
    aspect_terms = list(plan.canonical_aspects or [])
    expanded_terms = list(plan.expanded_aspects or [])
    section_terms = [term for term in expanded_terms if _looks_like_section_target(term) or any(token in term for token in ["责任", "处罚", "程序", "监督", "登记", "职责", "措施", "主体", "禁止", "限制"])]
    for subject in plan.subjects:
        source = _normalize_filename_for_match(subject.source or "")
        if not source:
            continue
        title = _source_display_title(source)
        subject_hint = subject.clean_text or subject.raw_text or title
        raw_text_query = " ".join(_compare_unique_texts([title, subject_hint] + aspect_terms + expanded_terms, limit=8))
        section_query = " ".join(_compare_unique_texts([title] + section_terms + aspect_terms, limit=6))
        doc_prior_query = " ".join(_compare_unique_texts([subject_hint, title] + aspect_terms, limit=6))
        out[source] = {
            "raw_text_query": raw_text_query or title,
            "section_query": section_query or raw_text_query or title,
            "doc_prior_query": doc_prior_query or raw_text_query or title,
        }
    if plan.route == "single_doc_compare" and len(plan.matched_sources) == 1:
        source = plan.matched_sources[0]
        title = _source_display_title(source)
        topic_terms = list(plan.topic_pair or aspect_terms or expanded_terms)
        out[source] = {
            "raw_text_query": " ".join(_compare_unique_texts([title] + topic_terms + expanded_terms, limit=8)) or title,
            "section_query": " ".join(_compare_unique_texts([title] + topic_terms, limit=6)) or title,
            "doc_prior_query": " ".join(_compare_unique_texts([title] + topic_terms[:2], limit=6)) or title,
        }
    return out


def _resolve_compare_subject_source(subject: str) -> Dict[str, Any]:
    target = _normalize_query(subject)
    if not target:
        return {"subject": "", "source": "", "match_kind": "", "doc_like": False, "prior": 0.0}
    doc_like = _looks_like_compare_document_target(target)
    if not doc_like:
        return {
            "subject": target,
            "source": "",
            "match_kind": "",
            "doc_like": False,
            "prior": 0.0,
        }
    region_hint = _extract_region_hint(target)
    strong_matches = _extract_strong_title_source_matches(target, limit=3)
    for top in strong_matches or []:
        src = _normalize_filename_for_match(top.get("source") or "")
        if not src:
            continue
        if region_hint:
            title = _normalize_query(_source_display_title(src))
            if title and region_hint not in title:
                continue
        return {
            "subject": target,
            "source": src,
            "match_kind": top.get("match_kind") or "",
            "doc_like": True,
            "prior": 1.0,
        }
    ranked_matches = _rank_title_source_matches(target, limit=3, include_topic_like=True)
    if doc_like and ranked_matches:
        for top in ranked_matches:
            src = _normalize_filename_for_match(top.get("source") or "")
            if not src:
                continue
            if region_hint:
                title = _normalize_query(_source_display_title(src))
                if title and region_hint not in title:
                    continue
            if (top.get("match_kind") or "") == "topic_like_title" and float(top.get("score") or 0.0) >= 2.5:
                return {
                    "subject": target,
                    "source": src,
                    "match_kind": top.get("match_kind") or "",
                    "doc_like": True,
                    "prior": min(float(top.get("score") or 0.0) / 4.0, 1.0),
                }
    if doc_like:
        plan = _build_doc_recall_plan(target, limit=3)
        for entry in plan:
            reasons = set((entry or {}).get("reasons") or [])
            prior = float((entry or {}).get("prior") or 0.0)
            title_score = float((entry or {}).get("title_score") or 0.0)
            src = _normalize_filename_for_match((entry or {}).get("source") or "")
            if not src:
                continue
            if region_hint:
                title = _normalize_query(_source_display_title(src))
                if title and region_hint not in title:
                    continue
            if title_score > 0 or (reasons.intersection({"documents_fts", "doc_term_overlap", "title_alias_substring"}) and prior >= 0.45):
                return {
                    "subject": target,
                    "source": src,
                    "match_kind": "doc_recall",
                    "doc_like": True,
                    "prior": prior,
                }
    return {
        "subject": target,
        "source": "",
        "match_kind": "",
        "doc_like": doc_like,
        "prior": 0.0,
    }


def _build_compare_plan(query: str) -> ComparePlan:
    q = _normalize_query(query)
    plan = ComparePlan(raw_query=q, has_intent=bool(q and _query_has_compare_intent(q)))
    if not plan.has_intent:
        return plan
    subject_info = _extract_compare_subject_spans(q)
    plan.subjects = list(subject_info.get("subjects") or [])
    plan.subject_zone = subject_info.get("subject_zone") or ""
    plan.tail_span = subject_info.get("tail_span") or ""
    for subject in plan.subjects:
        match = _resolve_compare_subject_source(subject.clean_text or subject.raw_text)
        subject.doc_like = bool(match.get("doc_like"))
        subject.source = _normalize_filename_for_match(match.get("source") or "")
        subject.match_kind = match.get("match_kind") or ""
        subject.prior = float(match.get("prior") or 0.0)
    plan.doc_like_subjects = _compare_unique_texts([subject.clean_text or subject.raw_text for subject in plan.subjects if subject.doc_like])
    plan.missing_targets = _compare_unique_texts([subject.clean_text or subject.raw_text for subject in plan.subjects if subject.doc_like and not subject.source])
    whole_query_title_matches = _extract_strong_title_source_matches(q, limit=2)
    plan.whole_query_sources = [
        _normalize_filename_for_match(item.get("source") or "")
        for item in whole_query_title_matches
        if _normalize_filename_for_match(item.get("source") or "")
    ]
    plan.matched_sources = _compare_unique_texts([subject.source for subject in plan.subjects if subject.source])
    plan.route = "open_topic_compare"
    plan.required = False
    plan.resolved = False
    plan.reason = "not_needed"
    if len(plan.doc_like_subjects) >= 2:
        if len(plan.matched_sources) >= 2:
            plan.route = "multi_doc_compare"
            plan.resolved = True
        elif len(plan.matched_sources) == 1:
            plan.route = "single_doc_compare"
            plan.resolved = True
            plan.reason = "compare_target_not_found_degraded"
        else:
            plan.route = "compare_targets_not_found"
            plan.required = True
            plan.reason = "compare_targets_not_found"
    elif len(plan.doc_like_subjects) == 1:
        if len(plan.matched_sources) == 1:
            plan.route = "single_doc_compare"
            plan.resolved = True
        else:
            plan.route = "compare_target_not_found"
            plan.required = True
            plan.reason = "compare_target_not_found"
    elif len(plan.whole_query_sources) == 1:
        plan.route = "single_doc_compare"
        plan.matched_sources = _compare_unique_texts(plan.whole_query_sources)
        plan.resolved = True
    if plan.route == "multi_doc_compare":
        plan.common_aspects = _extract_compare_common_aspects(plan)
    elif plan.route == "single_doc_compare" and len(plan.matched_sources) == 1:
        plan.topic_pair = _extract_single_doc_compare_topic_pair(q, plan.matched_sources[0])
    elif plan.route == "open_topic_compare":
        plan.topic_pair = _compare_unique_texts([subject.clean_text or subject.raw_text for subject in plan.subjects], limit=2)
    raw_aspects = list(plan.common_aspects or plan.topic_pair)
    plan.canonical_aspects = _compare_unique_texts([_canonicalize_compare_aspect(item) for item in raw_aspects], limit=4)
    plan.expanded_aspects = _expand_compare_aspects(plan.canonical_aspects, limit=8)
    plan.source_subqueries = _build_compare_source_subqueries(plan)
    if plan.route == "multi_doc_compare":
        plan.compare_status = "plan_ready"
    elif plan.route == "single_doc_compare":
        plan.compare_status = "single_doc_ready"
    elif plan.route in {"compare_target_not_found", "compare_targets_not_found"}:
        plan.compare_status = "target_missing"
    elif plan.route == "open_topic_compare":
        plan.compare_status = "open_topic"
    return plan


def _analyze_compare_route(query: str) -> Dict[str, Any]:
    return _compare_plan_to_dict(_build_compare_plan(query))


def _strip_section_question_tail(query: str) -> str:
    q = _normalize_query(query)
    if not q:
        return ""
    text = re.sub(r"^(关于|根据|请问|请|查询|说明|咨询|问下|想了解|我想了解)\s*", "", q)
    text = re.sub(r"[，。；;：:?？!！\s]+$", "", text)
    tail_patterns = [
        r"(有哪些规定|有哪些要求|有哪些内容|有什么要求|什么要求|是什么内容|是什么|如何处理|怎么处理|怎么办|如何|怎么|是否可行|是否可以|是否|吗|呢)$",
    ]
    changed = True
    while changed and text:
        changed = False
        for pat in tail_patterns:
            new_text = re.sub(pat, "", text)
            if new_text != text:
                text = re.sub(r"[，。；;：:?？!！\s]+$", "", new_text)
                changed = True
    return _normalize_query(text)


def _local_validate_section_targets(targets: List[str], limit: int = 5) -> List[str]:
    out: List[str] = []
    generic_targets = {
        "违法行为", "违法处理", "怎么处理", "处罚", "罚款", "法律责任", "责任", "管理职责", "政府职责",
        "监督检查", "执法措施", "安全责任", "登记", "程序", "审议", "表决", "公布",
        "绿化建设", "绿地保护", "砍伐树木", "消防安全",
    }
    for raw in targets or []:
        target = _normalize_query(raw)
        if not target:
            continue
        if len(target) < 2 or len(target) > 20:
            continue
        if target in generic_targets:
            continue
        # 仅允许章节/条款级 target，禁止法规名级输出。
        if any(marker in target for marker in ["条例", "办法", "法规", "管理条例", "实施办法"]):
            continue
        if not _looks_like_section_target(target):
            continue
        if target not in out:
            out.append(target)
        if len(out) >= max(1, int(limit)):
            break
    return out


def _sqlite_reverse_lookup_sources_by_targets(targets: List[str], limit: int = 3) -> List[str]:
    if not targets:
        return []
    conn = _lex_db_connect()
    source_score: Dict[str, float] = {}
    source_target_hits: Dict[str, set[str]] = {}
    row_limit = max(120, int(limit) * 120)
    for target in targets:
        section_rows = conn.execute(
            "SELECT source, section_title, section_path FROM document_sections WHERE section_title LIKE ? LIMIT ?",
            (f"%{target}%", row_limit),
        ).fetchall()
        for source, section_title, section_path in section_rows:
            src = _normalize_filename_for_match(source or "")
            if not src:
                continue
            if not _source_state(src).get("visible"):
                continue
            score = 3.4 if target in str(section_title or "") else 2.2
            try:
                path_items = json.loads(section_path or "[]")
            except Exception:
                path_items = []
            if any(target in str(item or "") for item in path_items):
                score += 1.2
            source_score[src] = source_score.get(src, 0.0) + score
            source_target_hits.setdefault(src, set()).add(target)
        pattern = f"%{target}%"
        rows = conn.execute(
            "SELECT m.source, m.section, m.metadata, f.text "
            "FROM chunks_meta m JOIN chunks_fts f ON f.rowid = m.id "
            "WHERE m.section LIKE ? OR f.text LIKE ? LIMIT ?",
            (pattern, pattern, row_limit),
        ).fetchall()
        for source, section, metadata, text in rows:
            src = _normalize_filename_for_match(source or "")
            if not src:
                continue
            if not _source_state(src).get("visible"):
                continue
            score = 0.0
            section_text = (section or "")
            if target in section_text:
                score += 3.0
            md: Dict[str, Any] = {}
            try:
                md = json.loads(metadata or "{}")
            except Exception:
                md = {}
            for field in ("section_title", "section", "section_path"):
                v = md.get(field)
                if isinstance(v, list):
                    if any(target in str(item) for item in v):
                        score += 2.4
                        break
                else:
                    if target in str(v or ""):
                        score += 2.4
                        break
            if target in (text or ""):
                score += 1.2
            if score <= 0:
                continue
            source_score[src] = source_score.get(src, 0.0) + score
            source_target_hits.setdefault(src, set()).add(target)
    ranked = sorted(
        source_score.items(),
        key=lambda item: (-item[1], -len(source_target_hits.get(item[0], set())), item[0]),
    )
    out: List[str] = []
    for src, _ in ranked:
        if src not in out:
            out.append(src)
        if len(out) >= max(1, int(limit)):
            break
    return out


def _llm_extract_section_targets(query: str, limit: int = 5) -> List[str]:
    if not (config.LLM_CHAT_COMPLETIONS_URL or config.LLM_API_BASE):
        return []
    def _extra_body() -> Dict[str, Any]:
        raw = (config.LLM_EXTRA_BODY or "").strip()
        if not raw:
            return {}
        try:
            obj = json.loads(raw)
            return obj if isinstance(obj, dict) else {}
        except Exception:
            return {}

    def _llm_text_from_response(data: Dict[str, Any]) -> str:
        choices = data.get("choices") or []
        if not choices or not isinstance(choices, list):
            return ""
        first = choices[0] or {}
        msg = first.get("message") or {}
        content = msg.get("content")
        if isinstance(content, str) and content.strip():
            return content.strip()
        reasoning = msg.get("reasoning")
        if isinstance(reasoning, str) and reasoning.strip():
            return reasoning.strip()
        text = first.get("text")
        if isinstance(text, str) and text.strip():
            return text.strip()
        return ""
    prompt = (
        "请从用户问题中抽取“章节/节/条标题级”短语，并严格输出 JSON 数组字符串。\n"
        "约束：\n"
        "1) 只输出 target，不输出解释。\n"
        "2) 不允许输出法规文档名（如包含条例/办法/法规等）。\n"
        "3) 每个 target 2-20 字，最多 5 个。\n"
        "4) 无法抽取时返回 []。\n\n"
        f"用户问题：{query}"
    )

    def _chat_url_candidates() -> List[str]:
        if config.LLM_CHAT_COMPLETIONS_URL:
            return [config.LLM_CHAT_COMPLETIONS_URL]
        base = (config.LLM_API_BASE or "").rstrip("/")
        candidates: List[str] = []
        if base:
            candidates.append(f"{base}/chat/completions")
            if not base.endswith("/v1"):
                candidates.append(f"{base}/v1/chat/completions")
            if base.endswith("/v1"):
                candidates.append(f"{base[:-3].rstrip('/')}/chat/completions")
        return candidates

    headers = {"Content-Type": "application/json"}
    if config.LLM_API_KEY:
        headers["Authorization"] = f"Bearer {config.LLM_API_KEY}"
    payload = {
        "model": config.LLM_MODEL,
        "messages": [
            {"role": "system", "content": "你是信息抽取器。仅返回 JSON 数组。"},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.0,
        "top_p": 1.0,
        "max_tokens": 120,
    }
    extra = _extra_body()
    if extra:
        for k, v in extra.items():
            if k not in payload:
                payload[k] = v

    try:
        import requests

        for url in _chat_url_candidates():
            try:
                resp = requests.post(url, headers=headers, json=payload, timeout=max(3, int(config.LLM_TIMEOUT)))
                if resp.status_code == 404:
                    continue
                resp.raise_for_status()
                data = resp.json()
                content = _llm_text_from_response(data)
                if not content:
                    continue
                match = re.search(r"\[[\s\S]*\]", content)
                raw_json = match.group(0) if match else content
                arr = json.loads(raw_json)
                if not isinstance(arr, list):
                    continue
                raw_targets = [str(item).strip() for item in arr if str(item).strip()]
                return _local_validate_section_targets(raw_targets, limit=limit)
            except Exception:
                continue
    except Exception:
        return []
    return []


_LLM_QUERY_PARSE_CACHE: Dict[str, Dict[str, Any]] = {}


def _llm_query_parse_cache_get(key: str) -> Optional[Dict[str, Any]]:
    cached = _LLM_QUERY_PARSE_CACHE.get(key)
    return cached if isinstance(cached, dict) else None


def _llm_query_parse_cache_set(key: str, value: Dict[str, Any]) -> None:
    try:
        max_items = int(getattr(config, "QUERY_PARSE_CACHE_SIZE", 512))
    except Exception:
        max_items = 512
    if max_items <= 0:
        return
    if len(_LLM_QUERY_PARSE_CACHE) >= max_items:
        _LLM_QUERY_PARSE_CACHE.clear()
    _LLM_QUERY_PARSE_CACHE[key] = value


def _llm_parse_query(user_query: str, locked_title: str = "") -> Dict[str, Any]:
    if not (config.LLM_CHAT_COMPLETIONS_URL or config.LLM_API_BASE):
        return {}
    uq = (user_query or "").strip()
    if not uq:
        return {}

    def _extra_body() -> Dict[str, Any]:
        raw = (config.LLM_EXTRA_BODY or "").strip()
        if not raw:
            return {}
        try:
            obj = json.loads(raw)
            return obj if isinstance(obj, dict) else {}
        except Exception:
            return {}

    def _llm_text_from_response(data: Dict[str, Any]) -> str:
        choices = data.get("choices") or []
        if not choices or not isinstance(choices, list):
            return ""
        first = choices[0] or {}
        msg = first.get("message") or {}
        content = msg.get("content")
        if isinstance(content, str) and content.strip():
            return content.strip()
        reasoning = msg.get("reasoning")
        if isinstance(reasoning, str) and reasoning.strip():
            return reasoning.strip()
        text = first.get("text")
        if isinstance(text, str) and text.strip():
            return text.strip()
        return ""

    def _chat_url_candidates() -> List[str]:
        if config.LLM_CHAT_COMPLETIONS_URL:
            return [config.LLM_CHAT_COMPLETIONS_URL]
        base = (config.LLM_API_BASE or "").rstrip("/")
        candidates: List[str] = []
        if base:
            candidates.append(f"{base}/chat/completions")
            if not base.endswith("/v1"):
                candidates.append(f"{base}/v1/chat/completions")
            if base.endswith("/v1"):
                candidates.append(f"{base[:-3].rstrip('/')}/chat/completions")
        return candidates

    max_tokens = int(getattr(config, "QUERY_PARSE_MAX_TOKENS", 260))
    max_docs = int(getattr(config, "QUERY_PARSE_MAX_DOCS", 2))
    max_anchors = int(getattr(config, "QUERY_PARSE_MAX_ANCHORS", 1))
    max_aspects = int(getattr(config, "QUERY_PARSE_MAX_ASPECTS", 4))
    max_sections = int(getattr(config, "QUERY_PARSE_MAX_SECTION_TARGETS", 4))

    title = (locked_title or "").strip()
    prompt = (
        "你是法规问答系统的“查询解析器”。你需要理解用户问题，输出干净可用的结构化 JSON。\n\n"
        "仅输出 JSON 对象，不要输出任何解释文字。\n\n"
        "输出字段：\n"
        f"- documents: 显式出现的法规/文件标题列表（最多 {max_docs} 个；只在用户明确提到时输出；禁止猜测）。\n"
        "- retrieval_query: 用于检索的干净问句（删除法规名、年份版本、比较词、客套口语）。\n"
        "- dense_query: 用于向量检索的等价改写（更偏法言法语；不含法规名/年份版本）。\n"
        f"- anchors: 用于跑题判定的核心内容锚点（最多 {max_anchors} 个短语）。\n"
        f"- aspects: 用于内容覆盖判定的要点（最多 {max_aspects} 个短语；尽量只保留核心名词短语）。\n"
        f"- section_targets: 若问题在问“第几条/总则/法律责任/罚则/附则”等章节定位，则输出章节目标（最多 {max_sections} 个）。\n"
        "- is_comparison: 是否对比/比较类问题（true/false）。\n\n"
        "强约束：\n"
        "1) anchors/aspects/section_targets 中禁止包含法规名（条例/办法/规定/规则/细则/决定/通知/法律/法规等）与年份版本（如 2018/2021/2025/修订/现行有效/最新版）。\n"
        "2) anchors 必须是能在条文中直接命中的“核心概念”，不要输出问句包装词（怎么/如何/哪些/什么/是否/请问）。\n"
        "3) 如果无法抽取某字段，输出空数组/空字符串，禁止省略字段。\n\n"
        f"已锁定法规标题（仅供理解上下文，禁止写入 documents/anchors/aspects）：{title}\n"
        f"用户问题：{uq}\n\n"
        "JSON："
    )

    headers = {"Content-Type": "application/json"}
    if config.LLM_API_KEY:
        headers["Authorization"] = f"Bearer {config.LLM_API_KEY}"
    payload: Dict[str, Any] = {
        "model": config.LLM_MODEL,
        "messages": [
            {"role": "system", "content": "你是信息抽取器。仅返回 JSON 对象。"},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.0,
        "top_p": 1.0,
        "max_tokens": max_tokens,
    }
    extra = _extra_body()
    if extra:
        for k, v in extra.items():
            if k not in payload:
                payload[k] = v

    def _clean_terms(values: Any, limit: int) -> List[str]:
        if not isinstance(values, list):
            return []
        out: List[str] = []
        for item in values:
            v = _normalize_query(str(item or ""))
            if not v:
                continue
            v = re.sub(r"^[，,、;；:：\s]+", "", v).strip()
            v = re.sub(r"[，,、;；:：\s]+$", "", v).strip()
            if len(v) < 2 or len(v) > 18:
                continue
            if re.search(r"\d{4}", v):
                continue
            if any(token in v for token in ["修订", "现行有效", "最新版", "最新"]):
                continue
            if any(token in v for token in ["条例", "办法", "规定", "规则", "细则", "决定", "通知", "法律", "法规"]):
                continue
            if any(token in v for token in ["对比", "比较", "区别", "差异", "不同"]):
                continue
            if any(token in v for token in ["请问", "怎么", "如何", "哪些", "什么", "是否", "吗", "呢"]):
                continue
            if any(v in kept or kept in v for kept in out):
                continue
            out.append(v)
            if len(out) >= max(0, int(limit)):
                break
        return out

    def _clean_documents(values: Any, limit: int) -> List[str]:
        if not isinstance(values, list):
            return []
        out: List[str] = []
        for item in values:
            t = str(item or "").strip()
            if not t:
                continue
            t = re.sub(r"[“”\"'<>《》]", "", t).strip()
            t = re.sub(r"\s+", "", t).strip()
            if len(t) < 3 or len(t) > 60:
                continue
            if any(x in t for x in ["怎么", "如何", "哪些", "什么", "是否"]):
                continue
            if t not in out:
                out.append(t)
            if len(out) >= max(0, int(limit)):
                break
        return out

    try:
        import requests
        for url in _chat_url_candidates():
            try:
                resp = requests.post(url, headers=headers, json=payload, timeout=max(3, min(20, int(config.LLM_TIMEOUT))))
                if resp.status_code == 404:
                    continue
                resp.raise_for_status()
                data = resp.json()
                text = _llm_text_from_response(data)
                if not text:
                    continue
                match = re.search(r"\{[\s\S]*\}", text)
                raw_json = match.group(0) if match else text
                obj = json.loads(raw_json)
                if not isinstance(obj, dict):
                    continue
                documents = _clean_documents(obj.get("documents"), limit=max_docs)
                retrieval_query = _normalize_query(str(obj.get("retrieval_query") or ""))
                dense_query = _normalize_query(str(obj.get("dense_query") or "")) or retrieval_query
                anchors = _clean_terms(obj.get("anchors"), limit=max_anchors)
                aspects = _clean_terms(obj.get("aspects"), limit=max_aspects)
                section_targets = _clean_terms(obj.get("section_targets"), limit=max_sections)
                is_comparison = bool(obj.get("is_comparison"))
                return {
                    "documents": documents,
                    "retrieval_query": retrieval_query,
                    "dense_query": dense_query,
                    "anchors": anchors,
                    "aspects": aspects,
                    "section_targets": section_targets,
                    "is_comparison": is_comparison,
                }
            except Exception:
                continue
    except Exception:
        return {}
    return {}


def _llm_parse_query_cached(user_query: str, locked_title: str = "") -> Dict[str, Any]:
    key = f"{(locked_title or '').strip()}||{(user_query or '').strip()}"
    cached = _llm_query_parse_cache_get(key)
    if cached is not None:
        return cached
    parsed = _llm_parse_query(user_query, locked_title=locked_title)
    if isinstance(parsed, dict):
        _llm_query_parse_cache_set(key, parsed)
        _llm_query_parse_cache_set((user_query or "").strip(), parsed)
        return parsed
    _llm_query_parse_cache_set(key, {})
    return {}


def _llm_extract_clean_keywords(user_query: str, evidence_query: str, locked_title: str = "") -> Dict[str, Any]:
    if not (config.LLM_CHAT_COMPLETIONS_URL or config.LLM_API_BASE):
        return {}
    def _extra_body() -> Dict[str, Any]:
        raw = (config.LLM_EXTRA_BODY or "").strip()
        if not raw:
            return {}
        try:
            obj = json.loads(raw)
            return obj if isinstance(obj, dict) else {}
        except Exception:
            return {}

    def _llm_text_from_response(data: Dict[str, Any]) -> str:
        choices = data.get("choices") or []
        if not choices or not isinstance(choices, list):
            return ""
        first = choices[0] or {}
        msg = first.get("message") or {}
        content = msg.get("content")
        if isinstance(content, str) and content.strip():
            return content.strip()
        reasoning = msg.get("reasoning")
        if isinstance(reasoning, str) and reasoning.strip():
            return reasoning.strip()
        text = first.get("text")
        if isinstance(text, str) and text.strip():
            return text.strip()
        return ""
    uq = (user_query or "").strip()
    eq = (evidence_query or "").strip()
    title = (locked_title or "").strip()
    if not eq:
        return {}
    max_items = max(3, int(getattr(config, "KEYWORD_PARSE_MAX_ITEMS", 8)))
    max_anchors = min(max_items, int(getattr(config, "KEYWORD_PARSE_MAX_ANCHORS", 2)))
    max_aspects = min(max_items, int(getattr(config, "KEYWORD_PARSE_MAX_ASPECTS", 4)))
    prompt = (
        "你是法规问答的关键词解析器。\n"
        "目标：理解用户问题，输出用于“条文检索/证据覆盖判定”的干净关键词。\n\n"
        "输出格式：仅输出 JSON 对象，且必须包含 anchors 和 aspects 两个字段。\n"
        "字段含义：\n"
        "- anchors：用于判定是否跑题的内容锚点（3~8 个短语）。\n"
        "- aspects：用于证据覆盖度判定的要点（可与 anchors 相同）。\n\n"
        "硬约束：\n"
        f"1) 每个短语 2~14 字；anchors 最多 {max_anchors} 个，aspects 最多 {max_aspects} 个。\n"
        "1.1) 禁止输出近义重复项/同词不同尾巴（如“认定/界定/类型/范围/情形”等扩写），每类概念只保留最核心的 1 个短语。\n"
        "2) 禁止输出法规/文件名（含“条例/办法/规定/规则/细则/决定/通知/法律/法规”等）。\n"
        "3) 禁止输出年份/版本（如 2018/2021/2025/修订/现行有效/最新版）。\n"
        "4) 禁止输出比较词（对比/区别/差异/有什么不同）。\n"
        "5) 禁止输出问句包装词（请问/怎么/如何/哪些/什么/是否/吗/呢）。\n"
        "6) 输出必须可直接用于检索；如果抽取不到，输出 {\"anchors\":[],\"aspects\":[]}。\n\n"
        f"已锁定法规标题（可参考但禁止输出）：{title}\n"
        f"用户原问：{uq}\n"
        f"检索问句（已净化）：{eq}\n\n"
        "JSON："
    )

    def _chat_url_candidates() -> List[str]:
        if config.LLM_CHAT_COMPLETIONS_URL:
            return [config.LLM_CHAT_COMPLETIONS_URL]
        base = (config.LLM_API_BASE or "").rstrip("/")
        candidates: List[str] = []
        if base:
            candidates.append(f"{base}/chat/completions")
            if not base.endswith("/v1"):
                candidates.append(f"{base}/v1/chat/completions")
            if base.endswith("/v1"):
                candidates.append(f"{base[:-3].rstrip('/')}/chat/completions")
        return candidates

    headers = {"Content-Type": "application/json"}
    if config.LLM_API_KEY:
        headers["Authorization"] = f"Bearer {config.LLM_API_KEY}"
    payload = {
        "model": config.LLM_MODEL,
        "messages": [
            {"role": "system", "content": "你是信息抽取器。仅返回 JSON 对象。"},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.0,
        "top_p": 1.0,
        "max_tokens": int(getattr(config, "KEYWORD_PARSE_MAX_TOKENS", 160)),
    }
    extra = _extra_body()
    if extra:
        for k, v in extra.items():
            if k not in payload:
                payload[k] = v

    def _clean_items(values: Any, limit: int) -> List[str]:
        if not isinstance(values, list):
            return []
        out: List[str] = []
        suffixes = ["范围", "类型", "界定", "认定", "情形", "条件", "标准", "程序", "规定"]
        for item in values:
            v = _normalize_query(str(item or ""))
            if not v:
                continue
            v = re.sub(r"^[，,、;；:：\s]+", "", v).strip()
            v = re.sub(r"[，,、;；:：\s]+$", "", v).strip()
            if len(v) < 2:
                continue
            if re.search(r"\d{4}", v):
                continue
            if any(token in v for token in ["修订", "现行有效", "最新版", "最新"]):
                continue
            if any(token in v for token in ["条例", "办法", "规定", "规则", "细则", "决定", "通知", "法律", "法规"]):
                continue
            if any(token in v for token in ["对比", "比较", "区别", "差异", "不同"]):
                continue
            if any(token in v for token in ["请问", "怎么", "如何", "哪些", "什么", "是否", "吗", "呢"]):
                continue
            # drop obvious near-duplicate expansions if base already kept
            dropped = False
            for suf in suffixes:
                if v.endswith(suf) and len(v) > len(suf) + 1:
                    base = _normalize_query(v[: -len(suf)])
                    if base and base in out:
                        dropped = True
                        break
            if dropped:
                continue
            # substring dedupe (keep shorter core term)
            if any(v in kept or kept in v for kept in out):
                continue
            if v not in out:
                out.append(v)
            if len(out) >= max(1, int(limit)):
                break
        return out

    try:
        import requests

        for url in _chat_url_candidates():
            try:
                resp = requests.post(url, headers=headers, json=payload, timeout=max(3, min(12, int(config.LLM_TIMEOUT))))
                if resp.status_code == 404:
                    continue
                resp.raise_for_status()
                data = resp.json()
                content = _llm_text_from_response(data)
                if not content:
                    continue
                match = re.search(r"\{[\s\S]*\}", content)
                raw_json = match.group(0) if match else content
                obj = json.loads(raw_json)
                if not isinstance(obj, dict):
                    continue
                anchors = _clean_items(obj.get("anchors"), limit=max_anchors)
                aspects = _clean_items(obj.get("aspects"), limit=max_aspects)
                return {"anchors": anchors, "aspects": aspects}
            except Exception:
                continue
    except Exception:
        return {}
    return {}


def _purify_retrieval_query_shallow(query: str) -> str:
    q = (query or "").strip()
    if not q:
        return ""
    q = re.sub(r"[《》「」\"“”]", " ", q)
    q = re.sub(r"\s+", " ", q).strip()
    q = re.sub(r"^[，。；;：:?？、,.\s]+", "", q).strip()
    q = re.sub(r"^(在)?(中|里|内)[，,。\s]+", "", q).strip()
    q = re.sub(r"^(在)?(中|里|内)(?=[A-Za-z0-9\u4e00-\u9fff])", "", q).strip()
    q = re.sub(r"^(关于|对于|有关|根据|请问|请|查询|说明|咨询|问下|想了解|我想了解)\s*", "", q).strip()
    q = re.sub(r"^(对|在)\s*", "", q).strip()
    q = re.sub(r"^(本|该|这个|上述)\s*", "", q).strip()
    q = re.sub(r"\s+", " ", q).strip()
    q = re.sub(r"^[，。；;：:?？、,.\s]+", "", q).strip()
    return q


def _llm_purify_retrieval_query(user_query: str, retrieval_query: str, locked_title: str = "") -> str:
    if not (config.LLM_CHAT_COMPLETIONS_URL or config.LLM_API_BASE):
        return ""
    def _extra_body() -> Dict[str, Any]:
        raw = (config.LLM_EXTRA_BODY or "").strip()
        if not raw:
            return {}
        try:
            obj = json.loads(raw)
            return obj if isinstance(obj, dict) else {}
        except Exception:
            return {}

    def _llm_text_from_response(data: Dict[str, Any]) -> str:
        choices = data.get("choices") or []
        if not choices or not isinstance(choices, list):
            return ""
        first = choices[0] or {}
        msg = first.get("message") or {}
        content = msg.get("content")
        if isinstance(content, str) and content.strip():
            return content.strip()
        reasoning = msg.get("reasoning")
        if isinstance(reasoning, str) and reasoning.strip():
            return reasoning.strip()
        text = first.get("text")
        if isinstance(text, str) and text.strip():
            return text.strip()
        return ""
    uq = (user_query or "").strip()
    rq = (retrieval_query or "").strip()
    if not rq:
        return ""
    title = (locked_title or "").strip()
    prompt = (
        "你是法规条文检索的 Query 净化器。\n"
        "目标：把问题改写成用于“在同一部法规条文中检索”的简短问句。\n\n"
        "规则：\n"
        "1) 只输出一行净化后的问句，不输出解释。\n"
        "2) 不要包含法规名称/文件名/年份版本信息（如 2010 版/2021 年修订）。\n"
        "3) 去掉引导词（如 关于/在…中/请问）。\n"
        "4) 保留核心概念词与约束条件（主体/行为/对象/情形/权限/职责等）。\n"
        "5) 输出 8~40 字；无法净化就原样输出“当前检索问句”。\n\n"
        f"已锁定法规标题（可参考但禁止输出）：{title}\n"
        f"用户原问：{uq}\n"
        f"当前检索问句：{rq}\n\n"
        "净化后问句："
    )

    def _chat_url_candidates() -> List[str]:
        if config.LLM_CHAT_COMPLETIONS_URL:
            return [config.LLM_CHAT_COMPLETIONS_URL]
        base = (config.LLM_API_BASE or "").rstrip("/")
        candidates: List[str] = []
        if base:
            candidates.append(f"{base}/chat/completions")
            if not base.endswith("/v1"):
                candidates.append(f"{base}/v1/chat/completions")
            if base.endswith("/v1"):
                candidates.append(f"{base[:-3].rstrip('/')}/chat/completions")
        return candidates

    headers = {"Content-Type": "application/json"}
    if config.LLM_API_KEY:
        headers["Authorization"] = f"Bearer {config.LLM_API_KEY}"
    payload = {
        "model": config.LLM_MODEL,
        "messages": [
            {"role": "system", "content": "你是 Query 净化器。只输出净化后的问句。"},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.0,
        "top_p": 1.0,
        "max_tokens": int(getattr(config, "QUERY_PURIFY_MAX_TOKENS", 60)),
    }
    extra = _extra_body()
    if extra:
        for k, v in extra.items():
            if k not in payload:
                payload[k] = v
    try:
        import requests
        for url in _chat_url_candidates():
            try:
                resp = requests.post(url, headers=headers, json=payload, timeout=max(3, min(12, int(config.LLM_TIMEOUT))))
                if resp.status_code == 404:
                    continue
                resp.raise_for_status()
                data = resp.json()
                content = _llm_text_from_response(data)
                content = re.sub(r"\s+", " ", content).strip()
                if content:
                    return content[:120]
            except Exception:
                continue
    except Exception:
        return ""
    return ""


def _rewrite_legalese_rule_fallback(query: str) -> str:
    q = _normalize_query(query)
    if not q:
        return ""
    pairs = [
        ("怎么处理", "处理规定 处罚措施"),
        ("如何处理", "处理规定 处罚措施"),
        ("怎么办", "处理规定"),
        ("不能做", "禁止行为"),
        ("不可以做", "禁止行为"),
        ("可以做吗", "是否允许"),
        ("能否", "是否可以"),
        ("流程", "程序"),
        ("审议流程", "审议程序"),
        ("办理流程", "办理程序"),
        ("是怎样的", "如何规定"),
        ("是怎么样的", "如何规定"),
        ("是什么", "定义"),
    ]
    out = q
    for a, b in pairs:
        if a in out:
            out = out.replace(a, b)
    out = re.sub(r"\s+", " ", out).strip()
    return out or q


def _llm_rewrite_legalese(user_query: str, retrieval_query: str, locked_title: str = "") -> str:
    if not (config.LLM_CHAT_COMPLETIONS_URL or config.LLM_API_BASE):
        return ""
    def _extra_body() -> Dict[str, Any]:
        raw = (config.LLM_EXTRA_BODY or "").strip()
        if not raw:
            return {}
        try:
            obj = json.loads(raw)
            return obj if isinstance(obj, dict) else {}
        except Exception:
            return {}

    def _llm_text_from_response(data: Dict[str, Any]) -> str:
        choices = data.get("choices") or []
        if not choices or not isinstance(choices, list):
            return ""
        first = choices[0] or {}
        msg = first.get("message") or {}
        content = msg.get("content")
        if isinstance(content, str) and content.strip():
            return content.strip()
        reasoning = msg.get("reasoning")
        if isinstance(reasoning, str) and reasoning.strip():
            return reasoning.strip()
        text = first.get("text")
        if isinstance(text, str) and text.strip():
            return text.strip()
        return ""
    uq = (user_query or "").strip()
    rq = (retrieval_query or "").strip()
    if not rq:
        return ""
    title = (locked_title or "").strip()
    prompt = (
        "你是法规问答系统的意图重写器。\n"
        "目标：把用户口语问题改写成更标准的“法言法语检索问句”，用于稠密向量检索。\n\n"
        "规则：\n"
        "1) 只输出一行重写后的问句，不输出解释。\n"
        "2) 不要输出法规名称/文件名/年份版本信息（如 2010 版/2021 年修订）。\n"
        "3) 保留核心概念、主体/行为/对象/条件/情形等约束。\n"
        "4) 优先把口语映射为法言法语：怎么处理→处罚措施/处理规定，不能做→禁止行为，流程→程序，是什么→定义。\n"
        "5) 输出 10~50 字；无法改写就输出“当前检索问句”的规范化版本。\n\n"
        f"已锁定法规标题（可参考但禁止输出）：{title}\n"
        f"用户原问：{uq}\n"
        f"当前检索问句：{rq}\n\n"
        "重写问句："
    )

    def _chat_url_candidates() -> List[str]:
        if config.LLM_CHAT_COMPLETIONS_URL:
            return [config.LLM_CHAT_COMPLETIONS_URL]
        base = (config.LLM_API_BASE or "").rstrip("/")
        candidates: List[str] = []
        if base:
            candidates.append(f"{base}/chat/completions")
            if not base.endswith("/v1"):
                candidates.append(f"{base}/v1/chat/completions")
            if base.endswith("/v1"):
                candidates.append(f"{base[:-3].rstrip('/')}/chat/completions")
        return candidates

    headers = {"Content-Type": "application/json"}
    if config.LLM_API_KEY:
        headers["Authorization"] = f"Bearer {config.LLM_API_KEY}"
    payload = {
        "model": config.LLM_MODEL,
        "messages": [
            {"role": "system", "content": "你是意图重写器。只输出重写后的问句。"},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.1,
        "top_p": 1.0,
        "max_tokens": int(getattr(config, "QUERY_REWRITE_MAX_TOKENS", 90)),
    }
    extra = _extra_body()
    if extra:
        for k, v in extra.items():
            if k not in payload:
                payload[k] = v
    try:
        import requests
        for url in _chat_url_candidates():
            try:
                resp = requests.post(url, headers=headers, json=payload, timeout=max(3, min(12, int(config.LLM_TIMEOUT))))
                if resp.status_code == 404:
                    continue
                resp.raise_for_status()
                data = resp.json()
                content = _llm_text_from_response(data)
                content = re.sub(r"\s+", " ", content).strip()
                if content:
                    return content[:160]
            except Exception:
                continue
    except Exception:
        return _rewrite_legalese_rule_fallback(rq)
    return _rewrite_legalese_rule_fallback(rq)


_EVIDENCE_CHECK_CACHE: Dict[str, bool] = {}


def _best_dense_relevance_for_locked_source(docs: List[Any], target_sources: Optional[List[str]]) -> float:
    allowed = [
        _normalize_filename_for_match(source or "")
        for source in (target_sources or [])
        if _normalize_filename_for_match(source or "")
    ]
    if len(allowed) != 1:
        return 0.0
    target = allowed[0]
    best = 0.0
    for doc in docs or []:
        if _normalize_filename_for_match(_hit_entity_source(doc) or "") != target:
            continue
        md = _hit_metadata(doc)
        if str(md.get("orig_score_mode") or "") != "distance":
            continue
        dist = float(md.get("orig_score") or 0.0)
        rel = max(0.0, 1.0 - dist)
        if rel > best:
            best = rel
    return float(best)


def _locked_source_evidence_window(docs: List[Any], target_sources: Optional[List[str]], limit: int) -> List[Any]:
    allowed = [
        _normalize_filename_for_match(source or "")
        for source in (target_sources or [])
        if _normalize_filename_for_match(source or "")
    ]
    if len(allowed) != 1:
        return []
    target = allowed[0]
    out: List[Any] = []
    for doc in docs or []:
        if _normalize_filename_for_match(_hit_entity_source(doc) or "") != target:
            continue
        if _is_heading_only_hit(doc):
            continue
        if not _hit_matches_source_state(doc, _source_state(target)):
            continue
        out.append(doc)
        if len(out) >= max(1, int(limit)):
            break
    return out


def _llm_evidence_core_concept_hit(query: str, window: List[Any]) -> bool:
    if not (config.LLM_CHAT_COMPLETIONS_URL or config.LLM_API_BASE):
        return False
    def _extra_body() -> Dict[str, Any]:
        raw = (config.LLM_EXTRA_BODY or "").strip()
        if not raw:
            return {}
        try:
            obj = json.loads(raw)
            return obj if isinstance(obj, dict) else {}
        except Exception:
            return {}

    def _llm_text_from_response(data: Dict[str, Any]) -> str:
        choices = data.get("choices") or []
        if not choices or not isinstance(choices, list):
            return ""
        first = choices[0] or {}
        msg = first.get("message") or {}
        content = msg.get("content")
        if isinstance(content, str) and content.strip():
            return content.strip()
        reasoning = msg.get("reasoning")
        if isinstance(reasoning, str) and reasoning.strip():
            return reasoning.strip()
        text = first.get("text")
        if isinstance(text, str) and text.strip():
            return text.strip()
        return ""
    q = _normalize_query(query)
    if not q or not window:
        return False

    parts: List[str] = []
    for doc in window[: max(1, int(getattr(config, "CONTEXT_TOP_N", 6)))]:
        section = _doc_section_name(doc)
        text = _hit_display_text(doc) or ""
        snippet = _normalize_query(f"{section}\n{text}".strip())
        if snippet:
            parts.append(snippet)
    evidence = "\n\n".join(parts)
    evidence = evidence[: max(200, int(getattr(config, "LLM_EVIDENCE_CHECK_MAX_CHARS", 1800)))]
    if not evidence:
        return False

    key = f"{q}||{hash(evidence)}"
    cached = _EVIDENCE_CHECK_CACHE.get(key)
    if cached is not None:
        return bool(cached)

    prompt = (
        "你是证据校验器。\n"
        "判断给定条文摘录是否包含回答该问题所需的核心概念或关键要点。\n"
        "只输出“是”或“否”。\n\n"
        f"问题：{q}\n\n"
        f"条文摘录：\n{evidence}\n\n"
        "答案："
    )

    def _chat_url_candidates() -> List[str]:
        if config.LLM_CHAT_COMPLETIONS_URL:
            return [config.LLM_CHAT_COMPLETIONS_URL]
        base = (config.LLM_API_BASE or "").rstrip("/")
        candidates: List[str] = []
        if base:
            candidates.append(f"{base}/chat/completions")
            if not base.endswith("/v1"):
                candidates.append(f"{base}/v1/chat/completions")
            if base.endswith("/v1"):
                candidates.append(f"{base[:-3].rstrip('/')}/chat/completions")
        return candidates

    headers = {"Content-Type": "application/json"}
    if config.LLM_API_KEY:
        headers["Authorization"] = f"Bearer {config.LLM_API_KEY}"
    payload = {
        "model": config.LLM_MODEL,
        "messages": [
            {"role": "system", "content": "你是证据校验器。只输出“是”或“否”。"},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.0,
        "top_p": 1.0,
        "max_tokens": 4,
    }
    extra = _extra_body()
    if extra:
        for k, v in extra.items():
            if k not in payload:
                payload[k] = v
    ok = False
    try:
        import requests
        for url in _chat_url_candidates():
            try:
                resp = requests.post(url, headers=headers, json=payload, timeout=max(3, min(12, int(config.LLM_TIMEOUT))))
                if resp.status_code == 404:
                    continue
                resp.raise_for_status()
                data = resp.json()
                content = _llm_text_from_response(data)
                content = re.sub(r"\s+", " ", content).strip()
                if content:
                    ok = content.startswith("是")
                    break
            except Exception:
                continue
    except Exception:
        ok = False

    if len(_EVIDENCE_CHECK_CACHE) >= 512:
        _EVIDENCE_CHECK_CACHE.clear()
    _EVIDENCE_CHECK_CACHE[key] = bool(ok)
    return bool(ok)


def _is_section_anchor_query(query: str) -> bool:
    q = _normalize_query(query)
    if not q:
        return False
    section_targets = _local_validate_section_targets(_extract_section_query_targets(q), limit=6)
    if section_targets:
        return True
    intent_keywords = [str(item).strip() for item in (_policy_get("rerank.section_lookup.intent_keywords", []) or [])]
    return any(keyword and keyword in q for keyword in intent_keywords)


def _doc_fallback_source_candidates(query: str, limit: int) -> Dict[str, Any]:
    route = _classify_query_route(query)
    if route not in {"weak_title_reference", "business_topic_qa", "open_regulation_qa"}:
        return {
            "sources": [],
            "section_anchor_hit": False,
            "section_targets": [],
        }
    section_targets = _local_validate_section_targets(_extract_section_query_targets(query), limit=max(2, int(limit) * 2))
    reverse_sources = _sqlite_reverse_lookup_sources_by_targets(section_targets, limit=max(1, int(limit))) if section_targets else []
    plan = _build_doc_recall_plan(query, limit=max(int(limit) * 4, 10))
    out: List[str] = list(reverse_sources)
    for entry in plan:
        source = _normalize_filename_for_match((entry or {}).get("source") or "")
        if not source or source in out:
            continue
        reasons = set((entry or {}).get("reasons") or [])
        prior = float((entry or {}).get("prior") or 0.0)
        title_score = float((entry or {}).get("title_score") or 0.0)
        # 接受带有明确文档级信号的候选；prior 高时允许兜底进入候选。
        if reasons.intersection({"title_alias_substring", "documents_fts", "doc_term_overlap"}) or title_score > 0 or prior >= 0.45:
            out.append(source)
        if len(out) >= max(1, int(limit)):
            break
    return {
        "sources": out[: max(1, int(limit))],
        "section_anchor_hit": bool(section_targets and reverse_sources),
        "section_targets": section_targets[:5],
    }


def _has_contextual_doc_reference(query: str) -> bool:
    q = _normalize_query(query)
    if not q:
        return False
    markers = [
        "该条例", "这个条例", "上述条例", "本条例",
        "该办法", "这个办法", "上述办法", "本办法",
        "该规定", "这个规定", "上述规定", "本规定",
        "该规则", "这个规则", "上述规则", "本规则",
        "该细则", "这个细则", "上述细则", "本细则",
        "这个文件", "该文件", "上述文件", "本文件",
        "该通知", "这个通知", "上述通知", "本通知",
    ]
    return any(marker in q for marker in markers)


def _source_lock_required(query: str, route: str) -> bool:
    if route == "compare_clarification":
        return True
    if route in {"exact_title_reference", "alias_title_reference", "weak_title_reference", "version_switch"}:
        return True
    q = _normalize_query(query)
    if route in {"business_topic_qa", "open_regulation_qa"}:
        if q and not _query_has_doc_identity_term(q) and not _extract_explicit_regulation_mentions(q) and not _extract_filename_candidates(q):
            if _query_has_strong_business_signal(q) or _query_quality_strong_topic_terms(q):
                return False
    strong_section_targets, _ = _split_section_targets(query)
    if strong_section_targets:
        return True
    if _has_contextual_doc_reference(query):
        return True
    if _is_generic_document_required_query(query):
        return True
    if _is_unlocked_content_query(query, route):
        return True
    if route in {"business_topic_qa", "open_regulation_qa"}:
        return False
    if _is_section_anchor_query(query):
        return True
    if _extract_section_query_targets(query):
        return True
    return False


def _is_contextual_followup_query(query: str) -> bool:
    q = _normalize_query(query)
    if not q:
        return False
    if _has_contextual_doc_reference(q):
        return True
    if re.search(r"第[一二三四五六七八九十百千0-9]+[章节条款]", q):
        return True
    follow_terms = ["法律责任", "奖励与处罚", "处罚条款", "监督检查", "申请程序", "管理职责", "安全责任"]
    return any(term in q for term in follow_terms)


def _is_pure_topic_question(query: str, route: str) -> bool:
    q = _normalize_query(query)
    if not q:
        return False
    if route not in {"business_topic_qa", "open_regulation_qa", "content_qa"}:
        return False
    if _query_has_doc_identity_term(q):
        return False
    if _extract_explicit_regulation_mentions(q) or _extract_filename_candidates(q):
        return False
    return bool(_query_anchor_terms(q)) and not _has_contextual_doc_reference(q)


def _build_source_resolution_result(
    *,
    route: str = "",
    required: bool,
    resolved: bool,
    sources: Optional[List[str]] = None,
    candidates: Optional[List[str]] = None,
    reason: str = "",
    strip_title_mentions: bool = False,
    clarification: str = "",
    target_text: str = "",
    lock_mode: str = "hard_lock",
    lock_confidence: float = 0.0,
    lock_message_prefix: str = "",
    source_lock_kind: str = "",
    source_resolution_trace: Optional[Dict[str, Any]] = None,
    inherited_from_context: bool = False,
    compare_subjects: Optional[List[str]] = None,
    compare_doc_like_subjects: Optional[List[str]] = None,
    compare_missing_targets: Optional[List[str]] = None,
    compare_common_aspects: Optional[List[str]] = None,
    compare_topic_pair: Optional[List[str]] = None,
    compare_canonical_aspects: Optional[List[str]] = None,
    compare_expanded_aspects: Optional[List[str]] = None,
    compare_source_subqueries: Optional[Dict[str, Any]] = None,
    compare_status: str = "not_compare",
    compare_plan: Optional[Dict[str, Any]] = None,
    retrieval_query_override: str = "",
) -> Dict[str, Any]:
    normalized_sources = list(sources or [])
    normalized_candidates = list(candidates or [])
    return {
        "route": route,
        "required": bool(required),
        "resolved": bool(resolved),
        "sources": normalized_sources,
        "candidates": normalized_candidates,
        "reason": reason,
        "strip_title_mentions": bool(strip_title_mentions),
        "clarification": clarification,
        "target_text": target_text,
        "lock_mode": lock_mode,
        "lock_confidence": float(lock_confidence or 0.0),
        "lock_message_prefix": lock_message_prefix,
        "source_lock_kind": source_lock_kind or lock_mode,
        "source_resolution_trace": dict(source_resolution_trace or {}),
        "inherited_from_context": bool(inherited_from_context),
        "compare_subjects": list(compare_subjects or []),
        "compare_doc_like_subjects": list(compare_doc_like_subjects or []),
        "compare_missing_targets": list(compare_missing_targets or []),
        "compare_common_aspects": list(compare_common_aspects or []),
        "compare_topic_pair": list(compare_topic_pair or []),
        "compare_canonical_aspects": list(compare_canonical_aspects or []),
        "compare_expanded_aspects": list(compare_expanded_aspects or []),
        "compare_source_subqueries": dict(compare_source_subqueries or {}),
        "compare_status": compare_status,
        "compare_plan": dict(compare_plan or {}),
        "retrieval_query_override": retrieval_query_override,
    }


def _resolve_soft_lock_candidate(query: str, route: str, candidate_sources: List[str]) -> Dict[str, Any]:
    candidates = _collapse_sources_by_canonical(candidate_sources, limit=3)
    if len(candidates) != 1:
        return {"resolved": False}
    source = candidates[0]
    title_ranked = _rank_title_source_matches(query, limit=6, include_topic_like=True)
    source_entry = next((entry for entry in title_ranked if _normalize_filename_for_match(entry.get("source") or "") == source), None)
    if not source_entry:
        return {"resolved": False}
    match_kind = str(source_entry.get("match_kind") or "")
    score = float(source_entry.get("score") or 0.0)
    soft_kinds = {"topic_like_title", "alias_title"}
    if match_kind not in soft_kinds:
        return {"resolved": False}
    if match_kind == "alias_title" and score >= 8.0:
        return {"resolved": False}
    if score < 4.0:
        return {"resolved": False}
    if not _query_has_doc_identity_term(query):
        return {"resolved": False}
    if not _source_supports_doc_identity_term(source, query):
        return {"resolved": False}
    if _is_pure_topic_question(query, route):
        return {"resolved": False}
    top_competitors = [entry for entry in title_ranked if _normalize_filename_for_match(entry.get("source") or "") != source]
    if top_competitors and float(top_competitors[0].get("score") or 0.0) >= max(3.8, score - 0.8):
        return {"resolved": False, "competition": [entry.get("source") for entry in top_competitors[:2] if entry.get("source")]}
    title = _doc_get(source).get("canonical_title") or _filename_stem(source) or source
    return {
        "resolved": True,
        "source": source,
        "reason": "soft_lock",
        "confidence": score / 10.0,
        "lock_message_prefix": f"我理解你查询的是《{title}》……\n",
        "trace": {
            "match_kind": match_kind,
            "title_score": score,
            "non_fallback": True,
            "doc_identity_terms": _query_doc_identity_terms(query),
            "competition": [entry.get("source") for entry in top_competitors[:2] if entry.get("source")],
        },
    }


def _text_overlap_ratio(left: str, right: str) -> float:
    lnorm = _normalize_reference_text(left)
    rnorm = _normalize_reference_text(right)
    if not lnorm or not rnorm:
        return 0.0
    lset = set(lnorm)
    rset = set(rnorm)
    overlap = len(lset & rset)
    denom = max(1, min(len(lset), len(rset)))
    return float(overlap) / float(denom)


def _edit_similarity_ratio(left: str, right: str) -> float:
    lnorm = _normalize_reference_text(left)
    rnorm = _normalize_reference_text(right)
    if not lnorm or not rnorm:
        return 0.0
    if lnorm == rnorm:
        return 1.0
    rows = len(lnorm) + 1
    cols = len(rnorm) + 1
    dp = list(range(cols))
    for i in range(1, rows):
        prev = dp[0]
        dp[0] = i
        for j in range(1, cols):
            current = dp[j]
            if lnorm[i - 1] == rnorm[j - 1]:
                dp[j] = prev
            else:
                dp[j] = min(prev, dp[j], dp[j - 1]) + 1
            prev = current
    distance = dp[-1]
    return 1.0 - float(distance) / float(max(len(lnorm), len(rnorm), 1))


def _source_profile_fields(source: str) -> Dict[str, Any]:
    safe_source = _normalize_filename_for_match(source or "")
    if not safe_source:
        return {}
    info = _doc_get(safe_source)
    version = int(info.get("active_version") or 0)
    if version > 0:
        profile = _doc_profile_get(safe_source, version)
        if profile:
            return profile
    return {}


def _geo_context_tokens(query: str, user_id: str) -> List[str]:
    out: List[str] = []
    for token in re.findall(r"[一-龥]{2,12}(?:省|市|区|县|州|盟|旗|乡|镇|新区|自治州|自治区)", _normalize_query(query)):
        normalized = _normalize_query(token)
        if normalized and normalized not in out:
            out.append(normalized)
    current = _get_current_locked_document(user_id) if user_id else None
    current_source = _normalize_filename_for_match((current or {}).get("source") or "")
    if current_source:
        profile = _source_profile_fields(current_source)
        region = _normalize_query(profile.get("region") or _extract_region_token(_source_display_title(current_source)))
        if region and region not in out:
            out.append(region)
    return out


def _resolve_geo_context_locked(query: str, user_id: str, candidate_sources: List[str]) -> Dict[str, Any]:
    geo_tokens = _geo_context_tokens(query, user_id)
    if not geo_tokens or len(candidate_sources) <= 1:
        return {"resolved": False}
    ranked: List[Tuple[int, str, str]] = []
    for source in candidate_sources:
        safe_source = _normalize_filename_for_match(source or "")
        if not safe_source:
            continue
        profile = _source_profile_fields(safe_source)
        region = _normalize_query(profile.get("region") or _extract_region_token(_source_display_title(safe_source)))
        if not region:
            continue
        best = 0
        for token in geo_tokens:
            if token == region or token in region or region in token:
                best = max(best, max(len(token), len(region)))
        if best > 0:
            ranked.append((best, safe_source, region))
    ranked.sort(key=lambda item: (-item[0], item[1]))
    if len(ranked) != 1:
        return {"resolved": False}
    best_score, source, region = ranked[0]
    title = _source_display_title(source)
    return {
        "resolved": True,
        "source": source,
        "reason": "geo_context_locked",
        "confidence": min(0.95, 0.72 + 0.03 * max(1, best_score)),
        "lock_message_prefix": f"我理解你查询的是《{title}》……\n",
        "trace": {"geo_context_tokens": geo_tokens, "matched_region": region},
    }


_TOPICAL_SUFFIX_TERMS = [
    "法律责任",
    "核心条款",
    "审议程序",
    "立法程序",
    "处罚规定",
    "处罚条款",
    "处罚标准",
    "收费标准",
    "处理规定",
    "监督检查",
    "管理职责",
    "奖励与处罚",
    "怎么处理",
    "怎么罚",
    "罚则",
    "要点",
    "综述",
    "概括",
]

_GENERIC_DOC_INTENT_TERMS = {
    "议事",
    "议事规则",
    "地方立法",
    "文明行为",
    "文明行为促进",
    "养犬管理",
    "河道管理",
    "红色文化资源保护",
    "红色文化资源保护利用",
    "非遗保护",
    "非物质文化遗产保护",
    "收费",
    "处罚",
    "管理",
}


def _topical_suffix_match(query: str) -> str:
    q = _normalize_query(query).replace(" ", "")
    if not q:
        return ""
    for term in _TOPICAL_SUFFIX_TERMS:
        if term.replace(" ", "") in q:
            return term
    match = re.search(r"(处罚|罚则|标准|程序|责任|条款|要求)$", q)
    return match.group(1) if match else ""


def _is_topical_suffix_query(query: str) -> bool:
    return bool(_topical_suffix_match(query))


def _query_doc_intent(query: str) -> str:
    mention = ""
    mentions = _extract_explicit_regulation_mentions(query)
    if mentions:
        mention = mentions[0]
    base_text = _normalize_query(mention or query)
    suffix = _topical_suffix_match(base_text)
    if suffix:
        base_text = base_text.replace(suffix, " ")
    base_text = re.sub(r"(处罚|罚则|标准|程序|责任|条款|要求)$", " ", base_text)
    base_text = re.sub(r"\s+", " ", base_text).strip()
    text = re.sub(r"[一-龥]{2,12}(?:省|市|区|县|州|盟|旗|乡|镇|新区|自治州|自治区)", " ", base_text)
    stripped_doc_type = re.sub(r"(管理条例|管理办法|实施办法|议事规则|条例|办法|规定|规则|细则|决定|通知|法规)$", " ", text)
    stripped_doc_type = re.sub(r"\s+", " ", stripped_doc_type).strip()
    text = stripped_doc_type or text or base_text
    text = re.sub(r"(关于|对于|有关|里的|里|中|的|有什么|有哪些|什么|如何|怎么|请问)", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return _normalize_query(text or base_text)


def _query_has_specific_doc_entity(query: str, doc_intent: str = "") -> bool:
    entities = _query_entity_families(query)
    for family in ("region", "version", "organization"):
        if entities.get(family):
            return True
    qnorm = _normalize_query(query)
    for value in entities.get("object", []):
        normalized = _normalize_query(value)
        if not normalized or normalized == _normalize_query(doc_intent):
            continue
        if _topical_suffix_match(normalized) or re.search(r"(处罚|罚则|标准|程序|责任|条款|要求)$", normalized):
            continue
        if normalized not in _GENERIC_DOC_INTENT_TERMS and len(normalized) >= 4 and normalized in qnorm:
            return True
    return False


def _multi_doc_topical_downgrade_allowed(query: str, candidate_sources: List[str]) -> Tuple[bool, Dict[str, Any]]:
    doc_intent = _query_doc_intent(query)
    if not doc_intent:
        return True, {"doc_intent": "", "blocked": False}
    generic_intent = doc_intent in _GENERIC_DOC_INTENT_TERMS
    specific_entity = _query_has_specific_doc_entity(query, doc_intent=doc_intent)
    blocked = generic_intent and not specific_entity
    return (not blocked), {
        "doc_intent": doc_intent,
        "generic_doc_intent": generic_intent,
        "specific_entity_present": specific_entity,
        "candidate_count": len(candidate_sources or []),
        "blocked": blocked,
    }


def _query_matches_source_region_or_landmark(query: str, source: str) -> bool:
    qnorm = _normalize_query(query)
    if not qnorm:
        return False
    title = _source_display_title(source)
    region = _normalize_query(_source_profile_fields(source).get("region") or _extract_region_token(title))
    compact_region = _normalize_query(_strip_region_admin_tokens(region))
    if region and (region in qnorm or (compact_region and compact_region in qnorm)):
        return True
    for entity in _canonical_source_core_entities(source):
        normalized = _normalize_query(entity)
        if not normalized or normalized in _GENERIC_DOC_INTENT_TERMS:
            continue
        if len(normalized) >= 4 and normalized in qnorm:
            return True
    return False


def _is_pseudo_singleton_soft_lock(query: str, source: str) -> bool:
    safe_source = _normalize_filename_for_match(source or "")
    if not safe_source:
        return False
    title = _source_display_title(safe_source)
    region = _normalize_query(_source_profile_fields(safe_source).get("region") or _extract_region_token(title))
    if not region:
        return False
    return not _query_matches_source_region_or_landmark(query, safe_source)


def _resolve_unique_weak_match_upgrade(query: str, candidate_sources: List[str]) -> Dict[str, Any]:
    candidates = _collapse_sources_by_canonical(candidate_sources, limit=5)
    if len(candidates) != 1:
        return {"resolved": False}
    source = candidates[0]
    title = _source_display_title(source)
    overlap = _text_overlap_ratio(query, title)
    edit_sim = _edit_similarity_ratio(query, title)
    score = max(overlap, edit_sim)
    if score < 0.70:
        return {"resolved": False}
    same_title_candidates = _find_same_title_candidates(title, exclude_source=source)
    visible_competitors = [item for item in same_title_candidates if _visible_document_exists(item)]
    if visible_competitors:
        return {"resolved": False, "competition": visible_competitors[:3]}
    if _is_pseudo_singleton_soft_lock(query, source):
        return {
            "resolved": False,
            "blocked_reason": "pseudo_singleton_region_mismatch",
            "candidate": source,
        }
    return {
        "resolved": True,
        "source": source,
        "reason": "soft_lock_unique",
        "confidence": score,
        "lock_message_prefix": f"我理解你查询的是《{title}》……\n",
        "trace": {"overlap_ratio": overlap, "edit_similarity": edit_sim},
    }


def _resolve_topical_suffix_multi_doc(query: str, candidate_sources: List[str]) -> Dict[str, Any]:
    candidates = _collapse_sources_by_canonical(candidate_sources, limit=3)
    if len(candidates) <= 1 or not _is_topical_suffix_query(query):
        return {"resolved": False}
    allowed, gate_trace = _multi_doc_topical_downgrade_allowed(query, candidates)
    if not allowed:
        return {"resolved": False, "blocked_reason": "generic_doc_intent", "trace": gate_trace}
    doc_types = {_normalize_query(_source_profile_fields(src).get("doc_type") or _doc_get(src).get("doc_type") or "") for src in candidates}
    doc_types = {item for item in doc_types if item}
    if len(doc_types) > 1:
        return {"resolved": False}
    return {
        "resolved": True,
        "sources": candidates[:3],
        "reason": "topical_suffix_multi_doc",
        "trace": {"doc_types": list(doc_types), "topical": True, **gate_trace},
    }


def _regulation_identity_key(source: str) -> str:
    safe_source = _normalize_filename_for_match(source or "")
    if not safe_source:
        return ""
    title = _source_display_title(safe_source)
    normalized_title = _normalize_reference_text(re.sub(r"(?:\d{4}[-年]\d{1,2}(?:[-月]\d{1,2})?|\d{4}年|\d{4}版|现行有效|最新)", " ", title))
    region = _normalize_query(_source_profile_fields(safe_source).get("region") or _extract_region_token(title))
    return f"{region}|{normalized_title}".strip("|")


def _source_effective_rank(source: str) -> Tuple[int, int, int, int, str]:
    safe_source = _normalize_filename_for_match(source or "")
    if not safe_source:
        return (0, 0, 0, 0, "")
    info = _doc_get(safe_source)
    profile = _source_profile_fields(safe_source)
    display_title = _source_display_title(safe_source)
    version_label = _normalize_query(profile.get("doc_version_label") or "")
    effective_date = _normalize_query(profile.get("effective_date") or "")
    publish_date = _normalize_query(profile.get("publish_date") or "")
    current_marker = 1 if any(token in f"{display_title} {version_label}" for token in ["现行有效", "最新"]) else 0
    active_version = int(info.get("active_version") or 0)
    effective_sort = int(re.sub(r"\D", "", effective_date) or "0")
    publish_sort = int(re.sub(r"\D", "", publish_date) or "0")
    return (current_marker, effective_sort, publish_sort, active_version, safe_source)


def _prefer_latest_effective_sources(sources: List[str], limit: Optional[int] = None) -> List[str]:
    grouped: Dict[str, str] = {}
    ordered_keys: List[str] = []
    for source in sources or []:
        safe_source = _normalize_filename_for_match(source or "")
        if not safe_source:
            continue
        identity_key = _regulation_identity_key(safe_source) or f"source:{safe_source}"
        if identity_key not in grouped:
            grouped[identity_key] = safe_source
            ordered_keys.append(identity_key)
            continue
        if _source_effective_rank(safe_source) > _source_effective_rank(grouped[identity_key]):
            grouped[identity_key] = safe_source
    out: List[str] = []
    for key in ordered_keys:
        source = grouped.get(key) or ""
        if source:
            out.append(source)
        if limit and len(out) >= max(1, int(limit)):
            break
    return out


def _strip_reference_text_from_query(query: str, references: List[str]) -> str:
    text = _normalize_query(query)
    for reference in references or []:
        normalized = _normalize_query(reference)
        if normalized:
            text = text.replace(normalized, " ")
        raw = (reference or "").strip()
        if raw:
            text = text.replace(raw, " ")
    return _normalize_query(re.sub(r"\s+", " ", text))


def _explicit_content_query(query: str, regulation_mentions: List[str]) -> str:
    stripped = _strip_reference_text_from_query(query, regulation_mentions)
    stripped = re.sub(r"^(关于|对于|有关)", " ", stripped)
    stripped = re.sub(r"\s+", " ", stripped).strip()
    return stripped or _normalize_query(query)


def _geo_filtered_sources(query: str, user_id: str, candidate_sources: List[str]) -> List[str]:
    geo_tokens = _geo_context_tokens(query, user_id)
    if not geo_tokens:
        return []
    matched: List[str] = []
    for source in candidate_sources or []:
        safe_source = _normalize_filename_for_match(source or "")
        if not safe_source:
            continue
        profile = _source_profile_fields(safe_source)
        region = _normalize_query(profile.get("region") or _extract_region_token(_source_display_title(safe_source)))
        if not region:
            continue
        if any(token == region or token in region or region in token for token in geo_tokens):
            if safe_source not in matched:
                matched.append(safe_source)
    return matched


def _prepare_explicit_regulation_candidates(candidate_sources: List[str], limit: int = 5) -> List[str]:
    unique_sources = [_normalize_filename_for_match(source or "") for source in candidate_sources if _normalize_filename_for_match(source or "")]
    unique_sources = list(dict.fromkeys(unique_sources))
    return _prefer_latest_effective_sources(unique_sources, limit=limit)


def _latest_effective_equivalent_source(source: str) -> str:
    safe_source = _normalize_filename_for_match(source or "")
    if not safe_source:
        return ""
    identity_key = _regulation_identity_key(safe_source)
    sibling_sources = [safe_source]
    if identity_key:
        conn = _lex_db_connect()
        rows = conn.execute("SELECT source FROM documents").fetchall()
        for row in rows:
            candidate = _normalize_filename_for_match((row[0] if row else "") or "")
            if not candidate or candidate == safe_source or not _visible_document_exists(candidate):
                continue
            if _regulation_identity_key(candidate) == identity_key:
                sibling_sources.append(candidate)
    else:
        title = _source_display_title(safe_source)
        sibling_sources.extend(_find_same_title_candidates(title, exclude_source=safe_source))
    latest = _prefer_latest_effective_sources(sibling_sources, limit=1)
    return latest[0] if latest else safe_source


def _resolve_contextual_inherited_lock(query: str, route: str, user_id: str) -> Dict[str, Any]:
    if not user_id or route in {"multi_doc_compare", "single_doc_compare", "compare_clarification"}:
        return {"resolved": False}
    q = _normalize_query(query)
    if not q or _extract_filename_candidates(q) or _extract_explicit_regulation_mentions(q):
        return {"resolved": False}
    generic_followup = bool(
        _query_has_strong_business_signal(q)
        and len(q) <= 20
        and not _extract_section_query_targets(q)
        and not re.search(r"[一-龥]{2,12}(?:省|市|区|县|州|盟|旗|乡|镇|新区|自治州|自治区)", q)
    )
    if not _is_contextual_followup_query(q) and not generic_followup:
        return {"resolved": False}
    current = _get_current_locked_document(user_id)
    if not current:
        return {"resolved": False}
    source = _normalize_filename_for_match(current.get("source") or "")
    if not source or not current.get("reliable"):
        return {"resolved": False}
    lock_mode = str(current.get("lock_mode") or "hard_lock")
    if lock_mode not in {"hard_lock", "soft_lock"}:
        return {"resolved": False}
    topic_anchors = _open_topic_anchor_terms(q)
    if generic_followup and topic_anchors and not _source_has_topic_anchor(source, topic_anchors):
        return {"resolved": False}
    title = _doc_get(source).get("canonical_title") or _filename_stem(source) or source
    return {
        "resolved": True,
        "source": source,
        "reason": "context_inherited_lock",
        "lock_mode": lock_mode,
        "confidence": 0.88 if lock_mode == "hard_lock" else 0.76,
        "lock_message_prefix": f"我理解你查询的是《{title}》……\n" if lock_mode == "soft_lock" else "",
        "trace": {"inherited_from": source, "previous_reason": current.get("reason") or ""},
    }


def _open_topic_anchor_terms(query: str) -> List[str]:
    anchors = _query_content_anchor_terms(query, qfilters=None, source_title_terms=[])
    if anchors:
        return anchors
    strong_terms = _query_quality_strong_topic_terms(query)
    if strong_terms:
        return strong_terms
    q = _normalize_query(query)
    fallback_markers = [
        "养犬", "犬只", "携犬", "携犬出户", "河道", "河流", "河湖", "水域", "渔业资源",
        "停车", "违法停车", "违规停放", "象山港",
    ]
    out: List[str] = []
    for marker in fallback_markers:
        if marker in q and marker not in out:
            out.append(marker)
    return out


def _resolve_open_topic_multi_doc(query: str, route: str, candidate_sources: List[str]) -> Dict[str, Any]:
    if not bool(getattr(config, "ENABLE_OPEN_TOPIC_MULTI_DOC", True)):
        return {"resolved": False}
    if route not in {"business_topic_qa", "open_regulation_qa"}:
        return {"resolved": False}
    max_sources = max(2, int(getattr(config, "OPEN_TOPIC_MULTI_DOC_MAX_SOURCES", 3)))
    min_prior = float(getattr(config, "OPEN_TOPIC_MULTI_DOC_MIN_PRIOR", 0.32))

    def _single_source_short_circuit(source: str, confidence: float, basis: str, trace: Dict[str, Any]) -> Dict[str, Any]:
        return {
            "resolved": True,
            "sources": [source],
            "reason": "open_topic_single_source",
            "lock_confidence": confidence,
            "single_source_locked": True,
            "trace": {
                **dict(trace or {}),
                "single_source_short_circuit": True,
                "single_source_basis": basis,
                "chosen_source": source,
                "lock_confidence": confidence,
            },
        }

    topic_anchors = _open_topic_anchor_terms(query)
    selected = [
        source
        for source in _collapse_sources_by_canonical(candidate_sources or [], limit=max_sources * 2)
        if _source_has_topic_anchor(source, topic_anchors)
    ][:max_sources]
    trace: Dict[str, Any] = {"candidate_sources": selected[:max_sources]}
    if len(selected) == 1:
        return _single_source_short_circuit(selected[0], 0.95, "candidate_sources", trace)
    if len(selected) < 2:
        plan = _build_doc_recall_plan(query, limit=max(max_sources * 3, 8))
        picked: List[str] = []
        first_prior = 0.0
        for entry in plan:
            source = _normalize_filename_for_match((entry or {}).get("source") or "")
            if not source:
                continue
            reasons = set((entry or {}).get("reasons") or [])
            prior = float((entry or {}).get("prior") or 0.0)
            if prior < min_prior:
                continue
            if not reasons.intersection({"title_alias_substring", "documents_fts", "doc_term_overlap", "doc_label_overlap"}):
                continue
            if not _source_has_topic_anchor(source, topic_anchors):
                continue
            if not picked:
                first_prior = prior
            if source not in picked:
                picked.append(source)
            if len(picked) >= max_sources:
                break
        selected = _collapse_sources_by_canonical(picked, limit=max_sources)
        trace["doc_recall_plan"] = [
            {
                "source": (item or {}).get("source"),
                "prior": (item or {}).get("prior"),
                "reasons": (item or {}).get("reasons"),
            }
            for item in plan[:max_sources]
        ]
        if len(selected) == 1:
            return _single_source_short_circuit(selected[0], max(0.91, min(0.97, first_prior + 0.58)), "doc_recall_plan", trace)
    if len(selected) < 2 and topic_anchors:
        catalog_ranked: List[Tuple[int, str]] = []
        for source, display_title, probe_text in _dense_title_probe_entries():
            haystack = _normalize_query("\n".join([display_title, probe_text]))
            match_count = 0
            for anchor in topic_anchors:
                if any(variant and variant in haystack for variant in _legal_term_cluster_variants(anchor)):
                    match_count += 1
            if match_count > 0:
                catalog_ranked.append((match_count, source))
        catalog_ranked.sort(key=lambda item: (-item[0], item[1]))
        selected = _collapse_sources_by_canonical([source for _, source in catalog_ranked], limit=max_sources)
        trace["catalog_candidates"] = [source for _, source in catalog_ranked[:max_sources]]
    if len(selected) < 2:
        return {"resolved": False, "trace": trace}
    return {
        "resolved": True,
        "sources": selected[:max_sources],
        "reason": "open_topic_multi_doc",
        "trace": trace,
    }


def _auto_lock_unique_latest_candidate(
    query: str,
    route: str,
    candidates: List[str],
    *,
    target_text: str = "",
    matched_text: str = "",
    strip_title_mentions: bool = True,
    source_lock_kind: str = "auto_latest_effective",
) -> Optional[Dict[str, Any]]:
    safe_candidates = [_normalize_filename_for_match(x or "") for x in (candidates or []) if _normalize_filename_for_match(x or "")]
    safe_candidates = list(dict.fromkeys(safe_candidates))
    if len(safe_candidates) < 2:
        return None
    latest_candidates: List[str] = []
    for src in safe_candidates:
        resolved = _latest_effective_equivalent_source(src) or src
        resolved = _normalize_filename_for_match(resolved)
        if resolved and resolved not in latest_candidates:
            latest_candidates.append(resolved)
    latest_candidates = list(dict.fromkeys(latest_candidates))
    if len(latest_candidates) != 1:
        usable = [s for s in safe_candidates if _doc_searchable_flag(s) == 1]
        if len(usable) == 1:
            latest_candidates = usable
        else:
            return None
    chosen = latest_candidates[0]
    if _doc_searchable_flag(chosen) != 1:
        return None
    return _build_source_resolution_result(
        route=route,
        required=True,
        resolved=True,
        sources=[chosen],
        candidates=safe_candidates[:3],
        reason="auto_latest_effective_unique",
        strip_title_mentions=bool(strip_title_mentions),
        clarification="",
        target_text=matched_text or target_text or chosen,
        lock_mode="hard_lock",
        lock_confidence=1.0,
        source_lock_kind=source_lock_kind,
        source_resolution_trace={
            "auto_lock": True,
            "candidate_sources": safe_candidates[:5],
            "chosen_source": chosen,
        },
    )


def _resolve_query_target_sources(query: str, fnames: Optional[List[str]] = None, user_id: str = "anonymous") -> Dict[str, Any]:
    compare_resolution = _analyze_compare_route(query)
    if compare_resolution.get("is_compare"):
        return _build_source_resolution_result(
            route=compare_resolution.get("route") or "open_topic_compare",
            required=bool(compare_resolution.get("required")),
            resolved=bool(compare_resolution.get("resolved")),
            sources=list(compare_resolution.get("sources") or []),
            candidates=list(compare_resolution.get("sources") or []),
            reason=compare_resolution.get("reason") or "not_needed",
            strip_title_mentions=bool(compare_resolution.get("strip_title_mentions")),
            clarification=compare_resolution.get("clarification") or "",
            target_text=compare_resolution.get("target_text") or "",
            lock_mode="hard_lock" if compare_resolution.get("resolved") else "none",
            lock_confidence=1.0 if compare_resolution.get("resolved") else 0.0,
            source_lock_kind="compare_lock",
            compare_subjects=list(compare_resolution.get("subjects") or []),
            compare_doc_like_subjects=list(compare_resolution.get("doc_like_subjects") or []),
            compare_missing_targets=list(compare_resolution.get("missing_doc_targets") or []),
            compare_common_aspects=list(compare_resolution.get("common_aspects") or []),
            compare_topic_pair=list(compare_resolution.get("topic_pair") or []),
            compare_canonical_aspects=list(compare_resolution.get("canonical_aspects") or []),
            compare_expanded_aspects=list(compare_resolution.get("expanded_aspects") or []),
            compare_source_subqueries=dict(compare_resolution.get("source_subqueries") or {}),
            compare_status=compare_resolution.get("compare_status") or "not_compare",
            compare_plan=dict(compare_resolution.get("compare_plan") or {}),
        )

    explicit_resolution = _resolve_explicit_reference_sources(query, fnames, user_id=user_id)
    if explicit_resolution.get("route") in {"explicit_doc_reference", "explicit_regulation_reference"}:
        return _build_source_resolution_result(
            route=explicit_resolution.get("route") or "",
            required=bool(explicit_resolution.get("required")),
            resolved=bool(explicit_resolution.get("resolved")),
            sources=list(explicit_resolution.get("sources") or []),
            candidates=list(explicit_resolution.get("candidates") or []),
            reason=explicit_resolution.get("reason") or "",
            strip_title_mentions=bool(explicit_resolution.get("strip_title_mentions")),
            clarification=explicit_resolution.get("clarification") or "",
            target_text=explicit_resolution.get("target_text") or "",
            lock_mode=explicit_resolution.get("lock_mode") or ("hard_lock" if explicit_resolution.get("resolved") else "none"),
            lock_confidence=float(explicit_resolution.get("lock_confidence") or (1.0 if explicit_resolution.get("resolved") else 0.0)),
            lock_message_prefix=explicit_resolution.get("lock_message_prefix") or "",
            source_lock_kind=explicit_resolution.get("source_lock_kind") or "explicit_reference",
            source_resolution_trace=dict(explicit_resolution.get("source_resolution_trace") or {}),
            retrieval_query_override=explicit_resolution.get("retrieval_query_override") or "",
        )
    if explicit_resolution.get("route") == "multi_doc_query":
        return _build_source_resolution_result(
            route="multi_doc_query",
            required=False,
            resolved=False,
            sources=list(explicit_resolution.get("sources") or []),
            candidates=list(explicit_resolution.get("candidates") or []),
            reason=explicit_resolution.get("reason") or "topical_suffix_multi_doc",
            strip_title_mentions=bool(explicit_resolution.get("strip_title_mentions")),
            clarification=explicit_resolution.get("clarification") or "",
            target_text=explicit_resolution.get("target_text") or "",
            lock_mode="none",
            source_lock_kind=explicit_resolution.get("source_lock_kind") or "topical_suffix_multi_doc",
            source_resolution_trace=dict(explicit_resolution.get("source_resolution_trace") or {}),
            retrieval_query_override=explicit_resolution.get("retrieval_query_override") or "",
        )

    route = _classify_query_route(query, fnames)
    contextual_doc_ref = _has_contextual_doc_reference(query)
    if not _source_lock_required(query, route):
        return _build_source_resolution_result(
            route=route,
            required=False,
            resolved=False,
            sources=[],
            candidates=[],
            reason="not_needed",
            strip_title_mentions=False,
            clarification="",
            target_text="",
            lock_mode="none",
        )

    inherited_lock = _resolve_contextual_inherited_lock(query, route, user_id)
    if inherited_lock.get("resolved"):
        inherited_source = str(inherited_lock.get("source") or "")
        return _build_source_resolution_result(
            route=route,
            required=True,
            resolved=True,
            sources=[inherited_source],
            candidates=[inherited_source],
            reason=str(inherited_lock.get("reason") or "context_inherited_lock"),
            strip_title_mentions=False,
            clarification="",
            target_text=inherited_source,
            lock_mode=str(inherited_lock.get("lock_mode") or "hard_lock"),
            lock_confidence=float(inherited_lock.get("confidence") or 0.0),
            lock_message_prefix=str(inherited_lock.get("lock_message_prefix") or ""),
            source_lock_kind="context_inherited",
            source_resolution_trace=dict(inherited_lock.get("trace") or {}),
            inherited_from_context=True,
        )

    strong_title_matches = _extract_strong_title_source_matches(
        query,
        limit=max(1, int(_policy_get("source_resolution.title_candidate_limit", 5))),
    )
    title_sources = _collapse_sources_by_canonical(
        [entry.get("source") for entry in strong_title_matches if entry.get("source")],
        limit=max(1, int(_policy_get("source_resolution.title_candidate_limit", 5))),
    )
    if len(title_sources) == 1:
        match_kind = (strong_title_matches[0].get("match_kind") or "alias_title").strip()
        matched_text = (strong_title_matches[0].get("matched_text") or "").strip()
        if match_kind == "alias_title" and not _short_alias_lock_allowed(query, title_sources[0], matched_text or query):
            unique_lock = _unique_alias_lock_resolution(query, title_sources, matched_text or query, ranked_matches=strong_title_matches)
            if unique_lock.get("resolved"):
                locked_source = str(unique_lock.get("source") or "")
                return _build_source_resolution_result(
                    route=route,
                    required=True,
                    resolved=True,
                    sources=[locked_source],
                    candidates=title_sources,
                    reason=str(unique_lock.get("reason") or "unique_alias_lock"),
                    strip_title_mentions=True,
                    clarification="",
                    target_text=matched_text or locked_source,
                    lock_mode="hard_lock",
                    lock_confidence=1.0,
                    source_lock_kind="unique_alias",
                )
            soft_lock = _resolve_soft_lock_candidate(query, route, title_sources)
            if soft_lock.get("resolved"):
                locked_source = str(soft_lock.get("source") or "")
                return _build_source_resolution_result(
                    route=route,
                    required=True,
                    resolved=True,
                    sources=[locked_source],
                    candidates=title_sources,
                    reason=str(soft_lock.get("reason") or "soft_lock"),
                    strip_title_mentions=True,
                    clarification="",
                    target_text=matched_text or locked_source,
                    lock_mode="soft_lock",
                    lock_confidence=float(soft_lock.get("confidence") or 0.0),
                    lock_message_prefix=str(soft_lock.get("lock_message_prefix") or ""),
                    source_lock_kind="soft_lock",
                    source_resolution_trace=dict(soft_lock.get("trace") or {}),
                )
            auto_locked = _auto_lock_unique_latest_candidate(
                query,
                route,
                title_sources,
                target_text=matched_text or title_sources[0],
                matched_text=matched_text,
                strip_title_mentions=True,
                source_lock_kind="auto_latest_effective",
            )
            if auto_locked:
                return auto_locked
            return _build_source_resolution_result(
                route=route,
                required=True,
                resolved=False,
                sources=[],
                candidates=title_sources,
                reason="document_ambiguous",
                strip_title_mentions=False,
                clarification=_build_document_clarification_prompt(title_sources),
                target_text=matched_text or title_sources[0],
                lock_mode="none",
            )
        return _build_source_resolution_result(
            route=route,
            required=True,
            resolved=True,
            sources=title_sources,
            candidates=title_sources,
            reason="exact_title_unique" if match_kind == "exact_title" else "title_alias_unique",
            strip_title_mentions=True,
            clarification="",
            target_text=title_sources[0],
            lock_mode="hard_lock",
            lock_confidence=1.0,
            source_lock_kind="title_unique",
        )
    if contextual_doc_ref and not title_sources:
        return _build_source_resolution_result(
            route=route,
            required=True,
            resolved=False,
            sources=[],
            candidates=[],
            reason="document_target_required",
            strip_title_mentions=False,
            clarification="请先说明你指的是哪一部法规或文件。",
            target_text="",
            lock_mode="none",
        )

    rule_section_targets = _local_validate_section_targets(_extract_section_query_targets(query), limit=6)
    section_anchor_query = bool(rule_section_targets) or _is_section_anchor_query(query)
    candidate_sources = list(title_sources)
    if not candidate_sources:
        fallback = _doc_fallback_source_candidates(
            query,
            limit=max(1, int(_policy_get("source_resolution.fallback_candidate_limit", 3))),
        )
        candidate_sources = list((fallback or {}).get("sources") or [])
        section_anchor_query = section_anchor_query or bool((fallback or {}).get("section_anchor_hit"))
    if not candidate_sources:
        candidate_sources = _normalized_query_title_candidate_sources(
            query,
            limit=max(1, int(_policy_get("source_resolution.fallback_candidate_limit", 3))),
        )
    # 规则层抽不到且即将触发 document_target_required 时，才进入 LLM target fallback。
    if not candidate_sources and not rule_section_targets:
        llm_targets = _llm_extract_section_targets(
            query,
            limit=max(2, int(_policy_get("source_resolution.fallback_candidate_limit", 3))),
        )
        llm_targets = _local_validate_section_targets(llm_targets, limit=5)
        if llm_targets:
            reverse_sources = _sqlite_reverse_lookup_sources_by_targets(
                llm_targets,
                limit=max(1, int(_policy_get("source_resolution.fallback_candidate_limit", 3))),
            )
            if reverse_sources:
                candidate_sources = reverse_sources
                section_anchor_query = True

    candidate_sources = _collapse_sources_by_canonical(candidate_sources, limit=3)
    entity_scored_candidates = _score_source_candidates_by_entities(query, candidate_sources)
    if entity_scored_candidates:
        top_entity = entity_scored_candidates[0]
        top_entity_source = _normalize_filename_for_match(top_entity.get("source") or "")
        if top_entity_source and top_entity.get("exclusive_families") and _source_supports_doc_identity_term(top_entity_source, query):
            return _build_source_resolution_result(
                route=route,
                required=True,
                resolved=True,
                sources=[top_entity_source],
                candidates=candidate_sources,
                reason="exclusive_entity_unique",
                strip_title_mentions=False,
                clarification="",
                target_text=top_entity_source,
                lock_mode="hard_lock",
                lock_confidence=min(1.0, 0.75 + 0.05 * len(top_entity.get("exclusive_families") or [])),
                source_lock_kind="exclusive_entity_unique",
                source_resolution_trace={"entity_scoring": entity_scored_candidates[:3]},
            )
        if len(entity_scored_candidates) > 1 and abs(float(entity_scored_candidates[0].get("score") or 0.0) - float(entity_scored_candidates[1].get("score") or 0.0)) <= 0.35:
            candidate_hint_sources = [str(item.get("source") or "") for item in entity_scored_candidates[:3] if str(item.get("source") or "")]
            auto_locked = _auto_lock_unique_latest_candidate(
                query,
                route,
                candidate_hint_sources,
                target_text="",
                matched_text="",
                strip_title_mentions=True,
                source_lock_kind="auto_latest_effective",
            )
            if auto_locked:
                return auto_locked
            return _build_source_resolution_result(
                route=route,
                required=True,
                resolved=False,
                sources=[],
                candidates=candidate_hint_sources,
                reason="document_ambiguous",
                strip_title_mentions=False,
                clarification=_build_document_clarification_prompt(candidate_hint_sources),
                target_text="",
                lock_mode="none",
                source_lock_kind="candidate_hint",
                source_resolution_trace={"entity_scoring": entity_scored_candidates[:3]},
            )
    unique_lock = _unique_alias_lock_resolution(
        query,
        candidate_sources,
        (strong_title_matches[0].get("matched_text") or query).strip() if strong_title_matches else query,
        ranked_matches=strong_title_matches,
    )
    if unique_lock.get("resolved"):
        locked_source = str(unique_lock.get("source") or "")
        return _build_source_resolution_result(
            route=route,
            required=True,
            resolved=True,
            sources=[locked_source],
            candidates=candidate_sources,
            reason=str(unique_lock.get("reason") or "unique_alias_lock"),
            strip_title_mentions=True,
            clarification="",
            target_text=locked_source,
            lock_mode="hard_lock",
            lock_confidence=1.0,
            source_lock_kind="unique_alias",
            source_resolution_trace={"entity_scoring": entity_scored_candidates[:3]},
        )
    soft_lock = _resolve_soft_lock_candidate(query, route, candidate_sources)
    if soft_lock.get("resolved"):
        locked_source = str(soft_lock.get("source") or "")
        return _build_source_resolution_result(
            route=route,
            required=True,
            resolved=True,
            sources=[locked_source],
            candidates=candidate_sources,
            reason=str(soft_lock.get("reason") or "soft_lock"),
            strip_title_mentions=True,
            clarification="",
            target_text=locked_source,
            lock_mode="soft_lock",
            lock_confidence=float(soft_lock.get("confidence") or 0.0),
            lock_message_prefix=str(soft_lock.get("lock_message_prefix") or ""),
            source_lock_kind="soft_lock",
            source_resolution_trace={
                "entity_scoring": entity_scored_candidates[:3],
                **dict(soft_lock.get("trace") or {}),
            },
        )
    if section_anchor_query and len(candidate_sources) == 1:
        return _build_source_resolution_result(
            route=route,
            required=True,
            resolved=True,
            sources=candidate_sources,
            candidates=candidate_sources,
            reason="section_anchor_unique",
            strip_title_mentions=False,
            clarification="",
            target_text=candidate_sources[0],
            lock_mode="hard_lock",
            lock_confidence=0.92,
            source_lock_kind="section_anchor_unique",
            source_resolution_trace={"entity_scoring": entity_scored_candidates[:3]},
        )
    open_topic_multi = _resolve_open_topic_multi_doc(query, route, candidate_sources)
    if open_topic_multi.get("resolved"):
        if len(list(open_topic_multi.get("sources") or [])) == 1:
            locked_source = str((open_topic_multi.get("sources") or [""])[0] or "")
            return _build_source_resolution_result(
                route=route,
                required=False,
                resolved=True,
                sources=[locked_source],
                candidates=[locked_source],
                reason=str(open_topic_multi.get("reason") or "open_topic_single_source"),
                strip_title_mentions=False,
                clarification="",
                target_text=locked_source,
                lock_mode="hard_lock",
                lock_confidence=float(open_topic_multi.get("lock_confidence") or 0.91),
                source_lock_kind="open_topic_single_source",
                source_resolution_trace=dict(open_topic_multi.get("trace") or {}),
            )
        return _build_source_resolution_result(
            route="multi_doc_query",
            required=False,
            resolved=False,
            sources=list(open_topic_multi.get("sources") or []),
            candidates=list(open_topic_multi.get("sources") or []),
            reason=str(open_topic_multi.get("reason") or "open_topic_multi_doc"),
            strip_title_mentions=False,
            clarification="",
            target_text="",
            lock_mode="none",
            source_lock_kind="open_topic_multi_doc",
            source_resolution_trace=dict(open_topic_multi.get("trace") or {}),
        )
    qnorm = _normalize_query(query)
    if route in {"business_topic_qa", "open_regulation_qa"} and candidate_sources and not section_anchor_query:
        if qnorm and not _query_has_doc_identity_term(qnorm) and not _extract_explicit_regulation_mentions(qnorm) and not _extract_filename_candidates(qnorm):
            if _query_has_strong_business_signal(qnorm) or _query_quality_strong_topic_terms(qnorm):
                return _build_source_resolution_result(
                    route=route,
                    required=False,
                    resolved=False,
                    sources=[],
                    candidates=candidate_sources,
                    reason="tier2_soft_confirm",
                    strip_title_mentions=False,
                    clarification=_build_document_clarification_prompt(candidate_sources),
                    target_text="",
                    lock_mode="none",
                    source_lock_kind="candidate_hint",
                    source_resolution_trace={"entity_scoring": entity_scored_candidates[:3]},
                )
    return _build_source_resolution_result(
        route=route,
        required=True,
        resolved=False,
        sources=[],
        candidates=candidate_sources,
        reason="section_anchor_ambiguous" if section_anchor_query and len(candidate_sources) > 1 else "document_target_required",
        strip_title_mentions=False,
        clarification=_build_document_clarification_prompt(candidate_sources),
        target_text="",
        lock_mode="none",
        source_lock_kind="candidate_hint" if candidate_sources else "none",
        source_resolution_trace={"entity_scoring": entity_scored_candidates[:3]},
    )


def _strip_source_title_mentions(query: str, sources: List[str]) -> str:
    text = _normalize_query(query)
    candidates: List[str] = []
    for source in sources or []:
        for candidate in _doc_title_alias_candidates(source):
            cand = (candidate or "").strip()
            if len(cand) >= 4 and cand not in candidates:
                candidates.append(cand)
    for token in sorted(candidates, key=len, reverse=True):
        text = text.replace(token, " ")
    return _normalize_query(text)


def _doc_recall_fallback(query: str, limit: int, source_filter: Optional[str] = None) -> List[str]:
    conn = _lex_db_connect()
    rows = conn.execute("SELECT filename, title, aliases, doc_type, topic, filename_stem FROM documents_fts").fetchall()
    ranked: List[Tuple[float, str]] = []
    for filename, title, aliases, doc_type, topic, filename_stem in rows:
        source = _normalize_filename_for_match(filename or "")
        if not source:
            continue
        if source_filter and source != source_filter:
            continue
        state = _source_state(source)
        if not state.get("visible"):
            continue
        title_text = "\n".join([title or "", aliases or "", filename_stem or "", source])
        score = _token_overlap_score(query, title_text)
        score += _doc_title_alias_score(source, query)
        if source_filter and source == source_filter:
            score += 2.0
        if score > 0:
            ranked.append((score, source))
    ranked.sort(key=lambda x: (-x[0], x[1]))
    out: List[str] = []
    for _, source in ranked:
        if source not in out:
            out.append(source)
        if len(out) >= int(limit):
            break
    return out


def _is_chinese_query(query: str) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", query or ""))


def _should_use_doc_fallback(query: str, fnames: List[str], query_route: Optional[str] = None) -> bool:
    if not bool(getattr(config, "ENABLE_DOC_FALLBACK", True)):
        return False
    route = query_route or _classify_query_route(query, fnames)
    if route in {"existence", "visibility_probe", "explicit_doc_reference", "explicit_regulation_reference", "exact_title_reference", "alias_title_reference"}:
        return False
    q = _normalize_query(query)
    if not q or not _is_chinese_query(q):
        return False
    if route in {"weak_title_reference", "topic_like_title"}:
        return bool(_extract_title_source_candidates(q, limit=1))
    if route == "business_topic_qa":
        return True
    if route == "open_regulation_qa":
        return bool(_query_match_terms(q))
    if fnames:
        return True
    return False


def _doc_term_overlap_recall(query: str, limit: int, source_filter: Optional[str] = None) -> Dict[str, Dict[str, Any]]:
    conn = _lex_db_connect()
    terms = _query_match_terms(query)
    profile_scores = _profile_source_recall(query, limit=max(int(limit) * 2, 12), source_filter=source_filter)
    if not terms:
        return profile_scores
    where_parts = []
    params: List[Any] = []
    if source_filter:
        where_parts.append("m.source = ?")
        params.append(source_filter)
    like_parts = []
    for term in terms:
        like_parts.append("f.text LIKE ?")
        params.append(f"%{term}%")
    if like_parts:
        where_parts.append("(" + " OR ".join(like_parts) + ")")
    where_sql = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""
    sql = (
        "SELECT m.source, f.text "
        "FROM chunks_meta m JOIN chunks_fts f ON f.rowid = m.id"
        f"{where_sql} LIMIT ?"
    )
    params.append(max(int(limit) * 40, int(getattr(config, "DOC_FALLBACK_CHUNK_SCAN_LIMIT", 400))))
    rows = conn.execute(sql, tuple(params)).fetchall()
    by_src: Dict[str, Dict[str, Any]] = {}
    for source, text in rows:
        src = _normalize_filename_for_match(source or "")
        if not src:
            continue
        state = _source_state(src)
        if not state.get("visible"):
            continue
        matched_terms = [term for term in terms if term in (text or "")]
        if not matched_terms:
            continue
        info = by_src.setdefault(src, {"matched_terms": [], "hit_count": 0, "score": 0.0})
        info["hit_count"] += 1
        for term in matched_terms:
            if term not in info["matched_terms"]:
                info["matched_terms"].append(term)
                info["score"] += max(0.8, min(len(term), 8) / 3.0)
        info["score"] += min(len(matched_terms), 3) * 0.15
    ranked: Dict[str, Dict[str, Any]] = {}
    for src, info in by_src.items():
        coverage = float(len(info["matched_terms"])) / float(max(len(terms), 1))
        ranked[src] = {
            "score": float(info["score"] + coverage),
            "matched_terms": info["matched_terms"],
            "hit_count": int(info["hit_count"]),
            "coverage": coverage,
        }
    for src, info in profile_scores.items():
        cur = ranked.setdefault(src, {"score": 0.0, "matched_terms": [], "hit_count": 0, "coverage": 0.0})
        cur["score"] = float(cur.get("score", 0.0)) + float(info.get("score", 0.0))
        cur["hit_count"] = int(cur.get("hit_count", 0)) + int(info.get("hit_count", 0))
        cur["coverage"] = max(float(cur.get("coverage", 0.0)), float(info.get("coverage", 0.0)))
        for term in info.get("matched_terms") or []:
            if term not in cur["matched_terms"]:
                cur["matched_terms"].append(term)
        if info.get("reasons"):
            cur["reasons"] = list(dict.fromkeys(list(cur.get("reasons") or []) + list(info.get("reasons") or [])))
    return {
        src: info
        for src, info in sorted(ranked.items(), key=lambda item: (-float(item[1].get("score", 0.0)), item[0]))[: int(limit)]
    }


def _build_doc_recall_plan(query: str, limit: int, source_filter: Optional[str] = None) -> List[Dict[str, Any]]:
    conn = _lex_db_connect()
    q = _normalize_query(query)
    if not q:
        return []
    indexed_rank: Dict[str, int] = {}
    try:
        rows = conn.execute(
            "SELECT filename FROM documents_fts WHERE documents_fts MATCH ? ORDER BY rank LIMIT ?",
            (q, max(int(limit) * 3, 12)),
        ).fetchall()
        for idx, (filename,) in enumerate(rows):
            src = _normalize_filename_for_match(filename or "")
            if src and src not in indexed_rank:
                indexed_rank[src] = idx
    except Exception:
        pass
    overlap_scores = _doc_term_overlap_recall(q, limit=max(int(limit) * 2, 12), source_filter=source_filter)
    rows = conn.execute("SELECT filename, title, aliases, doc_type, topic, filename_stem FROM documents_fts").fetchall()
    plan: List[Dict[str, Any]] = []
    for filename, title, aliases, doc_type, topic, filename_stem in rows:
        src = _normalize_filename_for_match(filename or "")
        if not src:
            continue
        if source_filter and src != source_filter:
            continue
        state = _source_state(src)
        if not state.get("visible"):
            continue
        title_text = "\n".join([title or "", aliases or "", filename_stem or "", doc_type or "", topic or "", src])
        title_score = _doc_title_alias_score(src, q)
        label_overlap = _token_overlap_score(q, title_text)
        indexed_score = _rrf(indexed_rank.get(src), max(10, int(limit) * 6)) if src in indexed_rank else 0.0
        overlap_info = overlap_scores.get(src) or {}
        overlap_score = float(overlap_info.get("score", 0.0))
        reasons: List[str] = []
        if title_score > 0:
            reasons.append("title_alias_substring")
        if label_overlap > 0:
            reasons.append("doc_label_overlap")
        if src in indexed_rank:
            reasons.append("documents_fts")
        if overlap_info:
            reasons.append("doc_term_overlap")
        for reason in overlap_info.get("reasons") or []:
            if reason not in reasons:
                reasons.append(reason)
        raw_score = (
            min(title_score / 6.0, 1.0) * 1.25
            + min(label_overlap / 8.0, 1.0) * 0.55
            + min(overlap_score / 8.0, 1.0) * 1.10
            + indexed_score * 8.0
        )
        if source_filter and src == source_filter:
            raw_score += 0.2
        prior = min(raw_score / 2.8, 1.0)
        if prior < float(getattr(config, "DOC_FALLBACK_MIN_PRIOR", 0.18)):
            continue
        plan.append({
            "source": src,
            "prior": prior,
            "raw_score": raw_score,
            "reasons": reasons,
            "title_score": title_score,
            "label_overlap": label_overlap,
            "indexed_rank": indexed_rank.get(src),
            "matched_terms": overlap_info.get("matched_terms") or [],
            "term_overlap_score": overlap_score,
            "term_overlap_hits": int(overlap_info.get("hit_count", 0)),
        })
    plan.sort(key=lambda item: (-float(item.get("prior", 0.0)), -float(item.get("raw_score", 0.0)), item.get("source") or ""))
    return plan[: int(limit)]


def _synthetic_doc_title_hit(source: str, query: str, score: float = 1.0, metadata_updates: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    info = _doc_get(source)
    text = "\n".join([
        info.get("canonical_title") or "",
        info.get("aliases") or "",
        info.get("filename_stem") or _filename_stem(source),
    ]).strip() or source
    ent = {
        "source": source,
        "text": text,
        "metadata": {
            "section": "document_title",
            "doc_type": info.get("doc_type") or "",
            "topic": info.get("topic") or "",
            "title_hit": True,
            "lexical_signal": "title_direct",
            "query": query,
        },
    }
    if metadata_updates:
        ent["metadata"].update(metadata_updates)
    return {"entity": ent, "score": float(score)}


def _lexical_recall_fallback(query: str, limit: int, source_filter: Optional[str] = None) -> List[Dict[str, Any]]:
    conn = _lex_db_connect()
    terms = _query_match_terms(query)
    if not terms:
        return []
    where_parts = []
    params: List[Any] = []
    if source_filter:
        where_parts.append("m.source = ?")
        params.append(source_filter)
    like_parts = []
    for term in terms:
        pattern = f"%{term}%"
        like_parts.append("f.text LIKE ?")
        params.append(pattern)
    if like_parts:
        where_parts.append("(" + " OR ".join(like_parts) + ")")
    where_sql = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""
    sql = (
        "SELECT m.id, f.text, m.source, m.section, m.metadata "
        "FROM chunks_meta m JOIN chunks_fts f ON f.rowid = m.id"
        f"{where_sql} LIMIT ?"
    )
    params.append(max(50, min(int(limit) * 6, 1200)))
    rows = conn.execute(sql, tuple(params)).fetchall()
    texts = [row[1] or "" for row in rows]
    bm25 = _bm25_scores(query, texts)
    ranked: List[Tuple[float, Dict[str, Any]]] = []
    for idx, row in enumerate(rows):
        rid, text, source, section, metadata = row
        md = {}
        try:
            md = json.loads(metadata or "{}")
        except Exception:
            md = {}
        source_name = _normalize_filename_for_match(source or "")
        if not source_name:
            continue
        state = _source_state(source_name)
        ent = {
            "source": source_name,
            "text": text or "",
            "metadata": {**md, "section": section or "", "lexical_signal": md.get("lexical_signal") or "chunk_fallback"},
        }
        hit = {"entity": ent, "score": 0.0}
        if not _hit_matches_source_state(hit, state):
            continue
        score = _token_overlap_score(query, text or "") + (bm25[idx] if idx < len(bm25) else 0.0)
        if score <= 0:
            continue
        ranked.append((score, hit))
    ranked.sort(key=lambda x: x[0], reverse=True)
    return [hit for _, hit in ranked[: int(limit)]]


def _bootstrap_runtime_state_from_legacy_or_milvus():
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    os.makedirs(DATABASE_DIR, exist_ok=True)
    legacy_db = os.path.join(LEGACY_UPLOAD_DIR, "lexical_index.db")
    runtime_db_exists = os.path.exists(LEXICAL_DB_FILE)
    runtime_tasks_exists = os.path.exists(TASKS_FILE)
    if os.path.abspath(UPLOAD_DIR) != os.path.abspath(LEGACY_UPLOAD_DIR) and (not runtime_tasks_exists) and os.path.exists(os.path.join(LEGACY_UPLOAD_DIR, "tasks.json")):
        try:
            with open(os.path.join(LEGACY_UPLOAD_DIR, "tasks.json"), "r", encoding="utf-8") as src, open(TASKS_FILE, "w", encoding="utf-8") as dst:
                dst.write(src.read())
        except Exception:
            pass
    if (not runtime_db_exists) and os.path.exists(legacy_db):
        try:
            with open(legacy_db, "rb") as src, open(LEXICAL_DB_FILE, "wb") as dst:
                dst.write(src.read())
        except Exception:
            pass


def _get_source_lock(source: str) -> threading.Lock:
    s = (source or "").strip()
    if not s:
        s = "__unknown__"
    lock = _SOURCE_LOCKS.get(s)
    if lock is None:
        lock = threading.Lock()
        _SOURCE_LOCKS[s] = lock
    return lock


def _register_source_async_task(source: str, task: asyncio.Task):
    safe = _normalize_filename_for_match(source or "") or "__unknown__"
    bucket = _SOURCE_ASYNC_TASKS.setdefault(safe, set())
    bucket.add(task)

    def _cleanup(_task: asyncio.Task):
        tasks = _SOURCE_ASYNC_TASKS.get(safe)
        if not tasks:
            return
        tasks.discard(_task)
        if not tasks:
            _SOURCE_ASYNC_TASKS.pop(safe, None)

    task.add_done_callback(_cleanup)
    return task


async def _cancel_source_async_tasks(source: str) -> int:
    safe = _normalize_filename_for_match(source or "") or "__unknown__"
    tasks = [task for task in list(_SOURCE_ASYNC_TASKS.get(safe) or set()) if not task.done()]
    if not tasks:
        return 0
    for task in tasks:
        task.cancel()
    await asyncio.gather(*tasks, return_exceptions=True)
    return len(tasks)


def _uploaded_artifact_candidates(source: str) -> List[str]:
    safe = _safe_filename(source)
    stem = os.path.splitext(safe)[0]
    candidates = []
    for root in {UPLOAD_DIR, LEGACY_UPLOAD_DIR}:
        if not root or not os.path.isdir(root):
            continue
        try:
            for name in os.listdir(root):
                if name in {os.path.basename(TASKS_FILE), os.path.basename(LEXICAL_DB_FILE)}:
                    continue
                matches = (
                    name == safe
                    or name.endswith(f"__{safe}")
                    or name == stem
                    or name.startswith(f"{safe}__")
                    or name.startswith(f"{stem}__")
                )
                if matches:
                    candidates.append(os.path.join(root, name))
        except Exception:
            continue
    for task in TASKS.values():
        if _normalize_filename_for_match(task.get("filename") or "") != safe:
            continue
        path = (task.get("path") or "").strip()
        if path:
            candidates.append(path)
    out: List[str] = []
    for item in candidates:
        if item and item not in out:
            out.append(item)
    return out


def _delete_uploaded_artifacts(source: str) -> Dict[str, Any]:
    removed: List[str] = []
    missing: List[str] = []
    failed: List[str] = []
    for path in _uploaded_artifact_candidates(source):
        try:
            if os.path.isdir(path):
                shutil.rmtree(path)
                removed.append(path)
            elif os.path.isfile(path):
                os.remove(path)
                removed.append(path)
            else:
                missing.append(path)
        except FileNotFoundError:
            missing.append(path)
        except Exception:
            failed.append(path)
    return {"removed": removed, "missing": missing, "failed": failed}


def _enqueue_pending_delete(source: str, last_error: str = "", delete_files: bool = True, retry_count: int = 0):
    conn = _lex_db_connect()
    now = datetime.now()
    next_retry_at = now.isoformat()
    conn.execute(
        "INSERT INTO pending_delete_queue(source, enqueued_at, next_retry_at, retry_count, last_error, delete_files) VALUES (?,?,?,?,?,?) "
        "ON CONFLICT(source) DO UPDATE SET enqueued_at=excluded.enqueued_at, next_retry_at=excluded.next_retry_at, last_error=excluded.last_error, delete_files=excluded.delete_files",
        (_safe_filename(source), now.isoformat(), next_retry_at, int(retry_count), last_error or "", 1 if delete_files else 0),
    )
    conn.commit()


def _pending_delete_due(limit: int = 20) -> List[Dict[str, Any]]:
    conn = _lex_db_connect()
    rows = conn.execute(
        "SELECT source, retry_count, last_error, delete_files FROM pending_delete_queue WHERE next_retry_at <= ? ORDER BY enqueued_at ASC LIMIT ?",
        (datetime.now().isoformat(), int(limit)),
    ).fetchall()
    return [
        {"source": row[0], "retry_count": int(row[1] or 0), "last_error": row[2] or "", "delete_files": bool(row[3])}
        for row in rows
    ]


def _complete_pending_delete(source: str):
    conn = _lex_db_connect()
    conn.execute("DELETE FROM pending_delete_queue WHERE source = ?", (_safe_filename(source),))
    conn.commit()


def _reschedule_pending_delete(source: str, retry_count: int, last_error: str):
    conn = _lex_db_connect()
    delay_sec = max(60, int(os.getenv("DELETE_RETRY_INTERVAL_SEC", "3600")))
    next_retry_at = (datetime.now() + timedelta(seconds=delay_sec)).isoformat()
    conn.execute(
        "UPDATE pending_delete_queue SET retry_count = ?, last_error = ?, next_retry_at = ? WHERE source = ?",
        (int(retry_count), last_error or "", next_retry_at, _safe_filename(source)),
    )
    conn.commit()


async def _process_pending_delete_queue_once(limit: int = 10):
    deletions = _pending_delete_due(limit=limit)
    for item in deletions:
        src = _safe_filename(item.get("source") or "")
        if not src:
            continue
        lock = _get_source_lock(src)
        if not lock.acquire(blocking=False):
            continue
        try:
            vector_db = VectorDBService()
            vector_db.connect()
            vector_delete_info = _delete_milvus_document_object(vector_db, src)
            _lex_db_delete_source(src)
            file_cleanup = _delete_uploaded_artifacts(src) if item.get("delete_files", True) else {"removed": [], "missing": [], "failed": []}
            if file_cleanup.get("failed"):
                _doc_upsert(src, status="pending_delete", last_error="artifact_cleanup_failed")
                _lex_db_set_status(src, "pending_delete")
                _reschedule_pending_delete(src, int(item.get("retry_count", 0)) + 1, "artifact_cleanup_failed")
                continue
            _complete_pending_delete(src)
            logger.info(
                "pending_delete_completed: source=%s residual_ids=%s files_removed=%s",
                src,
                vector_delete_info.get("residual_ids_deleted"),
                len(file_cleanup.get("removed") or []),
            )
        except Exception as e:
            _doc_upsert(src, status="pending_delete", last_error=str(e))
            _lex_db_set_status(src, "pending_delete")
            _reschedule_pending_delete(src, int(item.get("retry_count", 0)) + 1, str(e))
        finally:
            try:
                lock.release()
            except Exception:
                pass
def _purge_source_all(source: str):
    """删除某来源在向量库与词面索引中的所有记录"""
    safe = _safe_filename(source)
    try:
        vector_db = VectorDBService()
        vector_db.connect()
        vector_db.client.delete(
            collection_name=vector_db.collection_name,
            filter=f"source == {json.dumps(safe, ensure_ascii=False)}",
        )
    except Exception:
        pass
    try:
        _lex_db_delete_source(safe)
    except Exception:
        pass

class DocumentRequest(BaseModel):
    """文档上传请求"""
    filename: str
    content: str
    metadata: Optional[Dict[str, Any]] = None


class QueryRequest(BaseModel):
    """查询请求"""
    query: str
    user_id: str = "anonymous"
    top_k: int = 10
    enable_rerank: bool = True


class QueryResponse(BaseModel):
    """查询响应"""
    answer: str
    sources: List[Dict[str, Any]] = Field(default_factory=list)
    metadata: Dict[str, Any] = Field(default_factory=dict)
    documents: List[Dict[str, Any]] = Field(default_factory=list)


IR_ELEMENT_TYPES = {
    "title", "heading", "paragraph", "list_item", "table", "figure", "caption",
    "key_value", "code_block", "formula", "sheet", "page_break"
}
IR_PARSER_VERSION = "1.0"
DOCX_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
WP_NS = "http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing"


# ==================== 配置 ====================

class Config:
    """应用配置"""
    APP_ENV = os.getenv("APP_ENV", "").lower()
    MILVUS_HOST = os.getenv("MILVUS_HOST", ("127.0.0.1" if APP_ENV == "test_local" else "milvus"))
    MILVUS_PORT = int(os.getenv("MILVUS_PORT", "19530"))
    MILVUS_USER = os.getenv("MILVUS_USER", "minioadmin")
    MILVUS_PASSWORD = os.getenv("MILVUS_PASSWORD", "minioadmin")
    MILVUS_SECURE = os.getenv("MILVUS_SECURE", "false").lower() == "true"
    
    EMBEDDING_URL = os.getenv("EMBEDDING_SERVICE_URL", ("http://127.0.0.1:8001" if APP_ENV == "test_local" else "http://embedding-service:8000"))
    RERANK_URL = os.getenv("RERANK_SERVICE_URL", ("http://127.0.0.1:8002" if APP_ENV == "test_local" else "http://rerank-service:8000"))
    OCR_SERVICE_URL = os.getenv("OCR_SERVICE_URL", "")
    OCR_MODE = os.getenv("OCR_MODE", "general")
    OCR_LANG = os.getenv("OCR_LANG", "auto")
    OCR_TIMEOUT = int(os.getenv("OCR_TIMEOUT", "60"))
    OCR_SHARED_CONTAINER_DIR = os.getenv("OCR_SHARED_CONTAINER_DIR", "")
    OCR_SHARED_HOST_DIR = os.getenv("OCR_SHARED_HOST_DIR", "")
    PDF_OCR_MAX_TEXT_CHARS_PER_PAGE = float(os.getenv("PDF_OCR_MAX_TEXT_CHARS_PER_PAGE", "300"))
    MAX_FILE_SIZE_MB = int(os.getenv("MAX_FILE_SIZE_MB", "100"))
    MAX_BATCH_TOTAL_SIZE_MB = int(os.getenv("MAX_BATCH_TOTAL_SIZE_MB", "200"))
    MAX_PDF_PAGES = int(os.getenv("MAX_PDF_PAGES", "300"))
    MAX_IMAGE_PIXELS = int(os.getenv("MAX_IMAGE_PIXELS", "40000000"))
    MAX_XLSX_ROWS = int(os.getenv("MAX_XLSX_ROWS", "20000"))
    MAX_XLSX_COLS = int(os.getenv("MAX_XLSX_COLS", "200"))
    MAX_XLSX_SHEETS = int(os.getenv("MAX_XLSX_SHEETS", "50"))
    MIN_PARSE_TEXT_CHARS = int(os.getenv("MIN_PARSE_TEXT_CHARS", "40"))
    MIN_PARSE_QUALITY_SCORE = float(os.getenv("MIN_PARSE_QUALITY_SCORE", "0.35"))
    
    LLM_PROVIDER = os.getenv("LLM_PROVIDER", "openai")
    LLM_MODEL = os.getenv("LLM_MODEL", "qwen2.5-7b-instruct")
    LLM_API_BASE = os.getenv("LLM_API_BASE", "http://ollama:11434/v1")
    LLM_API_KEY = os.getenv("LLM_API_KEY", "")
    LLM_CHAT_COMPLETIONS_URL = os.getenv("LLM_CHAT_COMPLETIONS_URL", "")
    LLM_TIMEOUT = int(os.getenv("LLM_TIMEOUT", "60"))
    LLM_TEMPERATURE = float(os.getenv("LLM_TEMPERATURE", "0.6"))
    LLM_TOP_P = float(os.getenv("LLM_TOP_P", "0.95"))
    LLM_MAX_TOKENS = int(os.getenv("LLM_MAX_TOKENS", "500"))
    LLM_PRESENCE_PENALTY = float(os.getenv("LLM_PRESENCE_PENALTY", "1.5"))
    LLM_EXTRA_BODY = os.getenv("LLM_EXTRA_BODY", "")

    ENABLE_LLM_QUERY_PARSE = os.getenv("ENABLE_LLM_QUERY_PARSE", "true").lower() == "true"
    QUERY_PARSE_MAX_TOKENS = int(os.getenv("QUERY_PARSE_MAX_TOKENS", "260"))
    QUERY_PARSE_CACHE_SIZE = int(os.getenv("QUERY_PARSE_CACHE_SIZE", "512"))
    QUERY_PARSE_MAX_DOCS = int(os.getenv("QUERY_PARSE_MAX_DOCS", "2"))
    QUERY_PARSE_MAX_ANCHORS = int(os.getenv("QUERY_PARSE_MAX_ANCHORS", "1"))
    QUERY_PARSE_MAX_ASPECTS = int(os.getenv("QUERY_PARSE_MAX_ASPECTS", "4"))
    QUERY_PARSE_MAX_SECTION_TARGETS = int(os.getenv("QUERY_PARSE_MAX_SECTION_TARGETS", "4"))

    ENABLE_COMPARE_INTENT_TAG = os.getenv("ENABLE_COMPARE_INTENT_TAG", "true").lower() == "true"

    ENABLE_LLM_EVIDENCE_CHECK = os.getenv("ENABLE_LLM_EVIDENCE_CHECK", "false").lower() == "true"
    LLM_EVIDENCE_CHECK_MAX_CHARS = int(os.getenv("LLM_EVIDENCE_CHECK_MAX_CHARS", "1800"))
    LLM_EVIDENCE_CHECK_MIN_DENSE_REL = float(os.getenv("LLM_EVIDENCE_CHECK_MIN_DENSE_REL", "0.0"))
    DENSE_BACKSTOP_MIN_REL = float(os.getenv("DENSE_BACKSTOP_MIN_REL", "0.58"))
    LOCKED_DOC_RECALL_K = int(os.getenv("LOCKED_DOC_RECALL_K", "60"))
    
    CHUNK_SIZE = int(os.getenv("CHUNK_SIZE", "500"))
    OVERLAP = int(os.getenv("OVERLAP", "100"))
    TOP_K = int(os.getenv("TOP_K", "80"))
    RERANK_TOP_K = int(os.getenv("RERANK_TOP_K", "10"))
    ENABLE_RERANK = os.getenv("ENABLE_RERANK", "true").lower() == "true"
    MIN_QUERY_CHARS = int(os.getenv("MIN_QUERY_CHARS", "2"))
    MAX_QUERY_CHARS = int(os.getenv("MAX_QUERY_CHARS", "800"))
    CONTEXT_TOP_N = int(os.getenv("CONTEXT_TOP_N", "6"))
    CONTEXT_DOC_MAX_CHARS = int(os.getenv("CONTEXT_DOC_MAX_CHARS", "1200"))
    CONTEXT_MAX_CHARS = int(os.getenv("CONTEXT_MAX_CHARS", "8000"))
    EVIDENCE_MAX_TOKENS = int(os.getenv("EVIDENCE_MAX_TOKENS", "6500"))
    MIN_RELEVANCE_SCORE = float(os.getenv("MIN_RELEVANCE_SCORE", "0.25"))
    MAX_RELEVANCE_DISTANCE = float(os.getenv("MAX_RELEVANCE_DISTANCE", "0.8"))
    MIN_EVIDENCE_SCORE = float(os.getenv("MIN_EVIDENCE_SCORE", "0.6"))
    MIN_SUBSTANTIVE_CHUNKS = int(os.getenv("MIN_SUBSTANTIVE_CHUNKS", "1"))
    MIN_RESCUE_SCORE = float(os.getenv("MIN_RESCUE_SCORE", "0.48"))
    ENABLE_SEMANTIC_SOFTENING = os.getenv("ENABLE_SEMANTIC_SOFTENING", "true").lower() == "true"
    SEMANTIC_SOFTENING_MIN_SIM = float(os.getenv("SEMANTIC_SOFTENING_MIN_SIM", "0.85"))
    SEMANTIC_SOFTENING_MAX_TEXT_CHARS = int(os.getenv("SEMANTIC_SOFTENING_MAX_TEXT_CHARS", "1200"))
    ENABLE_DENSE_TITLE_FALLBACK = os.getenv("ENABLE_DENSE_TITLE_FALLBACK", "true").lower() == "true"
    DENSE_TITLE_MATCH_MIN_SIM = float(os.getenv("DENSE_TITLE_MATCH_MIN_SIM", "0.84"))
    DENSE_TITLE_MATCH_MARGIN = float(os.getenv("DENSE_TITLE_MATCH_MARGIN", "0.03"))
    DENSE_TITLE_PROBE_MAX_CHARS = int(os.getenv("DENSE_TITLE_PROBE_MAX_CHARS", "160"))
    ENABLE_OPEN_TOPIC_MULTI_DOC = os.getenv("ENABLE_OPEN_TOPIC_MULTI_DOC", "true").lower() == "true"
    OPEN_TOPIC_MULTI_DOC_MIN_PRIOR = float(os.getenv("OPEN_TOPIC_MULTI_DOC_MIN_PRIOR", "0.32"))
    OPEN_TOPIC_MULTI_DOC_MAX_SOURCES = int(os.getenv("OPEN_TOPIC_MULTI_DOC_MAX_SOURCES", "3"))
    PARTIAL_TERM_RESCUE_RELAX_RATIO = float(os.getenv("PARTIAL_TERM_RESCUE_RELAX_RATIO", "0.9"))
    PARTIAL_TERM_RESCUE_MIN_SUBSTANTIVE_CHUNKS = int(os.getenv("PARTIAL_TERM_RESCUE_MIN_SUBSTANTIVE_CHUNKS", "3"))
    PARTIAL_TERM_RESCUE_MIN_FOCUS_SCORE = float(os.getenv("PARTIAL_TERM_RESCUE_MIN_FOCUS_SCORE", "0.72"))
    PRIMARY_EVIDENCE_TOPK = int(os.getenv("PRIMARY_EVIDENCE_TOPK", "5"))
    RESCUE_EVIDENCE_TOPK = int(os.getenv("RESCUE_EVIDENCE_TOPK", "15"))
    ALLOW_GUARDED_FULL = os.getenv("ALLOW_GUARDED_FULL", "true").lower() == "true"
    REQUIRE_EVIDENCE = os.getenv("REQUIRE_EVIDENCE", "false").lower() == "true"
    RECALL_TOP_K = int(os.getenv("RECALL_TOP_K", "20"))
    RECALL_RELATIVE_SCORE_RATIO = float(os.getenv("RECALL_RELATIVE_SCORE_RATIO", "0.72"))
    RECALL_MIN_KEEP_N = int(os.getenv("RECALL_MIN_KEEP_N", "3"))
    RERANK_KEEP_N = int(os.getenv("RERANK_KEEP_N", "8"))
    FINAL_CONTEXT_N = int(os.getenv("FINAL_CONTEXT_N", "10"))
    FINAL_CONTEXT_N_MAX = int(os.getenv("FINAL_CONTEXT_N_MAX", "10"))
    MAX_SNIPPETS_PER_SOURCE = int(os.getenv("MAX_SNIPPETS_PER_SOURCE", "2"))
    SECTION_MAX_CHARS = int(os.getenv("SECTION_MAX_CHARS", "1600"))
    ANSWER_MAX_POINTS = int(os.getenv("ANSWER_MAX_POINTS", "5"))
    LLM_MAX_TOKENS_DEF = int(os.getenv("LLM_MAX_TOKENS_DEF", "220"))
    LLM_MAX_TOKENS_SUMMARY = int(os.getenv("LLM_MAX_TOKENS_SUMMARY", "260"))
    LLM_MAX_TOKENS_HOWTO = int(os.getenv("LLM_MAX_TOKENS_HOWTO", "320"))
    LLM_MAX_TOKENS_COMPARE = int(os.getenv("LLM_MAX_TOKENS_COMPARE", "320"))
    LLM_MAX_TOKENS_OTHER = int(os.getenv("LLM_MAX_TOKENS_OTHER", "260"))
    LLM_MAX_TOKENS_ARCH = int(os.getenv("LLM_MAX_TOKENS_ARCH", "300"))
    SOURCES_MAX_DISTANCE_ADD = float(os.getenv("SOURCES_MAX_DISTANCE_ADD", "0.08"))
    SOURCES_MIN_SCORE_RATIO = float(os.getenv("SOURCES_MIN_SCORE_RATIO", "0.6"))
    # 检索融合与召回配置
    FUSION_ALPHA = float(os.getenv("FUSION_ALPHA", "0.5"))
    LEXICAL_RECALL_LIMIT = int(os.getenv("LEXICAL_RECALL_LIMIT", "1000"))
    DISPLAY_SCORE_RATIO = float(os.getenv("DISPLAY_SCORE_RATIO", "0.8"))
    DISPLAY_DISTANCE_MARGIN = float(os.getenv("DISPLAY_DISTANCE_MARGIN", "0.02"))
    TEST_LEX_ONLY = os.getenv("TEST_LEX_ONLY", "false").lower() == "true"
    RRF_K = int(os.getenv("RRF_K", "60"))
    FUSION_W_DENSE = float(os.getenv("FUSION_W_DENSE", "0.80"))
    FUSION_W_LEX = float(os.getenv("FUSION_W_LEX", "0.12"))
    FUSION_W_PRIOR = float(os.getenv("FUSION_W_PRIOR", "0.002"))
    FUSION_W_DOC_PRIOR = float(os.getenv("FUSION_W_DOC_PRIOR", "0.003"))
    FUSION_M_TERM = float(os.getenv("FUSION_M_TERM", "1.08"))
    FUSION_M_TITLE = float(os.getenv("FUSION_M_TITLE", "1.35"))
    FUSION_M_DOC_RECALL = float(os.getenv("FUSION_M_DOC_RECALL", "1.2"))
    FUSION_M_AGREEMENT = float(os.getenv("FUSION_M_AGREEMENT", "1.12"))
    FUSION_MUD_SCORE = float(os.getenv("FUSION_MUD_SCORE", "0.018"))
    DENSE_BACKSTOP_MIN_SCORE = float(os.getenv("DENSE_BACKSTOP_MIN_SCORE", "0.55"))
    QUERY_ANCHOR_DENSE_BYPASS_MIN_SCORE = float(os.getenv("QUERY_ANCHOR_DENSE_BYPASS_MIN_SCORE", "0.6"))
    RETRIEVAL_CANDIDATE_K = int(os.getenv("RETRIEVAL_CANDIDATE_K", "60"))
    CHUNK_RERANK_KEEP_N = int(os.getenv("CHUNK_RERANK_KEEP_N", "18"))
    CHUNK_RERANK_POOL_N = int(os.getenv("CHUNK_RERANK_POOL_N", "60"))
    SOURCE_RERANK_KEEP_N = int(os.getenv("SOURCE_RERANK_KEEP_N", "6"))
    ENABLE_CHUNK_RERANK = os.getenv("ENABLE_CHUNK_RERANK", "true").lower() == "true"
    RERANK_LOW_CONF_ONLY = os.getenv("RERANK_LOW_CONF_ONLY", "true").lower() == "true"
    RERANK_SOURCE_SCORE_GAP = float(os.getenv("RERANK_SOURCE_SCORE_GAP", "0.04"))
    HYBRID_STRUCT_W_SECTION_TERM = float(os.getenv("HYBRID_STRUCT_W_SECTION_TERM", "0.24"))
    HYBRID_STRUCT_W_TEXT_TERM = float(os.getenv("HYBRID_STRUCT_W_TEXT_TERM", "0.18"))
    HYBRID_STRUCT_W_SECTION_OVERLAP = float(os.getenv("HYBRID_STRUCT_W_SECTION_OVERLAP", "0.22"))
    HYBRID_STRUCT_W_KEYWORD = float(os.getenv("HYBRID_STRUCT_W_KEYWORD", "0.18"))
    HYBRID_STRUCT_W_TITLE = float(os.getenv("HYBRID_STRUCT_W_TITLE", "0.10"))
    HYBRID_STRUCT_W_BASE = float(os.getenv("HYBRID_STRUCT_W_BASE", "0.08"))
    HYBRID_STRUCT_FOLLOW_BONUS = float(os.getenv("HYBRID_STRUCT_FOLLOW_BONUS", "0.16"))
    HYBRID_STRUCT_FOLLOW_WINDOW = int(os.getenv("HYBRID_STRUCT_FOLLOW_WINDOW", "3"))
    HYBRID_STRUCT_GENERIC_PENALTY = float(os.getenv("HYBRID_STRUCT_GENERIC_PENALTY", "0.18"))
    HYBRID_STRUCT_SECTION_MATCH_BONUS = float(os.getenv("HYBRID_STRUCT_SECTION_MATCH_BONUS", "0.22"))
    HYBRID_STRUCT_SECTION_MISMATCH_PENALTY = float(os.getenv("HYBRID_STRUCT_SECTION_MISMATCH_PENALTY", "0.12"))
    WEAK_QUERY_MAX_CHARS = int(os.getenv("WEAK_QUERY_MAX_CHARS", "18"))
    WEAK_QUERY_DOC_LIMIT = int(os.getenv("WEAK_QUERY_DOC_LIMIT", "6"))
    WEAK_QUERY_EXPANSION_LIMIT = int(os.getenv("WEAK_QUERY_EXPANSION_LIMIT", "3"))
    ENABLE_DOC_FALLBACK = os.getenv("ENABLE_DOC_FALLBACK", "true").lower() == "true"
    DOC_FALLBACK_SOURCE_LIMIT = int(os.getenv("DOC_FALLBACK_SOURCE_LIMIT", "6"))
    DOC_FALLBACK_CHUNK_SCAN_LIMIT = int(os.getenv("DOC_FALLBACK_CHUNK_SCAN_LIMIT", "400"))
    DOC_FALLBACK_MIN_PRIOR = float(os.getenv("DOC_FALLBACK_MIN_PRIOR", "0.18"))
    FUSION_W_DOC_PRIOR = float(os.getenv("FUSION_W_DOC_PRIOR", "0.003"))
    TITLE_CONSTRAINT_BOOST = float(os.getenv("TITLE_CONSTRAINT_BOOST", "1.08"))
    TITLE_CONSTRAINT_PENALTY = float(os.getenv("TITLE_CONSTRAINT_PENALTY", "0.82"))
    CONTEXTUAL_PREV_CHARS = int(os.getenv("CONTEXTUAL_PREV_CHARS", "120"))
    CONTEXTUAL_NEXT_CHARS = int(os.getenv("CONTEXTUAL_NEXT_CHARS", "120"))


SUPPORTED_FILE_EXTENSIONS = {
    ".txt", ".md", ".markdown", ".csv", ".json", ".log",
    ".pdf", ".docx", ".xlsx",
    ".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".gif"
}

_TEXT_LIKE_EXTENSIONS = {".txt", ".md", ".markdown", ".csv", ".json", ".log"}
_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".gif"}
_EXTENSION_EQUIVALENTS = {
    ".jpeg": ".jpg",
    ".jpg": ".jpg",
    ".tif": ".tiff",
    ".tiff": ".tiff",
    ".markdown": ".md",
    ".md": ".md",
}
_ALLOWED_MIME_BY_EXTENSION = {
    ".pdf": {"application/pdf"},
    ".docx": {"application/vnd.openxmlformats-officedocument.wordprocessingml.document", "application/zip"},
    ".xlsx": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/zip"},
    ".json": {"application/json", "text/plain"},
    ".csv": {"text/csv", "application/csv", "text/plain", "application/vnd.ms-excel"},
    ".txt": {"text/plain"},
    ".md": {"text/plain", "text/markdown"},
    ".log": {"text/plain"},
    ".png": {"image/png"},
    ".jpg": {"image/jpeg"},
    ".bmp": {"image/bmp"},
    ".gif": {"image/gif"},
    ".tiff": {"image/tiff"},
}


config = Config()
_load_tasks()
_lex_db_init()


def _normalize_query(query: str) -> str:
    q = (query or "").strip()
    q = " ".join(q.split())
    return q


def _blocked_reason(query: str) -> str:
    q = (query or "").lower()
    triggers = [
        "ignore previous",
        "system prompt",
        "developer message",
        "reveal",
        "api key",
        "token",
        "密钥",
        "系统提示词",
        "开发者消息",
        "忽略之前",
        "越狱",
        "jailbreak",
    ]
    for t in triggers:
        if t in q:
            return "blocked_prompt_injection"
    return ""


def _query_has_repeated_noise(query: str) -> bool:
    q = _normalize_query(query)
    compact = re.sub(r"\s+", "", q)
    if not compact:
        return False
    if len(compact) >= 4 and len(set(compact)) == 1:
        return True
    runs = re.findall(r"(.)\1{3,}", compact)
    if runs:
        return True
    meaningful = re.findall(r"[A-Za-z0-9\u4e00-\u9fff]", compact)
    if meaningful and len(set(meaningful)) <= 2 and len(compact) >= 6:
        return True
    return False


def _query_quality_strong_topic_terms(query: str) -> List[str]:
    q = _normalize_query(query)
    if not q:
        return []
    domain_markers = [
        "养犬", "犬只", "携犬", "携犬出户", "非遗", "非物质文化遗产", "传承人", "传承", "消防", "绿化", "绿地", "树木", "砍伐", "修剪",
        "地方立法", "立法", "法规案", "草案", "审议", "出租房", "出租人", "承租人", "行政执法", "城市管理", "执法",
        "违法行为", "监督检查", "法律责任", "奖励与处罚", "行政处罚", "登记", "免疫", "申请程序", "扶持措施", "安全责任", "管理职责",
        "建设要求", "执法措施", "禁止性规定", "公共安全", "区划管理", "预拌混凝土", "现场搅拌", "散装水泥",
    ]
    weak_business_markers = {
        "规定", "条例", "办法", "管理", "处罚", "责任",
        "要求", "法规", "规则", "条款", "程序", "条件", "标准",
        "文件", "文档", "问题", "内容",
    }
    filler_terms = {
        "什么", "哪些", "怎么", "如何", "是否", "有没有", "请问",
        "有关", "相关", "关于", "涉及", "一下", "一下子", "帮我", "帮忙",
        "查询", "检索", "看看", "查查", "说说", "问问",
    }
    cleaned_terms: List[str] = []
    for term in _query_anchor_terms(q):
        token = re.sub(r"^(什么|哪些|怎么|如何|是否|有没有|请问|关于|有关|相关|涉及|帮我|帮忙)", "", term).strip()
        token = re.sub(r"(什么|哪些|怎么|如何|是否|有没有|请问)$", "", token).strip()
        token = token.replace("相关", "").strip()
        if len(token) < 2:
            continue
        if token in weak_business_markers or token in filler_terms:
            continue
        if not any(marker in token for marker in domain_markers):
            continue
        if token not in cleaned_terms:
            cleaned_terms.append(token)
    return cleaned_terms


def _query_has_specific_regulation_title_signal(query: str) -> bool:
    q = _normalize_query(query)
    if not q:
        return False
    suffixes = ("条例", "办法", "规定", "规则", "细则", "决定", "通知")
    generic_title_parts = [
        "条例", "办法", "规定", "规则", "细则", "决定", "通知",
        "管理", "处罚", "责任", "要求", "程序", "条件", "标准", "法规", "文件", "文档", "相关",
        "什么", "哪些", "怎么", "如何", "是否", "有没有", "请问", "测试", "一下", "看看", "查一下", "查查",
    ]
    domain_markers = [
        "养犬", "犬只", "非遗", "非物质文化遗产", "传承", "消防", "绿化", "绿地", "树木", "砍伐", "修剪",
        "地方立法", "立法", "法规案", "草案", "审议", "出租房", "出租人", "承租人", "行政执法", "城市管理", "执法",
        "违法行为", "监督检查", "法律责任", "奖励与处罚", "登记", "免疫", "申请程序", "扶持措施", "安全责任", "管理职责",
        "建设要求", "执法措施", "禁止性规定", "公共安全", "区划管理", "城市绿化",
    ]
    region_markers = ["省", "市", "区", "县", "州", "盟", "旗", "乡", "镇", "村"]
    for mention in _extract_explicit_regulation_mentions(q):
        if _query_has_noise_prefix(mention):
            continue
        core = mention.strip()
        for suffix in suffixes:
            if core.endswith(suffix):
                core = core[: -len(suffix)]
                break
        for part in generic_title_parts:
            core = core.replace(part, "")
        core = core.strip()
        if len(re.findall(r"[\u4e00-\u9fff]", core)) < 2:
            continue
        if any(marker in core for marker in domain_markers) or any(marker in core for marker in region_markers):
            return True
    return False


def _query_has_strong_business_signal(query: str) -> bool:
    q = _normalize_query(query)
    if not q:
        return False
    if _extract_filename_candidates(q):
        return True
    if _query_has_specific_regulation_title_signal(q):
        return True
    if _extract_strong_title_source_matches(_strip_query_intent_phrases(q), limit=1):
        return True
    return bool(_query_quality_strong_topic_terms(q))


def _query_has_weak_business_signal(query: str) -> bool:
    q = _normalize_query(query)
    if not q:
        return False
    weak_business_markers = ["规定", "条例", "办法", "管理", "处罚", "责任", "要求", "法规"]
    return any(marker in q for marker in weak_business_markers)


def _query_has_business_signal(query: str) -> bool:
    return _query_has_strong_business_signal(query)


def _query_has_gate_intent_signal(query: str) -> bool:
    q = _normalize_query(query)
    intent_markers = ["什么", "哪些", "怎么", "如何", "是否", "有没有", "请问"]
    return any(marker in q for marker in intent_markers)


def _query_has_noise_prefix(query: str) -> bool:
    compact = re.sub(r"\s+", "", _normalize_query(query).lower())
    if not compact:
        return False
    noise_prefixes = [
        "测试", "测试一下", "随便", "随便问问", "随便查一下", "随便搜搜",
        "看看", "查一下", "查查", "这个呢", "这个文件",
    ]
    return any(compact.startswith(prefix) for prefix in noise_prefixes)


def _query_has_english_gibberish(query: str) -> bool:
    q = _normalize_query(query)
    if not q or _query_has_strong_business_signal(q):
        return False
    ascii_runs = re.findall(r"[A-Za-z0-9]+", q)
    if not ascii_runs:
        return False
    ascii_chars = re.findall(r"[A-Za-z0-9]", q)
    chinese_chars = re.findall(r"[\u4e00-\u9fff]", q)
    if any(len(run) >= 5 for run in ascii_runs):
        return True
    return bool(ascii_chars) and len(ascii_chars) >= max(4, len(chinese_chars) * 2)


def _query_has_multi_doc_rescue(query: str) -> bool:
    q = _normalize_query(query)
    if not q:
        return False
    if _classify_question_type(q) == "screening":
        return True
    multi_doc_markers = [
        "哪些文档", "哪些条例", "哪些文件", "哪些法规", "库里有哪些", "找一下涉及",
        "涉及哪些", "提到", "相关文档", "相关条例", "相关文件",
    ]
    return any(marker in q for marker in multi_doc_markers)


def _query_has_compare_rescue(query: str) -> bool:
    q = _normalize_query(query)
    if not q:
        return False
    return _query_has_compare_intent(q)


def _query_has_clarification_rescue(query: str) -> bool:
    q = _normalize_query(query)
    if not q:
        return False
    if _has_contextual_doc_reference(q):
        return True
    if _query_has_gate_intent_signal(q) and _is_generic_document_required_query(q):
        return True
    if _query_has_gate_intent_signal(q):
        mentions = _extract_explicit_regulation_mentions(q)
        if mentions and not _query_has_specific_regulation_title_signal(q):
            return True
    clarification_targets = [
        "处罚条款", "法律责任", "禁止性规定", "管理职责", "监督检查", "申请程序",
        "扶持措施", "安全责任", "建设要求", "执法措施", "违法行为", "奖励与处罚",
    ]
    return _query_has_gate_intent_signal(q) and any(target in q for target in clarification_targets)


def _query_has_business_topic_rescue(query: str) -> bool:
    q = _normalize_query(query)
    if not q:
        return False
    business_topic_markers = [
        "养犬", "犬只", "携犬", "携犬出户", "非遗", "非物质文化遗产", "传承人", "传承", "消防", "绿化", "绿地", "树木", "砍伐", "修剪",
        "地方立法", "立法", "法规案", "草案", "审议", "公开征求意见", "表决程序",
        "出租房", "出租人", "承租人", "行政执法", "城市管理", "执法", "违法行为", "监督检查", "法律责任", "行政处罚",
        "奖励与处罚", "登记", "免疫", "申请程序", "扶持措施", "安全责任", "管理职责", "政府职责",
        "建设要求", "执法措施", "禁止性规定", "公共安全", "区划管理", "保护措施", "管理要求", "监督机制",
    ]
    if not any(marker in q for marker in business_topic_markers):
        return False
    if _query_has_gate_intent_signal(q):
        return True
    if _query_has_multi_doc_rescue(q) or _query_has_compare_rescue(q):
        return True
    return any(marker in q for marker in ["规定", "要求", "处罚", "责任", "程序", "条件", "标准", "管理", "建设", "扶持"])


def _query_has_lockable_doc_rescue(query: str) -> bool:
    q = _normalize_query(query)
    if not q:
        return False
    if _extract_filename_candidates(q):
        return True
    if _query_has_specific_regulation_title_signal(q):
        return True
    if _extract_strong_title_source_matches(_strip_query_intent_phrases(q), limit=1):
        return True
    return bool(_normalized_query_title_candidate_sources(q, limit=1))


def _query_has_route_rescue_value(query: str) -> bool:
    q = _normalize_query(query)
    if not q:
        return False
    return any([
        _query_has_lockable_doc_rescue(q),
        _query_has_clarification_rescue(q),
        _query_has_multi_doc_rescue(q),
        _query_has_compare_rescue(q),
        _query_has_business_topic_rescue(q),
    ])


def _query_is_out_of_domain(query: str) -> bool:
    q = _normalize_query(query)
    if not q:
        return False
    if _query_has_strong_business_signal(q) or _query_has_weak_business_signal(q):
        return False
    compact = re.sub(r"\s+", "", q)
    if len(compact) <= 4 and not _query_has_gate_intent_signal(q):
        return False
    return True


def _query_is_only_generic_terms(query: str) -> bool:
    q = _normalize_query(query)
    if not q or _query_has_strong_business_signal(q):
        return False
    if not _query_has_weak_business_signal(q):
        return False
    stripped = re.sub(r"\s+", "", q)
    for phrase in ["什么", "哪些", "怎么", "如何", "是否", "有没有", "请问", "一下", "呢", "吗", "呀", "啊", "的"]:
        stripped = stripped.replace(phrase, "")
    weak_business_markers = ["规定", "条例", "办法", "管理", "处罚", "责任", "要求", "法规", "规则", "条款", "程序", "条件", "标准"]
    if not stripped:
        return False
    while True:
        replaced = stripped
        for marker in weak_business_markers:
            replaced = replaced.replace(marker, "")
        if replaced == stripped:
            break
        stripped = replaced
    return not stripped


def _query_has_intent_signal(query: str) -> bool:
    q = _normalize_query(query)
    intent_markers = [
        "什么", "哪些", "如何", "怎么", "为何", "是否", "规定", "要求", "处罚", "责任",
        "程序", "条件", "标准", "管理", "建设", "扶持", "条款", "章节", "办法",
        "条例", "规则", "细则", "决定", "通知", "问", "请问",
    ]
    return any(marker in q for marker in intent_markers)


def _query_intent_tier_state(
    query: str,
    llm_parse: Optional[Dict[str, Any]] = None,
    source_resolution: Optional[Dict[str, Any]] = None,
) -> Dict[str, str]:
    q = _normalize_query(query)
    static_state = _query_static_quality_state(q)
    if static_state["reason"] in {"invalid_query", "garbage_query"}:
        return {**static_state, "tier": "", "label": static_state["reason"]}

    parsed = llm_parse if isinstance(llm_parse, dict) else {}
    resolved = source_resolution if isinstance(source_resolution, dict) else {}
    documents = [
        _normalize_query(str(item or ""))
        for item in (parsed.get("documents") or [])
        if _normalize_query(str(item or ""))
    ]
    retrieval_query = _normalize_query(str(parsed.get("retrieval_query") or "")) or q
    deep_query = retrieval_query or q
    compact = re.sub(r"\s+", "", deep_query)
    route = str(resolved.get("route") or "")

    explicit_title_route = route in {"exact_title_reference", "alias_title_reference", "weak_title_reference", "version_switch"}
    explicit_title_signal = any([
        bool(_extract_filename_candidates(q)),
        bool(_extract_explicit_regulation_mentions(q)),
        bool(_query_has_specific_regulation_title_signal(q)),
        bool(_extract_strong_title_source_matches(_strip_query_intent_phrases(q), limit=1)),
    ])
    if documents or resolved.get("resolved") or explicit_title_route or explicit_title_signal:
        return {"quality": "valid", "reason": "", "tier": "tier_1", "label": "explicit_source"}

    if (not _query_has_strong_business_signal(deep_query)) and _query_is_out_of_domain(deep_query):
        return {"quality": "out_of_domain", "reason": "out_of_domain_query", "tier": "", "label": "out_of_domain"}

    if _query_is_only_generic_terms(deep_query):
        return {"quality": "low_information", "reason": "low_information_query", "tier": "", "label": "generic_only"}

    if _query_has_strong_business_signal(deep_query) or _query_quality_strong_topic_terms(deep_query):
        return {"quality": "valid", "reason": "", "tier": "tier_2", "label": "strong_topic"}

    weak_summary_ready = bool(_clarification_probe_terms(deep_query)) and len(compact) > 4
    if _query_has_weak_business_signal(deep_query) and weak_summary_ready:
        return {"quality": "valid", "reason": "", "tier": "tier_3", "label": "weak_topic_summary"}

    if len(compact) <= 4 and (not _query_has_strong_business_signal(deep_query)):
        return {"quality": "low_information", "reason": "low_information_query", "tier": "", "label": "too_short"}

    if (not _query_has_strong_business_signal(deep_query)) and (not _query_has_gate_intent_signal(deep_query)):
        return {"quality": "low_information", "reason": "low_information_query", "tier": "", "label": "missing_intent"}

    if _query_has_weak_business_signal(deep_query):
        return {"quality": "valid", "reason": "", "tier": "tier_3", "label": "weak_topic_summary"}

    return {"quality": "valid", "reason": "", "tier": "tier_2", "label": "route_rescue" if _query_has_route_rescue_value(deep_query) else "default_valid"}


def _query_static_quality_state(query: str) -> Dict[str, str]:
    q = _normalize_query(query)
    if not q:
        return {"quality": "invalid", "reason": "invalid_query"}

    compact = re.sub(r"\s+", "", q)
    if _query_has_repeated_noise(q):
        return {"quality": "garbage", "reason": "garbage_query"}

    if _query_has_english_gibberish(q):
        return {"quality": "garbage", "reason": "garbage_query"}

    known_noise = ["随便问问", "随便", "测试", "asdf", "abcde", "hello", "哈哈", "呵呵"]
    if any(token in q.lower() for token in known_noise) and not _query_has_strong_business_signal(q):
        return {"quality": "garbage", "reason": "garbage_query"}

    if _query_has_noise_prefix(q) and not _query_has_strong_business_signal(q):
        return {"quality": "garbage", "reason": "garbage_query"}

    if len(compact) <= 4 and (not _query_has_gate_intent_signal(q)) and (not _query_has_compare_intent(q)):
        return {"quality": "low_information", "reason": "low_information_query"}

    return {"quality": "valid", "reason": ""}


def _query_deep_quality_state(
    query: str,
    llm_parse: Optional[Dict[str, Any]] = None,
    source_resolution: Optional[Dict[str, Any]] = None,
) -> Dict[str, str]:
    intent_state = _query_intent_tier_state(query, llm_parse=llm_parse, source_resolution=source_resolution)
    return {"quality": intent_state.get("quality") or "valid", "reason": intent_state.get("reason") or "", "tier": intent_state.get("tier") or ""}


def _query_quality_state(query: str) -> Dict[str, str]:
    return _query_deep_quality_state(query)


def _invalid_query_message(reason: str) -> str:
    if reason in {"invalid_query", "garbage_query", "low_information_query"}:
        return "当前问题缺少可检索的有效主题，请补充法规名称、业务主题或具体问题。"
    if reason == "out_of_domain_query":
        return "当前问题不在法规知识库问答范围内，请改为询问法规、文件或相关业务主题。"
    return "请提供更具体的问题描述。"


def _hit_entity_text(hit: Any) -> str:
    if isinstance(hit, dict):
        ent = hit.get("entity") or {}
        if isinstance(ent, dict):
            return ent.get("text") or ""
        return ""
    ent = getattr(hit, "entity", None)
    if isinstance(ent, dict):
        return ent.get("text") or ""
    return ""


def _hit_display_text(hit: Any) -> str:
    md = _hit_metadata(hit)
    raw = (md.get("raw_text") or "").strip()
    if raw:
        return raw
    return _hit_entity_text(hit)


def _hit_entity_source(hit: Any) -> str:
    if isinstance(hit, dict):
        ent = hit.get("entity") or {}
        if isinstance(ent, dict):
            return ent.get("source") or ""
        return ""
    ent = getattr(hit, "entity", None)
    if isinstance(ent, dict):
        return ent.get("source") or ""
    return ""


def _hit_score(hit: Any) -> float:
    if isinstance(hit, dict):
        for k in ("score", "distance"):
            if k in hit and hit.get(k) is not None:
                try:
                    return float(hit.get(k))
                except Exception:
                    pass
    for k in ("score", "distance"):
        v = getattr(hit, k, None)
        if v is not None:
            try:
                return float(v)
            except Exception:
                pass
    return 0.0


def _hit_score_mode(hit: Any) -> str:
    if isinstance(hit, dict):
        if hit.get("score") is not None:
            return "score"
        if hit.get("distance") is not None:
            return "distance"
    if getattr(hit, "score", None) is not None:
        return "score"
    if getattr(hit, "distance", None) is not None:
        return "distance"
    return "score"


def _hit_metadata(hit: Any) -> Dict[str, Any]:
    if isinstance(hit, dict):
        ent = hit.get("entity") or {}
        if isinstance(ent, dict):
            md = ent.get("metadata") or {}
            return md if isinstance(md, dict) else {}
        return {}
    ent = getattr(hit, "entity", None)
    if isinstance(ent, dict):
        md = ent.get("metadata") or {}
        return md if isinstance(md, dict) else {}
    return {}

def _rrf(rank: Optional[int], k: int) -> float:
    try:
        r = int(rank) if rank is not None else None
        if r is None:
            return 0.0
        return 1.0 / float(k + r + 1)
    except Exception:
        return 0.0


def _is_doc_existence_query(query: str) -> bool:
    q = _normalize_query(query).lower()
    keys = ["文件是否存在", "是否存在", "查询文件", "检查存在文件", "存在性"]
    return any(k in q for k in keys)


def _is_deleted_visibility_query(query: str) -> bool:
    q = _normalize_query(query).lower().replace(" ", "")
    direct_keys = ["删除校验", "删除后是否不可见", "删除后检索结果为空是否正确", "删除后的文档可见性测试"]
    if any(k in q for k in direct_keys):
        return True
    delete_keys = ["删除后", "被删除", "删除的", "删除后的"]
    visibility_keys = ["不可见", "还可检索", "还出现在sources", "还出现在source", "还可见", "是否可见"]
    return any(k in q for k in delete_keys) and any(k in q for k in visibility_keys)


def _strip_query_intent_phrases(query: str) -> str:
    q = (query or "").strip()
    parsed = _llm_query_parse_cache_get(q) or {}
    rq = _normalize_query(str(parsed.get("retrieval_query") or ""))
    return rq or _normalize_query(query)


def _query_anchor_terms(query: str) -> List[str]:
    q = (query or "").strip()
    parsed = _llm_query_parse_cache_get(q) or {}
    anchors = parsed.get("anchors")
    if isinstance(anchors, list) and anchors:
        return [_normalize_query(str(x or "")) for x in anchors if _normalize_query(str(x or ""))]
    aspects = parsed.get("aspects")
    if isinstance(aspects, list) and aspects:
        return [_normalize_query(str(x or "")) for x in aspects if _normalize_query(str(x or ""))]
    return []


def _is_generic_regulation_query(query: str) -> bool:
    if _extract_filename_candidates(query):
        return False
    if _is_doc_existence_query(query) or _is_deleted_visibility_query(query):
        return False
    q = _normalize_query(query)
    has_regulation_marker = any(k in q for k in ["条例", "办法", "规定", "法规", "核心条款", "条款", "章节", "要点"])
    return has_regulation_marker and not _query_anchor_terms(query)


def _normalize_reference_text(text: str) -> str:
    return "".join(
        ch for ch in (text or "").strip().replace("《", "").replace("》", "")
        if ch.isalnum() or "\u4e00" <= ch <= "\u9fff"
    )


def _extract_explicit_regulation_mentions(query: str) -> List[str]:
    q = (query or "").strip()
    parsed = _llm_query_parse_cache_get(q) or {}
    docs = parsed.get("documents")
    if not isinstance(docs, list):
        return []
    out: List[str] = []
    for item in docs:
        t = str(item or "").strip()
        if not t:
            continue
        if t not in out:
            out.append(t)
        if len(out) >= 5:
            break
    return out


def _match_sources_for_explicit_title(title: str, limit: int = 5) -> List[str]:
    target = _normalize_reference_text(title)
    if len(target) < 4:
        return []
    conn = _lex_db_connect()
    rows = conn.execute("SELECT source FROM documents").fetchall()
    ranked: List[Tuple[float, str]] = []
    for row in rows:
        source = _normalize_filename_for_match((row[0] if row else "") or "")
        if not source or not _visible_document_exists(source):
            continue
        best = 0.0
        for alias in _doc_title_alias_candidates(source):
            alias_norm = _normalize_reference_text(alias)
            if len(alias_norm) < 4:
                continue
            if not _explicit_title_region_compatible(title, alias):
                continue
            if target == alias_norm:
                best = max(best, 10.0)
            elif target in alias_norm and _region_tokens_compatible(title, alias):
                best = max(best, 8.0)
            elif alias_norm in target and _region_tokens_compatible(title, alias):
                best = max(best, 7.0)
        if best > 0:
            ranked.append((best, source))
    ranked.sort(key=lambda item: (-item[0], item[1]))
    out: List[str] = []
    for _, source in ranked:
        if source not in out:
            out.append(source)
        if len(out) >= max(1, int(limit)):
            break
    return out


def _exact_title_or_alias_source_matches(text: str, limit: int = 5) -> List[Dict[str, Any]]:
    target = _normalize_reference_text(text)
    if len(target) < 4:
        return []
    conn = _lex_db_connect()
    rows = conn.execute("SELECT source FROM documents").fetchall()
    ranked: List[Tuple[float, str, str]] = []
    for row in rows:
        source = _normalize_filename_for_match((row[0] if row else "") or "")
        if not source or not _visible_document_exists(source):
            continue
        info = _doc_get(source)
        canonical_title = (info.get("canonical_title") or "").strip()
        canonical_norm = _normalize_reference_text(canonical_title)
        if canonical_norm and target == canonical_norm:
            ranked.append((10.0, source, "exact_title_unique"))
            continue
        alias_hit = False
        for alias in _doc_title_alias_candidates(source):
            alias_norm = _normalize_reference_text(alias)
            if len(alias_norm) < 4:
                continue
            if not _explicit_title_region_compatible(text, alias):
                continue
            if target == alias_norm:
                alias_hit = True
                break
        if alias_hit:
            ranked.append((9.0, source, "exact_alias_unique"))
    ranked.sort(key=lambda item: (-item[0], item[1]))
    out: List[Dict[str, Any]] = []
    for _, source, reason in ranked:
        if any(entry.get("source") == source for entry in out):
            continue
        out.append({"source": source, "reason": reason})
        if len(out) >= max(1, int(limit)):
            break
    return out


def _source_entity_families(source: str) -> Dict[str, List[str]]:
    info = _doc_get(source)
    title = _normalize_query(info.get("canonical_title") or _filename_stem(source) or source)
    if not title:
        return {}
    entity_map: Dict[str, List[str]] = {
        "region": [],
        "object": [],
        "doc_type": [],
        "version": [],
        "organization": [],
    }

    for match in re.findall(r"[一-龥]{2,12}(?:省|市|区|县|州|盟|旗|乡|镇|新区|自治州|自治区)", title):
        value = _normalize_query(match)
        if len(value) >= 2 and value not in entity_map["region"]:
            entity_map["region"].append(value)
    stripped = _normalize_query(_strip_leading_region_prefix(title))
    doc_types = ["管理条例", "管理办法", "实施办法", "议事规则", "条例", "办法", "规定", "规则", "细则", "决定", "通知"]
    for doc_type in doc_types:
        if doc_type in title and doc_type not in entity_map["doc_type"]:
            entity_map["doc_type"].append(doc_type)
    object_text = stripped
    for doc_type in sorted(doc_types, key=len, reverse=True):
        object_text = object_text.replace(_normalize_query(doc_type), " ")
    object_text = re.sub(r"\d{4}[-年]\d{1,2}(?:[-月]\d{1,2})?", " ", object_text)
    object_text = re.sub(r"\s+", " ", object_text).strip()
    if len(object_text) >= 2:
        entity_map["object"].append(object_text)
    for match in re.findall(r"(?:\d{4}[-年]\d{1,2}(?:[-月]\d{1,2})?|\d{4}年|\d{4}版|现行有效|最新)", title):
        value = _normalize_query(match)
        if len(value) >= 2 and value not in entity_map["version"]:
            entity_map["version"].append(value)
    for match in re.findall(r"[一-龥]{2,20}(?:人民政府|人大常委会|管理局|委员会|局|厅|部门)", title):
        value = _normalize_query(match)
        if len(value) >= 2 and value not in entity_map["organization"]:
            entity_map["organization"].append(value)
    return {key: values for key, values in entity_map.items() if values}


def _query_entity_families(query: str) -> Dict[str, List[str]]:
    q = _normalize_query(query)
    if not q:
        return {}
    entity_map: Dict[str, List[str]] = {
        "region": [],
        "object": [],
        "doc_type": [],
        "version": [],
        "organization": [],
    }
    for match in re.findall(r"[一-龥]{2,12}(?:省|市|区|县|州|盟|旗|乡|镇|新区|自治州|自治区)", q):
        value = _normalize_query(match)
        if len(value) >= 2 and value not in entity_map["region"]:
            entity_map["region"].append(value)
    for doc_type in ["管理条例", "管理办法", "实施办法", "议事规则", "条例", "办法", "规定", "规则", "细则", "决定", "通知"]:
        if doc_type in q and doc_type not in entity_map["doc_type"]:
            entity_map["doc_type"].append(doc_type)
    for match in re.findall(r"(?:\d{4}[-年]\d{1,2}(?:[-月]\d{1,2})?|\d{4}年|\d{4}版|现行有效|最新)", q):
        value = _normalize_query(match)
        if len(value) >= 2 and value not in entity_map["version"]:
            entity_map["version"].append(value)
    for match in re.findall(r"[一-龥]{2,20}(?:人民政府|人大常委会|管理局|委员会|局|厅|部门)", q):
        value = _normalize_query(match)
        if len(value) >= 2 and value not in entity_map["organization"]:
            entity_map["organization"].append(value)
    stripped = _strip_query_intent_phrases(q) or q
    stripped = re.sub(r"[一-龥]{2,12}(?:省|市|区|县|州|盟|旗|乡|镇|新区|自治州|自治区)", " ", stripped)
    stripped = re.sub(r"(?:管理条例|管理办法|实施办法|议事规则|条例|办法|规定|规则|细则|决定|通知|法规)", " ", stripped)
    stripped = re.sub(r"(?:\d{4}[-年]\d{1,2}(?:[-月]\d{1,2})?|\d{4}年|\d{4}版|现行有效|最新)", " ", stripped)
    stripped = re.sub(r"[一-龥]{2,20}(?:人民政府|人大常委会|管理局|委员会|局|厅|部门)", " ", stripped)
    stripped = re.sub(r"(有哪些|有什么|什么|如何|怎么|请问|中|里|关于|对于|有关)", " ", stripped)
    stripped = re.sub(r"\s+", " ", stripped).strip()
    if len(stripped) >= 2:
        entity_map["object"].append(_normalize_query(stripped))
    return {key: values for key, values in entity_map.items() if values}


def _score_source_candidates_by_entities(query: str, candidate_sources: List[str]) -> List[Dict[str, Any]]:
    query_entities = _query_entity_families(query)
    if not query_entities:
        return []
    family_weights = {"region": 2.6, "object": 2.4, "doc_type": 1.6, "version": 1.3, "organization": 1.4}
    out: List[Dict[str, Any]] = []
    for source in candidate_sources:
        safe_source = _normalize_filename_for_match(source or "")
        if not safe_source:
            continue
        source_entities = _source_entity_families(safe_source)
        if not source_entities:
            continue
        score = 0.0
        matched_families: List[str] = []
        exclusive_families: List[str] = []
        for family, weight in family_weights.items():
            query_values = [_normalize_query(item) for item in query_entities.get(family, []) if _normalize_query(item)]
            source_values = [_normalize_query(item) for item in source_entities.get(family, []) if _normalize_query(item)]
            if not query_values or not source_values:
                continue
            family_hit = False
            family_exclusive = False
            for query_value in query_values:
                for source_value in source_values:
                    if query_value == source_value or query_value in source_value or source_value in query_value:
                        family_hit = True
                        score += weight
                        if family not in matched_families:
                            matched_families.append(family)
                        if _entity_is_exclusive_to_source(query_value, safe_source):
                            family_exclusive = True
                        break
                if family_hit:
                    break
            if family_exclusive and family not in exclusive_families:
                exclusive_families.append(family)
        if score > 0:
            out.append({
                "source": safe_source,
                "score": float(score),
                "matched_families": matched_families,
                "exclusive_families": exclusive_families,
            })
    out.sort(key=lambda item: (-float(item.get("score", 0.0)), item.get("source") or ""))
    return out


def _entity_is_exclusive_to_source(entity: str, source: str) -> bool:
    target = _normalize_query(entity)
    safe_source = _normalize_filename_for_match(source or "")
    if len(target) < 3 or not safe_source:
        return False
    conn = _lex_db_connect()
    rows = conn.execute("SELECT source FROM documents").fetchall()
    seen_target = False
    for row in rows:
        candidate_source = _normalize_filename_for_match((row[0] if row else "") or "")
        if not candidate_source or not _visible_document_exists(candidate_source):
            continue
        if target not in [_normalize_query(item) for item in _canonical_source_core_entities(candidate_source) if _normalize_query(item)]:
            continue
        if candidate_source != safe_source:
            return False
        seen_target = True
    return seen_target


def _exclusive_entity_source_matches(text: str, limit: int = 5) -> List[Dict[str, Any]]:
    target = _normalize_query(text)
    if len(target) < 3:
        return []
    conn = _lex_db_connect()
    rows = conn.execute("SELECT source FROM documents").fetchall()
    ranked: List[Tuple[int, str]] = []
    for row in rows:
        source = _normalize_filename_for_match((row[0] if row else "") or "")
        if not source or not _visible_document_exists(source):
            continue
        best = 0
        for entity in _canonical_source_core_entities(source):
            entity_norm = _normalize_query(entity)
            if len(entity_norm) < 3 or entity_norm not in target:
                continue
            if not _entity_is_exclusive_to_source(entity_norm, source):
                continue
            best = max(best, len(entity_norm))
        if best > 0:
            ranked.append((best, source))
    ranked.sort(key=lambda item: (-item[0], item[1]))
    out: List[Dict[str, Any]] = []
    for _, source in ranked:
        if any(entry.get("source") == source for entry in out):
            continue
        out.append({"source": source, "reason": "exclusive_entity_unique"})
        if len(out) >= max(1, int(limit)):
            break
    return out


def _visible_document_exists(source: str) -> bool:
    state = _source_state(source)
    return bool(state.get("visible"))


def _document_existence_matches(fnames: List[str]) -> List[str]:
    milvus_stats: Optional[Dict[str, Dict[str, Any]]] = None
    out: List[str] = []
    for name in fnames:
        safe = _normalize_filename_for_match(name)
        if not safe or safe in out:
            continue
        if _visible_document_exists(safe):
            out.append(safe)
            continue
        if milvus_stats is None:
            milvus_stats = _milvus_source_stats()
        if safe in (milvus_stats or {}):
            out.append(safe)
    return out


def _resolve_explicit_reference_sources(query: str, fnames: Optional[List[str]] = None, user_id: str = "anonymous") -> Dict[str, Any]:
    explicit_sources = [_normalize_filename_for_match(name or "") for name in (fnames or []) if _normalize_filename_for_match(name or "")]
    if explicit_sources:
        unique_sources = _collapse_sources_by_canonical(list(dict.fromkeys(explicit_sources)), limit=3)
        matched_sources = _document_existence_matches(unique_sources)
        if len(unique_sources) == 1 and matched_sources:
            return {
                "route": "explicit_doc_reference",
                "required": True,
                "resolved": True,
                "sources": matched_sources,
                "candidates": matched_sources,
                "reason": "explicit_filename_unique",
                "strip_title_mentions": True,
                "clarification": "",
                "target_text": unique_sources[0],
            }
        if len(unique_sources) > 1:
            return {
                "route": "explicit_doc_reference",
                "required": True,
                "resolved": False,
                "sources": [],
                "candidates": unique_sources[:3],
                "reason": "document_ambiguous",
                "strip_title_mentions": False,
                "clarification": _build_document_clarification_prompt(unique_sources[:3]),
                "target_text": "、".join(unique_sources[:3]),
            }
        return {
            "route": "explicit_doc_reference",
            "required": True,
            "resolved": False,
            "sources": [],
            "candidates": [],
            "reason": "document_not_found",
            "strip_title_mentions": False,
            "clarification": "",
            "target_text": unique_sources[0],
        }

    regulation_mentions = _extract_explicit_regulation_mentions(query)
    if not regulation_mentions:
        return {
            "route": "",
            "required": False,
            "resolved": False,
            "sources": [],
            "candidates": [],
            "reason": "not_explicit_reference",
            "strip_title_mentions": False,
            "clarification": "",
            "target_text": "",
        }

    content_query = _explicit_content_query(query, regulation_mentions)

    def _resolve_prepared_candidates(
        raw_candidates: List[str],
        *,
        allow_soft_lock: bool = False,
        trace_label: str = "",
    ) -> Optional[Dict[str, Any]]:
        prepared_sources = _prepare_explicit_regulation_candidates(raw_candidates, limit=5)
        if len(prepared_sources) == 1:
            base_source = prepared_sources[0]
            resolved_source = _latest_effective_equivalent_source(base_source) or base_source
            if _is_pseudo_singleton_soft_lock(query, resolved_source):
                if _extract_region_token(_normalize_query(query)) or _geo_context_tokens(query, user_id):
                    return {
                        "route": "explicit_regulation_reference",
                        "required": True,
                        "resolved": False,
                        "sources": [],
                        "candidates": [resolved_source],
                        "reason": "document_ambiguous",
                        "strip_title_mentions": False,
                        "clarification": _build_document_clarification_prompt([resolved_source]),
                        "target_text": regulation_mentions[0],
                        "retrieval_query_override": content_query,
                        "source_resolution_trace": {
                            "trace_label": trace_label,
                            "raw_candidates": raw_candidates[:5],
                            "prepared_candidates": prepared_sources[:5],
                            "resolved_source": resolved_source,
                            "blocked_reason": "pseudo_singleton_region_mismatch",
                        },
                    }
            return {
                "route": "explicit_regulation_reference",
                "required": True,
                "resolved": True,
                "sources": [resolved_source],
                "candidates": [resolved_source],
                "reason": "latest_effective_unique" if (len(raw_candidates) > 1 or resolved_source != base_source) else "explicit_regulation_unique",
                "strip_title_mentions": True,
                "clarification": "",
                "target_text": regulation_mentions[0],
                "retrieval_query_override": content_query,
                "source_resolution_trace": {
                    "trace_label": trace_label,
                    "raw_candidates": raw_candidates[:5],
                    "prepared_candidates": prepared_sources[:5],
                    "resolved_source": resolved_source,
                },
            }
        geo_filtered = _prepare_explicit_regulation_candidates(_geo_filtered_sources(query, user_id, prepared_sources), limit=5)
        if len(geo_filtered) == 1:
            resolved_source = geo_filtered[0]
            return {
                "route": "explicit_regulation_reference",
                "required": True,
                "resolved": True,
                "sources": [resolved_source],
                "candidates": prepared_sources[:3],
                "reason": "geo_context_locked",
                "strip_title_mentions": True,
                "clarification": "",
                "target_text": regulation_mentions[0],
                "lock_mode": "soft_lock",
                "lock_confidence": 0.82,
                "lock_message_prefix": f"我理解你查询的是《{_source_display_title(resolved_source)}》……\n",
                "source_lock_kind": "geo_context_locked",
                "source_resolution_trace": {
                    "trace_label": trace_label,
                    "raw_candidates": raw_candidates[:5],
                    "prepared_candidates": prepared_sources[:5],
                    "geo_filtered_candidates": geo_filtered[:5],
                },
                "retrieval_query_override": content_query,
            }
        if allow_soft_lock:
            unique_weak = _resolve_unique_weak_match_upgrade(query, prepared_sources)
            if unique_weak.get("resolved"):
                resolved_source = str(unique_weak.get("source") or "")
                return {
                    "route": "explicit_regulation_reference",
                    "required": True,
                    "resolved": True,
                    "sources": [resolved_source],
                    "candidates": [resolved_source],
                    "reason": str(unique_weak.get("reason") or "soft_lock_unique"),
                    "strip_title_mentions": True,
                    "clarification": "",
                    "target_text": regulation_mentions[0],
                    "lock_mode": "soft_lock",
                    "lock_confidence": float(unique_weak.get("confidence") or 0.0),
                    "lock_message_prefix": str(unique_weak.get("lock_message_prefix") or ""),
                    "source_lock_kind": "soft_lock_unique",
                    "source_resolution_trace": {
                        "trace_label": trace_label,
                        "raw_candidates": raw_candidates[:5],
                        "prepared_candidates": prepared_sources[:5],
                        **dict(unique_weak.get("trace") or {}),
                    },
                    "retrieval_query_override": content_query,
                }
        if prepared_sources:
            topical_multi = _resolve_topical_suffix_multi_doc(query, prepared_sources)
            if topical_multi.get("resolved"):
                return {
                    "route": "multi_doc_query",
                    "required": False,
                    "resolved": False,
                    "sources": list(topical_multi.get("sources") or []),
                    "candidates": list(topical_multi.get("sources") or []),
                    "reason": str(topical_multi.get("reason") or "topical_suffix_multi_doc"),
                    "strip_title_mentions": False,
                    "clarification": "",
                    "target_text": regulation_mentions[0],
                    "lock_mode": "none",
                    "source_lock_kind": "topical_suffix_multi_doc",
                    "source_resolution_trace": {
                        "trace_label": trace_label,
                        "raw_candidates": raw_candidates[:5],
                        "prepared_candidates": prepared_sources[:5],
                        **dict(topical_multi.get("trace") or {}),
                    },
                    "retrieval_query_override": content_query,
                }
            return {
                "route": "explicit_regulation_reference",
                "required": True,
                "resolved": False,
                "sources": [],
                "candidates": prepared_sources[:3],
                "reason": "document_ambiguous",
                "strip_title_mentions": False,
                "clarification": _build_document_clarification_prompt(prepared_sources[:3]),
                "target_text": regulation_mentions[0],
                "retrieval_query_override": content_query,
                "source_resolution_trace": {
                    "trace_label": trace_label,
                    "raw_candidates": raw_candidates[:5],
                    "prepared_candidates": prepared_sources[:5],
                    "geo_filtered_candidates": geo_filtered[:5],
                },
            }
        return None

    strong_matches: List[Dict[str, Any]] = []
    for mention in regulation_mentions:
        for entry in _exact_title_or_alias_source_matches(mention, limit=5):
            source = _normalize_filename_for_match(entry.get("source") or "")
            if source and not any(item.get("source") == source for item in strong_matches):
                strong_matches.append({"source": source, "reason": str(entry.get("reason") or "explicit_regulation_unique")})
    strong_candidate_sources = [str(entry.get("source") or "") for entry in strong_matches]
    strong_resolution = _resolve_prepared_candidates(strong_candidate_sources, trace_label="strong_match")
    if strong_resolution:
        resolved_source = (strong_resolution.get("sources") or [""])[0]
        resolved_reason = next(
            (str(entry.get("reason") or "explicit_regulation_unique") for entry in strong_matches if str(entry.get("source") or "") == resolved_source),
            strong_resolution.get("reason") or "explicit_regulation_unique",
        )
        if strong_resolution.get("resolved"):
            strong_resolution["reason"] = resolved_reason if strong_resolution.get("reason") != "latest_effective_unique" else strong_resolution.get("reason")
        return strong_resolution

    entity_matches: List[Dict[str, Any]] = []
    for mention in regulation_mentions:
        for entry in _exclusive_entity_source_matches(mention, limit=5):
            source = _normalize_filename_for_match(entry.get("source") or "")
            if source and not any(item.get("source") == source for item in entity_matches):
                entity_matches.append({"source": source, "reason": str(entry.get("reason") or "exclusive_entity_unique")})
    entity_candidate_sources = [str(entry.get("source") or "") for entry in entity_matches]
    entity_resolution = _resolve_prepared_candidates(entity_candidate_sources, trace_label="entity_match")
    if entity_resolution:
        resolved_source = (entity_resolution.get("sources") or [""])[0]
        resolved_reason = next(
            (str(entry.get("reason") or "exclusive_entity_unique") for entry in entity_matches if str(entry.get("source") or "") == resolved_source),
            entity_resolution.get("reason") or "exclusive_entity_unique",
        )
        if entity_resolution.get("resolved"):
            entity_resolution["reason"] = resolved_reason if entity_resolution.get("reason") != "latest_effective_unique" else entity_resolution.get("reason")
        return entity_resolution

    candidate_sources: List[str] = []
    for mention in regulation_mentions:
        for source in _match_sources_for_explicit_title(mention, limit=5):
            if source not in candidate_sources:
                candidate_sources.append(source)
    prepared_candidate_resolution = _resolve_prepared_candidates(candidate_sources, trace_label="explicit_title_match")
    if prepared_candidate_resolution:
        return prepared_candidate_resolution
    related_candidates: List[str] = []
    for mention in regulation_mentions:
        if "相关" not in mention:
            continue
        normalized_mention = mention.replace("相关", "").strip()
        for source in _extract_title_source_candidates(normalized_mention, limit=5):
            if source not in related_candidates:
                related_candidates.append(source)
    related_resolution = _resolve_prepared_candidates(related_candidates, trace_label="related_title_match")
    if related_resolution:
        return related_resolution
    normalized_title_candidates = _normalized_title_candidate_sources(regulation_mentions[0], limit=5)
    if normalized_title_candidates:
        normalized_resolution = _resolve_prepared_candidates(normalized_title_candidates, allow_soft_lock=True, trace_label="normalized_title_match")
        if normalized_resolution:
            return normalized_resolution
    fallback_candidates = _extract_title_source_candidates(regulation_mentions[0], limit=5)
    if fallback_candidates:
        fallback_resolution = _resolve_prepared_candidates(fallback_candidates, allow_soft_lock=True, trace_label="fallback_title_match")
        if fallback_resolution:
            return fallback_resolution
    dense_title_matches = _dense_title_source_matches(regulation_mentions[0], limit=5)
    if dense_title_matches:
        top_match = dense_title_matches[0]
        top_score = float(top_match.get("score") or 0.0)
        second_score = float(dense_title_matches[1].get("score") or 0.0) if len(dense_title_matches) > 1 else 0.0
        min_sim = float(getattr(config, "DENSE_TITLE_MATCH_MIN_SIM", 0.84))
        min_margin = float(getattr(config, "DENSE_TITLE_MATCH_MARGIN", 0.03))
        if top_score >= min_sim and (top_score - second_score >= min_margin or top_score >= min_sim + 0.05):
            dense_resolution = _resolve_prepared_candidates([str(top_match.get("source") or "")], allow_soft_lock=True, trace_label="dense_title_match")
            if dense_resolution:
                trace = dict(dense_resolution.get("source_resolution_trace") or {})
                trace["dense_title_scores"] = [
                    {
                        "source": item.get("source"),
                        "title": item.get("title"),
                        "score": item.get("score"),
                    }
                    for item in dense_title_matches[:3]
                ]
                dense_resolution["source_resolution_trace"] = trace
                if dense_resolution.get("resolved"):
                    dense_resolution["reason"] = "dense_title_unique"
                return dense_resolution
    return {
        "route": "explicit_regulation_reference",
        "required": True,
        "resolved": False,
        "sources": [],
        "candidates": [],
        "reason": "document_not_found",
        "strip_title_mentions": False,
        "clarification": "",
        "target_text": regulation_mentions[0],
    }


def _synthetic_documents_for_sources(sources: List[str], text: str) -> List[Dict[str, Any]]:
    documents: List[Dict[str, Any]] = []
    for src in sources:
        info = _doc_get(src)
        metadata = {
            "doc_type": info.get("doc_type") or "",
            "active_version": info.get("active_version"),
            "status": info.get("status") or "completed",
        }
        documents.append({
            "source": src,
            "score": 1.0,
            "text": text,
            "metadata": metadata,
            "chunk_range": None,
        })
    return documents


def _synthetic_source_items(sources: List[str], text: str) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    for src in sources:
        info = _doc_get(src)
        ent = {
            "source": src,
            "text": text,
            "metadata": {
                "doc_type": info.get("doc_type") or "",
                "active_version": info.get("active_version"),
                "status": info.get("status") or "completed",
            },
        }
        items.append({"entity": ent, "score": 1.0})
    return items


def _docs_cover_query_anchor_terms(query: str, docs: List[Any]) -> bool:
    terms = _query_anchor_terms(query)
    if not terms:
        return True
    haystacks: List[str] = []
    for doc in docs[: min(len(docs), 5)]:
        haystacks.append(f"{_hit_entity_source(doc) or ''}\n{_hit_display_text(doc) or ''}")
    for term in terms:
        if any(term in hay for hay in haystacks):
            return True
    return False


def _negative_clean_refusal_reason(
    query: str,
    docs: List[Any],
    fnames: List[str],
    top_source_dense_score: float = 0.0,
) -> Optional[str]:
    if _is_deleted_visibility_query(query):
        return "deleted_visibility_probe"
    if _is_doc_existence_query(query):
        return None
    if fnames:
        return None
    if docs and not _docs_cover_query_anchor_terms(query, docs):
        dense_bypass = float(getattr(config, "QUERY_ANCHOR_DENSE_BYPASS_MIN_SCORE", 0.6))
        if float(top_source_dense_score or 0.0) >= dense_bypass:
            return None
        return "query_anchor_miss"
    return None


def _classify_query_scope(query: str, fnames: List[str], query_route: Optional[str] = None) -> str:
    route = query_route or _classify_query_route(query, fnames)
    q = _normalize_query(query)
    anchored_markers = ["条例", "办法", "规定", "规则", "条款", "要点", "章节", "版本切换"]
    if fnames or route in {"version_switch", "explicit_doc_reference", "explicit_regulation_reference", "exact_title_reference", "alias_title_reference"}:
        return "anchored_question"
    if _has_contextual_doc_reference(q):
        return "anchored_question"
    if route not in {"business_topic_qa", "open_regulation_qa"} and _is_section_anchor_query(q):
        return "anchored_question"
    if route in {"weak_title_reference", "exact_title_reference", "alias_title_reference"} and any(marker in q for marker in anchored_markers):
        return "anchored_question"
    return "open_question"


def _strip_filename_mentions(query: str, fnames: List[str]) -> str:
    text = _normalize_query(query)
    for name in fnames or []:
        token = (name or "").strip()
        if not token:
            continue
        text = text.replace(token, " ")
    return _normalize_query(text)


def _should_allow_llm_fallback(query: str, query_route: str, refusal_reason: str) -> bool:
    q = _normalize_query(query)
    if not q:
        return False
    if _query_has_lockable_doc_rescue(q):
        return False
    if query_route in {"existence", "visibility_probe", "explicit_doc_reference", "explicit_regulation_reference", "exact_title_reference", "alias_title_reference"}:
        return False
    if refusal_reason:
        return False
    return True


def _answer_mode_for_sources(target_sources: List[str], selected_docs: List[Any]) -> str:
    if not selected_docs:
        return "llm_fallback"
    if not target_sources:
        return "target_hit"
    selected_sources = [
        _normalize_filename_for_match(_hit_entity_source(doc) or "")
        for doc in selected_docs
        if _hit_entity_source(doc)
    ]
    target_set = [_normalize_filename_for_match(src) for src in target_sources if src]
    if any(_sources_equivalent(selected, target) for selected in selected_sources for target in target_set):
        return "target_hit"
    return "rag_related_doc"


def _build_related_doc_grounded_answer(selected_docs: List[Any]) -> str:
    if not selected_docs:
        return "未在知识库中找到足够相关的证据来回答该问题。"
    top_hit = selected_docs[0]
    snippet = re.sub(r"\s+", " ", _hit_display_text(top_hit)).strip()
    if len(snippet) > 140:
        snippet = snippet[:140].rstrip() + "..."
    section = (_hit_metadata(top_hit).get("section") or "").strip()
    lead = "根据相关文档证据"
    if section:
        lead += f"（{section}）"
    if snippet:
        return f"{lead}，可确认：{snippet}[1]"
    return "未在知识库中找到足够相关的证据来回答该问题。"


def _compare_focus_text(compare_plan: Optional[Dict[str, Any]] = None) -> str:
    plan = dict(compare_plan or {})
    focus_terms: List[str] = []
    for candidate in (
        plan.get("common_aspects") or [],
        plan.get("topic_pair") or [],
        plan.get("canonical_aspects") or [],
        plan.get("expanded_aspects") or [],
    ):
        for term in candidate:
            normalized = _normalize_coverage_aspect(term) or _normalize_query(term)
            if normalized and normalized not in focus_terms:
                focus_terms.append(normalized)
    return "、".join(focus_terms[:2]) or "相关事项"


def _compare_answer_snippet(doc: Any, limit: int = 120) -> str:
    snippet = re.sub(r"\s+", " ", _hit_display_text(doc) or "").strip()
    if len(snippet) > limit:
        snippet = snippet[:limit].rstrip() + "..."
    return snippet


def _format_compare_evidence(
    source_groups: List[Dict[str, Any]],
    query: str,
    score_mode: str,
    compare_plan: Optional[Dict[str, Any]] = None,
    compare_source_statuses: Optional[List[Dict[str, Any]]] = None,
) -> Tuple[str, List[Dict[str, Any]]]:
    lines: List[str] = []
    refs: List[Dict[str, Any]] = []
    total_tokens = 0
    token_budget = max(256, int(getattr(config, "EVIDENCE_MAX_TOKENS", 6500)))
    focus_text = _compare_focus_text(compare_plan)
    status_map = {
        _normalize_filename_for_match((item or {}).get("source") or ""): str((item or {}).get("status") or "")
        for item in (compare_source_statuses or [])
        if _normalize_filename_for_match((item or {}).get("source") or "")
    }
    header = f"对比焦点：{focus_text}"
    header_tokens = _estimate_token_count(header) + 2
    if header_tokens <= token_budget:
        lines.append(header)
        total_tokens += header_tokens
    flattened_docs = [doc for group in source_groups or [] for doc in (group.get("docs") or [])]
    best_score = _hit_score(flattened_docs[0]) if flattened_docs else 0.0
    evidence_index = 1
    for group in source_groups or []:
        source = _normalize_filename_for_match((group or {}).get("source") or "")
        title = _source_display_title(source) if source else "未知文档"
        evidence_query = _normalize_query((group or {}).get("evidence_query") or query) or query
        group_head = f"【文档】{title} | 对比子问题：{evidence_query}"
        group_head_tokens = _estimate_token_count(group_head) + 2
        if total_tokens + group_head_tokens <= token_budget:
            lines.append(group_head)
            total_tokens += group_head_tokens
        first_ref_recorded = False
        group_blocks: List[str] = []
        for doc in group.get("docs") or []:
            src = _hit_entity_source(doc) or source or "unknown"
            content = (_hit_display_text(doc) or "").strip()
            if not content:
                evidence_index += 1
                continue
            section = _doc_section_name(doc)
            chunk_range = _hit_chunk_range(doc)
            relevance = _evidence_relevance(doc, score_mode, best_score)
            parts = [f"来源：{src}", f"文档：{title}", f"相关性：{relevance:.2f}"]
            if section:
                parts.append(f"章节：{section}")
            if chunk_range:
                parts.append(f"位置：chunk_{chunk_range}")
            block = f"[证据 {evidence_index}] " + " | ".join(parts) + "\n" + content
            block_tokens = _estimate_token_count(block) + 2
            if total_tokens + block_tokens > token_budget:
                return "\n\n".join(lines), refs
            group_blocks.append(block)
            if not first_ref_recorded:
                refs.append({
                    "index": evidence_index,
                    "source": source,
                    "title": title,
                    "section": section,
                    "snippet": _compare_answer_snippet(doc),
                    "evidence_query": evidence_query,
                })
                first_ref_recorded = True
            evidence_index += 1
        group_summary = _summarize_compare_source_blocks(
            title=title,
            evidence_query=evidence_query,
            focus_text=focus_text,
            blocks=group_blocks,
            status=status_map.get(source, ""),
        )
        summary_tokens = _estimate_token_count(group_summary) + 2
        if total_tokens + summary_tokens > token_budget:
            return "\n\n".join(lines), refs
        lines.append(group_summary)
        total_tokens += summary_tokens
    return "\n\n".join(lines), refs


def _format_single_doc_compare_evidence(
    docs: List[Any],
    query: str,
    score_mode: str,
    compare_plan: Optional[Dict[str, Any]] = None,
) -> Tuple[str, List[Dict[str, Any]]]:
    lines: List[str] = []
    refs: List[Dict[str, Any]] = []
    total_tokens = 0
    token_budget = max(256, int(getattr(config, "EVIDENCE_MAX_TOKENS", 6500)))
    plan = dict(compare_plan or {})
    topic_pair = [str(item).strip() for item in (plan.get("topic_pair") or []) if str(item).strip()]
    focus_text = "、".join(topic_pair[:2]) or _compare_focus_text(plan)
    header = f"文档内对比焦点：{focus_text}"
    header_tokens = _estimate_token_count(header) + 2
    if header_tokens <= token_budget:
        lines.append(header)
        total_tokens += header_tokens
    best_score = _hit_score(docs[0]) if docs else 0.0
    for i, doc in enumerate(docs, start=1):
        src = _hit_entity_source(doc) or "unknown"
        title = _source_display_title(_normalize_filename_for_match(src) or src)
        content = (_hit_display_text(doc) or "").strip()
        if not content:
            continue
        section = _doc_section_name(doc)
        chunk_range = _hit_chunk_range(doc)
        relevance = _evidence_relevance(doc, score_mode, best_score)
        label = topic_pair[min(len(topic_pair) - 1, len(refs))] if topic_pair else (section or f"要点{i}")
        parts = [f"来源：{src}", f"文档：{title}", f"相关性：{relevance:.2f}", f"主题：{label}"]
        if section:
            parts.append(f"章节：{section}")
        if chunk_range:
            parts.append(f"位置：chunk_{chunk_range}")
        block = f"[证据 {i}] " + " | ".join(parts) + "\n" + content
        block_tokens = _estimate_token_count(block) + 2
        if total_tokens + block_tokens > token_budget:
            break
        lines.append(block)
        total_tokens += block_tokens
        if len(refs) < 2:
            refs.append({
                "index": i,
                "title": title,
                "section": section,
                "snippet": _compare_answer_snippet(doc),
                "label": label,
            })
    return "\n\n".join(lines), refs


def _build_multi_doc_compare_grounded_answer(source_refs: List[Dict[str, Any]], compare_plan: Optional[Dict[str, Any]] = None) -> str:
    if len(source_refs) < 2:
        return "未在知识库中找到足够相关的证据来回答该问题。"
    focus_text = _compare_focus_text(compare_plan)
    lines = [f"围绕{focus_text}，两份文档的直接证据如下：[{source_refs[0]['index']}][{source_refs[1]['index']}]"]
    for ref in source_refs[:2]:
        section_text = f"{ref['section']}显示" if ref.get("section") else "可见"
        snippet = ref.get("snippet") or "未提取到可引用片段"
        lines.append(f"- {ref['title']}：{section_text}{snippet}[{ref['index']}]")
    return "\n".join(lines)


def _compare_status_display_label(status: str) -> str:
    label_map = {
        "answerable": "已找到可对比证据",
        "guarded_full": "已找到可对比证据",
        "comparable_partial": "存在相关但不完全对齐的证据",
        "evidence_insufficient": "未检索到与焦点直接对应的条文证据",
        "not_found": "未找到相关证据",
    }
    return label_map.get(status or "", "证据不足")


def _compare_source_status_prompt_lines(source_statuses: List[Dict[str, Any]]) -> str:
    lines: List[str] = []
    for item in source_statuses or []:
        title = (item.get("title") or item.get("source") or "目标文档").strip()
        status = _compare_status_display_label(str(item.get("status") or ""))
        lines.append(f"- {title}：{status}")
    return "\n".join(lines)


def _fallback_compare_source_summary(title: str, status: str, blocks: List[str]) -> str:
    if status == "not_found":
        return f"【文档摘要】{title}\n- 未检索到与对比焦点直接对应的证据。"
    if not blocks:
        return f"【文档摘要】{title}\n- 暂未提纯出可直接比较的事实。"
    lines = [f"【文档摘要】{title}"]
    for block in blocks[:2]:
        compact = re.sub(r"\s+", " ", block).strip()
        compact = re.sub(r"^\[证据\s*([0-9]+)\]\s*", r"[\1] ", compact)
        if len(compact) > 220:
            compact = compact[:220].rstrip() + "..."
        lines.append(f"- {compact}")
    return "\n".join(lines)


def _summarize_compare_source_blocks(
    title: str,
    evidence_query: str,
    focus_text: str,
    blocks: List[str],
    status: str,
) -> str:
    if not bool(getattr(config, "ENABLE_COMPARE_MAP_REDUCE", True)):
        return _fallback_compare_source_summary(title, status, blocks)
    if status == "not_found":
        return f"【文档摘要】{title}\n- 未检索到与{focus_text or evidence_query or '对比焦点'}直接对应的证据。"
    if not blocks:
        return _fallback_compare_source_summary(title, status, blocks)
    evidence_text = "\n\n".join(blocks)

    def _call_summary_llm() -> str:
        import requests

        headers = {"Content-Type": "application/json"}
        if config.LLM_API_KEY:
            headers["Authorization"] = f"Bearer {config.LLM_API_KEY}"

        payload = {
            "model": config.LLM_MODEL,
            "messages": [
                {"role": "system", "content": "你是法规证据提纯助手"},
                {"role": "user", "content": f"""请只依据给定证据，为单个文档提炼可用于最终对比的关键事实。

规则：
1) 只输出 2 到 4 行要点，每行都必须保留原证据编号，如 [1] 或 [2]
2) 只保留与“{focus_text or evidence_query or '对比焦点'}”直接相关的事实，不要复述无关背景
3) 如果该文档没有直接涉及焦点，也要明确写出“未直接涉及该内容”或“未检索到相关信息”
4) 严禁编造，严禁合并不同证据编号为一个新编号
5) 输出必须以“【文档摘要】{title}”开头

文档：{title}
对比子问题：{evidence_query or focus_text or '相关事项'}
当前状态：{_compare_status_display_label(status)}

证据：
{evidence_text}

输出："""},
            ],
            "temperature": 0.2,
            "top_p": config.LLM_TOP_P,
            "max_tokens": int(getattr(config, "COMPARE_MAP_REDUCE_MAX_TOKENS", 220)),
            "presence_penalty": config.LLM_PRESENCE_PENALTY,
        }

        raw = (config.LLM_EXTRA_BODY or "").strip()
        if raw:
            try:
                extra = json.loads(raw)
                if isinstance(extra, dict):
                    for key, value in extra.items():
                        if key not in payload:
                            payload[key] = value
            except Exception:
                pass

        urls = []
        if config.LLM_CHAT_COMPLETIONS_URL:
            urls.append(config.LLM_CHAT_COMPLETIONS_URL)
        else:
            base = (config.LLM_API_BASE or "").rstrip("/")
            if base:
                urls.append(f"{base}/chat/completions")
                if not base.endswith("/v1"):
                    urls.append(f"{base}/v1/chat/completions")
                if base.endswith("/v1"):
                    urls.append(f"{base[:-3].rstrip('/')}/chat/completions")

        last_exc = None
        for url in urls:
            try:
                resp = requests.post(url, headers=headers, json=payload, timeout=max(3, min(20, int(config.LLM_TIMEOUT))))
                resp.raise_for_status()
                data = resp.json()
                choices = data.get("choices") or []
                if choices and isinstance(choices, list):
                    msg = (choices[0] or {}).get("message") or {}
                    content = msg.get("content") if isinstance(msg.get("content"), str) else ""
                    content = (content or "").strip()
                    if not content:
                        text_val = (choices[0] or {}).get("text")
                        content = (text_val or "").strip() if isinstance(text_val, str) else ""
                    if content:
                        return content
            except Exception as exc:
                last_exc = exc
                continue
        raise last_exc or RuntimeError("compare source summary failed")

    try:
        summary = _call_summary_llm().strip()
    except Exception:
        summary = ""
    if not summary:
        return _fallback_compare_source_summary(title, status, blocks)
    if "【文档摘要】" not in summary:
        summary = f"【文档摘要】{title}\n" + summary
    return summary


def _fallback_compare_refs_from_docs(selected_docs: List[Any], limit: int = 2) -> List[Dict[str, Any]]:
    refs: List[Dict[str, Any]] = []
    seen_sources: List[str] = []
    for idx, doc in enumerate(selected_docs or [], start=1):
        source = _normalize_filename_for_match(_hit_entity_source(doc) or "")
        if not source or source in seen_sources:
            continue
        seen_sources.append(source)
        refs.append({
            "index": idx,
            "source": source,
            "title": _source_display_title(source),
            "section": _doc_section_name(doc),
            "snippet": _compare_answer_snippet(doc),
        })
        if len(refs) >= limit:
            break
    return refs


def _build_single_doc_compare_grounded_answer(topic_refs: List[Dict[str, Any]], compare_plan: Optional[Dict[str, Any]] = None) -> str:
    if len(topic_refs) < 2:
        return "未在知识库中找到足够相关的证据来回答该问题。"
    plan = dict(compare_plan or {})
    topic_pair = [str(item).strip() for item in (plan.get("topic_pair") or []) if str(item).strip()]
    focus_text = "、".join(topic_pair[:2]) or _compare_focus_text(plan)
    lines = [f"该文档内围绕{focus_text}的直接证据如下：[{topic_refs[0]['index']}][{topic_refs[1]['index']}]"]
    missing_targets = [str(item).strip() for item in (plan.get("missing_targets") or []) if str(item).strip()]
    if missing_targets:
        missing_text = "、".join([f"《{item}》" if not (item.startswith("《") and item.endswith("》")) else item for item in missing_targets[:4]])
        lines.append(f"未在知识库中检索到{missing_text}，暂无法为您提供对比分析。")
    for ref in topic_refs[:2]:
        label = ref.get("label") or ref.get("section") or "相关要点"
        section_text = f"{ref['section']}显示" if ref.get("section") else "可见"
        snippet = ref.get("snippet") or "未提取到可引用片段"
        lines.append(f"- {label}：{section_text}{snippet}[{ref['index']}]")
    return "\n".join(lines)


def _is_version_switch_query(query: str) -> bool:
    q = _normalize_query(query)
    return ("版本切换后只读新版本" in q) or ("只读新版本" in q)


def _is_business_topic_query(query: str) -> bool:
    q = _normalize_query(query)
    if not q or _extract_filename_candidates(q) or _extract_explicit_regulation_mentions(q):
        return False
    if _is_doc_existence_query(q) or _is_deleted_visibility_query(q) or _is_version_switch_query(q):
        return False
    if _is_generic_document_required_query(q):
        return False
    action_markers = ["哪些", "什么", "如何", "怎么", "规定", "要求", "处罚", "责任", "程序", "条件", "标准", "管理", "建设", "扶持"]
    if not any(marker in q for marker in action_markers):
        return False
    generic_terms = {"规定", "要求", "处罚", "责任", "程序", "条件", "标准", "办法", "条例", "规则", "细则", "决定", "通知"}
    topic_terms = [term for term in _query_anchor_terms(q) if term not in generic_terms]
    return bool(topic_terms)


def _is_open_regulation_query(query: str) -> bool:
    q = _normalize_query(query)
    if not q or _extract_filename_candidates(q) or _extract_explicit_regulation_mentions(q):
        return False
    if _is_doc_existence_query(q) or _is_deleted_visibility_query(q) or _is_version_switch_query(q):
        return False
    if _is_business_topic_query(q):
        return False
    regulation_markers = ["条例", "办法", "规定", "规则", "细则", "决定", "通知", "法规"]
    return any(marker in q for marker in regulation_markers) and _query_has_intent_signal(q)


def _classify_query_route(query: str, fnames: Optional[List[str]] = None) -> str:
    q = _normalize_query(query)
    names = fnames if fnames is not None else _extract_filename_candidates(q)
    compare_resolution = _analyze_compare_route(q)
    if compare_resolution.get("is_compare"):
        return compare_resolution.get("route") or "open_topic_compare"
    explicit_resolution = _resolve_explicit_reference_sources(q, names)
    title_route = _classify_title_reference_route(q, names)
    strong_title_route = title_route if title_route in {"exact_title_reference", "alias_title_reference"} else ""
    weak_title_route = title_route if title_route in {"weak_title_reference", "topic_like_title"} else ""
    for route_name in _policy_get("query_route.order", ["existence", "visibility_probe", "version_switch", "explicit_doc_reference", "explicit_regulation_reference", "weak_title_reference", "content_qa"]):
        if route_name == "existence" and _is_doc_existence_query(q):
            return "existence"
        if route_name == "visibility_probe" and _is_deleted_visibility_query(q):
            return "visibility_probe"
        if route_name == "version_switch" and _is_version_switch_query(q):
            return "version_switch"
        if route_name == "explicit_doc_reference" and explicit_resolution.get("route") == "explicit_doc_reference":
            return "explicit_doc_reference"
        if route_name == "explicit_regulation_reference" and explicit_resolution.get("route") == "explicit_regulation_reference":
            return "explicit_regulation_reference"
        if route_name == "weak_title_reference":
            require_no_filenames = bool(_policy_get("query_route.weak_title_reference.require_no_filenames", True))
            if ((not require_no_filenames) or (not names)) and strong_title_route:
                return strong_title_route
        if route_name == "business_topic_qa" and _is_business_topic_query(q):
            return "business_topic_qa"
        if route_name == "open_regulation_qa" and _is_open_regulation_query(q):
            return "open_regulation_qa"
        if route_name == "content_qa":
            if weak_title_route:
                return weak_title_route
            return "content_qa"
    return "content_qa"


def _metadata_with_query_route(metadata: Dict[str, Any], query_route: str) -> Dict[str, Any]:
    out = dict(metadata or {})
    out["query_route"] = query_route
    return out


def _target_status(
    query_route: str,
    source_lock_required: bool,
    source_lock_resolved: bool,
    source_lock_reason: str,
    target_sources: List[str],
) -> str:
    if source_lock_required:
        if source_lock_resolved or target_sources:
            return "resolved"
        if source_lock_reason == "document_not_found":
            return "document_not_found"
        if source_lock_reason in {"document_ambiguous", "section_anchor_ambiguous"}:
            return "ambiguous"
        return "required_unresolved"
    if target_sources:
        return "resolved"
    if query_route in {"explicit_doc_reference", "explicit_regulation_reference", "exact_title_reference", "alias_title_reference", "weak_title_reference", "version_switch"}:
        return "document_required"
    return "open"


def _control_status(final_channel: str, blocked: Optional[str], target_status: str, refusal_reason: Optional[str]) -> str:
    if blocked:
        return blocked
    if final_channel == "document_not_found" or target_status == "document_not_found":
        return "document_not_found"
    if final_channel in {"document_ambiguous", "document_clarification"} or target_status in {"ambiguous", "required_unresolved"}:
        return "source_lock_failed"
    if refusal_reason:
        return "evidence_insufficient"
    if target_status == "resolved":
        return "source_locked"
    return "answerable"


def _refusal_stage(blocked: Optional[str], source_lock_required: bool, source_lock_resolved: bool, refusal_reason: Optional[str]) -> Optional[str]:
    if blocked:
        return "query_validation"
    if source_lock_required and not source_lock_resolved:
        return "source_lock"
    if refusal_reason:
        return "evidence"
    return None


def _build_control_plane_metadata(
    *,
    query: str,
    user_id: str,
    final_channel: str,
    query_route: Optional[str] = None,
    internal_route: Optional[str] = None,
    fnames: Optional[List[str]] = None,
    recall: Optional[Dict[str, Any]] = None,
    blocked: Optional[str] = None,
    refusal_reason: Optional[str] = None,
    query_quality: Optional[str] = None,
    docs_returned: Optional[int] = None,
    question_type: Optional[str] = None,
    answer_mode: Optional[str] = None,
    extra: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    recall = dict(recall or {})
    normalized_targets = [
        _normalize_filename_for_match(item)
        for item in (recall.get("target_sources") or fnames or [])
        if _normalize_filename_for_match(item)
    ]
    route_name = query_route or recall.get("query_route") or "content_qa"
    internal_name = internal_route or recall.get("query_route") or route_name
    source_lock_required = bool(recall.get("source_lock_required"))
    source_lock_resolved = bool(recall.get("resolved_source_lock") or (source_lock_required and normalized_targets))
    source_lock_reason = (recall.get("source_lock_reason") or "").strip()
    target_status = _target_status(route_name, source_lock_required, source_lock_resolved, source_lock_reason, normalized_targets)
    effective_query_quality = query_quality or ("valid" if not blocked else "invalid")
    effective_refusal_reason = refusal_reason or blocked
    scope = _classify_query_scope(query, normalized_targets or list(fnames or []), internal_name)
    llm_fallback_allowed = False
    if not blocked and not (source_lock_required and not source_lock_resolved):
        llm_fallback_allowed = _should_allow_llm_fallback(query, internal_name, refusal_reason)
    doc_fallback_enabled = False
    if not blocked and final_channel == "light_rag":
        doc_fallback_enabled = _should_use_doc_fallback(query, normalized_targets, internal_name)
    metadata = {
        "query": query,
        "user_id": user_id,
        "query_route": route_name,
        "internal_route": internal_name,
        "final_channel": final_channel,
        "query_quality": effective_query_quality,
        "source_lock_required": source_lock_required,
        "source_lock_resolved": source_lock_resolved,
        "source_lock_reason": source_lock_reason,
        "target_status": target_status,
        "target_sources": normalized_targets,
        "doc_fallback_enabled": doc_fallback_enabled,
        "llm_fallback_allowed": llm_fallback_allowed,
        "scope": scope,
        "refusal_stage": _refusal_stage(blocked, source_lock_required, source_lock_resolved, refusal_reason),
        "refusal_reason": effective_refusal_reason,
        "control_status": _control_status(final_channel, blocked, target_status, refusal_reason),
        "control_plane": "light",
        "lock_mode": recall.get("lock_mode") or "",
        "lock_confidence": float(recall.get("lock_confidence") or 0.0),
        "lock_message_prefix": recall.get("lock_message_prefix") or "",
        "source_lock_kind": recall.get("source_lock_kind") or "",
        "source_resolution_trace": dict(recall.get("source_resolution_trace") or {}),
        "inherited_from_context": bool(recall.get("inherited_from_context")),
    }
    if docs_returned is not None:
        metadata["docs_returned"] = docs_returned
    if question_type:
        metadata["question_type"] = question_type
    if answer_mode:
        metadata["answer_mode"] = answer_mode
    if extra:
        metadata.update(extra)
    return metadata


def _doc_title_alias_hit(source: str, query: str) -> bool:
    return _doc_title_alias_score(source, query) >= 2.0


def _annotate_lexical_hit(query: str, hit: Dict[str, Any], allowed_set: set[str], doc_recall_map: Optional[Dict[str, Dict[str, Any]]] = None) -> Dict[str, Any]:
    src = _normalize_filename_for_match(_hit_entity_source(hit) or "")
    if not src:
        return hit
    ent = dict((hit or {}).get("entity") or {})
    md = dict(ent.get("metadata") or {})
    title_score = _doc_title_alias_score(src, query)
    if title_score > 0:
        md["title_match_score"] = title_score
    if title_score >= 2.0 or md.get("title_hit"):
        md["title_hit"] = True
    if src in allowed_set:
        md["doc_recall_hit"] = True
    plan_entry = (doc_recall_map or {}).get(src) or {}
    if plan_entry:
        md["doc_prior"] = float(plan_entry.get("prior", 0.0))
        md["doc_recall_reasons"] = list(plan_entry.get("reasons") or [])
        md["doc_recall_rank"] = int(plan_entry.get("rank", 0))
    if not md.get("lexical_signal"):
        if md.get("title_hit") or (md.get("section") or "") == "document_title":
            md["lexical_signal"] = "title_direct"
        else:
            md["lexical_signal"] = "indexed_fts"
    ent["metadata"] = md
    out = dict(hit)
    out["entity"] = ent
    return out


def _build_source_signal_map(query: str, lex_items: List[Dict[str, Any]], doc_recall_plan: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    signal_map: Dict[str, Dict[str, Any]] = {}
    for rank, entry in enumerate(doc_recall_plan or []):
        src = _normalize_filename_for_match((entry or {}).get("source") or "")
        if not src:
            continue
        signals = signal_map.setdefault(src, {})
        signals["doc_recall"] = True
        signals["doc_prior"] = float((entry or {}).get("prior", 0.0))
        signals["doc_recall_rank"] = rank
        signals["doc_recall_reasons"] = list((entry or {}).get("reasons") or [])
        if "title_alias_substring" in signals["doc_recall_reasons"]:
            signals["title_hit"] = True
    for item in lex_items or []:
        src = _normalize_filename_for_match(_hit_entity_source(item) or "")
        if not src:
            continue
        signals = signal_map.setdefault(src, {})
        signals["lexical_hit"] = True
        md = _hit_metadata(item)
        lexical_signal = (md.get("lexical_signal") or "").strip()
        if lexical_signal:
            signals[lexical_signal] = True
        if md.get("doc_recall_hit"):
            signals["doc_recall"] = True
        if md.get("doc_prior") is not None:
            signals["doc_prior"] = max(float(signals.get("doc_prior", 0.0)), float(md.get("doc_prior", 0.0)))
        if md.get("title_hit") or _doc_title_alias_hit(src, query):
            signals["title_hit"] = True
    for src, signals in signal_map.items():
        if _doc_title_alias_hit(src, query):
            signals["title_hit"] = True
    return signal_map


def _fusion_source_score(
    src: str,
    query: str,
    dense_rank_map: Dict[str, int],
    lex_rank_map: Dict[str, int],
    source_count: Dict[str, int],
    source_signals: Dict[str, Dict[str, Any]],
    fname_set: set,
    allowed_set: set,
    weak_query: bool,
) -> float:
    K = int(getattr(config, "RRF_K", 60))
    dense_score = float(getattr(config, "FUSION_W_DENSE", 0.72)) * _rrf(dense_rank_map.get(src), K)
    lex_score = float(getattr(config, "FUSION_W_LEX", 0.20)) * _rrf(lex_rank_map.get(src), K)
    base_score = dense_score + lex_score
    signals = source_signals.get(src, {})
    multiplier = 1.0
    if signals.get("title_hit"):
        multiplier *= max(1.0, float(getattr(config, "FUSION_M_TITLE", 1.35)))
    if signals.get("doc_recall"):
        multiplier *= max(1.0, float(getattr(config, "FUSION_M_DOC_RECALL", 1.2)))
    if signals.get("lexical_hit"):
        multiplier *= max(1.0, float(getattr(config, "FUSION_M_TERM", 1.08)))
    if dense_rank_map.get(src) is not None and (signals.get("lexical_hit") or signals.get("doc_recall")):
        multiplier *= max(1.0, float(getattr(config, "FUSION_M_AGREEMENT", 1.12)))
    doc_prior_bonus = min(max(float(signals.get("doc_prior", 0.0)), 0.0), 1.0) * float(getattr(config, "FUSION_W_DOC_PRIOR", 0.003))
    prior_bonus = min(source_count.get(src, 0) / 20.0, 1.0) * float(getattr(config, "FUSION_W_PRIOR", 0.002))
    score = (base_score + doc_prior_bonus + prior_bonus) * multiplier
    return score * _source_constraint_multiplier(src, query, fname_set, allowed_set, weak_query)


def _clone_hit_with_score(hit: Any, score: float) -> Dict[str, Any]:
    ent = dict((hit or {}).get("entity") or {}) if isinstance(hit, dict) else dict(getattr(hit, "entity", None) or {})
    md = dict(ent.get("metadata") or {})
    if "orig_score" not in md:
        md["orig_score"] = float(_hit_score(hit) or 0.0)
        md["orig_score_mode"] = _hit_score_mode(hit)
    md["fusion_score"] = float(score)
    ent["metadata"] = md
    return {"entity": ent, "score": float(score)}


def _structured_chunk_hit(source: str, chunk: Dict[str, Any], score: float, reason: str) -> Dict[str, Any]:
    metadata = dict(chunk.get("metadata") or {})
    metadata.update({
        "heading_expanded": True,
        "heading_expansion_reason": reason,
    })
    ent = {
        "source": source,
        "text": chunk.get("text") or chunk.get("raw_text") or "",
        "metadata": metadata,
    }
    return {"entity": ent, "score": float(score)}


def _expand_heading_hits_to_article_hits(query: str, source: str, hits: List[Any], limit: int = 6) -> List[Any]:
    if not hits:
        return hits
    targets = _local_validate_section_targets(_extract_section_query_targets(query), limit=max(2, int(limit)))
    if not targets:
        return hits
    top_hits = list(hits[: max(int(limit), 4)])
    if any((_hit_metadata(hit).get("article_no") or "").strip() or _hit_metadata(hit).get("chunk_role") == "article" for hit in top_hits):
        return hits

    heading_section_ids: set[str] = set()
    heading_titles: List[str] = []
    for hit in top_hits:
        md = _hit_metadata(hit)
        section_title = _normalize_query(str(md.get("section_title") or md.get("section") or ""))
        section_path = [_normalize_query(str(item or "")) for item in (md.get("section_path") or []) if str(item or "").strip()]
        chunk_role = str(md.get("chunk_role") or "")
        target_aligned = any(
            target in section_title or any(target in item for item in section_path)
            for target in targets
        )
        if chunk_role in {"title", "chapter_heading", "section_heading", "toc", "toc_heading"} or target_aligned:
            section_node_id = str(md.get("section_node_id") or "").strip()
            parent_section_id = str(md.get("parent_section_id") or "").strip()
            if section_node_id:
                heading_section_ids.add(section_node_id)
            if parent_section_id:
                heading_section_ids.add(parent_section_id)
            if section_title and section_title not in heading_titles:
                heading_titles.append(section_title)
            for item in section_path:
                if item and item not in heading_titles:
                    heading_titles.append(item)
    if not heading_section_ids and not heading_titles:
        return hits

    synthesized: List[Dict[str, Any]] = []
    seen_keys: set[Tuple[str, str]] = set()
    for chunk in _get_chunks_for_source(source):
        md = dict(chunk.get("metadata") or {})
        article_no = str(md.get("article_no") or md.get("clause_label") or "").strip()
        chunk_role = str(md.get("chunk_role") or "")
        if not article_no and chunk_role != "article":
            continue
        section_title = _normalize_query(str(md.get("section_title") or md.get("section") or ""))
        parent_section_id = str(md.get("parent_section_id") or "").strip()
        section_node_id = str(md.get("section_node_id") or "").strip()
        section_path = [_normalize_query(str(item or "")) for item in (md.get("section_path") or []) if str(item or "").strip()]
        haystacks = [
            section_title,
            _normalize_query(article_no),
            _normalize_query(str(chunk.get("raw_text") or chunk.get("text") or "")),
        ] + section_path
        score = 0.0
        if section_node_id and section_node_id in heading_section_ids:
            score += 2.8
        if parent_section_id and parent_section_id in heading_section_ids:
            score += 2.8
        if any(title and (title == section_title or title in section_path) for title in heading_titles):
            score += 1.8
        for target in targets:
            if any(target and target in item for item in haystacks if item):
                score += 1.6
        if score <= 0:
            continue
        key = (source, (chunk.get("text") or chunk.get("raw_text") or "")[:120])
        if key in seen_keys:
            continue
        seen_keys.add(key)
        synthesized.append(_structured_chunk_hit(source, chunk, score + 0.5, "heading_to_article"))

    if not synthesized:
        return hits
    synthesized.sort(key=lambda item: -float(item.get("score", 0.0)))
    existing = list(hits)
    deduped: List[Any] = []
    merged_keys: set[Tuple[str, str]] = set()
    for hit in synthesized + existing:
        key = (_hit_entity_source(hit) or source, (_hit_entity_text(hit) or "")[:120])
        if key in merged_keys:
            continue
        merged_keys.add(key)
        deduped.append(hit)
    return deduped


def _top_ranked_source(rank_map: Dict[str, int]) -> Optional[str]:
    if not rank_map:
        return None
    return min(rank_map.items(), key=lambda item: item[1])[0]


def _should_apply_chunk_rerank(
    hits: List[Any],
    dense_rank_map: Dict[str, int],
    lex_rank_map: Dict[str, int],
    source_signals: Dict[str, Dict[str, Any]],
    enable_rerank: bool,
) -> bool:
    if (not hits) or (not enable_rerank) or (not config.ENABLE_RERANK):
        return False
    if not bool(getattr(config, "ENABLE_CHUNK_RERANK", False)):
        return False
    if len(_distinct_hit_sources(hits)) <= 1:
        return True
    if not bool(getattr(config, "RERANK_LOW_CONF_ONLY", True)):
        return True
    top_dense = _top_ranked_source(dense_rank_map)
    top_lex = _top_ranked_source(lex_rank_map)
    if top_dense and top_lex and top_dense != top_lex:
        return True
    top_src = top_dense or top_lex
    if top_src and source_signals.get(top_src, {}).get("title_hit"):
        return False
    return False


def _source_score_gap(src_scores: Dict[str, float]) -> float:
    if len(src_scores) <= 1:
        return 1.0
    ordered = sorted(src_scores.values(), reverse=True)
    return float(ordered[0] - ordered[1])
def _doc_title_alias_hit(source: str, query: str) -> bool:
    q = _normalize_query(query)
    for candidate in _doc_title_alias_candidates(source):
        a = (candidate or "").strip()
        if a and (a in q):
            return True
    return False
def _doc_cluster_accept(query: str, hits: List[Any], fnames: List[str]) -> bool:
    if not hits:
        return False
    by_src: Dict[str, List[Any]] = {}
    for h in hits:
        s = _normalize_filename_for_match(_hit_entity_source(h) or "")
        by_src.setdefault(s, []).append(h)
    top_src = None
    sec_src = None
    counts = sorted([(s, len(v)) for s, v in by_src.items()], key=lambda x: x[1], reverse=True)
    if counts:
        top_src = counts[0][0]
        if len(counts) > 1:
            sec_src = counts[1][0]
    total = len(hits)
    top_count = len(by_src.get(top_src or "", []))
    top_share = (float(top_count) / float(max(total, 1))) if top_src else 0.0
    adj_cnt = 0
    items = by_src.get(top_src or "", [])
    cids = []
    for it in items:
        md = _hit_metadata(it)
        try:
            cid = int(md.get("chunk_id")) if md.get("chunk_id") is not None else None
            if cid is not None:
                cids.append(cid)
        except Exception:
            pass
    cids = sorted(list(set(cids)))
    for i in range(1, len(cids)):
        if abs(cids[i] - cids[i-1]) <= 1:
            adj_cnt += 1
    if _is_doc_existence_query(query) and top_src:
        doc = _doc_get(top_src)
        st = doc.get("status") or ""
        av = doc.get("active_version")
        if (st == "completed") and (av is not None):
            return True
    if _doc_title_alias_hit(top_src or "", query) and (top_count >= 2):
        return True
    if top_share >= 0.5 and (top_count >= 2):
        return True
    if adj_cnt >= 1:
        return True
    if fnames and top_src and (_normalize_filename_for_match(fnames[0]) == top_src) and (top_count >= 1):
        return True
    return False


def _related_doc_admission(query: str, hits: List[Any], query_route: str) -> Optional[Dict[str, Any]]:
    if query_route in {"existence", "visibility_probe"}:
        return None
    if _is_deleted_visibility_query(query) or _is_doc_existence_query(query):
        return None
    if not hits:
        return None
    window = hits[: min(len(hits), 10)]
    by_src: Dict[str, List[Any]] = {}
    for hit in window:
        src = _normalize_filename_for_match(_hit_entity_source(hit) or "")
        if not src:
            continue
        by_src.setdefault(src, []).append(hit)
    if not by_src:
        return None
    counts = sorted(((src, len(items)) for src, items in by_src.items()), key=lambda item: (-item[1], item[0]))
    top_src, top_count = counts[0]
    second_count = counts[1][1] if len(counts) > 1 else 0
    top_share = float(top_count) / float(max(len(window), 1))
    if top_count < 4 or top_share < 0.7 or second_count > 2:
        return None
    top_hits = by_src.get(top_src) or []
    if not _doc_cluster_accept(query, top_hits, []):
        return None
    anchor_terms = _query_anchor_terms(query)
    if anchor_terms and not _docs_cover_query_anchor_terms(query, top_hits):
        return None
    has_section_evidence = any((_hit_metadata(hit).get("section") or "").strip() for hit in top_hits)
    if (not has_section_evidence) and (not anchor_terms):
        return None
    return {
        "source": top_src,
        "top_count": top_count,
        "second_count": second_count,
        "top_share": top_share,
    }


def _dense_source_score_map(hits: List[Any]) -> Dict[str, float]:
    out: Dict[str, float] = {}
    for hit in hits or []:
        src = _normalize_filename_for_match(_hit_entity_source(hit) or "")
        if not src:
            continue
        out[src] = max(float(out.get(src, 0.0)), float(_hit_score(hit)))
    return out


def _source_dense_tiebreak_score(hit: Any, dense_source_scores: Dict[str, float]) -> float:
    src = _normalize_filename_for_match(_hit_entity_source(hit) or "")
    if not src:
        return 0.0
    return float(dense_source_scores.get(src, 0.0))


def _source_entropy_metrics(hits: List[Any], top_n: int = 5) -> Dict[str, Any]:
    window = list(hits[: max(1, min(len(hits), top_n))]) if hits else []
    if not window:
        return {
            "top_source": None,
            "top_source_share": 0.0,
            "source_entropy": 0.0,
            "source_entropy_normalized": 0.0,
        }
    counts: Dict[str, int] = {}
    for hit in window:
        src = _normalize_filename_for_match(_hit_entity_source(hit) or "")
        if not src:
            continue
        counts[src] = counts.get(src, 0) + 1
    if not counts:
        return {
            "top_source": None,
            "top_source_share": 0.0,
            "source_entropy": 0.0,
            "source_entropy_normalized": 0.0,
        }
    total = float(sum(counts.values()))
    top_source, top_count = max(counts.items(), key=lambda item: item[1])
    entropy = 0.0
    for count in counts.values():
        prob = float(count) / total
        if prob > 0:
            entropy -= prob * math.log(prob, 2)
    max_entropy = math.log(len(counts), 2) if len(counts) > 1 else 0.0
    normalized = (entropy / max_entropy) if max_entropy > 0 else 0.0
    return {
        "top_source": top_source,
        "top_source_share": float(top_count) / total,
        "source_entropy": entropy,
        "source_entropy_normalized": normalized,
    }


def _final_admission_backstop(selected_docs: List[Any], dense_source_scores: Dict[str, float]) -> Dict[str, Any]:
    metrics = _source_entropy_metrics(selected_docs, top_n=5)
    top_source = metrics.get("top_source")
    top_source_docs = [
        hit for hit in (selected_docs or [])
        if _normalize_filename_for_match(_hit_entity_source(hit) or "") == top_source
    ]
    hybrid_score = max([_hit_score(hit) for hit in top_source_docs] or [0.0])
    dense_score = float(dense_source_scores.get(top_source or "", 0.0))
    mud_line = float(getattr(config, "FUSION_MUD_SCORE", 0.018))
    dense_floor = float(getattr(config, "DENSE_BACKSTOP_MIN_SCORE", 0.55))
    triggered = bool(top_source and (hybrid_score < mud_line) and (dense_score < dense_floor))
    logger.info(
        f"DEBUG GATE -> top_source={top_source} hybrid={hybrid_score:.4f} dense={dense_score:.4f} "
        f"mud_line={mud_line:.4f} dense_floor={dense_floor:.4f} blocked={triggered}"
    )
    metrics.update({
        "top_source_hybrid_score": hybrid_score,
        "top_source_dense_score": dense_score,
        "mud_line": mud_line,
        "dense_floor": dense_floor,
        "triggered": triggered,
    })
    return metrics


def _should_defer_cluster_to_backstop(docs: List[Any], score_mode: str, thr: Dict[str, float]) -> bool:
    if (not docs) or score_mode != "score":
        return False
    best = max([_hit_score(d) for d in docs[: min(3, len(docs))]] or [0.0])
    min_score = float(thr.get("min_score", 0.0))
    mud_line = float(getattr(config, "FUSION_MUD_SCORE", 0.018))
    return min_score <= best <= mud_line


HARD_REFUSAL_REASONS = {
    "no_relevant_evidence",
    "empty_evidence",
    "wrong_source",
    "version_mismatch",
    "deleted_or_invisible",
    "heading_only_after_expand",
    "all_generic_no_body",
    "low_relevance_after_rescue",
    "off_topic_in_document",
    "no_body_clause_evidence",
}


RESCUABLE_EVIDENCE_REASONS = {
    "heading_only_evidence",
    "insufficient_evidence",
    "low_evidence_relevance",
    "section_not_hit",
    "topic_not_hit",
    "generic_only",
    "partial_term_coverage",
}


LEGAL_TERM_CLUSTERS = {
    "处罚": ["处罚", "罚款", "警告", "责令改正", "责令限期改正", "没收", "吊销", "追究责任", "法律责任", "依法处理"],
    "奖励": ["奖励", "表彰", "鼓励", "扶持", "补助", "资助", "给予表彰", "给予奖励"],
    "登记": ["登记", "备案", "注册", "申请登记", "办理登记"],
    "管理要求": ["应当", "不得", "禁止", "规范", "要求", "义务", "责任", "管理", "监督"],
    "限制": ["不得", "禁止", "限制", "不得携带", "不得进入", "限期", "区域", "时间"],
    "许可": ["许可", "审批", "批准", "申请", "核准", "备案"],
    "执法权限": ["执法权限", "执法权", "行政执法权限", "行政执法权", "职权", "权限范围", "职责权限"],
    "物业服务主体": ["物业服务企业", "物业服务人"],
    "违法停车": ["违法停车", "违规停车", "违规停放", "违停", "不按规定停车", "不按规定停放", "机动车违法停放"],
    "养犬": ["养犬", "犬只", "携犬", "携犬出户", "遛犬"],
    "河道": ["河道", "河流", "河湖", "水域", "河道管理", "河道保护"],
}


def _evidence_reason_is_hard_refusal(reason: Optional[str]) -> bool:
    return bool(reason) and str(reason) in HARD_REFUSAL_REASONS


def _evidence_reason_is_rescuable(reason: Optional[str]) -> bool:
    return bool(reason) and str(reason) in RESCUABLE_EVIDENCE_REASONS


def _text_has_legal_action_signal(text: str) -> bool:
    content = _normalize_query(text)
    if not content:
        return False
    action_terms = [
        "应当", "不得", "可以", "责令", "处罚", "罚款", "警告", "没收", "吊销", "追究", "登记", "备案", "批准", "许可",
        "申请", "办理", "改正", "表彰", "奖励", "扶持", "监督", "管理", "法律责任",
    ]
    if any(term in content for term in action_terms):
        return True
    return bool(re.search(r"第[一二三四五六七八九十百千0-9]+[章节条款]", content))


def _is_substantive_short_legal_evidence(doc: Any) -> bool:
    text = (_hit_display_text(doc) or "").strip()
    if not text:
        return False
    md = _hit_metadata(doc)
    chunk_role = str(md.get("chunk_role") or "").strip()
    if chunk_role in {"title", "chapter_heading", "section_heading", "toc", "toc_heading", "appendix_heading"}:
        return False
    if len(text) >= 24:
        return True
    if str(md.get("article_no") or md.get("clause_label") or "").strip():
        return True
    return _text_has_legal_action_signal(text)


def _lightweight_evidence_reason(query: str, docs: List[Any]) -> Optional[str]:
    evidence_docs = [doc for doc in docs if not _is_heading_only_hit(doc)]
    if not evidence_docs:
        return "heading_only_evidence" if docs else "no_relevant_evidence"
    window_limit = max(3, int(getattr(config, "PRIMARY_EVIDENCE_TOPK", 5)))
    window = evidence_docs[: min(len(evidence_docs), window_limit)]
    if not window:
        return "no_relevant_evidence"
    if not any((_hit_display_text(doc) or "").strip() for doc in window):
        return "empty_evidence"
    evidence_chars = sum(len((_hit_display_text(doc) or "").strip()) for doc in window)
    if evidence_chars < 24 and not any(_is_substantive_short_legal_evidence(doc) for doc in window):
        return "insufficient_evidence"
    return None


def _doc_section_name(hit: Any) -> str:
    md = _hit_metadata(hit)
    return (md.get("section_title") or md.get("section") or "").strip()


def _is_heading_only_hit(hit: Any) -> bool:
    md = _hit_metadata(hit)
    chunk_role = str(md.get("chunk_role") or "").strip()
    return chunk_role in {"title", "chapter_heading", "section_heading", "toc", "toc_heading", "appendix_heading"}


def _has_clause_like_body_evidence(hit: Any) -> bool:
    if _is_heading_only_hit(hit):
        return False
    md = _hit_metadata(hit)
    chunk_role = str(md.get("chunk_role") or "").strip()
    if chunk_role in {"article", "clause"}:
        return True
    if str(md.get("article_no") or md.get("clause_label") or "").strip():
        return True
    text = _hit_display_text(hit) or ""
    return bool(re.search(r"第[一二三四五六七八九十百千0-9]+条", text))


def _source_title_aspect_terms(target_sources: Optional[List[str]]) -> List[str]:
    out: List[str] = []
    for source in target_sources or []:
        safe_source = _normalize_filename_for_match(source or "")
        if not safe_source:
            continue
        for candidate in _doc_title_alias_candidates(safe_source):
            value = _normalize_query(candidate)
            if len(value) >= 2 and value not in out:
                out.append(value)
    return out


def _canonical_source_core_entities(source: str) -> List[str]:
    info = _doc_get(source)
    title = _normalize_query(info.get("canonical_title") or _filename_stem(source) or source)
    if not title:
        return []
    variants: List[str] = []
    stripped = _normalize_query(_strip_leading_region_prefix(title))
    for candidate in [title, stripped]:
        value = _normalize_query(candidate)
        if not value:
            continue
        value = re.sub(r"(条例|办法|规定|规则|细则|通知|通告|决定|议事规则|实施办法|管理条例|管理办法)$", "", value)
        value = _normalize_query(value)
        if len(value) >= 3 and value not in variants:
            variants.append(value)
    return variants


def _source_alias_entity_hit(source: str, matched_text: str) -> bool:
    needle = _normalize_reference_text(matched_text)
    if len(needle) < 4:
        return False
    for alias in _doc_title_alias_candidates(source):
        alias_norm = _normalize_reference_text(alias)
        if not alias_norm:
            continue
        if needle == alias_norm or needle in alias_norm or alias_norm in needle:
            return True
    return False


def _source_has_topic_anchor(source: str, anchors: List[str]) -> bool:
    if not anchors:
        return True
    haystacks = _doc_profile_topic_terms(source) + _doc_profile_section_titles(source) + _doc_title_alias_candidates(source)
    hay = "\n".join(_normalize_query(item) for item in haystacks if _normalize_query(item))
    if not hay:
        return False
    for anchor in anchors:
        variants = _legal_term_cluster_variants(anchor)
        if any(variant and variant in hay for variant in variants):
            return True
    return False


def _source_body_anchor_match_count(source: str, anchors: List[str]) -> int:
    safe_source = _normalize_filename_for_match(source or "")
    if not safe_source or not anchors:
        return 0
    if not _source_state(safe_source).get("visible"):
        return 0
    conn = _lex_db_connect()
    matched = 0
    for anchor in anchors:
        hit = False
        variants = [variant for variant in _coverage_aspect_variants(anchor) if len(_normalize_query(variant)) >= 2]
        for variant in variants[:6]:
            pattern = f"%{variant}%"
            section_row = conn.execute(
                "SELECT 1 FROM document_sections WHERE source = ? AND (section_title LIKE ? OR section_path LIKE ?) LIMIT 1",
                (safe_source, pattern, pattern),
            ).fetchone()
            if section_row:
                hit = True
                break
            body_row = conn.execute(
                "SELECT 1 FROM chunks_meta m JOIN chunks_fts f ON f.rowid = m.id "
                "WHERE m.source = ? AND (m.section LIKE ? OR f.text LIKE ?) LIMIT 1",
                (safe_source, pattern, pattern),
            ).fetchone()
            if body_row:
                hit = True
                break
        if hit:
            matched += 1
    return matched


def _unique_alias_lock_resolution(
    query: str,
    candidate_sources: List[str],
    matched_text: str,
    ranked_matches: Optional[List[Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    alias_text = _normalize_query(matched_text)
    candidates = [
        _normalize_filename_for_match(source or "")
        for source in (candidate_sources or [])
        if _normalize_filename_for_match(source or "")
    ]
    candidates = list(dict.fromkeys(candidates))[:3]
    if len(alias_text) >= 8 or not candidates:
        return {"resolved": False}

    title_score_map: Dict[str, float] = {}
    for entry in ranked_matches or []:
        source = _normalize_filename_for_match((entry or {}).get("source") or "")
        if source in candidates:
            title_score_map[source] = max(title_score_map.get(source, 0.0), float((entry or {}).get("score") or 0.0))
    for probe in [matched_text, query]:
        if not probe:
            continue
        for entry in _rank_title_source_matches(probe, limit=max(len(candidates) * 2, 6), include_topic_like=False):
            source = _normalize_filename_for_match((entry or {}).get("source") or "")
            if source in candidates:
                title_score_map[source] = max(title_score_map.get(source, 0.0), float((entry or {}).get("score") or 0.0))

    scored: List[Dict[str, Any]] = []
    for source in candidates:
        source_title_terms = _source_title_aspect_terms([source])
        content_anchors = _query_content_anchor_terms(query, qfilters=None, source_title_terms=source_title_terms)
        section_anchors = _local_validate_section_targets(_extract_section_query_targets(query), limit=4)
        anchors: List[str] = []
        for anchor in content_anchors + section_anchors:
            normalized = _normalize_query(anchor)
            if len(normalized) >= 2 and normalized not in anchors:
                anchors.append(normalized)
        title_score = float(title_score_map.get(source, 0.0))
        entity_hit = _source_alias_entity_hit(source, matched_text)
        topic_hit = _source_has_topic_anchor(source, anchors)
        body_anchor_hits = _source_body_anchor_match_count(source, anchors)
        anchor_ratio = float(body_anchor_hits) / float(max(1, len(anchors))) if anchors else 0.0
        total_score = min(title_score / 10.0, 1.0)
        if entity_hit:
            total_score += 0.45
        if topic_hit:
            total_score += 0.65
        total_score += 0.95 * min(anchor_ratio, 1.0)
        scored.append({
            "source": source,
            "title_score": title_score,
            "entity_hit": entity_hit,
            "topic_hit": topic_hit,
            "body_anchor_hits": body_anchor_hits,
            "anchor_count": len(anchors),
            "score": total_score,
        })

    scored.sort(key=lambda item: (-float(item["score"]), -float(item["title_score"]), item["source"]))
    top = scored[0] if scored else None
    second = scored[1] if len(scored) > 1 else None
    if not top:
        return {"resolved": False}
    margin = float(top["score"]) - float(second["score"]) if second else float(top["score"])
    if float(top["title_score"]) < 7.0:
        return {"resolved": False, "candidates": scored}
    if not bool(top["entity_hit"]):
        return {"resolved": False, "candidates": scored}
    if int(top["body_anchor_hits"]) <= 0:
        return {"resolved": False, "candidates": scored}
    if not bool(top["topic_hit"]) and int(top["body_anchor_hits"]) < max(1, int(top["anchor_count"])):
        return {"resolved": False, "candidates": scored}
    if float(top["score"]) < 1.95:
        return {"resolved": False, "candidates": scored}
    if second and margin < 0.35:
        return {"resolved": False, "candidates": scored}
    return {
        "resolved": True,
        "source": top["source"],
        "reason": "unique_alias_lock",
        "score": float(top["score"]),
        "margin": float(margin),
        "candidates": scored,
    }


def _short_alias_lock_allowed(query: str, source: str, matched_text: str) -> bool:
    alias_text = _normalize_query(matched_text)
    if len(alias_text) >= 8:
        return True
    core_entities = _canonical_source_core_entities(source)
    if core_entities and not any(entity and (entity in alias_text or alias_text in entity) for entity in core_entities):
        return False
    anchors = _query_content_anchor_terms(query, qfilters=None, source_title_terms=_source_title_aspect_terms([source]))
    return _source_has_topic_anchor(source, anchors)


def _query_content_anchor_terms(query: str, qfilters: Optional[Dict[str, Any]], source_title_terms: List[str]) -> List[str]:
    llm_override: Optional[List[str]] = None
    llm_extra: List[str] = []
    if isinstance(qfilters, dict):
        override = qfilters.get("_llm_anchor_override")
        if isinstance(override, list):
            llm_override = [str(x).strip() for x in override if str(x).strip()]
        extra = qfilters.get("_llm_anchor_extra")
        if isinstance(extra, list):
            llm_extra = [str(x).strip() for x in extra if str(x).strip()]
    semantic_terms = llm_override if llm_override is not None else (_query_semantic_aspects(query, qfilters=qfilters).get("terms") or [])
    generic_tail_terms = {
        "规定", "要求", "内容", "申请", "办理", "流程", "程序", "处罚", "奖励", "限制", "管理", "责任",
    }
    anchors: List[str] = []
    for term in list(semantic_terms) + list(llm_extra):
        normalized = _normalize_coverage_aspect(term) or _normalize_query(term)
        if len(normalized) < 2 or _is_question_wrapper_aspect(normalized):
            continue
        if _aspect_is_subsumed_by_covered(normalized, source_title_terms):
            continue
        if _looks_like_section_target(normalized):
            continue
        reduced = re.sub(r"(怎么|如何|怎样|是否|能否)?(申请|办理|规定|要求|流程|程序|处罚|奖励|限制|管理|责任)$", "", normalized)
        reduced = re.sub(r"(有哪些|有什么|是什么|内容|规定)$", "", reduced)
        reduced = _normalize_query(reduced) or normalized
        if len(reduced) < 2 or reduced in generic_tail_terms:
            continue
        if reduced not in anchors:
            anchors.append(reduced)
    return anchors


_MEANINGLESS_ASPECT_TERMS = {
    "里",
    "中",
    "内",
    "关于",
    "对于",
    "对",
    "有关",
    "以及",
    "与",
    "和",
    "或",
    "及",
    "在",
    "的",
    "版",
    "版里",
    "版中",
    "版内",
}


def _is_meaningless_coverage_aspect(term: str) -> bool:
    t = _normalize_query(term)
    if not t:
        return True
    if t in _MEANINGLESS_ASPECT_TERMS:
        return True
    if len(t) <= 2 and any(marker in t for marker in ["里", "中", "内", "的"]):
        return True
    if re.fullmatch(r"[0-9]{4}", t):
        return True
    if re.fullmatch(r"[0-9]{2,4}版", t):
        return True
    if re.fullmatch(r"[0-9]{2,4}年", t):
        return True
    if re.fullmatch(r"[0-9]{4}(年)?(版里|版中|版内)", t):
        return True
    if re.fullmatch(r"[0-9]{4}(年)?(版|版本|修订|修正|修订版|修正版|发布|公布|颁布)", t):
        return True
    if re.fullmatch(r"(现行|最新)(有效)?(版|版本)?", t):
        return True
    return False


def _source_identity_terms_for_validation(target_sources: Optional[List[str]]) -> List[str]:
    out: List[str] = []
    for src in target_sources or []:
        safe = _normalize_filename_for_match(src or "")
        if not safe:
            continue
        for t in _doc_title_alias_candidates(safe):
            v = _normalize_query(t)
            if len(v) >= 2 and v not in out:
                out.append(v)
        for ent in _canonical_source_core_entities(safe):
            v = _normalize_query(ent)
            if len(v) >= 2 and v not in out:
                out.append(v)
    return out


def _definition_intent_query(query: str) -> bool:
    q = _normalize_query(query)
    if not q:
        return False
    patterns = [
        r"哪些情形属于",
        r"什么是",
        r"何为",
        r"定义",
        r"是指什么",
        r"包括哪些情形",
        r"(权限|职权|职责).{0,6}(有哪些|是什么|包括|范围|如何)",
    ]
    return any(re.search(p, q) for p in patterns)


def _definition_chunk_bonus(query: str, hit: Any) -> float:
    if not _definition_intent_query(query):
        return 0.0
    section = _doc_section_name(hit)
    text = _hit_display_text(hit) or ""
    hay = f"{section}\n{text}"
    bonus = 0.0
    if any(marker in section for marker in ["总则", "名词解释", "术语", "定义", "附则", "执法权限", "职责", "职权", "权限"]):
        bonus += 0.06
    if any(marker in hay for marker in ["是指", "指", "包括下列", "包含下列", "包括：", "包含：", "是指："]):
        bonus += 0.06
    if re.search(r"第[一二三四五六七八九十百千0-9]+条.*(是指|指)", hay):
        bonus += 0.05
    return float(min(0.18, bonus))


def _delegated_to_external_source_snippet(query: str, window: List[Any]) -> str:
    if not window:
        return ""
    q = _normalize_query(query)
    if not q:
        return ""
    intent_markers = ["标准", "办法", "细则", "规定", "程序", "条件", "范围", "情形", "权限", "职责", "认定"]
    if not any(marker in q for marker in intent_markers):
        return ""
    direct_answer_signals: List[str] = []
    if "情形" in q:
        direct_answer_signals.extend(["属于下列情形之一", "下列情形", "可以在", "不得在"])
    if "条件" in q:
        direct_answer_signals.extend(["符合下列条件", "下列条件", "应当符合", "可以在"])
    if "职责" in q or "责任" in q:
        direct_answer_signals.extend(["负责", "职责", "责任", "主管部门", "应当"])
    if direct_answer_signals:
        for doc in window:
            if _is_heading_only_hit(doc):
                continue
            hay = "\n".join([_doc_section_name(doc), _hit_display_text(doc) or ""])
            if any(signal in hay for signal in direct_answer_signals):
                return ""
    patterns = [
        r"另行制定",
        r"另行规定",
        r"另行公布",
        r"另行发布",
        r"另行确定",
        r"另行明确",
        r"另行通知",
        r"由[^。；\n]{0,24}(制定|规定|公布|发布|确定|明确)",
        r"具体[^。；\n]{0,8}(办法|标准|规定|要求)[^。；\n]{0,6}由[^。；\n]{0,24}(制定|规定|公布|发布|确定|明确)",
    ]
    for doc in window:
        if _is_heading_only_hit(doc):
            continue
        text = (_hit_display_text(doc) or "").strip()
        if not text:
            continue
        for pat in patterns:
            m = re.search(pat, text)
            if m:
                snippet = text[max(0, m.start() - 30): m.end() + 30]
                snippet = re.sub(r"\s+", " ", snippet).strip()
                return snippet[:240]
    return ""


def _delegated_chunk_bonus(query: str, hit: Any) -> float:
    q = _normalize_query(query)
    if not q:
        return 0.0
    intent_markers = ["标准", "办法", "细则", "规定", "程序", "条件", "范围", "情形", "权限", "职责", "认定"]
    if not any(marker in q for marker in intent_markers):
        return 0.0
    section = _doc_section_name(hit)
    text = _hit_display_text(hit) or ""
    hay = f"{section}\n{text}"
    patterns = [
        r"另行制定",
        r"另行规定",
        r"另行公布",
        r"另行发布",
        r"另行确定",
        r"另行明确",
        r"另行通知",
        r"由[^。；\n]{0,24}(制定|规定|公布|发布|确定|明确)",
        r"具体[^。；\n]{0,8}(办法|标准|规定|要求)[^。；\n]{0,6}由[^。；\n]{0,24}(制定|规定|公布|发布|确定|明确)",
    ]
    if any(re.search(pat, hay) for pat in patterns):
        return 0.08
    return 0.0


def _body_anchor_hits(query: str, docs: List[Any], qfilters: Optional[Dict[str, Any]], target_sources: Optional[List[str]]) -> Dict[str, Any]:
    source_title_terms = _source_title_aspect_terms(target_sources)
    anchors = _query_content_anchor_terms(query, qfilters, source_title_terms)
    body_text = "\n".join(
        _hit_display_text(doc) or ""
        for doc in (docs or [])
        if _has_clause_like_body_evidence(doc)
    )
    covered: List[str] = []
    for anchor in anchors:
        variants = _legal_term_cluster_variants(anchor)
        if any(variant and variant in body_text for variant in variants):
            covered.append(anchor)
    return {
        "anchors": anchors,
        "covered": covered,
        "all_covered": bool(anchors) and len(covered) == len(anchors),
        "body_anchor_hit": bool(covered),
    }


def _is_wide_topic_query(query: str, qfilters: Optional[Dict[str, Any]] = None, target_sources: Optional[List[str]] = None) -> bool:
    q = _normalize_query(query)
    if not q:
        return False
    if _extract_section_query_targets(q):
        return False
    anchors = _query_content_anchor_terms(q, qfilters, _source_title_aspect_terms(target_sources))
    if len(anchors) >= 2:
        return True
    broad_patterns = [
        r"有哪些规定$",
        r"有哪些要求$",
        r"有哪些内容$",
        r"怎么规定$",
        r"如何规定$",
    ]
    return any(re.search(pattern, q) for pattern in broad_patterns)


def _is_off_topic_locked_document_query(
    query: str,
    docs: List[Any],
    qfilters: Optional[Dict[str, Any]] = None,
    target_sources: Optional[List[str]] = None,
    source_lock_resolved: bool = False,
) -> bool:
    if _is_doc_existence_query(query) or _is_deleted_visibility_query(query) or _is_version_switch_query(query):
        return False
    normalized_targets = [
        _normalize_filename_for_match(source or "")
        for source in (target_sources or [])
        if _normalize_filename_for_match(source or "")
    ]
    if not source_lock_resolved or len(normalized_targets) != 1:
        return False
    source_title_terms = _source_title_aspect_terms(normalized_targets)
    content_anchors = _query_content_anchor_terms(query, qfilters, source_title_terms)
    if not content_anchors:
        return False
    target_source = normalized_targets[0]
    haystack_parts: List[str] = []
    for doc in docs or []:
        if _normalize_filename_for_match(_hit_entity_source(doc) or "") != target_source:
            continue
        if _is_heading_only_hit(doc):
            continue
        if not _hit_matches_source_state(doc, _source_state(target_source)):
            continue
        haystack_parts.append(_doc_section_name(doc))
        haystack_parts.append(_hit_display_text(doc) or "")
    haystack = "\n".join(part for part in haystack_parts if part)
    if not haystack:
        return False
    for anchor in content_anchors:
        if any(variant and variant in haystack for variant in _legal_term_cluster_variants(anchor)):
            return False
    if _semantic_softened_aspects(content_anchors, [doc for doc in docs or [] if _normalize_filename_for_match(_hit_entity_source(doc) or "") == target_source]):
        return False
    return True


def _semantic_softened_aspects(aspects: List[str], docs: List[Any]) -> List[str]:
    if not bool(getattr(config, "ENABLE_SEMANTIC_SOFTENING", True)):
        return []
    pending = [str(item).strip() for item in (aspects or []) if str(item).strip()]
    if not pending or not docs:
        return []
    max_chars = max(200, int(getattr(config, "SEMANTIC_SOFTENING_MAX_TEXT_CHARS", 1200)))
    min_sim = float(getattr(config, "SEMANTIC_SOFTENING_MIN_SIM", 0.85))
    haystack_parts: List[str] = []
    for doc in (docs or [])[:4]:
        if _is_heading_only_hit(doc):
            continue
        part = "\n".join([_doc_section_name(doc), _hit_display_text(doc) or ""]).strip()
        if part:
            haystack_parts.append(part)
    haystack = re.sub(r"\s+", " ", "\n".join(haystack_parts)).strip()
    if not haystack:
        return []
    doc_embedding = _embed_text_sync_cached(haystack[:max_chars])
    if not doc_embedding:
        return []
    softened: List[str] = []
    for aspect in pending[:4]:
        probe = (_normalize_coverage_aspect(aspect) or _normalize_query(aspect))[:max_chars]
        if len(probe) < 2:
            continue
        aspect_embedding = _embed_text_sync_cached(probe)
        if not aspect_embedding:
            continue
        similarity = _normalized_embedding_cosine(doc_embedding, aspect_embedding)
        if similarity >= min_sim and aspect not in softened:
            softened.append(aspect)
    return softened


def _is_document_state_query(query: str) -> bool:
    return _is_doc_existence_query(query) or _is_deleted_visibility_query(query) or _is_version_switch_query(query)


def _evidence_chunk_score(
    query: str,
    hit: Any,
    score_mode: str,
    qfilters: Optional[Dict[str, Any]] = None,
    target_sources: Optional[List[str]] = None,
) -> float:
    hybrid_score, _ = _hybrid_structural_chunk_score(
        query,
        hit,
        score_mode,
        profile=_infer_rerank_profile(query, _classify_question_type(query)),
    )
    md = _hit_metadata(hit)
    src = _normalize_filename_for_match(_hit_entity_source(hit) or "")
    section = _doc_section_name(hit)
    text = _hit_display_text(hit) or ""
    haystack = f"{section}\n{text}"
    allowed_sources = {
        _normalize_filename_for_match(source or "")
        for source in (target_sources or [])
        if _normalize_filename_for_match(source or "")
    }
    section_term_hits, text_term_hits, _, _, _, _, _ = _chunk_query_signal(query, hit, score_mode)
    anchor_count = max(1.0, float(len(_query_anchor_terms(query)) or 1))
    term_bonus = 0.12 * _clip01((section_term_hits + text_term_hits) / anchor_count)
    semantic_terms = _query_semantic_aspects(query, qfilters=qfilters).get("terms") or []
    semantic_hits = 0
    for term in semantic_terms:
        variants = _legal_term_cluster_variants(term)
        if any(variant and variant in haystack for variant in variants):
            semantic_hits += 1
    semantic_bonus = 0.14 * _clip01(float(semantic_hits) / float(max(1, len(semantic_terms))))
    section_align, section_exact = _section_target_alignment(section, query)
    section_bonus = 0.08 * section_align + 0.04 * section_exact
    topic_bonus = 0.0
    if (qfilters or {}).get("topic") and qfilters["topic"] in _normalize_topics(md.get("topics")):
        topic_bonus = 0.08
    source_bonus = 0.10 if allowed_sources and src in allowed_sources else 0.0
    definition_bonus = _definition_chunk_bonus(query, hit)
    delegated_bonus = _delegated_chunk_bonus(query, hit)
    heading_penalty = 0.24 if _is_heading_only_hit(hit) else 0.0
    return _clip01(
        _clip01(float(hybrid_score))
        + source_bonus
        + term_bonus
        + semantic_bonus
        + section_bonus
        + topic_bonus
        + definition_bonus
        + delegated_bonus
        - heading_penalty
    )


def _rescue_candidate_allowed(
    query: str,
    hit: Any,
    score_mode: str,
    qfilters: Optional[Dict[str, Any]] = None,
    target_sources: Optional[List[str]] = None,
    semantic_terms: Optional[List[str]] = None,
) -> bool:
    src = _normalize_filename_for_match(_hit_entity_source(hit) or "")
    if not src:
        return False
    allowed_sources = {
        _normalize_filename_for_match(source or "")
        for source in (target_sources or [])
        if _normalize_filename_for_match(source or "")
    }
    if allowed_sources and src not in allowed_sources:
        return False
    if not _hit_matches_source_state(hit, _source_state(src)):
        return False
    if _is_heading_only_hit(hit):
        return False

    section = _doc_section_name(hit)
    if not _is_generic_section_title(section):
        return True
    if _is_substantive_short_legal_evidence(hit):
        return True
    md = _hit_metadata(hit)
    if (qfilters or {}).get("topic") and qfilters["topic"] in _normalize_topics(md.get("topics")):
        return True
    section_align, _ = _section_target_alignment(section, query)
    if section_align > 0:
        return True
    section_term_hits, text_term_hits, section_score, keyword_score, _, title_signal, _ = _chunk_query_signal(query, hit, score_mode)
    if section_term_hits > 0 or text_term_hits > 0 or section_score >= 1.0 or keyword_score >= 1.0 or title_signal > 0:
        return True
    hay = f"{section}\n{_hit_display_text(hit) or ''}"
    for term in semantic_terms or []:
        if term and term in hay:
            return True
    return False


def _qualified_substantive_evidence_docs(
    query: str,
    docs: List[Any],
    qfilters: Optional[Dict[str, Any]] = None,
    target_sources: Optional[List[str]] = None,
    min_score_override: Optional[float] = None,
) -> List[Any]:
    evidence_docs = [doc for doc in (docs or []) if not _is_heading_only_hit(doc)]
    if not evidence_docs:
        return []
    score_mode = _hit_score_mode(evidence_docs[0]) if evidence_docs else "score"
    min_score = _clip01(float(min_score_override if min_score_override is not None else getattr(config, "MIN_EVIDENCE_SCORE", 0.6)))
    decorated = []
    for idx, doc in enumerate(evidence_docs):
        score = _evidence_chunk_score(query, doc, score_mode, qfilters=qfilters, target_sources=target_sources)
        if score >= min_score or _is_substantive_short_legal_evidence(doc):
            decorated.append((score, -idx, doc))
    decorated.sort(key=lambda item: (item[0], item[1]), reverse=True)
    return [doc for _, _, doc in decorated]


def _looks_like_section_target(term: str) -> bool:
    t = (term or "").strip()
    if len(t) < 2:
        return False
    if re.match(r"^第[一二三四五六七八九十0-9]+[章节条款编]$", t):
        return True
    named_targets = {
        "总则", "附则", "罚则", "法律责任", "奖励与处罚", "奖惩", "立法程序", "适用范围", "职责分工", "监督管理",
        "养犬行为规范", "养犬区划管理、免疫与登记",
    }
    if t in named_targets:
        return True
    if any(marker in t for marker in ["有哪些", "什么", "如何", "怎么", "是否"]):
        return False
    return any(token in t for token in ["规范", "管理", "登记", "免疫", "责任", "处罚", "奖励", "程序", "监督", "区划"])


def _query_semantic_aspects(query: str, qfilters: Optional[Dict[str, Any]] = None) -> Dict[str, List[str]]:
    aspects: List[str] = []
    q = _normalize_query(query)
    if isinstance(qfilters, dict):
        override = qfilters.get("_llm_aspects_override")
        if isinstance(override, list):
            raw_terms = [str(x).strip() for x in override if str(x).strip()]
            terms = [_normalize_coverage_aspect(x) for x in raw_terms]
            terms = [t for t in terms if t and not _is_question_wrapper_aspect(t)]
            topics: List[str] = []
            topic = ((qfilters or {}).get("topic") or "").strip()
            if topic:
                topics.append(topic)
            return {"terms": list(dict.fromkeys(terms))[:8], "topics": topics}
    stop_terms = {
        "条例", "办法", "规定", "规则", "条款", "内容", "相关", "什么", "哪些", "是否", "如何", "怎么",
        "说明", "根据", "请", "一下", "一下子", "内容是什么", "有哪些", "规定是什么",
    }

    def _add_aspect(term: str):
        t = (term or "").strip()
        t = _normalize_coverage_aspect(t)
        if len(t) < 2 or t in stop_terms:
            return
        if _is_question_wrapper_aspect(t):
            return
        if re.search(r"[和与及、]", t) and not _looks_like_section_target(t):
            return
        if t not in aspects:
            aspects.append(t)

    normalized = re.sub(r"[，。；;：:?？]", " ", q)
    for piece in re.split(r"\s+", normalized):
        part = piece.strip()
        if not part:
            continue
        part = re.sub(r"^(关于|根据|请问|请|查询|说明|第[一二三四五六七八九十0-9]+章的)", "", part)
        part = re.sub(r"(有哪些规定|有哪些要求|有哪些内容|是什么内容|是什么|如何处理|怎么处理|如何规定|怎么规定|怎么办|怎么做|吗|呢|有哪些)$", "", part)
        for sub in re.split(r"[、和及与]", part):
            _add_aspect(sub)

    for term in _extract_section_query_targets(q):
        if _looks_like_section_target(term):
            _add_aspect(term)
    for term in _query_match_terms(q):
        _add_aspect(term)

    if isinstance(qfilters, dict):
        extra = qfilters.get("_llm_aspects_extra")
        if isinstance(extra, list):
            for item in extra:
                _add_aspect(str(item))

    topics: List[str] = []
    topic = ((qfilters or {}).get("topic") or "").strip()
    if topic:
        topics.append(topic)

    return {"terms": aspects[:8], "topics": topics}


def _normalize_coverage_aspect(term: str) -> str:
    text = _normalize_query(term)
    if not text:
        return ""
    text = re.sub(r"^(关于|根据|请问|请|查询|说明|对)", "", text)
    text = re.sub(r"(有哪些规定|有哪些要求|有哪些内容|是什么内容|是什么|如何处理|怎么处理|如何规定|怎么规定|怎么办|怎么做|吗|呢|有哪些)$", "", text)
    if "的" in text:
        prefix, suffix = text.rsplit("的", 1)
        if len(suffix) >= 2 and any(marker in prefix for marker in ["条例", "办法", "规定", "规则", "法规"]):
            text = suffix
    return _normalize_query(text)


def _coverage_aspect_variants(term: str) -> List[str]:
    variants: List[str] = []
    for candidate in [term, _normalize_coverage_aspect(term)]:
        value = _normalize_query(candidate)
        if len(value) >= 2 and value not in variants:
            variants.append(value)
        for token in re.split(r"[、和及与/\s]+", value):
            token = _normalize_query(token)
            if len(token) >= 2 and token not in variants:
                variants.append(token)
    return variants


def _legal_term_cluster_variants(term: str) -> List[str]:
    variants = list(_coverage_aspect_variants(term))
    for cluster_key, cluster_terms in LEGAL_TERM_CLUSTERS.items():
        cluster_variants = []
        for item in [cluster_key] + list(cluster_terms or []):
            value = _normalize_query(item)
            if len(value) >= 2 and value not in cluster_variants:
                cluster_variants.append(value)
        if not cluster_variants:
            continue
        matched = False
        for left in variants:
            for right in cluster_variants:
                if left == right or left in right or right in left:
                    matched = True
                    break
            if matched:
                break
        if matched:
            for value in cluster_variants:
                if value not in variants:
                    variants.append(value)
    return variants


def _is_question_wrapper_aspect(term: str) -> bool:
    text = _normalize_query(term)
    if not text:
        return False
    if any(marker in text for marker in ["有哪些", "有何", "什么", "如何", "怎么", "是否"]):
        return True
    if re.match(r"^对?.+(限制|要求|规定|处罚|流程|程序)$", text):
        return True
    return False


STRONG_SECTION_TARGET_NAMES = {"总则", "附则", "罚则", "法律责任", "立法程序", "监督检查", "监督管理"}
WEAK_SECTION_TARGET_NAMES = {"奖励与处罚", "奖惩", "处罚规定", "行为规范", "管理要求", "办理流程", "登记要求", "限制措施", "处罚", "奖励", "登记", "程序", "流程"}


def _split_section_targets(query: str) -> Tuple[List[str], List[str]]:
    strong_targets: List[str] = []
    weak_raw: List[str] = []
    for target in _extract_section_query_targets(query):
        normalized = _normalize_query(target)
        if len(normalized) < 2:
            continue
        if re.match(r"^第[一二三四五六七八九十百千0-9]+[章节条款编]$", normalized) or normalized in STRONG_SECTION_TARGET_NAMES:
            if normalized not in strong_targets:
                strong_targets.append(normalized)
            continue
        weak_raw.append(normalized)
    weak_targets = _local_validate_section_targets(list(dict.fromkeys(weak_raw)), limit=10)
    return strong_targets, weak_targets


def _aspect_is_subsumed_by_covered(term: str, covered_terms: List[str]) -> bool:
    term_variants = _coverage_aspect_variants(term)
    if not term_variants:
        return False
    for covered in covered_terms or []:
        covered_variants = _coverage_aspect_variants(covered)
        for left in term_variants:
            for right in covered_variants:
                if left == right:
                    return True
                if len(left) >= 2 and left in right:
                    return True
                if len(right) >= 2 and right in left:
                    return True
    return False


def _filter_identity_noise_aspects(aspects: List[str], identity_terms: List[str]) -> List[str]:
    out: List[str] = []
    for aspect in aspects or []:
        normalized = _normalize_coverage_aspect(aspect) or _normalize_query(aspect)
        if not normalized or _is_meaningless_coverage_aspect(normalized):
            continue
        if _aspect_is_subsumed_by_covered(normalized, identity_terms):
            continue
        if normalized not in out:
            out.append(normalized)
    return out


def _same_chunk_relaxed_aspect_coverage(term: str, covered_terms: List[str], docs: List[Any]) -> bool:
    def _compact_text(value: str) -> str:
        return re.sub(r"\s+", "", _normalize_query(value))

    uncovered_variants = [
        _compact_text(variant)
        for variant in _legal_term_cluster_variants(term)
        if _compact_text(variant) and len(_compact_text(variant)) >= 2
    ]
    if not uncovered_variants:
        return False

    covered_variants: List[str] = []
    for covered in covered_terms or []:
        for variant in _legal_term_cluster_variants(covered):
            compact = _compact_text(variant)
            if compact and len(compact) >= 2 and compact not in covered_variants:
                covered_variants.append(compact)
    if not covered_variants:
        return False

    for doc in docs or []:
        if _is_heading_only_hit(doc):
            continue
        text_blob = _compact_text(f"{_doc_section_name(doc)}\n{_hit_display_text(doc) or ''}")
        if not text_blob:
            continue
        if not any(variant in text_blob for variant in uncovered_variants):
            continue
        if any(variant in text_blob for variant in covered_variants):
            return True
    return False


def _intra_doc_focus_score(docs: List[Any], window_size: int = 5) -> float:
    window = list(docs[: min(len(docs), max(1, window_size))]) if docs else []
    if not window:
        return 0.0

    source_counts: Dict[str, int] = {}
    section_counts: Dict[str, int] = {}
    top_source = None
    for doc in window:
        src = _normalize_filename_for_match(_hit_entity_source(doc) or "")
        if src:
            source_counts[src] = source_counts.get(src, 0) + 1
    if source_counts:
        top_source = max(source_counts.items(), key=lambda item: (item[1], item[0]))[0]
    for doc in window:
        src = _normalize_filename_for_match(_hit_entity_source(doc) or "")
        if top_source and src != top_source:
            continue
        section = _doc_section_name(doc) or "__no_section__"
        section_counts[section] = section_counts.get(section, 0) + 1

    top_source_share = (max(source_counts.values()) / float(len(window))) if source_counts else 0.0
    top_section_share = (max(section_counts.values()) / float(sum(section_counts.values()))) if section_counts else 0.0
    return round((0.65 * top_source_share) + (0.35 * top_section_share), 3)


def _dedupe_evidence_docs(docs: List[Any], limit: int) -> List[Any]:
    seen = set()
    out: List[Any] = []
    for doc in docs or []:
        md = _hit_metadata(doc)
        key = (
            _normalize_filename_for_match(_hit_entity_source(doc) or ""),
            int(md.get("chunk_id") or 0),
            (_hit_entity_text(doc) or "")[:96],
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(doc)
        if len(out) >= limit:
            break
    return out


def _build_evidence_candidate_pool(
    query: str,
    docs: List[Any],
    target_sources: Optional[List[str]] = None,
    qfilters: Optional[Dict[str, Any]] = None,
    min_score_override: Optional[float] = None,
) -> List[Any]:
    rescue_topk = max(5, int(getattr(config, "RESCUE_EVIDENCE_TOPK", 15)))
    pool = list((docs or [])[:rescue_topk])
    state_scoped: List[Any] = []
    for doc in pool:
        src = _normalize_filename_for_match(_hit_entity_source(doc) or "")
        if not src:
            continue
        if _hit_matches_source_state(doc, _source_state(src)):
            state_scoped.append(doc)
    if state_scoped:
        pool = state_scoped
    allowed_sources = [
        _normalize_filename_for_match(source or "")
        for source in (target_sources or [])
        if _normalize_filename_for_match(source or "")
    ]
    if allowed_sources:
        scoped = [doc for doc in pool if _normalize_filename_for_match(_hit_entity_source(doc) or "") in allowed_sources]
        if scoped:
            pool = scoped
    if pool and all(_is_heading_only_hit(doc) for doc in pool):
        expanded: List[Any] = []
        grouped: Dict[str, List[Any]] = {}
        for doc in pool:
            source = _normalize_filename_for_match(_hit_entity_source(doc) or "")
            if not source:
                continue
            grouped.setdefault(source, []).append(doc)
        for source, source_hits in grouped.items():
            expanded.extend(_expand_heading_hits_to_article_hits(query, source, source_hits, limit=rescue_topk))
        if expanded:
            pool = expanded
    body_docs = [doc for doc in pool if not _is_heading_only_hit(doc)]
    if not body_docs:
        return []
    score_mode = _hit_score_mode(body_docs[0]) if body_docs else "score"
    semantic_terms = _query_semantic_aspects(query, qfilters=qfilters).get("terms") or []
    allowed_body_docs = [
        doc for doc in body_docs
        if _rescue_candidate_allowed(
            query,
            doc,
            score_mode,
            qfilters=qfilters,
            target_sources=target_sources,
            semantic_terms=semantic_terms,
        )
    ]
    if allowed_body_docs:
        body_docs = allowed_body_docs
    rescue_min_score = min(
        _clip01(float(getattr(config, "MIN_EVIDENCE_SCORE", 0.6))),
        _clip01(float(min_score_override if min_score_override is not None else getattr(config, "MIN_RESCUE_SCORE", 0.48))),
    )
    filtered = [
        doc for doc in body_docs
        if _evidence_chunk_score(query, doc, score_mode, qfilters=qfilters, target_sources=target_sources) >= rescue_min_score
        or _is_substantive_short_legal_evidence(doc)
    ]
    if not filtered:
        filtered = body_docs
    reranked = _intra_doc_chunk_rerank(query, filtered, score_mode=score_mode, qtype=_classify_question_type(query))
    decorated = [
        (
            _evidence_chunk_score(query, doc, score_mode, qfilters=qfilters, target_sources=target_sources),
            -idx,
            doc,
        )
        for idx, doc in enumerate(reranked)
    ]
    decorated.sort(key=lambda item: (item[0], item[1]), reverse=True)
    return _dedupe_evidence_docs([doc for _, _, doc in decorated], rescue_topk)


def _rescue_min_score_for_reason(reason: str, is_comparison: bool = False) -> float:
    base_score = _clip01(float(getattr(config, "MIN_RESCUE_SCORE", 0.48)))
    if reason == "partial_term_coverage" and not is_comparison:
        relax_ratio = max(0.0, min(1.0, float(getattr(config, "PARTIAL_TERM_RESCUE_RELAX_RATIO", 0.9))))
        return _clip01(base_score * relax_ratio)
    return base_score


def _evaluate_evidence_window(
    query: str,
    docs: List[Any],
    qfilters: Optional[Dict[str, Any]] = None,
    target_sources: Optional[List[str]] = None,
    min_score_override: Optional[float] = None,
    source_lock_resolved: bool = False,
    source_lock_reason: str = "",
    is_comparison: bool = False,
    compare_missing_targets: Optional[List[str]] = None,
) -> Dict[str, Any]:
    evidence_docs = [doc for doc in (docs or []) if not _is_heading_only_hit(doc)]
    refusal_reason = _lightweight_evidence_reason(query, docs or evidence_docs)
    qualified_docs = _qualified_substantive_evidence_docs(
        query,
        evidence_docs,
        qfilters=qfilters,
        target_sources=target_sources,
        min_score_override=min_score_override,
    )
    min_substantive_chunks = max(1, int(getattr(config, "MIN_SUBSTANTIVE_CHUNKS", 1)))
    if not refusal_reason and len(qualified_docs) < min_substantive_chunks:
        refusal_reason = "low_evidence_relevance"

    primary_topk = max(3, int(getattr(config, "PRIMARY_EVIDENCE_TOPK", 5)))
    normalized_targets = [
        _normalize_filename_for_match(source or "")
        for source in (target_sources or [])
        if _normalize_filename_for_match(source or "")
    ]
    if source_lock_resolved and len(normalized_targets) == 1:
        primary_topk = max(primary_topk, 10)
    window = qualified_docs[: min(len(qualified_docs), primary_topk)] if qualified_docs else []
    semantic = _query_semantic_aspects(query, qfilters=qfilters)
    text_blob = "\n".join([_hit_display_text(doc) or "" for doc in window])
    def _compact_match_text(value: str) -> str:
        return re.sub(r"[\s，,、;；:：\-/（）()\[\]{}《》“”\"'<>]+", "", value or "")
    text_blob_compact = _compact_match_text(text_blob)
    source_title_terms: List[str] = []
    for doc in window:
        src = _normalize_filename_for_match(_hit_entity_source(doc) or "")
        if not src:
            continue
        for candidate in _doc_title_alias_candidates(src):
            value = _normalize_query(candidate)
            if len(value) >= 2 and value not in source_title_terms:
                source_title_terms.append(value)

    def _aspect_covered(term: str) -> bool:
        variants = _legal_term_cluster_variants(term)
        if not variants:
            return False
        base_term = _normalize_query(term)
        if "保护" in base_term and len(base_term) >= 4:
            left = _normalize_query(base_term.split("保护", 1)[0])
            if len(left) >= 2 and _compact_match_text(left) in text_blob_compact:
                if any(_compact_match_text(marker) in text_blob_compact for marker in ["保护", "保护区", "保护范围", "保护措施", "保护责任"]):
                    return True
        for variant in variants:
            v_compact = _compact_match_text(str(variant or ""))
            if not v_compact:
                continue
            neg_patterns = [
                rf"未说明.{{0,6}}{re.escape(v_compact)}",
                rf"未提及.{{0,6}}{re.escape(v_compact)}",
                rf"未明确.{{0,6}}{re.escape(v_compact)}",
                rf"不涉及.{{0,6}}{re.escape(v_compact)}",
                rf"未覆盖.{{0,6}}{re.escape(v_compact)}",
                rf"没有.{{0,6}}{re.escape(v_compact)}",
            ]
            if any(re.search(pattern, text_blob_compact) for pattern in neg_patterns):
                continue
            if v_compact in text_blob_compact:
                return True
            parts = [
                _compact_match_text(p)
                for p in re.split(r"[\s、,，;；/与和及]+", str(variant or ""))
                if _compact_match_text(p) and len(_compact_match_text(p)) >= 2
            ]
            if len(parts) >= 2 and all(p in text_blob_compact for p in parts):
                return True
            if re.fullmatch(r"[\u4e00-\u9fff]{4,10}", v_compact):
                suffixes = [
                    "法律责任", "违法行为", "行政处罚", "处罚决定",
                    "监督检查", "监督管理", "行政许可",
                    "处罚", "罚款", "责任", "行为",
                    "监督", "检查", "许可", "登记", "备案", "审批",
                    "程序", "流程", "条件", "标准", "措施", "要求",
                    "义务", "权利", "范围", "机关", "部门", "主体", "对象", "原则",
                ]
                for suf in sorted(suffixes, key=len, reverse=True):
                    if not v_compact.endswith(suf):
                        continue
                    prefix = v_compact[: -len(suf)]
                    if len(prefix) < 2:
                        continue
                    if prefix in text_blob_compact and suf in text_blob_compact:
                        return True
        return False

    terms = semantic["terms"]
    if source_lock_resolved and source_lock_reason == "exact_title_unique":
        identity_terms = _source_identity_terms_for_validation(target_sources)
        if identity_terms:
            terms = [
                term
                for term in terms
                if term
                and not _aspect_is_subsumed_by_covered(term, identity_terms)
                and not _aspect_is_subsumed_by_covered(_normalize_query(term).rstrip("里中内"), identity_terms)
            ]
    covered_terms = [term for term in terms if _aspect_covered(term)]
    uncovered_terms = [
        term for term in terms
        if term
        and term not in covered_terms
        and not _aspect_is_subsumed_by_covered(term, covered_terms)
        and not _aspect_is_subsumed_by_covered(term, source_title_terms)
    ]

    strong_section_targets, weak_section_targets = _split_section_targets(query)
    section_targets = strong_section_targets + [term for term in weak_section_targets if term not in strong_section_targets]

    section_hit = False
    weak_section_hit = False
    if section_targets:
        for doc in window:
            align, _ = _section_target_alignment(_doc_section_name(doc), query)
            if align > 0:
                section_hit = True
                weak_section_hit = True
                break
            doc_text = _hit_display_text(doc) or ""
            doc_section = _doc_section_name(doc)
            if any(target and (target in doc_section or target in doc_text) for target in strong_section_targets):
                section_hit = True
                break
            if any(target and (target in doc_section or target in doc_text or _aspect_covered(target)) for target in weak_section_targets):
                weak_section_hit = True

    topic_target = ((qfilters or {}).get("topic") or "").strip()
    topic_hit = False
    if topic_target:
        for doc in window:
            if topic_target in _normalize_topics(_hit_metadata(doc).get("topics")):
                topic_hit = True
                break

    allowed_sources = [
        _normalize_filename_for_match(source or "")
        for source in (target_sources or [])
        if _normalize_filename_for_match(source or "")
    ]
    if allowed_sources and window and not any(
        any(_sources_equivalent(_hit_entity_source(doc) or "", allowed) for allowed in allowed_sources)
        for doc in window
    ):
        refusal_reason = "wrong_source"

    if section_hit or weak_section_hit:
        uncovered_terms = [
            term for term in uncovered_terms
            if not _is_question_wrapper_aspect(term)
            and not _aspect_is_subsumed_by_covered(term, section_targets)
        ]
    uncovered_terms = [term for term in uncovered_terms if not _is_meaningless_coverage_aspect(term)]
    relaxed_covered_terms = [
        term for term in uncovered_terms
        if _same_chunk_relaxed_aspect_coverage(term, covered_terms, window)
    ]
    if relaxed_covered_terms:
        covered_terms.extend([term for term in relaxed_covered_terms if term not in covered_terms])
        uncovered_terms = [term for term in uncovered_terms if term not in relaxed_covered_terms]
    semantic_softened_terms = _semantic_softened_aspects(uncovered_terms, window)
    if semantic_softened_terms:
        covered_terms.extend([term for term in semantic_softened_terms if term not in covered_terms])
        uncovered_terms = [term for term in uncovered_terms if term not in semantic_softened_terms]
    trusted_empty_gap = not uncovered_terms

    generic_only = bool(window) and all(
        (_is_generic_section_title(_doc_section_name(doc)) or not _doc_section_name(doc))
        for doc in window
    )
    generic_but_substantive = bool(generic_only) and bool(covered_terms or any(_is_substantive_short_legal_evidence(doc) for doc in window))
    clause_like_body_hit = any(_has_clause_like_body_evidence(doc) for doc in window)
    wide_topic = _is_wide_topic_query(query, qfilters=qfilters, target_sources=target_sources)
    body_anchor_state = _body_anchor_hits(query, window, qfilters=qfilters, target_sources=target_sources)
    focus_score = _intra_doc_focus_score(window)
    delegated_snippet = _delegated_to_external_source_snippet(query, window)
    delegated_to_external = bool(delegated_snippet)
    compare_degraded = bool(is_comparison and bool(compare_missing_targets) and source_lock_resolved and len(allowed_sources) == 1)
    fully_covered_substantive_window = bool(
        trusted_empty_gap
        and covered_terms
        and int(len(qualified_docs)) >= 2
        and (clause_like_body_hit or any(_is_substantive_short_legal_evidence(doc) for doc in window))
    )

    if not refusal_reason and not trusted_empty_gap and not delegated_to_external and not compare_degraded:
        would_off_topic = _is_off_topic_locked_document_query(
            query,
            docs,
            qfilters=qfilters,
            target_sources=target_sources,
            source_lock_resolved=source_lock_resolved,
        )
        if would_off_topic:
            best_dense_rel = _best_dense_relevance_for_locked_source(docs, target_sources)
            if best_dense_rel >= float(getattr(config, "DENSE_BACKSTOP_MIN_REL", 0.58)):
                would_off_topic = False
            elif (
                source_lock_resolved
                and covered_terms
                and int(len(qualified_docs)) >= int(getattr(config, "PARTIAL_TERM_RESCUE_MIN_SUBSTANTIVE_CHUNKS", 3))
                and float(focus_score) >= float(getattr(config, "PARTIAL_TERM_RESCUE_MIN_FOCUS_SCORE", 0.72))
            ):
                would_off_topic = False
            elif (
                getattr(config, "ENABLE_LLM_EVIDENCE_CHECK", False)
                and _classify_question_type(query) != "compare"
                and (clause_like_body_hit or bool(_locked_source_evidence_window(docs, target_sources, primary_topk)))
                and best_dense_rel >= float(getattr(config, "LLM_EVIDENCE_CHECK_MIN_DENSE_REL", 0.0))
            ):
                llm_window = _locked_source_evidence_window(docs, target_sources, primary_topk) or window
                llm_hit = _llm_evidence_core_concept_hit(query, llm_window)
                if llm_hit:
                    would_off_topic = False
            if would_off_topic:
                refusal_reason = "off_topic_in_document"

    if delegated_to_external and not (
        refusal_reason in {"wrong_source", "query_too_short", "query_too_long", "blocked_prompt_injection"}
    ):
        coverage_reason = "delegated_to_external_source"
        answer_scope = "full"
    elif refusal_reason:
        coverage_reason = refusal_reason
        answer_scope = "partial" if _evidence_reason_is_rescuable(refusal_reason) else "refusal"
    elif _is_document_state_query(query) and source_lock_resolved and bool(docs):
        coverage_reason = "sufficient_evidence"
        answer_scope = "full"
    elif strong_section_targets and not section_hit:
        coverage_reason = "section_not_hit"
        answer_scope = "partial" if covered_terms or weak_section_hit else "refusal"
    elif topic_target and not topic_hit and not (
        source_lock_resolved and section_hit and focus_score >= 0.8 and len(qualified_docs) >= 8
    ):
        coverage_reason = "topic_not_hit"
        answer_scope = "partial" if covered_terms else "refusal"
    elif weak_section_targets and not weak_section_hit and not (
        source_lock_resolved and section_hit and focus_score >= 0.8 and len(qualified_docs) >= 8
    ):
        coverage_reason = "topic_not_hit"
        answer_scope = "partial" if covered_terms else "refusal"
    elif fully_covered_substantive_window:
        coverage_reason = "sufficient_evidence"
        answer_scope = "full"
    elif generic_only and not generic_but_substantive:
        coverage_reason = "all_generic_no_body"
        answer_scope = "refusal"
    elif generic_only:
        coverage_reason = "generic_only"
        answer_scope = "partial"
    elif wide_topic and not body_anchor_state["body_anchor_hit"]:
        coverage_reason = "topic_anchor_missing"
        answer_scope = "refusal"
    elif wide_topic and body_anchor_state["anchors"] and not body_anchor_state["all_covered"]:
        coverage_reason = "partial_term_coverage"
        answer_scope = "partial"
    elif uncovered_terms:
        coverage_reason = "partial_term_coverage"
        answer_scope = "partial"
    elif not clause_like_body_hit:
        coverage_reason = "no_body_clause_evidence"
        answer_scope = "refusal"
    else:
        coverage_reason = "sufficient_evidence"
        answer_scope = "full"

    covered_aspects: List[str] = []
    uncovered_aspects: List[str] = []
    if strong_section_targets:
        if section_hit:
            covered_aspects.extend(strong_section_targets[:3])
        else:
            uncovered_aspects.extend(strong_section_targets[:3])
    if weak_section_targets:
        if weak_section_hit:
            covered_aspects.extend([term for term in weak_section_targets[:3] if term not in covered_aspects])
        else:
            uncovered_aspects.extend([term for term in weak_section_targets[:3] if term not in uncovered_aspects])
    if topic_target:
        if topic_hit:
            covered_aspects.append(topic_target)
        else:
            uncovered_aspects.append(topic_target)
    covered_aspects.extend([
        _normalize_coverage_aspect(term) or term
        for term in covered_terms
        if (_normalize_coverage_aspect(term) or term) not in covered_aspects
    ])
    covered_aspects.extend([
        title for title in source_title_terms
        if title in _normalize_query(query) and title not in covered_aspects
    ])
    uncovered_aspects.extend([
        _normalize_coverage_aspect(term) or term
        for term in uncovered_terms
        if (_normalize_coverage_aspect(term) or term) not in uncovered_aspects
        and not _aspect_is_subsumed_by_covered(term, covered_aspects)
    ])

    return {
        "evidence_coverage_reason": coverage_reason,
        "section_hit": bool(section_hit or weak_section_hit),
        "topic_hit": bool(topic_hit or weak_section_hit),
        "heading_only": bool(docs) and not bool(evidence_docs),
        "qualified_substantive_chunks": len(qualified_docs),
        "generic_only": bool(generic_only),
        "insufficient": bool(refusal_reason),
        "intra_doc_focus_score": float(focus_score),
        "answer_scope": answer_scope,
        "covered_aspects": covered_aspects[:6],
        "uncovered_aspects": uncovered_aspects[:6],
        "delegated_to_external_source": bool(delegated_to_external),
        "delegated_snippet": delegated_snippet,
        "compare_degraded": bool(compare_degraded),
        "compare_missing_targets": list(compare_missing_targets or [])[:4],
    }


def _rescue_failure_reason(initial_reason: str, rescue_reason: Optional[str], rescued_docs: List[Any]) -> str:
    reason = str(rescue_reason or initial_reason or "evidence_insufficient")
    if initial_reason == "heading_only_evidence" and not rescued_docs:
        return "heading_only_after_expand"
    if reason in {"low_evidence_relevance", "insufficient_evidence", "no_relevant_evidence", "empty_evidence"}:
        return "low_relevance_after_rescue"
    return reason if _evidence_reason_is_hard_refusal(reason) else f"rescue_failed_{initial_reason}"


def _evidence_observations(
    query: str,
    docs: List[Any],
    qfilters: Optional[Dict[str, Any]] = None,
    candidate_docs: Optional[List[Any]] = None,
    target_sources: Optional[List[str]] = None,
    source_lock_resolved: bool = False,
    source_lock_reason: str = "",
    is_comparison: bool = False,
    compare_missing_targets: Optional[List[str]] = None,
) -> Dict[str, Any]:
    base = _evaluate_evidence_window(
        query,
        docs,
        qfilters=qfilters,
        target_sources=target_sources,
        source_lock_resolved=source_lock_resolved,
        source_lock_reason=source_lock_reason,
        is_comparison=is_comparison,
        compare_missing_targets=compare_missing_targets,
    )
    reason = str(base.get("evidence_coverage_reason") or "")
    if base.get("answer_scope") == "full" or _evidence_reason_is_hard_refusal(reason):
        return base
    if not _evidence_reason_is_rescuable(reason):
        return base
    normalized_targets = [
        _normalize_filename_for_match(source or "")
        for source in (target_sources or [])
        if _normalize_filename_for_match(source or "")
    ]
    if not source_lock_resolved or len(normalized_targets) != 1:
        return base

    rescue_min_score = _rescue_min_score_for_reason(reason, is_comparison=is_comparison)

    candidate_pool = _build_evidence_candidate_pool(
        query,
        candidate_docs or docs,
        target_sources=target_sources,
        qfilters=qfilters,
        min_score_override=rescue_min_score,
    )
    if not candidate_pool:
        return {
            **base,
            "evidence_coverage_reason": _rescue_failure_reason(reason, None, []),
            "answer_scope": "refusal",
            "rescue_attempted": True,
            "rescue_success": False,
        }

    rescued = _evaluate_evidence_window(
        query,
        candidate_pool,
        qfilters=qfilters,
        target_sources=target_sources,
        min_score_override=rescue_min_score,
        source_lock_resolved=source_lock_resolved,
        source_lock_reason=source_lock_reason,
        is_comparison=is_comparison,
        compare_missing_targets=compare_missing_targets,
    )
    rescued_reason = str(rescued.get("evidence_coverage_reason") or "")
    if rescued.get("answer_scope") == "full":
        rescued_scope = "guarded_full" if bool(getattr(config, "ALLOW_GUARDED_FULL", True)) else "full"
        return {
            **rescued,
            "answer_scope": rescued_scope,
            "evidence_coverage_reason": f"rescued_{reason}",
            "rescue_attempted": True,
            "rescue_success": True,
            "rescue_from_reason": reason,
        }

    if (
        reason == "partial_term_coverage"
        and not is_comparison
        and int(rescued.get("qualified_substantive_chunks") or 0) >= int(getattr(config, "PARTIAL_TERM_RESCUE_MIN_SUBSTANTIVE_CHUNKS", 3))
        and float(rescued.get("intra_doc_focus_score") or 0.0) >= float(getattr(config, "PARTIAL_TERM_RESCUE_MIN_FOCUS_SCORE", 0.72))
    ):
        rescued_scope = "guarded_full" if bool(getattr(config, "ALLOW_GUARDED_FULL", True)) else "full"
        return {
            **rescued,
            "answer_scope": rescued_scope,
            "evidence_coverage_reason": f"rescued_relaxed_{reason}",
            "rescue_attempted": True,
            "rescue_success": True,
            "rescue_from_reason": reason,
            "relaxed_partial_term_rescue": True,
        }

    return {
        **rescued,
        "answer_scope": "refusal",
        "evidence_coverage_reason": _rescue_failure_reason(reason, rescued_reason, candidate_pool),
        "rescue_attempted": True,
        "rescue_success": False,
        "rescue_from_reason": reason,
    }


def _select_retrieve_output_docs(docs: List[Any], top_k: int, default_n: int) -> List[Any]:
    if not docs:
        return []
    keep_n = min(len(docs), min(max(int(top_k or default_n), 3), 8))
    return docs[:keep_n]


def _select_process_output_docs(query: str, docs: List[Any], score_mode: str, qfilters: Dict[str, Any], default_n: int) -> List[Any]:
    if len(docs) <= 1:
        return docs[: min(len(docs), default_n)]

    window = docs[: min(len(docs), 6)]
    source_counts: Dict[str, int] = {}
    for doc in window:
        src = _normalize_filename_for_match(_hit_entity_source(doc) or "")
        if src:
            source_counts[src] = source_counts.get(src, 0) + 1
    dominant_source = None
    dominant_share = 0.0
    if source_counts:
        dominant_source, dominant_count = max(source_counts.items(), key=lambda item: (item[1], item[0]))
        dominant_share = float(dominant_count) / float(len(window))

    pool = docs[:]
    if dominant_source and dominant_share >= 0.5:
        same_source_docs = [doc for doc in docs if _normalize_filename_for_match(_hit_entity_source(doc) or "") == dominant_source]
        if same_source_docs:
            pool = same_source_docs

    aligned_docs = [doc for doc in pool if _section_target_alignment(_doc_section_name(doc), query)[0] > 0]
    if aligned_docs:
        pool = aligned_docs

    decorated = []
    for idx, doc in enumerate(pool):
        src = _normalize_filename_for_match(_hit_entity_source(doc) or "")
        section = _doc_section_name(doc)
        md = _hit_metadata(doc)
        base_rel = _chunk_base_relevance(doc, score_mode)
        align, exact = _section_target_alignment(section, query)
        topic_bonus = 0.0
        if qfilters.get("topic") and qfilters["topic"] in _normalize_topics(md.get("topics")):
            topic_bonus = 0.2
        source_bonus = 0.25 if dominant_source and src == dominant_source else 0.0
        section_bonus = 0.35 * align + 0.15 * exact
        generic_penalty = 0.18 if _is_generic_section_title(section) else 0.0
        focus_score = base_rel + source_bonus + section_bonus + topic_bonus - generic_penalty
        decorated.append((focus_score, -idx, doc))
    decorated.sort(key=lambda item: (item[0], item[1]), reverse=True)
    return [doc for _, _, doc in decorated[: min(len(decorated), default_n)]]


def _source_display_title(source: str) -> str:
    info = _doc_get(source)
    return (info.get("canonical_title") or "").strip() or _filename_stem(source) or source


def _merge_compare_source_doc_groups(source_groups: List[Dict[str, Any]], per_source_limit: int) -> List[Any]:
    merged: List[Any] = []
    seen_keys = set()
    limit = max(1, int(per_source_limit))
    for group in source_groups or []:
        for doc in (group.get("docs") or [])[:limit]:
            key = (_normalize_filename_for_match(_hit_entity_source(doc) or ""), (_hit_entity_text(doc) or "")[:96])
            if key in seen_keys:
                continue
            seen_keys.add(key)
            merged.append(doc)
    return merged


def _compare_source_status_entry(
    query: str,
    group: Dict[str, Any],
    qfilters: Optional[Dict[str, Any]],
    compare_identity_terms: List[str],
    min_substantive_chunks_for_compare_partial: int,
) -> Optional[Dict[str, Any]]:
    source = _normalize_filename_for_match((group or {}).get("source") or "")
    docs = list((group or {}).get("docs") or [])
    evidence_query = _normalize_query((group or {}).get("evidence_query") or query) or query
    if not source:
        return None
    if not docs:
        observations = {
            "evidence_coverage_reason": "compare_source_not_found",
            "answer_scope": "refusal",
            "covered_aspects": [],
            "uncovered_aspects": [],
        }
        status = "not_found"
    else:
        observations = _evidence_observations(evidence_query, docs, qfilters=qfilters)
        if observations.get("answer_scope") == "full":
            status = "answerable"
        elif observations.get("answer_scope") == "guarded_full":
            status = "guarded_full"
        else:
            status = "evidence_insufficient"
    filtered_covered = _filter_identity_noise_aspects(observations.get("covered_aspects") or [], compare_identity_terms)
    filtered_uncovered = _filter_identity_noise_aspects(observations.get("uncovered_aspects") or [], compare_identity_terms)
    if status == "evidence_insufficient" and docs and filtered_covered:
        status = "comparable_partial"
    elif (
        status == "evidence_insufficient"
        and docs
        and not filtered_covered
        and int(observations.get("qualified_substantive_chunks") or 0) >= min_substantive_chunks_for_compare_partial
    ):
        status = "comparable_partial"
    return {
        "source": source,
        "title": _source_display_title(source),
        "evidence_query": evidence_query,
        "status": status,
        "docs_count": len(docs),
        "observations": {
            **observations,
            "covered_aspects": filtered_covered,
            "uncovered_aspects": filtered_uncovered,
        },
    }


def _finalize_compare_evidence_observations(source_statuses: List[Dict[str, Any]]) -> Dict[str, Any]:
    covered_aspects: List[str] = []
    uncovered_aspects: List[str] = []
    for item in source_statuses:
        observations = dict(item.get("observations") or {})
        for aspect in observations.get("covered_aspects") or []:
            if aspect and aspect not in covered_aspects:
                covered_aspects.append(aspect)
        for aspect in observations.get("uncovered_aspects") or []:
            if aspect and aspect not in uncovered_aspects:
                uncovered_aspects.append(aspect)
    if source_statuses and all(item["status"] in {"answerable", "guarded_full", "comparable_partial"} for item in source_statuses):
        return {
            "evidence_coverage_reason": "sufficient_evidence",
            "answer_scope": "guarded_full" if any(item["status"] in {"guarded_full", "comparable_partial"} for item in source_statuses) else "full",
            "compare_status": "compare_ready",
            "compare_source_statuses": source_statuses,
            "covered_aspects": covered_aspects[:8],
            "uncovered_aspects": uncovered_aspects[:8],
        }
    if (
        source_statuses
        and any(item["status"] in {"answerable", "guarded_full", "comparable_partial"} for item in source_statuses)
        and any(item["status"] in {"not_found", "evidence_insufficient"} for item in source_statuses)
        and all(item["status"] in {"answerable", "guarded_full", "comparable_partial", "not_found", "evidence_insufficient"} for item in source_statuses)
    ):
        return {
            "evidence_coverage_reason": "sufficient_evidence",
            "answer_scope": "guarded_full",
            "compare_status": "compare_asymmetric",
            "compare_degraded": True,
            "compare_source_statuses": source_statuses,
            "covered_aspects": covered_aspects[:8],
            "uncovered_aspects": uncovered_aspects[:8],
        }
    if source_statuses and all(item["status"] == "not_found" for item in source_statuses):
        reason = "compare_targets_not_found"
        compare_status = "all_sources_missing"
    elif any(item["status"] == "not_found" for item in source_statuses):
        reason = "compare_source_missing"
        compare_status = "partial_sources_missing"
    else:
        reason = "compare_evidence_insufficient"
        compare_status = "evidence_insufficient"
    return {
        "evidence_coverage_reason": reason,
        "answer_scope": "refusal",
        "compare_status": compare_status,
        "compare_source_statuses": source_statuses,
        "covered_aspects": covered_aspects[:8],
        "uncovered_aspects": uncovered_aspects[:8],
    }


def _compare_evidence_observations(query: str, source_groups: List[Dict[str, Any]], qfilters: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    min_substantive_chunks_for_compare_partial = max(2, int(getattr(config, "COMPARE_PARTIAL_MIN_SUBSTANTIVE_CHUNKS", 2)))
    compare_identity_terms = _source_identity_terms_for_validation([
        _normalize_filename_for_match((group or {}).get("source") or "")
        for group in (source_groups or [])
        if _normalize_filename_for_match((group or {}).get("source") or "")
    ])
    source_statuses: List[Dict[str, Any]] = []
    for group in source_groups or []:
        entry = _compare_source_status_entry(
            query,
            group,
            qfilters=qfilters,
            compare_identity_terms=compare_identity_terms,
            min_substantive_chunks_for_compare_partial=min_substantive_chunks_for_compare_partial,
        )
        if entry:
            source_statuses.append(entry)
    return _finalize_compare_evidence_observations(source_statuses)


async def _compare_evidence_observations_async(query: str, source_groups: List[Dict[str, Any]], qfilters: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    min_substantive_chunks_for_compare_partial = max(2, int(getattr(config, "COMPARE_PARTIAL_MIN_SUBSTANTIVE_CHUNKS", 2)))
    compare_identity_terms = _source_identity_terms_for_validation([
        _normalize_filename_for_match((group or {}).get("source") or "")
        for group in (source_groups or [])
        if _normalize_filename_for_match((group or {}).get("source") or "")
    ])
    tasks = [
        asyncio.to_thread(
            _compare_source_status_entry,
            query,
            group,
            qfilters,
            compare_identity_terms,
            min_substantive_chunks_for_compare_partial,
        )
        for group in (source_groups or [])
    ]
    source_statuses = [item for item in await asyncio.gather(*tasks) if item]
    return _finalize_compare_evidence_observations(source_statuses)


def _build_compare_evidence_failure_prompt(source_statuses: List[Dict[str, Any]]) -> str:
    if not source_statuses:
        return "无法完成法规对比，因为当前没有可用的目标文档证据。"

    def _status_detail(item: Dict[str, Any]) -> str:
        title = (item.get("title") or item.get("source") or "目标文档").strip()
        status = str(item.get("status") or "")
        observations = dict(item.get("observations") or {})
        uncovered = [str(term).strip() for term in (observations.get("uncovered_aspects") or []) if str(term).strip()]
        focus_text = "、".join(uncovered[:2]) if uncovered else "当前对比焦点"
        if status == "not_found":
            return f"《{title}》未检索到可用于对比的相关条文"
        if status == "evidence_insufficient":
            if uncovered:
                return f"《{title}》在核心条文中未直接提及{focus_text}，暂缺直接证据"
            return f"《{title}》暂未检索到与当前对比焦点直接对应的条文证据"
        if status == "comparable_partial":
            return f"《{title}》仅检索到部分相关条文，但不足以稳定支撑结构化对比"
        return f"《{title}》已找到可对比证据"

    parts: List[str] = []
    for item in source_statuses:
        parts.append(_status_detail(item))

    if all(str(item.get("status") or "") == "not_found" for item in source_statuses):
        return "已为您检索相关法规，但当前未找到可用于本次对比的目标条文：" + "；".join(parts) + "。建议补充更准确的法规名称或年份后重试。"

    return "已为您检索相关法规，但暂时无法形成可靠的结构化对比：" + "；".join(parts) + "。如需继续，我可以基于已命中的文档先说明已明确规定的内容。"


def _append_answer_scope_semantics(answer: str, observations: Dict[str, Any]) -> str:
    scope = observations.get("answer_scope")
    if scope != "partial":
        return answer
    uncovered = list(observations.get("uncovered_aspects") or [])
    if not uncovered:
        uncovered = ["问题中的其余部分"]
    uncovered_text = "、".join(uncovered[:4])
    body = (answer or "").strip() or "已根据现有证据回答可覆盖的部分。"
    return f"已覆盖：{body}\n\n未覆盖：当前证据未覆盖 {uncovered_text}。"


def _hit_chunk_id(hit: Any) -> Optional[int]:
    md = _hit_metadata(hit)
    v = md.get("chunk_id")
    if v is None:
        return None
    try:
        return int(v)
    except Exception:
        return None


def _hit_chunk_range(hit: Any) -> str:
    md = _hit_metadata(hit)
    s = md.get("chunk_id_start")
    e = md.get("chunk_id_end")
    if s is not None and e is not None:
        try:
            s_i = int(s)
            e_i = int(e)
            return f"{s_i}-{e_i}" if s_i != e_i else f"{s_i}"
        except Exception:
            pass
    cid = _hit_chunk_id(hit)
    return f"{cid}" if cid is not None else ""


def _build_excerpt(text: str, query: str, max_chars: int) -> str:
    t = (text or "").strip()
    if not t:
        return ""
    q = _normalize_query(query)
    terms = [w for w in q.replace("，", " ").replace(",", " ").split() if len(w) >= 2]
    terms = list(dict.fromkeys(terms))[:8]
    if not terms or len(t) <= max_chars:
        return t[:max_chars]
    lower = t.lower()
    best_i = -1
    for term in terms:
        i = lower.find(term.lower())
        if i != -1:
            best_i = i
            break
    if best_i == -1:
        return t[:max_chars]
    half = max_chars // 2
    start = max(0, best_i - half)
    end = min(len(t), start + max_chars)
    start = max(0, end - max_chars)
    snippet = t[start:end].strip()
    if start > 0:
        snippet = "…" + snippet
    if end < len(t):
        snippet = snippet + "…"
    return snippet


def _source_state(source: str) -> Dict[str, Any]:
    src = _normalize_filename_for_match(source or "")
    doc = _doc_get(src)
    status = _public_task_status(doc.get("status")) if doc.get("status") else ""
    active_version = doc.get("active_version")
    pending_version = doc.get("pending_version")
    try:
        active_version = int(active_version) if active_version is not None else None
    except Exception:
        active_version = None
    try:
        pending_version = int(pending_version) if pending_version is not None else None
    except Exception:
        pending_version = None
    hidden_statuses = {"deleting", "pending_delete", "delete_failed"}
    if status in hidden_statuses:
        visible = False
    else:
        # Reindexing must keep the previously published active version visible.
        visible = bool(active_version is not None) or (not status) or (status == "completed")
    return {
        "source": src,
        "status": status,
        "active_version": active_version,
        "pending_version": pending_version,
        "visible": visible,
    }


def _hit_matches_source_state(hit: Any, state: Dict[str, Any]) -> bool:
    if not state.get("visible"):
        return False
    active_version = state.get("active_version")
    if active_version is None:
        status = state.get("status") or ""
        return status in {"", "completed"}
    md = _hit_metadata(hit)
    hit_version = md.get("doc_version")
    if hit_version is None:
        return True
    try:
        return int(hit_version) == int(active_version)
    except Exception:
        return True


def _filter_hits_by_source_state(hits: List[Any]) -> Dict[str, Any]:
    out: List[Any] = []
    dropped = 0
    states: Dict[str, Dict[str, Any]] = {}
    for hit in hits:
        src = _normalize_filename_for_match(_hit_entity_source(hit) or "")
        if src not in states:
            states[src] = _source_state(src)
        if _hit_matches_source_state(hit, states[src]):
            out.append(hit)
        else:
            dropped += 1
    return {"hits": out, "dropped": dropped, "states": states}


@lru_cache(maxsize=1)
def _evidence_token_encoder() -> Any:
    try:
        import tiktoken
    except Exception:
        return None
    model_name = (getattr(config, "LLM_MODEL", "") or "").strip()
    try:
        return tiktoken.encoding_for_model(model_name) if model_name else tiktoken.get_encoding("cl100k_base")
    except Exception:
        return tiktoken.get_encoding("cl100k_base")


def _estimate_token_count(text: str) -> int:
    if not text:
        return 0
    encoder = _evidence_token_encoder()
    if encoder is None:
        return max(1, math.ceil(len(text) / 4))
    try:
        return len(encoder.encode(text))
    except Exception:
        return max(1, math.ceil(len(text) / 4))


def _evidence_relevance(doc: Any, score_mode: str, best_score: float) -> float:
    score = float(_hit_score(doc))
    if score_mode == "distance":
        best_sim = 1.0 / (1.0 + max(best_score, 0.0))
        sim = 1.0 / (1.0 + max(score, 0.0))
        return min(max(sim / max(best_sim, 1e-9), 0.0), 1.0)
    if best_score <= 0:
        return 0.0
    return min(max(score / best_score, 0.0), 1.0)


def _format_evidence(docs: List[Any], query: str, score_mode: str) -> str:
    lines = []
    total_tokens = 0
    token_budget = max(256, int(getattr(config, "EVIDENCE_MAX_TOKENS", 6500)))
    best_score = _hit_score(docs[0]) if docs else 0.0
    for i, doc in enumerate(docs, start=1):
        src = _hit_entity_source(doc) or "unknown"
        content = (_hit_display_text(doc) or "").strip()
        if not content:
            continue
        md = _hit_metadata(doc)
        section = (md.get("section") or md.get("section_title") or "").strip()
        chunk_range = _hit_chunk_range(doc)
        relevance = _evidence_relevance(doc, score_mode, best_score)
        parts = [f"来源：{src}", f"相关性：{relevance:.2f}"]
        if section:
            parts.append(f"章节：{section}")
        if chunk_range:
            parts.append(f"位置：chunk_{chunk_range}")
        head = f"[证据 {i}] " + " | ".join(parts)
        block = head + "\n" + content
        block_tokens = _estimate_token_count(block) + 2
        if total_tokens + block_tokens > token_budget:
            break
        lines.append(block)
        total_tokens += block_tokens
    return "\n\n".join(lines)


def _split_sections(text: str) -> List[Dict[str, Any]]:
    t = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    lines = t.split("\n")
    sections: List[Dict[str, Any]] = []
    current_title = ""
    buf: List[str] = []

    def _flush():
        nonlocal buf, current_title
        body = "\n".join(buf).strip()
        if body:
            sections.append({"section": current_title.strip(), "text": body})
        buf = []

    for line in lines:
        s = line.strip()
        is_md_heading = False
        title = ""
        if s.startswith("#"):
            parts = s.lstrip("#").strip()
            if parts:
                is_md_heading = True
                title = parts
        if s.startswith("# Sheet: "):
            is_md_heading = True
            title = s[len("# Sheet: "):].strip()

        if is_md_heading:
            _flush()
            current_title = title
            continue
        buf.append(line)

    _flush()
    if not sections:
        return [{"section": "", "text": (text or "").strip()}]
    return sections


def _chapter_heading_title(text: str) -> str:
    s = (text or "").strip()
    if not s:
        return ""
    if s.startswith("# Sheet: "):
        return s[len("# Sheet: "):].strip()
    if s.startswith("#"):
        return s.lstrip("#").strip()
    if re.match(r"^第[一二三四五六七八九十百千0-9]+[章节编](?:\s|$)", s):
        return s
    return ""


def _safe_float(value: Any) -> Optional[float]:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except Exception:
        return None


def _median_number(values: List[Any]) -> Optional[float]:
    numbers = sorted([float(value) for value in values if _safe_float(value) is not None])
    if not numbers:
        return None
    mid = len(numbers) // 2
    if len(numbers) % 2 == 1:
        return numbers[mid]
    return (numbers[mid - 1] + numbers[mid]) / 2.0


def _bbox_from_points(points: Any) -> Optional[Dict[str, float]]:
    coords: List[tuple[float, float]] = []
    if not isinstance(points, list):
        return None
    for point in points:
        if isinstance(point, dict):
            x = _safe_float(point.get("x"))
            y = _safe_float(point.get("y"))
            if x is not None and y is not None:
                coords.append((x, y))
        elif isinstance(point, (list, tuple)) and len(point) >= 2:
            x = _safe_float(point[0])
            y = _safe_float(point[1])
            if x is not None and y is not None:
                coords.append((x, y))
    if not coords:
        return None
    xs = [coord[0] for coord in coords]
    ys = [coord[1] for coord in coords]
    return {"x0": min(xs), "y0": min(ys), "x1": max(xs), "y1": max(ys)}


def _normalize_bbox(value: Any) -> Optional[Dict[str, float]]:
    if isinstance(value, dict):
        x0 = _safe_float(value.get("x0"))
        y0 = _safe_float(value.get("y0"))
        x1 = _safe_float(value.get("x1"))
        y1 = _safe_float(value.get("y1"))
        if None not in {x0, y0, x1, y1}:
            return {"x0": x0, "y0": y0, "x1": x1, "y1": y1}
        for key in ("bbox", "box", "polygon", "points"):
            nested = value.get(key)
            nested_bbox = _normalize_bbox(nested)
            if nested_bbox is not None:
                return nested_bbox
        return None
    if isinstance(value, (list, tuple)) and len(value) >= 4:
        if all(not isinstance(item, (list, tuple, dict)) for item in value[:4]):
            x0 = _safe_float(value[0])
            y0 = _safe_float(value[1])
            x1 = _safe_float(value[2])
            y1 = _safe_float(value[3])
            if None not in {x0, y0, x1, y1}:
                return {"x0": x0, "y0": y0, "x1": x1, "y1": y1}
        polygon_bbox = _bbox_from_points(list(value))
        if polygon_bbox is not None:
            return polygon_bbox
    return None


def _bbox_layout_metrics(bbox: Optional[Dict[str, Any]], page_width: Optional[float] = None, page_height: Optional[float] = None) -> Dict[str, Any]:
    normalized = _normalize_bbox(bbox)
    if normalized is None:
        return {}
    width = max(0.0, float(normalized["x1"]) - float(normalized["x0"]))
    height = max(0.0, float(normalized["y1"]) - float(normalized["y0"]))
    center_x = float(normalized["x0"]) + width / 2.0
    metrics: Dict[str, Any] = {
        "bbox": normalized,
        "block_width": width,
        "block_height": height,
        "center_x": center_x,
    }
    if page_width and page_width > 0:
        metrics["page_width"] = page_width
        metrics["left_ratio"] = float(normalized["x0"]) / page_width
        metrics["width_ratio"] = width / page_width
        metrics["center_offset_ratio"] = abs(center_x - (page_width / 2.0)) / page_width
        metrics["is_centered"] = metrics["center_offset_ratio"] <= 0.08 and (metrics.get("width_ratio") or 1.0) <= 0.82
    if page_height and page_height > 0:
        metrics["page_height"] = page_height
        metrics["top_ratio"] = float(normalized["y0"]) / page_height
        metrics["height_ratio"] = height / page_height
    return metrics


def _normalize_heading_hint_text(text: str) -> str:
    compact = re.sub(r"\s+", "", text or "")
    compact = compact.replace("摇", "")
    compact = compact.replace("笫", "第").replace("總", "总").replace("則", "则")
    compact = re.sub(r"(?<=第)[\-—_~]+(?=[编章节条])", "一", compact)
    compact = compact.replace("附則", "附则")
    compact = compact.replace("葉", "叶").replace("曱", "曳")
    compact = re.sub(r"叶([^叶曳]{1,80})曳", r"《\1》", compact)
    compact = compact.replace("渊", "（").replace("冤", "）")
    return compact


def _normalize_heading_title(text: str) -> str:
    normalized = str(text or "").replace("\r", " ").replace("\n", " ")
    normalized = normalized.replace("摇", " ")
    normalized = normalized.replace("笫", "第").replace("總", "总").replace("則", "则")
    normalized = normalized.replace("附則", "附则")
    normalized = normalized.replace("渊", "（").replace("冤", "）")
    normalized = re.sub(r"叶([^叶曳]{1,80})曳", r"《\1》", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


def _normalized_heading_compact(text: str) -> str:
    return re.sub(r"\s+", "", _normalize_heading_title(text))


def _looks_like_toc_title(text: str) -> bool:
    compact = _normalized_heading_compact(text)
    return compact in {"目录", "目次"}


def _has_terminal_sentence_punctuation(text: str) -> bool:
    compact = _normalized_heading_compact(text)
    return bool(compact and re.search(r"[。．｡！？；;，,、：:遥袁尧]$", compact))


def _is_bare_legal_heading_text(text: str) -> bool:
    compact = _normalized_heading_compact(text)
    return compact in {
        "总则",
        "附则",
        "罚则",
        "总纲",
        "通则",
        "附录",
    }


def _law_semantic_truncation_label(text: str) -> str:
    compact = _normalized_heading_compact(text)
    if not compact:
        return ""
    patterns = [
        r"^关于.+的说明$",
        r"^关于.+审查情况的报告$",
        r"^关于.+审议结果的报告$",
        r"^关于.+审查结果的报告$",
        r"^关于.+的报告$",
        r"^关于批准.+(决定|决议)$",
    ]
    for pattern in patterns:
        if re.match(pattern, compact):
            return _normalize_heading_title(text)
    return ""


def _cid_garbled_char_ratio(text: str) -> float:
    raw = (text or "").strip()
    if not raw:
        return 0.0
    meaningful = re.findall(r"[A-Za-z0-9\u4e00-\u9fff]", raw)
    if not meaningful:
        return 0.0
    suspicious = re.findall(r"[摇遥袁叶曳渊冤]", raw)
    return len(suspicious) / max(1, len(meaningful))


def _looks_like_cid_garbled_text(text: str) -> bool:
    raw = (text or "").strip()
    if not raw:
        return False
    suspicious_count = len(re.findall(r"[摇遥袁叶曳渊冤]", raw))
    if suspicious_count < 3:
        return False
    return _cid_garbled_char_ratio(raw) >= 0.08


def _legal_heading_level(text: str) -> Optional[int]:
    normalized = _normalize_heading_hint_text(text)
    if not normalized:
        return None
    if re.match(r"^第[一二三四五六七八九十百千0-9]+编", normalized):
        return 1
    if re.match(r"^第[一二三四五六七八九十百千0-9]+章", normalized):
        return 2
    if re.match(r"^第[一二三四五六七八九十百千0-9]+节", normalized):
        return 3
    return None


def _pdf_is_toc_entry_text(text: str) -> bool:
    s = (text or "").strip()
    if not s:
        return False
    if _looks_like_toc_title(s):
        return True
    if re.search(r"[.．·•…]{2,}\s*\d{1,3}$", s):
        return True
    compact = _normalize_heading_hint_text(s)
    if _legal_heading_level(compact) and re.search(r"\d{1,3}$", compact):
        return True
    return False


def _build_visual_page_profile(entries: List[Dict[str, Any]], page_width: Optional[float] = None, page_height: Optional[float] = None) -> Dict[str, Any]:
    font_sizes: List[float] = []
    for entry in entries:
        layout = entry.get("layout") or {}
        font_size = _safe_float(layout.get("font_size") or layout.get("font_size_max") or layout.get("font_size_median"))
        if font_size is not None and font_size > 0:
            font_sizes.append(font_size)
    body_font = _median_number(font_sizes)
    return {
        "page_width": page_width,
        "page_height": page_height,
        "body_font_size": body_font,
        "max_font_size": max(font_sizes) if font_sizes else None,
    }


def _infer_visual_heading_level(text: str, layout: Dict[str, Any], page_profile: Dict[str, Any], next_text: str = "") -> Optional[int]:
    s = _normalize_heading_title(text)
    if not s or _clause_heading_label(s):
        return None
    compact = _normalized_heading_compact(s)
    if len(compact) > 30:
        return None
    if _has_terminal_sentence_punctuation(s):
        return None
    level = _legal_heading_level(s)
    font_size = _safe_float(layout.get("font_size") or layout.get("font_size_max") or layout.get("font_size_median"))
    body_font = _safe_float(page_profile.get("body_font_size"))
    font_ratio = (font_size / body_font) if font_size and body_font and body_font > 0 else 1.0
    is_centered = bool(layout.get("is_centered"))
    width_ratio = _safe_float(layout.get("width_ratio"))
    top_ratio = _safe_float(layout.get("top_ratio"))
    visual_strong = font_ratio >= 1.12 or is_centered or (width_ratio is not None and width_ratio <= 0.55)
    if level is not None:
        if visual_strong or len(compact) <= 18:
            return level
        return None
    if _is_bare_legal_heading_text(s) and visual_strong and (top_ratio is None or top_ratio <= 0.3):
        return 2
    return None


def _is_visual_title_candidate(text: str, layout: Dict[str, Any], page_profile: Dict[str, Any], page_no: int, has_elements: bool) -> bool:
    if has_elements or page_no != 1:
        return False
    s = _normalize_heading_title(text)
    if not s or _looks_like_toc_title(s) or _pdf_is_toc_entry_text(s) or _legal_heading_level(s) is not None or _clause_heading_label(s):
        return False
    compact = _normalized_heading_compact(s)
    if len(compact) > 40:
        return False
    font_size = _safe_float(layout.get("font_size") or layout.get("font_size_max") or layout.get("font_size_median"))
    body_font = _safe_float(page_profile.get("body_font_size"))
    font_ratio = (font_size / body_font) if font_size and body_font and body_font > 0 else 1.0
    top_ratio = _safe_float(layout.get("top_ratio"))
    width_ratio = _safe_float(layout.get("width_ratio"))
    is_centered = bool(layout.get("is_centered"))
    return font_ratio >= 1.25 or (is_centered and (top_ratio is None or top_ratio <= 0.22) and (width_ratio is None or width_ratio <= 0.82))


def _should_exit_visual_toc(text: str, next_text: str, layout: Dict[str, Any], page_profile: Dict[str, Any]) -> bool:
    s = _normalize_heading_title(text)
    if not s:
        return False
    if _clause_heading_label(s):
        return True
    heading_level = _infer_visual_heading_level(s, layout, page_profile, next_text)
    if heading_level is not None:
        return bool(next_text and (_clause_heading_label(next_text) or not _pdf_is_toc_entry_text(next_text)))
    return len(re.sub(r"\s+", "", s)) > 50 and not _pdf_is_toc_entry_text(s)


def _extract_ocr_lines(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    raw_items = payload.get("lines")
    if not isinstance(raw_items, list):
        raw_items = payload.get("texts")
    if not isinstance(raw_items, list):
        raw_items = payload.get("results")
    if not isinstance(raw_items, list):
        raw_items = [payload.get("texts")] if isinstance(payload.get("texts"), str) else []
    lines: List[Dict[str, Any]] = []
    for index, item in enumerate(raw_items or []):
        if isinstance(item, str):
            text = " ".join(item.split())
            if text:
                lines.append({"text": text, "line_index": index})
            continue
        if not isinstance(item, dict):
            continue
        text = str(item.get("text") or item.get("content") or item.get("value") or item.get("line_text") or "").strip()
        text = " ".join(text.split())
        if not text:
            continue
        bbox = None
        for key in ("bbox", "box", "polygon", "points"):
            bbox = _normalize_bbox(item.get(key))
            if bbox is not None:
                break
        line: Dict[str, Any] = {
            "text": text,
            "bbox": bbox,
            "line_index": int(item.get("line_index") if item.get("line_index") is not None else index),
        }
        confidence = _safe_float(item.get("confidence") or item.get("score"))
        if confidence is not None:
            line["confidence"] = confidence
        font_size = _safe_float(item.get("font_size") or item.get("size"))
        if font_size is not None:
            line["font_size"] = font_size
        if isinstance(item.get("meta"), dict):
            line["meta"] = dict(item.get("meta") or {})
        lines.append(line)
    return lines


def _summarize_ocr_meta(meta: Any) -> Dict[str, Any]:
    if not isinstance(meta, dict):
        return {}
    summary: Dict[str, Any] = {}
    preferred_keys = {
        "build_id",
        "engine",
        "model",
        "lang",
        "page_width",
        "page_height",
        "width",
        "height",
        "rotation",
    }
    for key in preferred_keys:
        value = meta.get(key)
        if isinstance(value, (str, int, float, bool)) or value is None:
            if isinstance(value, str) and len(value) > 120:
                value = value[:120]
            summary[key] = value
    if not summary:
        for key, value in meta.items():
            if isinstance(value, (str, int, float, bool)) or value is None:
                if isinstance(value, str) and len(value) > 120:
                    value = value[:120]
                summary[key] = value
            if len(summary) >= 8:
                break
    return summary


def _clause_heading_label(text: str) -> str:
    s = (text or "").strip()
    match = re.match(r"^(第[一二三四五六七八九十百千0-9]+条(?:之[一二三四五六七八九十百千0-9]+)?)", s)
    return match.group(1) if match else ""


def _normalize_section_path(section_path: Optional[List[Any]]) -> List[str]:
    normalized: List[str] = []
    for part in section_path or []:
        label = str(part or "").strip()
        if not label:
            continue
        if normalized and normalized[-1] == label:
            continue
        normalized.append(label)
    return normalized


def _section_path_label(section_path: Optional[List[Any]]) -> str:
    return " / ".join(_normalize_section_path(section_path))


def _section_node_id(section_path: Optional[List[Any]]) -> Optional[str]:
    normalized = _normalize_section_path(section_path)
    if not normalized:
        return None
    return "section::" + " > ".join(_normalize_ir_text(part) for part in normalized)


def _build_section_ref(
    section_path: Optional[List[Any]],
    *,
    fallback_section: str = "",
    section_id: Optional[Any] = None,
) -> Dict[str, Any]:
    normalized_path = _normalize_section_path(section_path or ([fallback_section] if fallback_section else []))
    path_label = _section_path_label(normalized_path)
    section = (fallback_section or "").strip() or path_label or (normalized_path[-1] if normalized_path else "")
    section_title = normalized_path[-1] if normalized_path else section
    parent_section_path = normalized_path[:-1]
    return {
        "section": section,
        "section_title": section_title,
        "section_id": section_id,
        "section_path": normalized_path,
        "section_node_id": _section_node_id(normalized_path),
        "parent_section_id": _section_node_id(parent_section_path),
        "parent_section_path": parent_section_path,
        "parent_section_title": parent_section_path[-1] if parent_section_path else None,
        "section_depth": len(normalized_path),
    }


def _resolved_ir_section_path(element: Dict[str, Any], raw_text: str) -> List[str]:
    explicit_path = _normalize_section_path(element.get("section_path"))
    chapter_title = _chapter_heading_title(raw_text)
    if not chapter_title:
        return explicit_path
    if explicit_path and explicit_path[-1] == chapter_title:
        return explicit_path
    if explicit_path and (element.get("element_type") or "") in {"heading", "sheet", "title"}:
        return explicit_path + [chapter_title]
    if explicit_path:
        return explicit_path
    return [chapter_title]


def _base_chunk_unit(
    *,
    section: str,
    clause_label: str,
    text: str,
    raw_text: Optional[str] = None,
    normalized_text: Optional[str] = None,
    fts_text: Optional[str] = None,
    section_path: Optional[List[str]] = None,
    page_no: Optional[int] = None,
    element_id: Optional[str] = None,
    element_type: Optional[str] = None,
    reading_order: Optional[int] = None,
    unit_kind: str = "paragraph",
    section_id: Optional[Any] = None,
    section_title: Optional[str] = None,
    section_node_id: Optional[str] = None,
    parent_section_id: Optional[str] = None,
    parent_section_path: Optional[List[str]] = None,
    parent_section_title: Optional[str] = None,
    section_depth: Optional[int] = None,
    article_no: Optional[str] = None,
) -> Dict[str, Any]:
    raw = (raw_text if raw_text is not None else text or "").strip()
    normalized = (normalized_text or _normalize_ir_text(raw)).strip() or raw
    fts = (fts_text or raw or normalized).strip() or normalized
    normalized_section_path = _normalize_section_path(section_path or ([] if not section else [section]))
    return {
        "section": (section or "").strip(),
        "section_title": (section_title or section or "").strip(),
        "section_id": section_id,
        "section_node_id": section_node_id,
        "clause_label": (clause_label or "").strip(),
        "text": (text or raw or normalized).strip(),
        "raw_text": raw,
        "normalized_text": normalized,
        "fts_text": fts,
        "section_path": normalized_section_path,
        "parent_section_id": parent_section_id,
        "parent_section_path": _normalize_section_path(parent_section_path),
        "parent_section_title": (parent_section_title or "").strip() or None,
        "section_depth": int(section_depth or len(normalized_section_path)),
        "article_no": (article_no or clause_label or "").strip(),
        "page_no": page_no,
        "element_id": element_id,
        "element_type": element_type,
        "reading_order": reading_order,
        "unit_kind": unit_kind,
    }


def _build_plain_chunk_units(text: str) -> List[Dict[str, Any]]:
    normalized = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    lines = normalized.split("\n")
    units: List[Dict[str, Any]] = []
    current_section_ref: Optional[Dict[str, Any]] = None
    current_clause = ""
    current_lines: List[str] = []
    section_ordinals: Dict[str, int] = {}

    def _section_ref_for_path(section_path: Optional[List[Any]], fallback_section: str = "") -> Dict[str, Any]:
        normalized_path = _normalize_section_path(section_path or ([fallback_section] if fallback_section else []))
        node_id = _section_node_id(normalized_path)
        ordinal = None
        if node_id:
            ordinal = section_ordinals.setdefault(node_id, len(section_ordinals) + 1)
        return _build_section_ref(normalized_path, fallback_section=fallback_section, section_id=ordinal)

    def _flush_current():
        nonlocal current_lines
        body = "\n".join(current_lines).strip()
        current_lines = []
        if not body:
            return
        section_ref = current_section_ref or _build_section_ref([], fallback_section="", section_id=None)
        units.append(
            _base_chunk_unit(
                section=section_ref.get("section") or "",
                section_title=section_ref.get("section_title"),
                section_id=section_ref.get("section_id"),
                section_node_id=section_ref.get("section_node_id"),
                clause_label=current_clause,
                text=body,
                raw_text=body,
                normalized_text=_normalize_ir_text(body),
                fts_text=body,
                section_path=section_ref.get("section_path") or [],
                parent_section_id=section_ref.get("parent_section_id"),
                parent_section_path=section_ref.get("parent_section_path") or [],
                parent_section_title=section_ref.get("parent_section_title"),
                section_depth=section_ref.get("section_depth"),
                article_no=current_clause,
                unit_kind="clause" if current_clause else "paragraph",
            )
        )

    for raw_line in lines:
        line = (raw_line or "").rstrip()
        stripped = line.strip()
        chapter_title = _chapter_heading_title(stripped)
        if chapter_title:
            _flush_current()
            current_section_ref = _section_ref_for_path([chapter_title], fallback_section=chapter_title)
            current_clause = ""
            continue
        clause_label = _clause_heading_label(stripped)
        if clause_label:
            _flush_current()
            current_clause = clause_label
            current_lines = [line]
            continue
        if not stripped:
            if current_lines:
                current_lines.append("")
            continue
        current_lines.append(line)

    _flush_current()
    if units:
        return units
    body = normalized.strip()
    if not body:
        return []
    root_section_ref = _build_section_ref([], fallback_section="", section_id=None)
    return [
        _base_chunk_unit(
            section="",
            section_title=root_section_ref.get("section_title"),
            section_id=root_section_ref.get("section_id"),
            section_node_id=root_section_ref.get("section_node_id"),
            clause_label="",
            text=body,
            raw_text=body,
            fts_text=body,
            section_path=[],
            parent_section_id=root_section_ref.get("parent_section_id"),
            parent_section_path=root_section_ref.get("parent_section_path") or [],
            parent_section_title=root_section_ref.get("parent_section_title"),
            section_depth=root_section_ref.get("section_depth"),
            unit_kind="paragraph",
        )
    ]


def _build_ir_chunk_units(document_ir: Dict[str, Any]) -> List[Dict[str, Any]]:
    units: List[Dict[str, Any]] = []
    current_section_ref: Optional[Dict[str, Any]] = None
    current_clause = ""
    current_elements: List[Dict[str, Any]] = []
    section_ordinals: Dict[str, int] = {}

    def _section_ref_for_path(section_path: Optional[List[Any]], fallback_section: str = "") -> Optional[Dict[str, Any]]:
        normalized_path = _normalize_section_path(section_path or ([fallback_section] if fallback_section else []))
        if not normalized_path and not (fallback_section or "").strip():
            return None
        node_id = _section_node_id(normalized_path)
        ordinal = None
        if node_id:
            ordinal = section_ordinals.setdefault(node_id, len(section_ordinals) + 1)
        return _build_section_ref(normalized_path, fallback_section=fallback_section, section_id=ordinal)

    def _flush_current():
        nonlocal current_elements
        if not current_elements:
            return
        serialized_parts: List[str] = []
        raw_parts: List[str] = []
        normalized_parts: List[str] = []
        fts_parts: List[str] = []
        for element in current_elements:
            serialized = _serialize_ir_element(element)
            if serialized["serialized_text"]:
                serialized_parts.append(serialized["serialized_text"])
            if serialized["raw_text"]:
                raw_parts.append(serialized["raw_text"])
            if serialized["normalized_text"]:
                normalized_parts.append(serialized["normalized_text"])
            if serialized["fts_text"]:
                fts_parts.append(serialized["fts_text"])
        primary = current_elements[0]
        combined_raw = "\n".join(raw_parts).strip()
        combined_text = "\n".join(serialized_parts).strip() or combined_raw
        combined_normalized = "\n".join(normalized_parts).strip() or _normalize_ir_text(combined_raw)
        combined_fts = "\n".join(fts_parts).strip() or combined_raw
        if not (combined_text or combined_raw or combined_normalized or combined_fts):
            current_elements = []
            return
        section_ref = current_section_ref or _build_section_ref([], fallback_section="", section_id=None)
        units.append(
            _base_chunk_unit(
                section=section_ref.get("section") or "",
                section_title=section_ref.get("section_title"),
                section_id=section_ref.get("section_id"),
                section_node_id=section_ref.get("section_node_id"),
                clause_label=current_clause,
                text=combined_text,
                raw_text=combined_raw,
                normalized_text=combined_normalized,
                fts_text=combined_fts,
                section_path=section_ref.get("section_path") or [],
                parent_section_id=section_ref.get("parent_section_id"),
                parent_section_path=section_ref.get("parent_section_path") or [],
                parent_section_title=section_ref.get("parent_section_title"),
                section_depth=section_ref.get("section_depth"),
                article_no=current_clause,
                page_no=primary.get("page_no"),
                element_id=primary.get("element_id"),
                element_type="clause_group" if current_clause else (primary.get("element_type") or "paragraph"),
                reading_order=primary.get("reading_order"),
                unit_kind="clause" if current_clause else "paragraph",
            )
        )
        current_elements = []

    for element in document_ir.get("elements") or []:
        if element.get("element_type") == "page_break":
            continue
        serialized = _serialize_ir_element(element)
        raw_text = (serialized.get("raw_text") or "").strip()
        chapter_title = _chapter_heading_title(raw_text)
        resolved_section_path = _resolved_ir_section_path(element, raw_text)
        resolved_section_label = _section_path_label(resolved_section_path) or (serialized.get("section_label") or "").strip()
        element_section_ref = _section_ref_for_path(resolved_section_path, fallback_section=resolved_section_label)
        if chapter_title:
            _flush_current()
            current_section_ref = _section_ref_for_path(resolved_section_path or [chapter_title], fallback_section=chapter_title)
            current_clause = ""
            continue
        if element_section_ref and (not current_section_ref or element_section_ref.get("section_node_id") != current_section_ref.get("section_node_id")):
            _flush_current()
            current_section_ref = element_section_ref
            current_clause = ""
        elif current_section_ref is None and element_section_ref is not None:
            current_section_ref = element_section_ref
        clause_label = _clause_heading_label(raw_text)
        if clause_label:
            _flush_current()
            current_clause = clause_label
            current_elements = [element]
            continue
        if not current_elements:
            current_clause = ""
            current_elements = [element]
        else:
            current_elements.append(element)

    _flush_current()
    return units


def _semantic_source_span(element: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "element_id": element.get("element_id"),
        "page_no": element.get("page_no"),
        "reading_order": element.get("reading_order"),
        "element_type": element.get("element_type"),
    }


def _semantic_unit_base(
    *,
    unit_id: str,
    unit_type: str,
    section_ref: Optional[Dict[str, Any]],
    text: str,
    raw_text: Optional[str] = None,
    normalized_text: Optional[str] = None,
    fts_text: Optional[str] = None,
    clause_label: str = "",
    unit_kind: str = "paragraph",
    page_span: Optional[List[int]] = None,
    source_spans: Optional[List[Dict[str, Any]]] = None,
    payload: Optional[Dict[str, Any]] = None,
    element_type: Optional[str] = None,
    article_no: str = "",
) -> Dict[str, Any]:
    section_ref = section_ref or _build_section_ref([], fallback_section="", section_id=None)
    text_value = (text or raw_text or normalized_text or "").strip()
    raw_value = (raw_text if raw_text is not None else text_value).strip()
    normalized_value = (normalized_text or _normalize_ir_text(raw_value)).strip() or raw_value
    fts_value = (fts_text or raw_value or normalized_value).strip() or normalized_value
    spans = list(source_spans or [])
    pages = [int(span.get("page_no")) for span in spans if span.get("page_no") is not None]
    span_range = list(page_span or ([] if not pages else [min(pages), max(pages)]))
    return {
        "unit_id": unit_id,
        "unit_type": unit_type,
        "unit_kind": unit_kind,
        "clause_label": (clause_label or "").strip(),
        "article_no": (article_no or clause_label or "").strip(),
        "section": section_ref.get("section") or "",
        "section_title": section_ref.get("section_title") or section_ref.get("section") or "",
        "section_id": section_ref.get("section_id"),
        "section_node_id": section_ref.get("section_node_id"),
        "section_path": section_ref.get("section_path") or [],
        "parent_section_id": section_ref.get("parent_section_id"),
        "parent_section_path": section_ref.get("parent_section_path") or [],
        "parent_section_title": section_ref.get("parent_section_title"),
        "section_depth": section_ref.get("section_depth") or 0,
        "text": text_value,
        "raw_text": raw_value,
        "normalized_text": normalized_value,
        "fts_text": fts_value,
        "page_no": span_range[0] if span_range else (pages[0] if pages else None),
        "page_span": span_range,
        "source_spans": spans,
        "payload": payload or {},
        "element_type": element_type or unit_type,
    }


def _normalize_semantic_table_payload(element: Dict[str, Any]) -> Dict[str, Any]:
    payload = dict(element.get("json_payload") or {})
    table_json = payload.get("table_json") if isinstance(payload.get("table_json"), dict) else {}
    headers = []
    rows = []
    merged_cells = []
    if isinstance(table_json.get("headers"), list):
        headers = [str(value or "").strip() for value in table_json.get("headers") or []]
    if isinstance(table_json.get("rows"), list):
        raw_rows = table_json.get("rows") or []
        if raw_rows and all(isinstance(row, list) for row in raw_rows):
            rows = [["" if cell is None else str(cell) for cell in row] for row in raw_rows]
        elif raw_rows and all(isinstance(row, dict) for row in raw_rows):
            for row in raw_rows:
                cells = row.get("cells") or []
                rendered = []
                for cell in cells:
                    if isinstance(cell, dict):
                        rendered.append("" if cell.get("display_value") is None else str(cell.get("display_value")))
                    else:
                        rendered.append("" if cell is None else str(cell))
                rows.append(rendered)
    if isinstance(table_json.get("merged_cells"), list):
        merged_cells = list(table_json.get("merged_cells") or [])
    if not rows and isinstance(payload.get("rows"), list):
        rows = [["" if cell is None else str(cell) for cell in row] for row in payload.get("rows") or [] if isinstance(row, list)]
    if not headers and rows:
        headers = [f"col_{idx + 1}" for idx in range(max(len(row) for row in rows))]
    table_text = (payload.get("table_text") or element.get("markdown") or element.get("text_raw") or "").strip()
    return {
        "table_id": element.get("element_id"),
        "caption": (payload.get("caption") or "").strip() or None,
        "headers": headers,
        "rows": rows,
        "merged_cells": merged_cells,
        "column_count": int(table_json.get("column_count") or len(headers) or (max((len(row) for row in rows), default=0))),
        "row_count": int(table_json.get("row_count") or len(rows)),
        "header_row_count": int(table_json.get("header_row_count") or (1 if headers else 0)),
        "table_text": table_text,
        "table_json": table_json,
    }


def _render_semantic_table_markdown(headers: List[str], rows: List[List[str]]) -> str:
    normalized_headers = [str(value or "").strip() for value in headers]
    width = max(len(normalized_headers), max((len(row) for row in rows), default=0))
    if width <= 0:
        return ""
    if not normalized_headers:
        normalized_headers = [f"col_{idx + 1}" for idx in range(width)]
    if len(normalized_headers) < width:
        normalized_headers.extend([f"col_{idx + 1}" for idx in range(len(normalized_headers), width)])
    normalized_rows = []
    for row in rows:
        current = [str(value or "").strip() for value in row]
        if len(current) < width:
            current.extend([""] * (width - len(current)))
        normalized_rows.append(current[:width])
    lines = ["| " + " | ".join(normalized_headers[:width]) + " |", "| " + " | ".join(["---"] * width) + " |"]
    for row in normalized_rows:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines).strip()


def _render_semantic_unit_text(unit: Dict[str, Any]) -> str:
    if (unit.get("unit_type") or "") == "table":
        payload = unit.get("payload") or {}
        caption = (payload.get("caption") or "").strip()
        rendered = _render_semantic_table_markdown(payload.get("headers") or [], payload.get("rows") or [])
        if not rendered:
            rendered = (payload.get("table_text") or unit.get("raw_text") or unit.get("text") or "").strip()
        if caption:
            return (caption + "\n" + rendered).strip()
        return rendered.strip()
    return (unit.get("text") or unit.get("raw_text") or "").strip()


def _semantic_unit_role(element: Dict[str, Any], raw_text: str) -> str:
    element_type = (element.get("element_type") or "paragraph").strip().lower()
    payload = element.get("json_payload") or {}
    pdf_role = str(payload.get("pdf_role") or payload.get("ocr_role") or "").strip().lower()
    section_path = _normalize_section_path(element.get("section_path"))
    if element_type == "title":
        return "title"
    if section_path and section_path[0] == "toc":
        return "toc"
    if pdf_role in {"toc_title", "toc_entry"}:
        return "toc"
    if pdf_role == "appendix_heading":
        return "appendix_heading"
    if element_type == "heading":
        if _chapter_heading_title(raw_text):
            return "chapter_heading"
        return "section_heading"
    if element_type == "sheet":
        return "section_heading"
    if element_type == "table":
        return "table"
    if element_type == "list_item":
        return "list"
    return "paragraph"


def _build_chunk_hydration_prefix(filename: str, units: List[Dict[str, Any]], total_chunks: Optional[int] = None, chunk_index: Optional[int] = None) -> List[str]:
    profile = _doc_title_profile(filename)
    title = (profile.get("canonical_title") or profile.get("stem") or _filename_stem(filename)).strip()
    first = units[0] if units else {}
    lines = [f"文档标题：{title}"]
    section_path = first.get("section_path") or []
    if section_path:
        lines.append(f"章节路径：{' > '.join(section_path)}")
    clause_label = (first.get("clause_label") or "").strip()
    if clause_label:
        lines.append(f"条文锚点：{clause_label}")
    page_values = [page for unit in units for page in (unit.get("page_span") or []) if page is not None]
    if page_values:
        lines.append(f"页码范围：{min(page_values)}-{max(page_values)}")
    chunk_role = (first.get("unit_type") or first.get("unit_kind") or "paragraph").strip()
    if chunk_role:
        lines.append(f"语义单元：{chunk_role}")
    if chunk_index is not None and total_chunks is not None and total_chunks > 0:
        lines.append(f"分块位置：第 {chunk_index + 1}/{total_chunks} 块")
    return lines


def _build_semantic_units_from_document_ir(document_ir: Dict[str, Any]) -> List[Dict[str, Any]]:
    units: List[Dict[str, Any]] = []
    current_section_ref: Optional[Dict[str, Any]] = None
    current_clause = ""
    current_elements: List[Dict[str, Any]] = []
    current_kind = "paragraph"
    section_ordinals: Dict[str, int] = {}

    def _section_ref_for_path(section_path: Optional[List[Any]], fallback_section: str = "") -> Optional[Dict[str, Any]]:
        normalized_path = _normalize_section_path(section_path or ([fallback_section] if fallback_section else []))
        if not normalized_path and not (fallback_section or "").strip():
            return None
        node_id = _section_node_id(normalized_path)
        ordinal = None
        if node_id:
            ordinal = section_ordinals.setdefault(node_id, len(section_ordinals) + 1)
        return _build_section_ref(normalized_path, fallback_section=fallback_section, section_id=ordinal)

    def _flush_text_unit():
        nonlocal current_elements, current_clause, current_kind
        if not current_elements:
            return
        serialized_parts: List[str] = []
        raw_parts: List[str] = []
        normalized_parts: List[str] = []
        fts_parts: List[str] = []
        source_spans: List[Dict[str, Any]] = []
        for element in current_elements:
            serialized = _serialize_ir_element(element)
            if serialized.get("serialized_text"):
                serialized_parts.append(serialized["serialized_text"])
            if serialized.get("raw_text"):
                raw_parts.append(serialized["raw_text"])
            if serialized.get("normalized_text"):
                normalized_parts.append(serialized["normalized_text"])
            if serialized.get("fts_text"):
                fts_parts.append(serialized["fts_text"])
            source_spans.append(_semantic_source_span(element))
        combined_text = "\n".join(serialized_parts).strip() or "\n".join(raw_parts).strip()
        combined_raw = "\n".join(raw_parts).strip()
        combined_normalized = "\n".join(normalized_parts).strip() or _normalize_ir_text(combined_raw)
        combined_fts = "\n".join(fts_parts).strip() or combined_raw
        if combined_text or combined_raw:
            unit_type = "article" if current_clause else current_kind
            units.append(
                _semantic_unit_base(
                    unit_id=f"su_{len(units):06d}",
                    unit_type=unit_type,
                    unit_kind="clause" if current_clause else current_kind,
                    section_ref=current_section_ref,
                    text=combined_text,
                    raw_text=combined_raw,
                    normalized_text=combined_normalized,
                    fts_text=combined_fts,
                    clause_label=current_clause,
                    article_no=current_clause,
                    source_spans=source_spans,
                    element_type=unit_type,
                )
            )
        current_elements = []
        current_clause = ""
        current_kind = "paragraph"

    for element in document_ir.get("elements") or []:
        if element.get("element_type") == "page_break":
            continue
        serialized = _serialize_ir_element(element)
        raw_text = (serialized.get("raw_text") or "").strip()
        chapter_title = _chapter_heading_title(raw_text)
        resolved_section_path = _resolved_ir_section_path(element, raw_text)
        resolved_section_label = _section_path_label(resolved_section_path) or (serialized.get("section_label") or "").strip()
        element_section_ref = _section_ref_for_path(resolved_section_path, fallback_section=resolved_section_label)
        element_type = (element.get("element_type") or "paragraph").strip().lower()
        unit_role = _semantic_unit_role(element, raw_text)
        if unit_role in {"title", "toc", "chapter_heading", "section_heading", "appendix_heading"}:
            _flush_text_unit()
            if chapter_title:
                current_section_ref = _section_ref_for_path(resolved_section_path or [chapter_title], fallback_section=chapter_title)
            elif element_section_ref and unit_role != "title":
                current_section_ref = element_section_ref
            units.append(
                _semantic_unit_base(
                    unit_id=f"su_{len(units):06d}",
                    unit_type=unit_role,
                    unit_kind=unit_role,
                    section_ref=current_section_ref or element_section_ref,
                    text=serialized.get("serialized_text") or raw_text,
                    raw_text=raw_text,
                    normalized_text=serialized.get("normalized_text") or _normalize_ir_text(raw_text),
                    fts_text=serialized.get("fts_text") or raw_text,
                    source_spans=[_semantic_source_span(element)],
                    payload=element.get("json_payload") or {},
                    element_type=element_type,
                )
            )
            continue
        if chapter_title:
            _flush_text_unit()
            current_section_ref = _section_ref_for_path(resolved_section_path or [chapter_title], fallback_section=chapter_title)
            continue
        if element_section_ref and (not current_section_ref or element_section_ref.get("section_node_id") != current_section_ref.get("section_node_id")):
            _flush_text_unit()
            current_section_ref = element_section_ref
        elif current_section_ref is None and element_section_ref is not None:
            current_section_ref = element_section_ref
        if element_type == "table":
            _flush_text_unit()
            payload = _normalize_semantic_table_payload(element)
            units.append(
                _semantic_unit_base(
                    unit_id=f"su_{len(units):06d}",
                    unit_type="table",
                    unit_kind="table",
                    section_ref=current_section_ref or element_section_ref,
                    text=(payload.get("table_text") or element.get("text_raw") or "").strip(),
                    raw_text=(payload.get("table_text") or element.get("text_raw") or "").strip(),
                    normalized_text=_normalize_ir_text(payload.get("table_text") or element.get("text_raw") or ""),
                    fts_text=(payload.get("table_text") or element.get("text_raw") or "").strip(),
                    source_spans=[_semantic_source_span(element)],
                    payload=payload,
                    element_type="table",
                )
            )
            continue
        clause_label = _clause_heading_label(raw_text)
        if clause_label:
            _flush_text_unit()
            current_clause = clause_label
            current_kind = "paragraph"
            current_elements = [element]
            continue
        if not current_elements:
            current_kind = "list" if element_type == "list_item" else "paragraph"
            current_elements = [element]
            continue
        if current_clause:
            current_elements.append(element)
            continue
        next_kind = "list" if element_type == "list_item" else "paragraph"
        same_kind = next_kind == current_kind
        same_section = bool(current_section_ref) or not element_section_ref
        if same_kind and same_section:
            current_elements.append(element)
        else:
            _flush_text_unit()
            current_kind = next_kind
            current_elements = [element]
    _flush_text_unit()
    return units


def _render_table_chunk_slice(unit: Dict[str, Any], row_start: int, row_end: int) -> Dict[str, Any]:
    payload = unit.get("payload") or {}
    rows = list(payload.get("rows") or [])
    headers = list(payload.get("headers") or [])
    selected_rows = rows[row_start:row_end]
    body = _render_semantic_table_markdown(headers, selected_rows)
    caption = (payload.get("caption") or "").strip()
    if caption:
        body = (caption + "\n" + body).strip()
    sliced_payload = {
        **payload,
        "rows": selected_rows,
        "row_start": row_start,
        "row_end": row_end,
    }
    return {
        **unit,
        "text": body,
        "raw_text": body,
        "normalized_text": _normalize_ir_text(body),
        "fts_text": body,
        "payload": sliced_payload,
    }


def _semantic_units_to_chunk_plans(filename: str, units: List[Dict[str, Any]], chunk_size: int, overlap: int) -> List[Dict[str, Any]]:
    plans: List[Dict[str, Any]] = []
    current_bucket: List[Dict[str, Any]] = []
    current_bucket_len = 0

    def _append_plan(unit_like: Dict[str, Any], *, text_value: str, raw_value: str, normalized_value: str, fts_value: str, payload: Optional[Dict[str, Any]] = None):
        first_span = (unit_like.get("source_spans") or [{}])[0]
        plans.append({
            "chunk_plan_id": f"cp_{len(plans):06d}",
            "chunk_role": unit_like.get("unit_type") or "body",
            "section": unit_like.get("section") or "",
            "section_title": unit_like.get("section_title") or unit_like.get("section") or "",
            "section_id": unit_like.get("section_id"),
            "section_node_id": unit_like.get("section_node_id"),
            "section_path": unit_like.get("section_path") or [],
            "parent_section_id": unit_like.get("parent_section_id"),
            "parent_section_path": unit_like.get("parent_section_path") or [],
            "parent_section_title": unit_like.get("parent_section_title"),
            "section_depth": unit_like.get("section_depth"),
            "clause_label": unit_like.get("clause_label") or "",
            "article_no": unit_like.get("article_no") or unit_like.get("clause_label") or "",
            "unit_kind": unit_like.get("unit_kind") or unit_like.get("unit_type") or "paragraph",
            "semantic_unit_ids": [unit_like.get("unit_id")] if unit_like.get("unit_id") else (unit_like.get("semantic_unit_ids") or []),
            "text": text_value,
            "raw_text": raw_value,
            "normalized_text": normalized_value,
            "fts_text": fts_value,
            "page_no": unit_like.get("page_no"),
            "page_span": unit_like.get("page_span") or [],
            "payload": payload if payload is not None else (unit_like.get("payload") or {}),
            "element_id": first_span.get("element_id"),
            "element_type": unit_like.get("element_type") or unit_like.get("unit_type") or "paragraph",
            "reading_order": first_span.get("reading_order"),
        })

    def _flush_bucket():
        nonlocal current_bucket, current_bucket_len
        if not current_bucket:
            return
        body = "\n".join([_render_semantic_unit_text(unit) for unit in current_bucket if _render_semantic_unit_text(unit)]).strip()
        if body:
            first = current_bucket[0]
            page_values = [page for unit in current_bucket for page in (unit.get("page_span") or []) if page is not None]
            combined_unit = {
                **first,
                "unit_type": "table" if first.get("unit_type") == "table" else "body",
                "unit_kind": first.get("unit_kind") or "paragraph",
                "semantic_unit_ids": [unit.get("unit_id") for unit in current_bucket],
                "page_no": min(page_values) if page_values else first.get("page_no"),
                "page_span": ([min(page_values), max(page_values)] if page_values else (first.get("page_span") or [])),
                "payload": {},
            }
            _add_plan_from_unit({
                **combined_unit,
                "text": body,
                "raw_text": body,
                "normalized_text": _normalize_ir_text(body),
                "fts_text": body,
            })
        current_bucket = []
        current_bucket_len = 0

    def _add_plan_from_unit(unit: Dict[str, Any]):
        body = (unit.get("text") or _render_semantic_unit_text(unit) or "").strip()
        raw_value = (unit.get("raw_text") or body).strip()
        normalized_value = (unit.get("normalized_text") or _normalize_ir_text(raw_value)).strip() or raw_value
        fts_value = (unit.get("fts_text") or raw_value or body).strip() or raw_value
        if len(raw_value) > chunk_size:
            text_chunks = split_text(body, chunk_size, overlap) or [body]
            raw_chunks = split_text(raw_value, chunk_size, overlap) or [raw_value]
            norm_chunks = split_text(normalized_value, chunk_size, overlap) or [normalized_value]
            fts_chunks = split_text(fts_value, chunk_size, overlap) or [fts_value]
            total_chunks = len(raw_chunks)
            for idx in range(total_chunks):
                split_payload = {**(unit.get("payload") or {}), "sub_chunk_index": idx, "sub_chunk_count": total_chunks}
                _append_plan(
                    unit,
                    text_value=text_chunks[min(idx, len(text_chunks) - 1)] if text_chunks else raw_chunks[idx],
                    raw_value=raw_chunks[idx],
                    normalized_value=norm_chunks[min(idx, len(norm_chunks) - 1)] if norm_chunks else raw_chunks[idx],
                    fts_value=fts_chunks[min(idx, len(fts_chunks) - 1)] if fts_chunks else raw_chunks[idx],
                    payload=split_payload,
                )
            return
        _append_plan(
            unit,
            text_value=body,
            raw_value=raw_value,
            normalized_value=normalized_value,
            fts_value=fts_value,
            payload=unit.get("payload") or {},
        )

    for unit in units:
        unit_type = (unit.get("unit_type") or "paragraph").strip()
        body = _render_semantic_unit_text(unit)
        body_len = len(body)
        if unit_type == "table":
            _flush_bucket()
            payload = unit.get("payload") or {}
            rows = list(payload.get("rows") or [])
            if rows and body_len > chunk_size:
                headers = payload.get("headers") or []
                header_markdown = _render_semantic_table_markdown(headers, [])
                approx_header = max(32, len(header_markdown))
                rows_per_chunk = max(1, (chunk_size - approx_header) // max(24, payload.get("column_count") or 1))
                for row_start in range(0, len(rows), rows_per_chunk):
                    sliced = _render_table_chunk_slice(unit, row_start, min(len(rows), row_start + rows_per_chunk))
                    _add_plan_from_unit(sliced)
            else:
                _add_plan_from_unit(unit)
            continue
        standalone = bool((unit.get("clause_label") or "").strip()) or (unit_type in {"title", "toc", "chapter_heading", "section_heading", "appendix_heading", "article", "table"})
        if standalone:
            _flush_bucket()
            _add_plan_from_unit(unit)
            continue
        if not current_bucket:
            current_bucket = [unit]
            current_bucket_len = body_len
            continue
        same_section = (current_bucket[0].get("section_node_id") or "") == (unit.get("section_node_id") or "")
        projected_len = current_bucket_len + (1 if current_bucket else 0) + body_len
        if same_section and projected_len <= chunk_size:
            current_bucket.append(unit)
            current_bucket_len = projected_len
        else:
            _flush_bucket()
            current_bucket = [unit]
            current_bucket_len = body_len
    _flush_bucket()
    total = len(plans)
    for idx, plan in enumerate(plans):
        plan["hydration_prefix_lines"] = _build_chunk_hydration_prefix(filename, [u for u in units if u.get("unit_id") in set(plan.get("semantic_unit_ids") or [])], total_chunks=total, chunk_index=idx)
    return plans


def _chunk_plans_to_items(plans: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    total = len(plans)
    for idx, plan in enumerate(plans):
        items.append({
            "chunk_id": idx,
            "prev_chunk_id": idx - 1 if idx > 0 else None,
            "next_chunk_id": idx + 1 if idx + 1 < total else None,
            "section": plan.get("section") or "",
            "section_title": plan.get("section_title") or plan.get("section") or "",
            "section_id": plan.get("section_id"),
            "section_node_id": plan.get("section_node_id"),
            "section_path": plan.get("section_path") or [],
            "parent_section_id": plan.get("parent_section_id"),
            "parent_section_path": plan.get("parent_section_path") or [],
            "parent_section_title": plan.get("parent_section_title"),
            "section_depth": plan.get("section_depth"),
            "clause_label": plan.get("clause_label") or "",
            "article_no": plan.get("article_no") or plan.get("clause_label") or "",
            "unit_kind": plan.get("unit_kind") or "paragraph",
            "chunk_role": plan.get("chunk_role") or "body",
            "semantic_unit_ids": plan.get("semantic_unit_ids") or [],
            "text": plan.get("text") or "",
            "raw_text": plan.get("raw_text") or plan.get("text") or "",
            "normalized_text": plan.get("normalized_text") or _normalize_ir_text(plan.get("raw_text") or plan.get("text") or ""),
            "fts_text": plan.get("fts_text") or plan.get("raw_text") or plan.get("text") or "",
            "page_no": plan.get("page_no"),
            "page_span": plan.get("page_span") or [],
            "payload": plan.get("payload") or {},
            "hydration_prefix_lines": plan.get("hydration_prefix_lines") or [],
            "element_id": plan.get("element_id"),
            "element_type": plan.get("element_type"),
            "reading_order": plan.get("reading_order"),
        })
    return items


def _merge_short_chunk_units(units: List[Dict[str, Any]], min_chars: int) -> List[Dict[str, Any]]:
    if not units:
        return []
    merged: List[Dict[str, Any]] = []
    buffer: Optional[Dict[str, Any]] = None

    def _merge_unit(target: Dict[str, Any], unit: Dict[str, Any]) -> Dict[str, Any]:
        joiner = "\n" if (target.get("raw_text") and unit.get("raw_text")) else ""
        target["text"] = (target.get("text") or "") + joiner + (unit.get("text") or "")
        target["raw_text"] = (target.get("raw_text") or "") + joiner + (unit.get("raw_text") or "")
        target["normalized_text"] = (target.get("normalized_text") or "") + joiner + (unit.get("normalized_text") or "")
        target["fts_text"] = (target.get("fts_text") or "") + joiner + (unit.get("fts_text") or "")
        return target

    def _flush_buffer():
        nonlocal buffer
        if buffer is not None:
            merged.append(buffer)
            buffer = None

    for unit in units:
        if (unit.get("clause_label") or "").strip():
            _flush_buffer()
            merged.append(unit)
            continue
        raw_len = len((unit.get("raw_text") or "").strip())
        if buffer is None:
            buffer = dict(unit)
            if raw_len >= min_chars:
                _flush_buffer()
            continue
        same_section = (buffer.get("section") or "") == (unit.get("section") or "")
        if same_section and len((buffer.get("raw_text") or "").strip()) < min_chars:
            buffer = _merge_unit(buffer, unit)
            if len((buffer.get("raw_text") or "").strip()) >= min_chars:
                _flush_buffer()
        else:
            _flush_buffer()
            buffer = dict(unit)
            if raw_len >= min_chars:
                _flush_buffer()

    _flush_buffer()
    return merged


def _split_overlong_chunk_units(units: List[Dict[str, Any]], chunk_size: int, overlap: int) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for unit in units:
        text_chunks = split_text(unit.get("text") or unit.get("raw_text") or "", chunk_size, overlap) or [unit.get("text") or unit.get("raw_text") or ""]
        raw_chunks = split_text(unit.get("raw_text") or unit.get("text") or "", chunk_size, overlap) or [unit.get("raw_text") or unit.get("text") or ""]
        norm_chunks = split_text(unit.get("normalized_text") or unit.get("raw_text") or "", chunk_size, overlap) or [unit.get("normalized_text") or unit.get("raw_text") or ""]
        fts_chunks = split_text(unit.get("fts_text") or unit.get("raw_text") or "", chunk_size, overlap) or [unit.get("fts_text") or unit.get("raw_text") or ""]
        total = len(text_chunks)
        for idx, text_chunk in enumerate(text_chunks):
            out.append({
                **unit,
                "text": text_chunk,
                "raw_text": raw_chunks[min(idx, len(raw_chunks) - 1)] if raw_chunks else text_chunk,
                "normalized_text": norm_chunks[min(idx, len(norm_chunks) - 1)] if norm_chunks else text_chunk,
                "fts_text": fts_chunks[min(idx, len(fts_chunks) - 1)] if fts_chunks else text_chunk,
                "sub_chunk_index": idx,
                "sub_chunk_count": total,
            })
    return out


def _attach_chunk_adjacency(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    total = len(items)
    for idx, item in enumerate(items):
        out.append({
            **item,
            "chunk_id": idx,
            "prev_chunk_id": idx - 1 if idx > 0 else None,
            "next_chunk_id": idx + 1 if idx + 1 < total else None,
        })
    return out


def _chunk_units_to_items(units: List[Dict[str, Any]], chunk_size: int, overlap: int) -> List[Dict[str, Any]]:
    min_chars = max(80, int(chunk_size * 0.35))
    merged = _merge_short_chunk_units(units, min_chars=min_chars)
    split_units = _split_overlong_chunk_units(merged, chunk_size=chunk_size, overlap=overlap)
    return _attach_chunk_adjacency(split_units)


def split_text_with_sections(filename: str, text: str, chunk_size: int, overlap: int) -> List[Dict[str, Any]]:
    del filename
    return _chunk_units_to_items(_build_plain_chunk_units(text), chunk_size=chunk_size, overlap=overlap)


def _contextualize_chunk_items(filename: str, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not items:
        return []
    profile = _doc_title_profile(filename)
    title = (profile.get("canonical_title") or profile.get("stem") or _filename_stem(filename)).strip()
    prev_chars = max(0, int(getattr(config, "CONTEXTUAL_PREV_CHARS", 120)))
    next_chars = max(0, int(getattr(config, "CONTEXTUAL_NEXT_CHARS", 120)))
    total = len(items)
    contextualized: List[Dict[str, Any]] = []
    for idx, item in enumerate(items):
        raw_text = (item.get("text") or "").strip()
        if not raw_text:
            continue
        prev_text = ((items[idx - 1].get("text") or "").strip() if idx > 0 else "")
        next_text = ((items[idx + 1].get("text") or "").strip() if idx + 1 < total else "")
        section = (item.get("section") or "").strip()
        prefix_lines = [str(line).strip() for line in (item.get("hydration_prefix_lines") or []) if str(line).strip()]
        parts = prefix_lines or [f"文档标题：{title}"]
        if not prefix_lines and section:
            parts.append(f"章节：{section}")
        if not any(str(line).startswith("分块位置：") for line in parts):
            parts.append(f"分块位置：第 {idx + 1}/{total} 块")
        if prev_text and prev_chars > 0:
            parts.append(f"上文：{prev_text[-prev_chars:].strip()}")
        parts.append(f"正文：{raw_text}")
        if next_text and next_chars > 0:
            parts.append(f"下文：{next_text[:next_chars].strip()}")
        contextualized.append({
            **item,
            "raw_text": raw_text,
            "text": "\n".join([p for p in parts if p and p.strip()]),
        })
    return contextualized


def _summarize_chunk_payload(payload: Any) -> Dict[str, Any]:
    if not isinstance(payload, dict):
        return {}
    if any(key in payload for key in ("ocr_meta", "ocr_line_meta", "ocr_role", "probe")):
        layout = payload.get("layout") if isinstance(payload.get("layout"), dict) else {}
        layout_summary: Dict[str, Any] = {}
        for key in ("font_size", "font_size_max", "font_size_median", "is_centered", "width_ratio", "top_ratio", "left_ratio", "center_offset_ratio", "line_count", "bold_ratio"):
            value = layout.get(key)
            if isinstance(value, (str, int, float, bool)) or value is None:
                layout_summary[key] = value
        probe = payload.get("probe") if isinstance(payload.get("probe"), dict) else {}
        probe_summary = {
            "route": probe.get("route"),
            "parser_backend": probe.get("parser_backend"),
        }
        summarized = {
            "probe": {k: v for k, v in probe_summary.items() if v is not None},
            "ocr_role": payload.get("ocr_role"),
            "pdf_role": payload.get("pdf_role"),
            "appendix_label": payload.get("appendix_label"),
            "heading_level": payload.get("heading_level"),
            "ocr_line_index": payload.get("ocr_line_index"),
            "ocr_meta": _summarize_ocr_meta(payload.get("ocr_meta")),
            "ocr_line_meta": _summarize_ocr_meta(payload.get("ocr_line_meta")),
            "layout": layout_summary,
        }
        return {k: v for k, v in summarized.items() if v not in (None, {}, [])}
    return payload


def _milvus_safe_metadata(metadata: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    if not isinstance(metadata, dict):
        return {}
    trimmed: Dict[str, Any] = {}
    for key, value in metadata.items():
        if key in {"raw_text", "text_normalized", "fts_text"}:
            continue
        if key == "payload":
            value = _summarize_chunk_payload(value)
        trimmed[key] = value
    return trimmed


def _normalize_ir_text(text: str) -> str:
    return " ".join((text or "").replace("\r", " ").replace("\n", " ").split())


def _new_document_ir(source: str, metadata: Optional[Dict[str, Any]] = None, doc_version: Optional[int] = None, parser_name: str = "plain_text") -> Dict[str, Any]:
    safe_source = _safe_filename(source)
    return {
        "doc_id": f"{safe_source}::{doc_version or 0}",
        "source": safe_source,
        "doc_version": doc_version,
        "metadata": dict(metadata or {}),
        "parser_name": parser_name,
        "parser_version": IR_PARSER_VERSION,
        "elements": [],
    }


def _append_ir_element(
    document_ir: Dict[str, Any],
    *,
    element_type: str,
    text_raw: str = "",
    text_normalized: Optional[str] = None,
    page_no: Optional[int] = None,
    section_path: Optional[List[str]] = None,
    bbox: Optional[Dict[str, Any]] = None,
    reading_order: Optional[int] = None,
    html: Optional[str] = None,
    markdown: Optional[str] = None,
    json_payload: Optional[Dict[str, Any]] = None,
    ocr_used: bool = False,
    ocr_confidence: Optional[float] = None,
    parser_name: Optional[str] = None,
):
    if element_type not in IR_ELEMENT_TYPES:
        element_type = "paragraph"
    elements = document_ir.setdefault("elements", [])
    idx = len(elements)
    raw = text_raw or ""
    normalized = text_normalized if text_normalized is not None else _normalize_ir_text(raw)
    elements.append({
        "element_id": f"el_{idx:06d}",
        "doc_id": document_ir.get("doc_id"),
        "source": document_ir.get("source"),
        "doc_version": document_ir.get("doc_version"),
        "page_no": page_no,
        "section_path": section_path or [],
        "element_type": element_type,
        "bbox": bbox,
        "reading_order": idx if reading_order is None else int(reading_order),
        "text_raw": raw,
        "text_normalized": normalized,
        "html": html,
        "markdown": markdown,
        "json_payload": json_payload,
        "ocr_used": bool(ocr_used),
        "ocr_confidence": ocr_confidence,
        "parser_name": parser_name or document_ir.get("parser_name") or "plain_text",
        "parser_version": document_ir.get("parser_version") or IR_PARSER_VERSION,
    })


def _document_ir_plain_text(document_ir: Dict[str, Any], normalized: bool = False) -> str:
    blocks: List[str] = []
    for element in document_ir.get("elements") or []:
        if element.get("element_type") == "page_break":
            continue
        text = (element.get("text_normalized") if normalized else element.get("text_raw")) or element.get("text_raw") or ""
        text = (text or "").strip()
        if text:
            blocks.append(text)
    return "\n\n".join(blocks).strip()


def _serialize_ir_element(element: Dict[str, Any]) -> Dict[str, str]:
    if _should_skip_ir_element_for_chunking(element):
        return {
            "serialized_text": "",
            "raw_text": "",
            "normalized_text": "",
            "fts_text": "",
            "section_label": "",
        }
    section_path = _normalize_section_path(element.get("section_path"))
    section_label = _section_path_label(section_path)
    raw_text = (element.get("text_raw") or "").strip()
    normalized = (element.get("text_normalized") or "").strip() or _normalize_ir_text(raw_text)
    payload = element.get("json_payload")
    payload_text = ""
    if payload:
        try:
            payload_text = json.dumps(payload, ensure_ascii=False)
        except Exception:
            payload_text = str(payload)
    content = normalized or raw_text or payload_text
    prefix_parts = []
    if section_label:
        prefix_parts.append(f"章节路径：{section_label}")
    if element.get("page_no") is not None:
        prefix_parts.append(f"页码：{element.get('page_no')}")
    prefix_parts.append(f"元素类型：{element.get('element_type') or 'paragraph'}")
    serialized = "\n".join(prefix_parts + [content]).strip()
    fts_text = raw_text
    if normalized and normalized != raw_text:
        fts_text = (fts_text + "\n" + normalized).strip()
    if payload_text and payload_text not in fts_text:
        fts_text = (fts_text + "\n" + payload_text).strip()
    return {
        "serialized_text": serialized,
        "raw_text": raw_text or content,
        "normalized_text": normalized or raw_text or content,
        "fts_text": fts_text or content,
        "section_label": section_label,
    }


def _store_document_ir(source: str, document_ir: Dict[str, Any]):
    conn = _lex_db_connect()
    safe_source = _safe_filename(source)
    version_key = int(document_ir.get("doc_version") or 0)
    conn.execute("DELETE FROM document_ir_meta WHERE source = ? AND doc_version = ?", (safe_source, version_key))
    conn.execute(
        "INSERT INTO document_ir_meta(source, doc_id, doc_version, metadata, parser_name, parser_version) VALUES (?,?,?,?,?,?)",
        (
            safe_source,
            document_ir.get("doc_id"),
            version_key,
            json.dumps(document_ir.get("metadata") or {}, ensure_ascii=False),
            document_ir.get("parser_name") or "plain_text",
            document_ir.get("parser_version") or IR_PARSER_VERSION,
        ),
    )
    conn.execute("DELETE FROM document_ir WHERE source = ? AND doc_version = ?", (safe_source, version_key))
    for element in document_ir.get("elements") or []:
        conn.execute(
            "INSERT INTO document_ir(source, doc_id, doc_version, element_id, page_no, section_path, element_type, bbox, reading_order, text_raw, text_normalized, html, markdown, json_payload, ocr_used, ocr_confidence, parser_name, parser_version) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                safe_source,
                element.get("doc_id"),
                element.get("doc_version"),
                element.get("element_id"),
                element.get("page_no"),
                json.dumps(element.get("section_path") or [], ensure_ascii=False),
                element.get("element_type"),
                json.dumps(element.get("bbox"), ensure_ascii=False) if element.get("bbox") is not None else None,
                element.get("reading_order"),
                element.get("text_raw") or "",
                element.get("text_normalized") or "",
                element.get("html"),
                element.get("markdown"),
                json.dumps(element.get("json_payload"), ensure_ascii=False) if element.get("json_payload") is not None else None,
                1 if element.get("ocr_used") else 0,
                element.get("ocr_confidence"),
                element.get("parser_name"),
                element.get("parser_version"),
            ),
        )


def _load_document_ir(source: str, doc_version: Optional[int] = None) -> Optional[Dict[str, Any]]:
    conn = _lex_db_connect()
    safe_source = _safe_filename(source)
    if doc_version is None:
        row = conn.execute("SELECT doc_id, doc_version, parser_name, parser_version FROM document_ir WHERE source = ? ORDER BY reading_order ASC LIMIT 1", (safe_source,)).fetchone()
        if not row:
            return None
        doc_id, doc_version, parser_name, parser_version = row
    else:
        row = conn.execute("SELECT doc_id, parser_name, parser_version FROM document_ir WHERE source = ? AND doc_version = ? ORDER BY reading_order ASC LIMIT 1", (safe_source, doc_version)).fetchone()
        if not row:
            return None
        doc_id, parser_name, parser_version = row
    rows = conn.execute(
        "SELECT element_id, page_no, section_path, element_type, bbox, reading_order, text_raw, text_normalized, html, markdown, json_payload, ocr_used, ocr_confidence, parser_name, parser_version "
        "FROM document_ir WHERE source = ? AND (? IS NULL OR doc_version = ?) ORDER BY reading_order ASC",
        (safe_source, doc_version, doc_version),
    ).fetchall()
    meta_row = conn.execute(
        "SELECT metadata FROM document_ir_meta WHERE source = ? AND doc_version = ? LIMIT 1",
        (safe_source, int(doc_version or 0)),
    ).fetchone()
    metadata = {}
    if meta_row and meta_row[0]:
        try:
            metadata = json.loads(meta_row[0])
        except Exception:
            metadata = {}
    document_ir = {
        "doc_id": doc_id,
        "source": safe_source,
        "doc_version": doc_version,
        "metadata": metadata,
        "parser_name": parser_name,
        "parser_version": parser_version,
        "elements": [],
    }
    for row in rows:
        section_path = []
        bbox = None
        json_payload = None
        try:
            section_path = json.loads(row[2] or "[]")
        except Exception:
            section_path = []
        try:
            bbox = json.loads(row[4]) if row[4] else None
        except Exception:
            bbox = None
        try:
            json_payload = json.loads(row[10]) if row[10] else None
        except Exception:
            json_payload = None
        document_ir["elements"].append({
            "element_id": row[0],
            "doc_id": doc_id,
            "source": safe_source,
            "doc_version": doc_version,
            "page_no": row[1],
            "section_path": section_path,
            "element_type": row[3],
            "bbox": bbox,
            "reading_order": row[5],
            "text_raw": row[6] or "",
            "text_normalized": row[7] or "",
            "html": row[8],
            "markdown": row[9],
            "json_payload": json_payload,
            "ocr_used": bool(row[11]),
            "ocr_confidence": row[12],
            "parser_name": row[13],
            "parser_version": row[14],
        })
    return document_ir


def _backfill_document_ir_from_chunks(source: str, doc_version: Optional[int] = None) -> Optional[Dict[str, Any]]:
    conn = _lex_db_connect()
    safe_source = _safe_filename(source)
    rows = conn.execute(
        "SELECT m.id, m.chunk_id, m.section, m.metadata, f.text FROM chunks_meta m LEFT JOIN chunks_fts f ON f.rowid = m.id WHERE m.source = ? ORDER BY m.chunk_id ASC",
        (safe_source,),
    ).fetchall()
    if not rows:
        return None
    target_version = doc_version or _get_active_version(safe_source) or _doc_get(safe_source).get("pending_version") or 1
    try:
        target_version = int(target_version)
    except Exception:
        target_version = 1
    filtered_rows = []
    for row in rows:
        metadata = {}
        try:
            metadata = json.loads(row[3] or "{}")
        except Exception:
            metadata = {}
        version = metadata.get("doc_version")
        try:
            version = int(version) if version is not None else None
        except Exception:
            version = None
        if version is None:
            version = target_version
        if version == target_version:
            filtered_rows.append((row, metadata))
    if not filtered_rows:
        return None
    document_ir = _new_document_ir(safe_source, metadata={"backfilled": True}, doc_version=target_version, parser_name="legacy_chunk_backfill")
    for ( _, chunk_id, section, _, fts_text), metadata in filtered_rows:
        raw_text = (metadata.get("raw_text") or fts_text or "").strip()
        normalized = (metadata.get("text_normalized") or _normalize_ir_text(raw_text)).strip()
        section_path = metadata.get("section_path") or ([section] if (section or "").strip() else [])
        _append_ir_element(
            document_ir,
            element_type=metadata.get("element_type") or "paragraph",
            text_raw=raw_text,
            text_normalized=normalized,
            page_no=metadata.get("page_no"),
            section_path=section_path,
            reading_order=int(metadata.get("reading_order") if metadata.get("reading_order") is not None else chunk_id or 0),
            parser_name=metadata.get("parser_name") or "legacy_chunk_backfill",
        )
    _store_document_ir(safe_source, document_ir)
    return document_ir


def _ensure_document_ir(source: str, doc_version: Optional[int] = None) -> Optional[Dict[str, Any]]:
    document_ir = _load_document_ir(source, doc_version)
    if document_ir and (document_ir.get("elements") or []):
        return document_ir
    return _backfill_document_ir_from_chunks(source, doc_version)


def _append_text_block_to_ir(document_ir: Dict[str, Any], text: str, *, page_no: Optional[int], base_section_path: Optional[List[str]] = None, parser_name: Optional[str] = None):
    lines = (text or "").replace("\r\n", "\n").replace("\r", "\n").split("\n")
    section_stack = list(base_section_path or [])
    paragraph_buf: List[str] = []
    in_code = False
    code_buf: List[str] = []

    def flush_paragraph():
        nonlocal paragraph_buf
        body = "\n".join(paragraph_buf).strip()
        paragraph_buf = []
        if not body:
            return
        if parser_name in {"pymupdf", "pymupdf4llm", "pypdf"} and _is_pdf_noise_text(body, page_no=page_no):
            return
        key_match = re.match(r"^([^:：\n]{1,80})[:：]\s*(.+)$", body)
        if key_match:
            _append_ir_element(
                document_ir,
                element_type="key_value",
                text_raw=body,
                text_normalized=f"{key_match.group(1).strip()}: {key_match.group(2).strip()}",
                page_no=page_no,
                section_path=section_stack,
                json_payload={"key": key_match.group(1).strip(), "value": key_match.group(2).strip()},
                parser_name=parser_name,
            )
        else:
            _append_ir_element(document_ir, element_type="paragraph", text_raw=body, page_no=page_no, section_path=section_stack, parser_name=parser_name)

    for raw_line in lines:
        line = (raw_line or "").rstrip()
        stripped = line.strip()
        if stripped.startswith("```"):
            flush_paragraph()
            if in_code:
                _append_ir_element(document_ir, element_type="code_block", text_raw="\n".join(code_buf).strip(), page_no=page_no, section_path=section_stack, parser_name=parser_name)
                code_buf = []
                in_code = False
            else:
                in_code = True
                code_buf = []
            continue
        if in_code:
            code_buf.append(line)
            continue
        if not stripped:
            flush_paragraph()
            continue
        if stripped == "---PAGE_BREAK---":
            flush_paragraph()
            _append_ir_element(document_ir, element_type="page_break", text_raw="", page_no=page_no, section_path=section_stack, parser_name=parser_name)
            continue
        if stripped.startswith("#"):
            flush_paragraph()
            level = len(stripped) - len(stripped.lstrip("#"))
            title = stripped.lstrip("#").strip()
            if level <= 1 and not document_ir.get("elements"):
                _append_ir_element(document_ir, element_type="title", text_raw=title, page_no=page_no, section_path=section_stack, parser_name=parser_name)
            else:
                section_stack = section_stack[:max(0, level - 2)] + [title]
                _append_ir_element(document_ir, element_type="heading", text_raw=title, page_no=page_no, section_path=section_stack[:-1], parser_name=parser_name)
            continue
        if stripped.startswith("# Sheet: "):
            flush_paragraph()
            sheet_name = stripped[len("# Sheet: "):].strip()
            section_stack = [sheet_name]
            _append_ir_element(document_ir, element_type="sheet", text_raw=sheet_name, page_no=page_no, section_path=[sheet_name], parser_name=parser_name)
            continue
        if re.match(r"^([\-\*•]|\d+[\.)])\s+", stripped):
            flush_paragraph()
            item_text = re.sub(r"^([\-\*•]|\d+[\.)])\s+", "", stripped).strip()
            _append_ir_element(document_ir, element_type="list_item", text_raw=item_text, page_no=page_no, section_path=section_stack, parser_name=parser_name)
            continue
        if re.match(r"^\$[^\n]+\$$", stripped):
            flush_paragraph()
            _append_ir_element(document_ir, element_type="formula", text_raw=stripped, page_no=page_no, section_path=section_stack, parser_name=parser_name)
            continue
        paragraph_buf.append(line)
    flush_paragraph()
    if in_code and code_buf:
        _append_ir_element(document_ir, element_type="code_block", text_raw="\n".join(code_buf).strip(), page_no=page_no, section_path=section_stack, parser_name=parser_name)


def _build_document_ir_from_text(filename: str, text: str, metadata: Optional[Dict[str, Any]] = None, parser_name: str = "plain_text", doc_version: Optional[int] = None) -> Dict[str, Any]:
    document_ir = _new_document_ir(filename, metadata=metadata, doc_version=doc_version, parser_name=parser_name)
    _append_text_block_to_ir(document_ir, text or "", page_no=1, parser_name=parser_name)
    if not document_ir.get("elements"):
        _append_ir_element(document_ir, element_type="paragraph", text_raw=(text or "").strip(), page_no=1, parser_name=parser_name)
    return document_ir


def _document_ir_to_structured_items(document_ir: Dict[str, Any], chunk_size: int, overlap: int) -> List[Dict[str, Any]]:
    semantic_units = _build_semantic_units_from_document_ir(document_ir)
    if not semantic_units:
        return []
    plans = _semantic_units_to_chunk_plans(document_ir.get("source") or "document", semantic_units, chunk_size, overlap)
    return _chunk_plans_to_items(plans)


def _looks_like_pdf_page_number(text: str, page_no: Optional[int] = None) -> bool:
    compact = re.sub(r"[\s\-—_·•.。，、:：/|]+", "", text or "")
    del page_no
    return bool(compact and re.fullmatch(r"\d{1,3}", compact))


def _is_pdf_noise_text(text: str, page_no: Optional[int] = None) -> bool:
    raw = (text or "").strip()
    if not raw:
        return True
    line_fragments = [re.sub(r"\s+", "", line) for line in raw.splitlines() if line.strip()]
    if len(line_fragments) >= 6 and line_fragments and max(len(line) for line in line_fragments) <= 1:
        return True
    compact = re.sub(r"\s+", "", raw)
    if not compact:
        return True
    if _looks_like_pdf_page_number(compact, page_no=page_no):
        return True
    meaningful_chars = re.findall(r"[A-Za-z0-9\u4e00-\u9fff]", compact)
    meaningful_runs = re.findall(r"[A-Za-z0-9\u4e00-\u9fff]+", raw)
    symbol_chars = len(compact) - len(meaningful_chars)
    if not meaningful_chars:
        return True
    if len(meaningful_chars) <= 2 and symbol_chars >= max(2, len(meaningful_chars)):
        return True
    if meaningful_runs and max(len(run) for run in meaningful_runs) <= 2 and symbol_chars >= len(meaningful_chars):
        return True
    if len(compact) >= 6 and symbol_chars / max(1, len(compact)) >= 0.6 and len(meaningful_chars) <= 3:
        return True
    return False


def _should_skip_ir_element_for_chunking(element: Dict[str, Any]) -> bool:
    element_type = (element.get("element_type") or "").strip().lower()
    raw_text = (element.get("text_raw") or "").strip()
    section_path = _normalize_section_path(element.get("section_path"))
    if section_path:
        root = section_path[0]
        if root == "toc" or root.startswith("header_") or root.startswith("footer_"):
            return True
    if element_type == "figure":
        payload = element.get("json_payload") or {}
        if not raw_text and isinstance(payload, dict) and payload.get("kind") == "image_block":
            return True
        if re.fullmatch(r"文本框\s*\d+", raw_text):
            return True
    if element_type in {"paragraph", "list_item", "key_value", "caption", "heading", "title"}:
        return _is_pdf_noise_text(raw_text, page_no=element.get("page_no"))
    return False


def _module_available(module_name: str) -> bool:
    try:
        return importlib.util.find_spec(module_name) is not None
    except Exception:
        return False


def _detect_parser_capabilities() -> Dict[str, bool]:
    return {
        "python_magic": _module_available("magic"),
        "pymupdf": _module_available("fitz"),
        "pymupdf4llm": _module_available("pymupdf4llm"),
        "docling": _module_available("docling"),
        "unstructured": _module_available("unstructured"),
        "paddleocr": _module_available("paddleocr"),
    }


def _bytes_look_like_text(content: bytes) -> bool:
    sample = content[:4096]
    if not sample:
        return True
    if b"\x00" in sample:
        return False
    textish = 0
    for byte in sample:
        if byte in (9, 10, 12, 13) or 32 <= byte <= 126 or byte >= 128:
            textish += 1
    return (textish / max(1, len(sample))) >= 0.85


def _sniff_zip_container_extension(content: bytes) -> Optional[str]:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            names = set(zf.namelist())
    except Exception:
        return None
    if "word/document.xml" in names:
        return ".docx"
    if "xl/workbook.xml" in names:
        return ".xlsx"
    return None


def _sniff_file_signature(content: bytes) -> Dict[str, Any]:
    sample = content[:32]
    if sample.startswith(b"%PDF"):
        return {"label": "pdf", "suggested_ext": ".pdf"}
    if sample.startswith((b"\xff\xd8\xff",)):
        return {"label": "jpeg", "suggested_ext": ".jpg"}
    if sample.startswith(b"\x89PNG\r\n\x1a\n"):
        return {"label": "png", "suggested_ext": ".png"}
    if sample.startswith((b"GIF87a", b"GIF89a")):
        return {"label": "gif", "suggested_ext": ".gif"}
    if sample.startswith((b"II*\x00", b"MM\x00*")):
        return {"label": "tiff", "suggested_ext": ".tiff"}
    if sample.startswith(b"BM"):
        return {"label": "bmp", "suggested_ext": ".bmp"}
    if sample.startswith((b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")):
        office_ext = _sniff_zip_container_extension(content)
        return {"label": "zip", "suggested_ext": office_ext}
    if sample.lstrip().startswith((b"{", b"[")):
        return {"label": "json", "suggested_ext": ".json"}
    if _bytes_look_like_text(content):
        return {"label": "text", "suggested_ext": None}
    return {"label": "binary", "suggested_ext": None}


def _sniff_mime_type(filename: str, content: bytes, signature_ext: Optional[str] = None) -> str:
    signature_mime = {
        ".pdf": "application/pdf",
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".json": "application/json",
        ".csv": "text/csv",
        ".txt": "text/plain",
        ".md": "text/markdown",
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".bmp": "image/bmp",
        ".gif": "image/gif",
        ".tif": "image/tiff",
        ".tiff": "image/tiff",
    }.get(signature_ext)
    mime_type = None
    if _module_available("magic"):
        try:
            import magic  # type: ignore

            mime_type = magic.from_buffer(content[:4096], mime=True)
        except Exception:
            mime_type = None
    if not mime_type:
        mime_type, _ = mimetypes.guess_type(filename)
    if signature_mime and (not mime_type or mime_type in {"text/plain", "application/octet-stream", "application/zip"}):
        mime_type = signature_mime
    if not mime_type and _bytes_look_like_text(content):
        mime_type = "text/plain"
    return mime_type or "application/octet-stream"


def _probe_pdf_layout_with_pymupdf(content: bytes) -> Dict[str, Any]:
    if not _module_available("fitz"):
        return {}
    try:
        import fitz  # type: ignore

        document = fitz.open(stream=content, filetype="pdf")
        page_count = getattr(document, "page_count", 0) or 0
        multi_column_pages = 0
        table_dense_pages = 0
        for page in document:
            blocks = [block for block in (page.get_text("blocks") or []) if len(block) >= 5 and str(block[4]).strip()]
            if blocks:
                left = sum(1 for block in blocks if float(block[0]) <= (page.rect.width * 0.45))
                right = sum(1 for block in blocks if float(block[0]) >= (page.rect.width * 0.55))
                if left and right:
                    multi_column_pages += 1
                dense = sum(1 for block in blocks if ("|" in str(block[4])) or ("\t" in str(block[4])) or re.search(r"\S+\s{2,}\S+", str(block[4])))
                if dense / max(1, len(blocks)) >= 0.25:
                    table_dense_pages += 1
        document.close()
        return {
            "multi_column": bool(page_count and multi_column_pages >= max(1, math.ceil(page_count / 2))),
            "table_dense": bool(page_count and table_dense_pages >= max(1, math.ceil(page_count / 3))),
            "layout_backend": "pymupdf",
        }
    except Exception:
        return {}


def _probe_pdf_document(content: bytes) -> Dict[str, Any]:
    try:
        from pypdf import PdfReader
    except ImportError:
        return {}
    try:
        reader = PdfReader(io.BytesIO(content))
        if getattr(reader, "is_encrypted", False):
            return {"page_count": len(reader.pages), "encrypted": True}
        text_lengths: List[int] = []
        image_pages = 0
        scanned_pages = 0
        garbled_pages = 0
        for page in reader.pages:
            page_text = (page.extract_text() or "").strip()
            text_lengths.append(len(page_text))
            if _looks_like_cid_garbled_text(page_text):
                garbled_pages += 1
            image_count = 0
            try:
                image_count = len(getattr(page, "images", []) or [])
            except Exception:
                image_count = 0
            if image_count > 0:
                image_pages += 1
            if len(page_text) < 40 and image_count > 0:
                scanned_pages += 1
        page_count = len(text_lengths)
        layout_probe = _probe_pdf_layout_with_pymupdf(content)
        return {
            "page_count": page_count,
            "is_scanned_pdf": bool(page_count and scanned_pages >= max(1, math.ceil(page_count * 0.6))),
            "image_page_majority": bool(page_count and image_pages >= max(1, math.ceil(page_count * 0.5))),
            "garbled_text_pages": garbled_pages,
            "garbled_text_majority": bool(page_count and garbled_pages >= max(1, math.ceil(page_count * 0.3))),
            "avg_text_chars_per_page": (sum(text_lengths) / page_count) if page_count else 0.0,
            **layout_probe,
        }
    except Exception:
        return {}


def _probe_docx_document(content: bytes) -> Dict[str, Any]:
    details: Dict[str, Any] = {}
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            names = zf.namelist()
            details["has_header"] = any(name.startswith("word/header") for name in names)
            details["has_footer"] = any(name.startswith("word/footer") for name in names)
            details["has_footnotes"] = "word/footnotes.xml" in names
            details["has_endnotes"] = "word/endnotes.xml" in names
            details["has_comments"] = "word/comments.xml" in names
            details["image_count"] = sum(1 for name in names if name.startswith("word/media/"))
            if "word/document.xml" in names:
                root = ElementTree.fromstring(zf.read("word/document.xml"))
                details["revision_insertions"] = len(root.findall(f".//{{{DOCX_NS}}}ins"))
                details["revision_deletions"] = len(root.findall(f".//{{{DOCX_NS}}}del"))
    except Exception:
        pass
    return details


def _probe_xlsx_document(content: bytes) -> Dict[str, Any]:
    try:
        import openpyxl
    except ImportError:
        return {}
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=False)
        sheet_count = len(workbook.worksheets)
        table_count = 0
        max_rows = 0
        max_cols = 0
        for sheet in workbook.worksheets:
            try:
                table_count += len(getattr(sheet, "tables", {}) or {})
            except Exception:
                pass
            try:
                max_rows = max(max_rows, int(getattr(sheet, "max_row", 0) or 0))
                max_cols = max(max_cols, int(getattr(sheet, "max_column", 0) or 0))
            except Exception:
                pass
        workbook.close()
        return {"sheet_count": sheet_count, "table_count": table_count, "max_rows": max_rows, "max_cols": max_cols}
    except Exception:
        return {}


def _probe_image_dimensions(content: bytes) -> Dict[str, Any]:
    if not _module_available("PIL"):
        return {}
    try:
        from PIL import Image  # type: ignore

        with Image.open(io.BytesIO(content)) as image:
            width, height = image.size
        return {
            "image_width": int(width),
            "image_height": int(height),
            "image_pixels": int(width) * int(height),
        }
    except Exception:
        return {}


def _choose_pdf_fast_backend(capabilities: Dict[str, bool]) -> str:
    if capabilities.get("pymupdf4llm"):
        return "pymupdf4llm"
    if capabilities.get("pymupdf"):
        return "pymupdf"
    if capabilities.get("docling"):
        return "docling"
    return "pypdf"


def _choose_ocr_backend(capabilities: Dict[str, bool]) -> Optional[str]:
    if (config.OCR_SERVICE_URL or "").strip():
        return "external_http_ocr"
    return None


def _ocr_backend_candidates(capabilities: Dict[str, bool]) -> List[str]:
    candidates: List[str] = []
    if (config.OCR_SERVICE_URL or "").strip():
        candidates.append("external_http_ocr")
    candidates.extend([name for name in ("docling", "unstructured", "paddleocr") if capabilities.get(name)])
    return candidates


def _should_route_pdf_to_ocr(probe: Dict[str, Any]) -> bool:
    if probe.get("is_scanned_pdf"):
        return True
    if probe.get("garbled_text_majority"):
        return True
    if not probe.get("image_page_majority"):
        return False
    try:
        avg_text_chars = float(probe.get("avg_text_chars_per_page") or 0.0)
    except Exception:
        avg_text_chars = 0.0
    return avg_text_chars <= float(config.PDF_OCR_MAX_TEXT_CHARS_PER_PAGE)


def _route_document_parser(probe: Dict[str, Any]) -> Dict[str, Any]:
    capabilities = _detect_parser_capabilities()
    detected_ext = (probe.get("detected_ext") or probe.get("extension") or "").lower()
    mime_type = (probe.get("mime_type") or "").lower()
    image_like = detected_ext in {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".gif"} or mime_type.startswith("image/")
    if detected_ext == ".pdf":
        if _should_route_pdf_to_ocr(probe):
            ocr_backend = _choose_ocr_backend(capabilities)
            return {
                "route": "pdf_ocr_layout",
                "parser_backend": ocr_backend or "fallback_no_ocr_backend",
                "backend_candidates": _ocr_backend_candidates(capabilities),
                "degraded": ocr_backend is None,
            }
        return {
            "route": "pdf_digital_fast",
            "parser_backend": _choose_pdf_fast_backend(capabilities),
            "backend_candidates": [name for name in ("pymupdf4llm", "pymupdf", "docling", "pypdf") if capabilities.get(name) or name == "pypdf"],
            "degraded": False,
        }
    if image_like:
        ocr_backend = _choose_ocr_backend(capabilities)
        return {
            "route": "image_ocr_layout",
            "parser_backend": ocr_backend or "fallback_no_ocr_backend",
            "backend_candidates": _ocr_backend_candidates(capabilities),
            "degraded": ocr_backend is None,
        }
    if detected_ext == ".docx":
        return {
            "route": "docx_structured",
            "parser_backend": "python-docx",
            "backend_candidates": [name for name in ("docling", "python-docx") if name == "python-docx" or capabilities.get("docling")],
            "degraded": False,
        }
    if detected_ext == ".xlsx":
        return {
            "route": "xlsx_structured",
            "parser_backend": "openpyxl",
            "backend_candidates": [name for name in ("docling", "openpyxl") if name == "openpyxl" or capabilities.get("docling")],
            "degraded": False,
        }
    if detected_ext == ".csv":
        return {"route": "csv_structured", "parser_backend": "csv", "backend_candidates": ["csv"], "degraded": False}
    if detected_ext == ".json":
        return {"route": "json_structured", "parser_backend": "json", "backend_candidates": ["json"], "degraded": False}
    return {"route": "plain_text", "parser_backend": "text", "backend_candidates": ["text"], "degraded": False}


def _probe_file_for_parser(filename: str, content: bytes) -> Dict[str, Any]:
    safe_name = _safe_filename(filename)
    extension = os.path.splitext(safe_name)[1].lower()
    signature = _sniff_file_signature(content)
    detected_ext = (signature.get("suggested_ext") or extension or "").lower()
    mime_type = _sniff_mime_type(safe_name, content, detected_ext)
    probe: Dict[str, Any] = {
        "filename": safe_name,
        "extension": extension,
        "detected_ext": detected_ext or extension,
        "mime_type": mime_type,
        "signature": signature.get("label") or "unknown",
        "file_size": len(content or b""),
        "page_count": None,
        "sheet_count": None,
        "is_scanned_pdf": False,
        "image_page_majority": False,
        "multi_column": False,
        "table_dense": False,
    }
    if probe["detected_ext"] == ".pdf":
        probe.update(_probe_pdf_document(content))
    elif probe["detected_ext"] == ".docx":
        probe.update(_probe_docx_document(content))
    elif probe["detected_ext"] == ".xlsx":
        probe.update(_probe_xlsx_document(content))
    elif probe["detected_ext"] in {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".gif"}:
        probe["is_scanned_pdf"] = True
        probe["image_page_majority"] = True
        probe["page_count"] = 1
        probe.update(_probe_image_dimensions(content))
    probe.update(_route_document_parser(probe))
    return probe


def _classify_question_type(query: str) -> str:
    q = _normalize_query(query).lower()
    for pattern in _policy_get("question_type_patterns", []):
        qtype = str((pattern or {}).get("type") or "").strip()
        for k in (pattern or {}).get("keywords") or []:
            if k in q:
                return qtype
    return "other"


def _answer_limits(qtype: str) -> Dict[str, Any]:
    points = min(max(config.ANSWER_MAX_POINTS, 3), 5)
    if qtype == "definition":
        return {"points": points, "max_tokens": config.LLM_MAX_TOKENS_DEF}
    if qtype == "summary":
        return {"points": points, "max_tokens": config.LLM_MAX_TOKENS_SUMMARY}
    if qtype == "arch":
        return {"points": points, "max_tokens": config.LLM_MAX_TOKENS_ARCH}
    if qtype == "howto":
        return {"points": min(max(points, 3), 8), "max_tokens": config.LLM_MAX_TOKENS_HOWTO}
    if qtype in {"compare", "compare_degraded"}:
        return {"points": points, "max_tokens": config.LLM_MAX_TOKENS_COMPARE}
    return {"points": points, "max_tokens": config.LLM_MAX_TOKENS_OTHER}


def _build_answer_prompt(
    query: str,
    evidence: str,
    qtype: str,
    answer_mode: str = "target_hit",
    compare_missing_targets: Optional[List[str]] = None,
    compare_source_status_hints: str = "",
) -> str:
    limits = _answer_limits(qtype)
    points = limits["points"]
    missing_items = [str(item).strip() for item in (compare_missing_targets or []) if str(item).strip()][:4]
    missing_text = "、".join([f"《{item}》" if not (item.startswith("《") and item.endswith("》")) else item for item in missing_items])

    if answer_mode == "compare_degraded":
        return f"""你是企业知识库问答助手。只允许依据“证据”回答，禁止凭空补全。

你正在处理“对比问法的单文档柔性降级”场景：用户想做法规对比，但知识库只检索到一侧文档。

规则：
1) 只依据证据回答已检索到的一侧文档；严禁对缺失文档进行推测或外推
2) 开头必须明确：已检索到一侧文档的规定，并明确缺失文档未检索到，暂无法提供对比分析
3) 每个关键结论必须带引用编号（如 [1] [2]），禁止输出推理过程
4) 若证据明确写明“另行制定/另行规定/由…制定公布”等外委情形，只能据此说明具体规则由相关机构另行制定或公布；不得补写未给出的具体标准/程序

输出格式（默认简洁）：
1) 第一段：围绕用户关注点，概括已检索到文档的可确认规定（不超过 40 字），末尾带引用
2) 列 {points} 个要点（每条 1 行），只写已检索到文档的证据结论，末尾带引用
3) 末尾追加 1 行固定提醒：未在知识库中检索到 {missing_text or "对比的另一部法规"}，暂无法为您提供对比分析。

证据：
{evidence}

缺失对比文档：{missing_text or "（未提供或未识别）"}

问题：{query}

回答："""

    if answer_mode == "compare_asymmetric":
        return f"""你是企业知识库问答助手。只允许依据“证据”回答，禁止凭空补全。

你正在处理“多文档不对称对比”场景：至少一部文档已找到可支撑对比的证据，但另一部文档未检索到直接对应的信息，或未明确涉及该焦点。

规则：
1) 必须正常完成对比，不得直接拒答
2) 已确认的文档事实必须带引用编号（如 [1] [2]）
3) 如果某部文档未找到直接对应信息，必须明确写出“未检索到相关信息”或“未直接涉及此内容”；这本身也是重要差异
4) 严禁把一部文档的规定外推到另一部文档
5) 不输出推理过程

输出格式（默认简洁）：
1) 先用一句话概括最核心的非对称差异（不超过 40 字）
2) 列 {points} 条对比要点；每条都要分别写明两侧文档的情况
3) 对有证据的一侧给出明确结论并带引用；对缺失或未涉及的一侧直接说明“未检索到相关信息/未直接涉及此内容”

当前各文档状态：
{compare_source_status_hints or '（未提供）'}

证据：
{evidence}

问题：{query}

回答："""

    if qtype == "fallback_brief":
        return f"""你是企业知识库问答助手。

当前没有可引用的知识库证据。你只能给出一个简短、保守的通用回答。

规则：
1) 回答第一行必须以【通用回答，未基于知识库证据】开头
2) 全文最多 3 句
3) 不得伪造知识库来源，不得添加引用编号
4) 如果问题明显高风险、需要精确依据或容易误导，直接回复：无法基于当前知识库证据安全回答该问题。

问题：{query}

回答："""

    if qtype == "howto":
        format_block = f"""输出格式（默认简洁）：
1) 先用一句话直接回答（不超过 30 字），末尾带引用，如：……[1]
2) 给出 3~8 个步骤（编号 1~N），每步 1 行，末尾带引用
3) 不写背景综述，不输出推理过程
4) 证据不足时，直接回复：未在知识库中找到足够相关的证据来回答该问题。"""
    elif qtype == "compare":
        format_block = f"""输出格式（默认简洁）：
1) 先用一句话给出核心区别（不超过 30 字），末尾带引用
2) 列 {points} 条对比要点；每条必须同时写出两侧文档各自的证据结论，且分别带引用
3) 严禁把一部文档的内容外推到另一部文档；没有证据的一侧不能补写
4) 不写背景综述，不输出推理过程
5) 证据不足时，直接回复：未在知识库中找到足够相关的证据来回答该问题。"""
    elif qtype == "summary":
        format_block = f"""输出格式（默认简洁）：
1) 先用一句话直接回答（不超过 30 字），末尾带引用
2) 列 {points} 个要点（每条 1 行），末尾带引用
3) 不主动展开无关内容，不写综述/背景，不输出推理过程
4) 证据不足时，直接回复：未在知识库中找到足够相关的证据来回答该问题。"""
    elif qtype == "arch":
        format_block = f"""输出格式（默认简洁）：
1) 先用一句话直接回答（不超过 30 字），末尾带引用
2) 分两组列要点（每条 1 行，末尾带引用）：
   存储：
   - （写 {max(2, points // 2)}~{points} 条）
   调用：
   - （写 {max(2, points // 2)}~{points} 条）
3) 不写背景综述，不输出推理过程
4) 证据不足时，直接回复：未在知识库中找到足够相关的证据来回答该问题。"""
    elif qtype == "regulation_execution":
        format_block = f"""输出格式：
1) 列出可执行条目（每条 1 行），每条包含：
   条目；责任主体；约束方式；处罚/后果（末尾带引用）
2) 严格按字段顺序与命名输出
3) 不输出背景与推理过程
4) 证据不足时，直接回复：未在知识库中找到足够相关的证据来回答该问题。"""
    elif qtype == "regulation_involved_docs":
        format_block = f"""输出格式：
1) 列出涉及的文档（每条 1 行），每条包含：
   文档名；命中理由；对应条文摘要（末尾带引用）
2) 同一条目内来源必须一致
3) 不输出背景与推理过程
4) 证据不足时，直接回复：未在知识库中找到足够相关的证据来回答该问题。"""
    else:
        if answer_mode == "target_hit":
            format_block = f"""输出格式（目标文档优先）：
1) 先用一句话总结目标文档中已确认的核心内容（不超过 40 字），末尾带引用
2) 列 {points} 个要点（每条 1 行），优先提炼目标文档正文；正文不足时，可结合章节标题、条款标题和文档标题归纳主题，末尾带引用
3) 如果证据只覆盖部分章节，先回答可确认部分，再明确说明未覆盖部分；不要因为覆盖不全就直接拒答
4) 只有当证据与问题明显无关时，才回复：未在知识库中找到足够相关的证据来回答该问题。"""
        else:
            format_block = f"""输出格式（默认简洁）：
1) 先用一句话直接回答（不超过 30 字），末尾带引用
2) 列 {points} 个要点（每条 1 行），末尾带引用
3) 证据存在部分对应的章节、条款或标题时，优先提炼可确认内容并明确未覆盖部分；章节标题、条款标题和文档标题可以作为回答线索
4) 只有当证据整体与问题明显无关时，才回复：未在知识库中找到足够相关的证据来回答该问题。"""

    return f"""你是企业知识库问答助手。只允许依据“证据”回答，禁止凭空补全。

规则：
1) 每个关键结论必须带引用编号（如 [1] [2]）
2) 禁止输出推理过程
3) 章节标题、条款标题、文档标题可以作为弱线索，用于概括主题，但不得超出证据可支持的范围
4) 若证据明确写明“另行制定/另行规定/由…制定公布”等外委情形，只能据此说明具体规则由相关机构另行制定或公布；不得补写未给出的具体标准/程序

{format_block}

证据：
{evidence}

问题：{query}

回答："""


def _merge_and_dedupe_hits(hits: List[Any], score_mode: str) -> List[Dict[str, Any]]:
    by_group: Dict[str, List[Any]] = {}
    for h in hits:
        src = _hit_entity_source(h) or "unknown"
        md = _hit_metadata(h)
        section_id = md.get("section_node_id") or md.get("section_id")
        section = (md.get("section") or "").strip()
        key = f"{src}||{section_id}||{section}"
        by_group.setdefault(key, []).append(h)

    merged: List[Dict[str, Any]] = []
    for key, hs in by_group.items():
        src, section_id, section = (key.split("||", 2) + ["", "", ""])[:3]
        with_id = []
        without_id = []
        for h in hs:
            cid = _hit_chunk_id(h)
            if cid is None:
                without_id.append(h)
            else:
                with_id.append((cid, h))

        with_id.sort(key=lambda x: x[0])
        cur = None
        for cid, h in with_id:
            text = (_hit_entity_text(h) or "").strip()
            if not text:
                continue
            score = _hit_score(h)
            if cur and cid == cur["end"] + 1:
                cur["text"] = (cur["text"] + "\n" + text).strip()
                if score_mode == "distance":
                    cur["rank"] = min(cur["rank"], score)
                else:
                    cur["rank"] = max(cur["rank"], score)
                cur["end"] = cid
            else:
                if cur:
                    merged.append(cur)
                cur = {
                    "source": src,
                    "section_id": section_id if section_id != "None" else None,
                    "section": section,
                    "start": cid,
                    "end": cid,
                    "text": text,
                    "rank": score
                }
        if cur:
            merged.append(cur)

        for h in without_id:
            text = (_hit_entity_text(h) or "").strip()
            if not text:
                continue
            merged.append({
                "source": src,
                "section_id": section_id if section_id != "None" else None,
                "section": section,
                "start": None,
                "end": None,
                "text": text,
                "rank": _hit_score(h)
            })

    best_by_text: Dict[str, Dict[str, Any]] = {}
    for m in merged:
        key = "".join((m["text"] or "").split()).lower()
        if not key:
            continue
        prev = best_by_text.get(key)
        if not prev:
            best_by_text[key] = m
            continue
        if score_mode == "distance":
            if m["rank"] < prev["rank"]:
                best_by_text[key] = m
        else:
            if m["rank"] > prev["rank"]:
                best_by_text[key] = m

    uniq = list(best_by_text.values())
    if score_mode == "distance":
        uniq.sort(key=lambda x: x["rank"])
    else:
        uniq.sort(key=lambda x: x["rank"], reverse=True)

    out: List[Dict[str, Any]] = []
    for m in uniq:
        md = {}
        if m["start"] is not None and m["end"] is not None:
            md["chunk_id_start"] = m["start"]
            md["chunk_id_end"] = m["end"]
        if m.get("section"):
            md["section"] = m.get("section")
        if m.get("section_id") is not None:
            md["section_id"] = m.get("section_id")
            if isinstance(m.get("section_id"), str) and str(m.get("section_id")).startswith("section::"):
                md["section_node_id"] = m.get("section_id")
        ent = {"source": m["source"], "text": m["text"], "metadata": md}
        item: Dict[str, Any] = {"entity": ent}
        if score_mode == "distance":
            item["distance"] = float(m["rank"])
        else:
            item["score"] = float(m["rank"])
        out.append(item)
    return out


def _aggregate_doc_sections(hits: List[Any], score_mode: str) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for hit in hits:
        src = _hit_entity_source(hit) or "unknown"
        text = (_hit_entity_text(hit) or "").strip()
        if not text:
            continue
        md = dict(_hit_metadata(hit) or {})
        section = (md.get("section") or "").strip()
        if section:
            md["section"] = section
            md.setdefault("section_title", section)
        ent = {"source": src, "text": text, "metadata": md}
        item: Dict[str, Any] = {"entity": ent}
        if score_mode == "distance":
            item["distance"] = float(_hit_score(hit))
        else:
            item["score"] = float(_hit_score(hit))
        out.append(item)

    if score_mode == "distance":
        out.sort(key=lambda x: _hit_score(x))
    else:
        out.sort(key=lambda x: _hit_score(x), reverse=True)
    return out


def _docs_for_query_context(qtype: str, merged_docs: List[Any], aggregated_docs: List[Any]) -> List[Any]:
    if len(merged_docs) <= 8:
        return merged_docs
    if qtype in {"single_doc_extract", "regulation_execution"}:
        return merged_docs
    return aggregated_docs


def _chunk_base_relevance(hit: Any, score_mode: str) -> float:
    score = float(_hit_score(hit))
    if score_mode == "distance":
        return 1.0 / (1.0 + max(score, 0.0))
    return score


def _clip01(value: float) -> float:
    return max(0.0, min(1.0, float(value)))


def _to_int(value: Any) -> Optional[int]:
    try:
        if value is None:
            return None
        return int(value)
    except Exception:
        return None


def _chunk_position_id(hit: Any) -> Optional[int]:
    md = _hit_metadata(hit)
    for key in ("chunk_id_start", "chunk_id", "reading_order", "order", "idx"):
        v = _to_int(md.get(key))
        if v is not None:
            return v
    v = _to_int(md.get("chunk_id_end"))
    if v is not None:
        return v
    return None


def _is_generic_section_title(section: str) -> bool:
    s = (section or "").strip().lower()
    if not s:
        return False
    generic_keys = [
        "总则", "附则", "适用范围", "职责", "机构职责", "部门职责", "工作职责", "目的", "依据", "原则", "术语", "定义",
        "general", "appendix", "scope", "responsibility", "overview", "introduction",
    ]
    return any(k in s for k in generic_keys)


def _extract_section_query_targets(query: str) -> List[str]:
    q = (query or "").strip()
    parsed = _llm_query_parse_cache_get(q) or {}
    targets = parsed.get("section_targets")
    if not isinstance(targets, list):
        return []
    out: List[str] = []
    for item in targets:
        v = _normalize_query(str(item or ""))
        if not v:
            continue
        if v not in out:
            out.append(v)
        if len(out) >= 8:
            break
    return out


def _is_generic_document_required_query(query: str) -> bool:
    q = _normalize_query(query)
    if not q or _extract_filename_candidates(q) or _extract_explicit_regulation_mentions(q):
        return False
    generic_only_patterns = [r"^有哪些规定[？?]?$", r"^怎么处罚[？?]?$", r"^如何处罚[？?]?$", r"^怎么处理[？?]?$", r"^如何处理[？?]?$"]
    if any(re.match(pattern, q) for pattern in generic_only_patterns):
        return True
    stripped = _strip_section_question_tail(q) or q
    stripped = re.sub(r"^(有哪[些]|有哪些|有什么|什么是|什么|如何|怎么|请问)", "", stripped).strip()
    stripped = re.sub(r"(有哪些|有什么|是什么|怎么规定|如何规定|怎么处理|如何处理|怎么|如何)$", "", stripped).strip()
    stripped = re.sub(r"[？?！!。；;，,：:\s]+$", "", stripped).strip()
    generic_targets = {
        "处罚条款", "法律责任", "管理职责", "监督检查", "违法行为", "申请程序",
        "禁止性规定", "扶持措施", "登记要求", "安全责任", "建设要求", "执法措施", "审议流程",
        "处罚规定", "奖励与处罚", "奖惩", "程序", "流程", "职责", "监督管理",
    }
    raw_direct_hits = [target for target in generic_targets if target in q]
    direct_hits = [target for target in generic_targets if target in stripped]
    section_targets = [target for target in _extract_section_query_targets(q) if target in generic_targets]
    if not raw_direct_hits and not direct_hits and not section_targets:
        return False
    remainder = stripped or q
    for target in sorted(generic_targets, key=len, reverse=True):
        remainder = remainder.replace(target, " ")
    nongeneric_topic_terms = {
        term for term in _query_anchor_terms(remainder)
        if term not in {"有哪些", "有什么", "什么", "如何", "怎么", "请问", "一般", "规定", "要求", "处罚", "责任", "程序", "条件", "标准", "管理", "建设", "扶持", "登记", "流程", "职责"}
    }
    return not nongeneric_topic_terms


def _is_unlocked_content_query(query: str, route: str) -> bool:
    if route in {
        "existence",
        "visibility_probe",
        "explicit_doc_reference",
        "explicit_regulation_reference",
        "exact_title_reference",
        "alias_title_reference",
        "weak_title_reference",
        "version_switch",
        "compare_clarification",
        "multi_doc_compare",
        "single_doc_compare",
    }:
        return False
    q = _normalize_query(query)
    if not q:
        return False
    if _is_generic_document_required_query(q):
        return True
    if _query_content_anchor_terms(q, qfilters=None, source_title_terms=[]):
        return True
    if _extract_section_query_targets(q):
        return True
    generic_content_patterns = [r"有哪些规定$", r"怎么处罚$", r"如何处罚$", r"怎么处理$", r"如何处理$", r"有哪些要求$"]
    return any(re.search(pattern, q) for pattern in generic_content_patterns)


def _section_target_alignment(section: str, query: str) -> tuple:
    sec = (section or "").strip().replace(" ", "")
    if not sec:
        return 0.0, 0.0
    targets = _extract_section_query_targets(query)
    if not targets:
        return 0.0, 0.0
    hits = 0.0
    exact = 0.0
    for t in targets:
        t_norm = (t or "").strip().replace(" ", "")
        if not t_norm:
            continue
        if t_norm in sec or sec in t_norm:
            hits += 1.0
            if t_norm == sec or (len(t_norm) >= 3 and t_norm in sec):
                exact = 1.0
    return _clip01(hits / max(1.0, float(len(targets)))), exact


def _is_section_lookup_query(query: str, qtype: str) -> bool:
    q = _normalize_query(query)
    if not q:
        return False
    if re.search(r"第[一二三四五六七八九十百千0-9]+[章节条款]", q):
        return True
    section_intent = _policy_keywords("rerank.section_lookup.intent_keywords")
    if any(k in q for k in section_intent):
        return True
    if any(k in q for k in _policy_keywords("rerank.section_lookup.restriction_keywords")) and any(k in q for k in _policy_keywords("rerank.section_lookup.restriction_subject_keywords")):
        return True
    if qtype in set(_policy_keywords("rerank.section_lookup.trigger_qtypes")):
        return True
    return False


def _infer_rerank_profile(query: str, qtype: str) -> str:
    q = _normalize_query(query)
    ql = q.lower()
    clause_keys = _policy_keywords("rerank.clause_lookup.keywords")
    broad_keys = _policy_keywords("rerank.broad.keywords")
    if _is_section_lookup_query(query, qtype):
        return "section_lookup"
    if qtype in set(_policy_keywords("rerank.clause_lookup.trigger_qtypes")):
        return "clause_lookup"
    if any(k in q for k in clause_keys):
        return "clause_lookup"
    if qtype in set(_policy_keywords("rerank.broad.trigger_qtypes")) or any(k in ql for k in broad_keys):
        return "broad"
    return "balanced"


def _rerank_profile_weights(profile: str) -> Dict[str, float]:
    base = {
        "section_term": float(getattr(config, "HYBRID_STRUCT_W_SECTION_TERM", 0.24)),
        "text_term": float(getattr(config, "HYBRID_STRUCT_W_TEXT_TERM", 0.18)),
        "section_overlap": float(getattr(config, "HYBRID_STRUCT_W_SECTION_OVERLAP", 0.22)),
        "keyword": float(getattr(config, "HYBRID_STRUCT_W_KEYWORD", 0.18)),
        "title": float(getattr(config, "HYBRID_STRUCT_W_TITLE", 0.10)),
        "base": float(getattr(config, "HYBRID_STRUCT_W_BASE", 0.08)),
    }
    if profile == "clause_lookup":
        return {
            "section_term": base["section_term"] + 0.10,
            "text_term": base["text_term"] + 0.08,
            "section_overlap": base["section_overlap"] + 0.06,
            "keyword": base["keyword"] + 0.08,
            "title": max(0.02, base["title"] - 0.04),
            "base": max(0.03, base["base"] - 0.06),
        }
    if profile == "section_lookup":
        return {
            "section_term": base["section_term"] + 0.22,
            "text_term": base["text_term"] + 0.02,
            "section_overlap": base["section_overlap"] + 0.18,
            "keyword": base["keyword"] + 0.02,
            "title": max(0.03, base["title"] - 0.03),
            "base": max(0.02, base["base"] - 0.06),
        }
    if profile == "broad":
        return {
            "section_term": max(0.10, base["section_term"] - 0.06),
            "text_term": base["text_term"] + 0.03,
            "section_overlap": base["section_overlap"] + 0.02,
            "keyword": base["keyword"] + 0.03,
            "title": base["title"] + 0.06,
            "base": base["base"] + 0.04,
        }
    return base


def _section_follow_bonus(section: str, pos: Optional[int], section_anchor_positions: Dict[str, List[int]], profile: str) -> float:
    sec = (section or "").strip()
    if not sec:
        return 0.0
    anchors = section_anchor_positions.get(sec) or []
    if not anchors:
        return 0.0
    base_bonus = float(getattr(config, "HYBRID_STRUCT_FOLLOW_BONUS", 0.16))
    if profile == "clause_lookup":
        base_bonus *= 1.15
    if profile == "section_lookup":
        base_bonus *= 1.35
    window = max(1, int(getattr(config, "HYBRID_STRUCT_FOLLOW_WINDOW", 3)))
    if profile == "section_lookup":
        window += 1
    if pos is None:
        return base_bonus * 0.45
    nearest = min(abs(pos - a) for a in anchors)
    if nearest > window:
        return 0.0
    return base_bonus * (1.0 - float(nearest) / float(window + 1))


def _generic_chunk_penalty(section: str, text: str, query: str, text_term_hits: float, section_term_hits: float, section_score: float, profile: str) -> float:
    if profile == "broad":
        return 0.0
    if not _is_generic_section_title(section):
        return 0.0
    q = _normalize_query(query)
    strict_need = any(k in q for k in ["处罚", "罚款", "流程", "程序", "标准", "时限", "材料", "限制", "禁止"])
    hay = f"{section} {(text or '')}"
    strict_hit = any(k in hay for k in ["处罚", "罚款", "罚则", "流程", "程序", "步骤", "标准", "时限", "材料", "限制", "禁止", "不得", "应当"])
    if profile == "section_lookup":
        align, _ = _section_target_alignment(section, query)
        if align > 0:
            return 0.0
        penalty = float(getattr(config, "HYBRID_STRUCT_GENERIC_PENALTY", 0.18)) * 1.20
        return penalty
    if not strict_need and (text_term_hits > 0 or section_term_hits > 0 or section_score >= 1.0):
        return 0.0
    if strict_need and strict_hit:
        return 0.0
    penalty = float(getattr(config, "HYBRID_STRUCT_GENERIC_PENALTY", 0.18))
    if strict_need:
        penalty *= 1.15
    return penalty


def _chunk_query_signal(query: str, hit: Any, score_mode: str) -> tuple:
    md = _hit_metadata(hit)
    src = _normalize_filename_for_match(_hit_entity_source(hit) or "")
    text = _hit_display_text(hit)
    section = (md.get("section_title") or md.get("section") or "").strip()
    title_signal = float(_doc_title_alias_score(src, query))
    section_score = _token_overlap_score(query, section)
    keyword_score = _token_overlap_score(query, text)
    anchor_terms = _query_anchor_terms(query)
    section_term_hits = sum(1 for term in anchor_terms if term and term in section)
    text_term_hits = sum(1 for term in anchor_terms if term and term in text)
    title_hit = 1 if md.get("title_hit") or section == "document_title" or title_signal > 0 else 0
    return (
        float(section_term_hits),
        float(text_term_hits),
        float(section_score),
        float(keyword_score),
        float(title_hit),
        float(title_signal),
        float(_chunk_base_relevance(hit, score_mode)),
    )


def _hybrid_structural_chunk_score(
    query: str,
    hit: Any,
    score_mode: str,
    profile: str = "balanced",
    section_anchor_positions: Optional[Dict[str, List[int]]] = None,
) -> tuple:
    section_term_hits, text_term_hits, section_score, keyword_score, title_hit, title_signal, base_rel = _chunk_query_signal(query, hit, score_mode)
    md = _hit_metadata(hit)
    section = (md.get("section_title") or md.get("section") or "").strip()
    text = _hit_display_text(hit)
    pos = _chunk_position_id(hit)
    anchor_terms = _query_anchor_terms(query)
    anchor_cnt = max(1.0, float(len(anchor_terms)))

    section_term_norm = _clip01(section_term_hits / anchor_cnt)
    text_term_norm = _clip01(text_term_hits / anchor_cnt)
    section_overlap_norm = _clip01(section_score / 12.0)
    keyword_overlap_norm = _clip01(keyword_score / 18.0)
    title_norm = _clip01((title_signal + 1.5 * title_hit) / 8.0)
    base_rel_norm = _clip01(base_rel)

    weights = _rerank_profile_weights(profile)
    w_section_term = weights["section_term"]
    w_text_term = weights["text_term"]
    w_section_overlap = weights["section_overlap"]
    w_keyword = weights["keyword"]
    w_title = weights["title"]
    w_base = weights["base"]

    hybrid_score = (
        w_section_term * section_term_norm
        + w_text_term * text_term_norm
        + w_section_overlap * section_overlap_norm
        + w_keyword * keyword_overlap_norm
        + w_title * title_norm
        + w_base * base_rel_norm
    )

    follow_bonus = _section_follow_bonus(section, pos, section_anchor_positions or {}, profile)
    generic_penalty = _generic_chunk_penalty(section, text, query, text_term_hits, section_term_hits, section_score, profile)
    section_align, section_exact = _section_target_alignment(section, query)
    section_match_bonus = 0.0
    section_mismatch_penalty = 0.0
    if profile == "section_lookup":
        base_match_bonus = float(getattr(config, "HYBRID_STRUCT_SECTION_MATCH_BONUS", 0.22))
        section_match_bonus = base_match_bonus * (0.4 + 0.6 * section_align + 0.2 * section_exact) if section_align > 0 else 0.0
        if _extract_section_query_targets(query) and section_align <= 0:
            section_mismatch_penalty = float(getattr(config, "HYBRID_STRUCT_SECTION_MISMATCH_PENALTY", 0.12))
    hybrid_score = hybrid_score + follow_bonus + section_match_bonus - generic_penalty - section_mismatch_penalty

    # Keep deterministic tie-breakers for stable reranking inside one source.
    tie_breaker = (
        float(section_term_hits),
        float(text_term_hits),
        float(section_score),
        float(keyword_score),
        float(title_hit),
        float(title_signal),
        float(base_rel),
        float(follow_bonus),
        float(-generic_penalty),
        float(section_match_bonus),
        float(-section_mismatch_penalty),
    )
    return float(hybrid_score), tie_breaker


def _intra_doc_chunk_rerank(query: str, hits: List[Any], score_mode: str, qtype: str = "other") -> List[Any]:
    if len(hits) <= 1:
        return hits
    profile = _infer_rerank_profile(query, qtype)
    grouped: Dict[str, List[Any]] = {}
    source_order: List[str] = []
    for hit in hits:
        src = _normalize_filename_for_match(_hit_entity_source(hit) or "")
        if src not in grouped:
            grouped[src] = []
            source_order.append(src)
        grouped[src].append(hit)

    reranked: List[Any] = []
    for src in source_order:
        group = grouped[src]
        section_anchor_positions: Dict[str, List[int]] = {}
        for hit in group:
            md = _hit_metadata(hit)
            section = (md.get("section_title") or md.get("section") or "").strip()
            if not section:
                continue
            section_term_hits, text_term_hits, section_score, _, _, _, _ = _chunk_query_signal(query, hit, score_mode)
            if section_term_hits > 0 or (section_score >= 1.0 and text_term_hits > 0):
                pos = _chunk_position_id(hit)
                if pos is not None:
                    section_anchor_positions.setdefault(section, []).append(pos)
        decorated = []
        for idx, hit in enumerate(group):
            hybrid_score, tie_breaker = _hybrid_structural_chunk_score(
                query,
                hit,
                score_mode,
                profile=profile,
                section_anchor_positions=section_anchor_positions,
            )
            decorated.append((hybrid_score, tie_breaker, -idx, hit))
        decorated.sort(key=lambda item: (item[0], item[1], item[2]), reverse=True)
        reranked.extend([hit for _, _, _, hit in decorated])
    return reranked


def _should_keep_structural_chunk(query: str, hit: Any, score_mode: str) -> bool:
    q = _normalize_query(query)
    if not q:
        return False
    section = _doc_section_name(hit)
    text = _hit_display_text(hit)
    section_align, _ = _section_target_alignment(section, q)
    section_term_hits, text_term_hits, _, keyword_score, title_hit, title_signal, _ = _chunk_query_signal(q, hit, score_mode)
    semantic_terms = _query_semantic_aspects(q).get("terms") or []
    hay = f"{section}\n{text}"
    if section_align > 0 or section_term_hits > 0:
        return True
    if text_term_hits > 0 or keyword_score > 0:
        return True
    if any(term and term in hay for term in semantic_terms):
        return True
    if title_hit > 0 and (title_signal > 0 or keyword_score > 0):
        return True
    return False


def _filter_low_relevance_sources(
    hits: List[Any],
    score_mode: str,
    query: str = "",
    minimum_keep: Optional[int] = None,
) -> List[Any]:
    if not hits:
        return []

    ratio = _clip01(getattr(config, "RECALL_RELATIVE_SCORE_RATIO", 0.72))
    fixed_min = max(0.0, float(getattr(config, "MIN_RELEVANCE_SCORE", 0.25)))
    min_keep = max(1, int(minimum_keep if minimum_keep is not None else getattr(config, "RECALL_MIN_KEEP_N", 3)))
    top1_rel = _chunk_base_relevance(hits[0], score_mode)
    cutoff_rel = max(fixed_min, top1_rel * ratio) if top1_rel > 0 else fixed_min

    kept: List[Any] = []
    seen = set()

    def _hit_key(hit: Any) -> tuple:
        md = _hit_metadata(hit)
        return (
            _normalize_filename_for_match(_hit_entity_source(hit) or ""),
            md.get("chunk_id_start"),
            md.get("chunk_id_end"),
            md.get("chunk_id"),
            (_hit_entity_text(hit) or "")[:80],
        )

    def _append_unique(hit: Any):
        key = _hit_key(hit)
        if key in seen:
            return
        seen.add(key)
        kept.append(hit)

    for hit in hits:
        if _chunk_base_relevance(hit, score_mode) >= cutoff_rel:
            _append_unique(hit)

    for hit in hits:
        if _should_keep_structural_chunk(query, hit, score_mode):
            _append_unique(hit)

    if len(kept) < min_keep:
        for hit in hits:
            _append_unique(hit)
            if len(kept) >= min_keep:
                break

    return kept if kept else hits[: min(len(hits), min_keep)]

def _classify_doc_type(filename: str, text: str) -> str:
    name = (filename or "").lower()
    t = (text or "").lower()
    reg_keys = ["条例", "办法", "规定", "规范", "准则", "细则", "应当", "不得", "处罚", "罚款", "罚则"]
    rep_keys = ["研究", "报告", "调研", "白皮书", "年度", "分析", "研究报告", "调研报告"]
    if any(k in name for k in ["条例", "办法", "规定"]) or any(k in t for k in reg_keys):
        return "regulation"
    if any(k in name for k in ["报告", "调研", "研究"]) or any(k in t for k in rep_keys):
        return "research_report"
    return "other"

def _infer_topics(text: str) -> List[str]:
    t = (text or "").lower()
    topics = []
    env_keys = ["环保", "环境", "污染", "治理", "制度", "监管", "处罚", "排放", "生态", "环境保护"]
    ai_keys = ["人工智能", "ai", "成熟度", "算法", "模型", "智能化", "数据", "企业", "调研", "研究"]
    if any(k in t for k in env_keys):
        topics.append("环保治理制度设计")
    if any(k in t for k in ai_keys):
        topics.append("AI成熟度研究")
    return topics

def _query_filters(query: str) -> Dict[str, Any]:
    q = (query or "").lower()
    doc_type = None
    topic = None
    for rule in _policy_get("query_filters.doc_type_rules", []):
        if _policy_match_rule(q, rule):
            doc_type = (rule or {}).get("value")
            break
    for rule in _policy_get("query_filters.topic_rules", []):
        if _policy_match_rule(q, rule):
            topic = (rule or {}).get("value")
            break
    return {"doc_type": doc_type, "topic": topic}


def _query_char_len(query: str) -> int:
    return len("".join((query or "").split()))


def _is_weak_reference_query(query: str) -> bool:
    q = _normalize_query(query)
    if not q or _extract_filename_candidates(q):
        return False
    generic_doc = any(k in q for k in _policy_keywords("weak_reference.generic_doc_markers"))
    generic_need = any(k in q for k in _policy_keywords("weak_reference.generic_need_markers"))
    return generic_doc or generic_need


def _build_controlled_expansion_queries(query: str, allowed_docs: List[str]) -> List[Dict[str, str]]:
    if not _is_weak_reference_query(query):
        return []
    expansions: List[Dict[str, str]] = []
    seen = {query}
    limit = max(0, int(getattr(config, "WEAK_QUERY_EXPANSION_LIMIT", 3)))
    for source in allowed_docs[:limit]:
        info = _doc_get(source)
        title = (info.get("canonical_title") or info.get("filename_stem") or _filename_stem(source)).strip()
        if not title:
            continue
        expanded = f"{title} {query}".strip()
        if expanded in seen:
            continue
        seen.add(expanded)
        expansions.append({
            "query": expanded,
            "source": _normalize_filename_for_match(source),
            "reason": "title_anchor",
        })
    return expansions


def _source_constraint_multiplier(src: str, query: str, fname_set: set, allowed_set: set, weak_query: bool) -> float:
    multiplier = 1.0
    if src in fname_set:
        return max(multiplier, 1.2)
    title_hit = _doc_title_alias_hit(src, query)
    if title_hit:
        multiplier = max(multiplier, float(getattr(config, "TITLE_CONSTRAINT_BOOST", 1.08)))
    if allowed_set:
        if src in allowed_set:
            multiplier = max(multiplier, float(getattr(config, "TITLE_CONSTRAINT_BOOST", 1.08)))
        elif weak_query or title_hit:
            multiplier *= float(getattr(config, "TITLE_CONSTRAINT_PENALTY", 0.82))
    return multiplier


def _collect_lexical_candidates(query: str, safe_names: List[str], doc_recall_plan: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    source_filter = safe_names[0] if len(safe_names) == 1 else None
    allowed_docs = [entry.get("source") for entry in (doc_recall_plan or []) if (entry or {}).get("source")]
    doc_recall_map = {entry["source"]: {**entry, "rank": idx} for idx, entry in enumerate(doc_recall_plan or []) if entry.get("source")}
    try:
        items.extend(_lexical_recall_indexed(query, getattr(config, "LEXICAL_RECALL_LIMIT", 1000), source_filter=source_filter))
    except Exception:
        pass
    if not items:
        try:
            items.extend(_lexical_recall_fallback(query, getattr(config, "LEXICAL_RECALL_LIMIT", 1000), source_filter=source_filter))
        except Exception:
            pass
    title_sources = [src for src in (allowed_docs or []) if src]
    if source_filter and source_filter not in title_sources:
        title_sources.insert(0, source_filter)
    for src in title_sources[: max(1, int(getattr(config, "WEAK_QUERY_DOC_LIMIT", 6)))]:
        plan_entry = doc_recall_map.get(src) or {}
        items.append(
            _synthetic_doc_title_hit(
                src,
                query,
                score=max(1.0, float(plan_entry.get("prior", 0.0)) + 1.0),
                metadata_updates={
                    "doc_recall_hit": True,
                    "doc_prior": float(plan_entry.get("prior", 0.0)),
                    "doc_recall_reasons": list(plan_entry.get("reasons") or []),
                    "doc_recall_rank": int(plan_entry.get("rank", 0)),
                },
            )
        )
    for expansion in _build_controlled_expansion_queries(query, allowed_docs):
        try:
            items.extend(
                _lexical_recall_indexed(
                    expansion["query"],
                    max(20, min(200, getattr(config, "LEXICAL_RECALL_LIMIT", 1000) // 4)),
                    source_filter=expansion.get("source"),
                )
            )
        except Exception:
            try:
                items.extend(
                    _lexical_recall_fallback(
                        expansion["query"],
                        max(20, min(200, getattr(config, "LEXICAL_RECALL_LIMIT", 1000) // 4)),
                        source_filter=expansion.get("source"),
                    )
                )
            except Exception:
                continue
    allowed_set = set([_normalize_filename_for_match(src) for src in (allowed_docs or []) if src])
    deduped: List[Dict[str, Any]] = []
    seen = set()
    for item in items:
        item = _annotate_lexical_hit(query, item, allowed_set, doc_recall_map=doc_recall_map)
        src = _normalize_filename_for_match(_hit_entity_source(item) or "")
        chunk_id = _hit_metadata(item).get("chunk_id")
        key = (src, chunk_id, (_hit_entity_text(item) or "")[:80])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)
    return deduped


def _distinct_hit_sources(hits: List[Any]) -> List[str]:
    out: List[str] = []
    for hit in hits:
        src = _normalize_filename_for_match(_hit_entity_source(hit) or "")
        if src and src not in out:
            out.append(src)
    return out


async def _chunk_level_rerank(rerank_service: Any, query: str, hits: List[Any], top_k: int, enable_rerank: bool) -> Dict[str, Any]:
    score_mode = _hit_score_mode(hits[0]) if hits else "score"
    if (not hits) or (not enable_rerank) or (not config.ENABLE_RERANK):
        kept = hits[:top_k] if top_k > 0 else hits[:]
        return {"hits": kept, "score_mode": score_mode, "used": False}
    rerank_n = min(len(hits), max(1, top_k))
    if rerank_n <= 1:
        kept = hits[:top_k] if top_k > 0 else hits[:]
        return {"hits": kept, "score_mode": score_mode, "used": False}
    try:
        reranked = await rerank_service.rerank(
            query=query,
            documents=[_hit_entity_text(hit) for hit in hits[:rerank_n]],
            top_k=rerank_n,
        )
    except Exception:
        kept = hits[:top_k] if top_k > 0 else hits[:]
        return {"hits": kept, "score_mode": score_mode, "used": False}
    reranked_hits: List[Dict[str, Any]] = []
    for item in reranked or []:
        idx = item.get("index") if isinstance(item, dict) else getattr(item, "index", None)
        score = item.get("score", 0.0) if isinstance(item, dict) else getattr(item, "score", 0.0)
        try:
            idx_i = int(idx)
        except Exception:
            continue
        if idx_i < 0 or idx_i >= rerank_n:
            continue
        base_hit = hits[idx_i]
        ent = base_hit.get("entity") if isinstance(base_hit, dict) else getattr(base_hit, "entity", None)
        reranked_hits.append({"entity": ent, "score": float(score)})
    if not reranked_hits:
        kept = hits[:top_k] if top_k > 0 else hits[:]
        return {"hits": kept, "score_mode": score_mode, "used": False}
    return {"hits": reranked_hits, "score_mode": "score", "used": True}


async def _source_level_rerank(
    rerank_service: Any,
    query: str,
    hits: List[Any],
    src_scores: Dict[str, float],
    keep_n: int,
    enable_rerank: bool,
    dense_rank_map: Optional[Dict[str, int]] = None,
    lex_rank_map: Optional[Dict[str, int]] = None,
    source_signals: Optional[Dict[str, Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    if (not src_scores) or (not hits) or (not enable_rerank) or (not config.ENABLE_RERANK):
        return {"scores": src_scores, "used": False}
    if len(src_scores) <= 1:
        return {"scores": src_scores, "used": False}
    if bool(getattr(config, "RERANK_LOW_CONF_ONLY", True)):
        dense_top = _top_ranked_source(dense_rank_map or {})
        lex_top = _top_ranked_source(lex_rank_map or {})
        current_top = max(src_scores.items(), key=lambda item: item[1])[0]
        gap = _source_score_gap(src_scores)
        anchored = bool((source_signals or {}).get(current_top, {}).get("title_hit"))
        if gap > float(getattr(config, "RERANK_SOURCE_SCORE_GAP", 0.04)):
            return {"scores": src_scores, "used": False}
        if anchored and dense_top == current_top and (lex_top in (None, current_top)):
            return {"scores": src_scores, "used": False}
    doc_sources = list(src_scores.keys())
    by_src: Dict[str, List[str]] = {}
    for hit in hits:
        src = _normalize_filename_for_match(_hit_entity_source(hit) or "")
        by_src.setdefault(src, []).append(_hit_entity_text(hit) or "")
    documents = []
    for src in doc_sources:
        snippets = [txt for txt in by_src.get(src, []) if txt]
        documents.append(("\n".join(snippets[:3])).strip() or src)
    try:
        reranked = await rerank_service.rerank(query=query, documents=documents, top_k=min(len(documents), max(1, keep_n)))
    except Exception:
        return {"scores": src_scores, "used": False}
    merged = dict(src_scores)
    weight = float(os.getenv("FUSION_W_RERANK_DOC", "0.3"))
    for item in reranked or []:
        idx = item.get("index") if isinstance(item, dict) else getattr(item, "index", None)
        score = item.get("score", 0.0) if isinstance(item, dict) else getattr(item, "score", 0.0)
        try:
            src = doc_sources[int(idx)]
        except Exception:
            continue
        merged[src] = merged.get(src, 0.0) + weight * float(score)
    return {"scores": merged, "used": True}


def _apply_retrieval_filters(docs: List[Any], qfilters: Dict[str, Any], fnames: List[str]) -> List[Any]:
    filtered_docs = docs[:]
    filtered_docs = [d for d in filtered_docs if not _is_heading_only_hit(d)]
    if qfilters.get("doc_type"):
        filtered_docs = [d for d in filtered_docs if (_hit_metadata(d).get("doc_type") or "") == qfilters["doc_type"]]
    if qfilters.get("topic"):
        filtered_docs = [d for d in filtered_docs if qfilters["topic"] in (_hit_metadata(d).get("topics") or [])]
    if fnames:
        sset = set([_normalize_filename_for_match(x) for x in fnames])
        filtered_docs = [d for d in filtered_docs if _normalize_filename_for_match(_hit_entity_source(d) or "") in sset]
    return filtered_docs


def _summarize_source_scores(
    docs: List[Any],
    dense_rank_map: Dict[str, int],
    lex_rank_map: Dict[str, int],
    source_count: Dict[str, int],
    source_signals: Dict[str, Dict[str, Any]],
    fname_set: set,
    allowed_set: set,
    weak_query: bool,
    query: str,
) -> Dict[str, float]:
    by_src: Dict[str, int] = {}
    for hit in docs:
        src = _normalize_filename_for_match(_hit_entity_source(hit) or "")
        by_src[src] = by_src.get(src, 0) + 1
    scores: Dict[str, float] = {}
    for src in by_src:
        scores[src] = _fusion_source_score(src, query, dense_rank_map, lex_rank_map, source_count, source_signals, fname_set, allowed_set, weak_query)
    return scores

def _normalize_source_name(name: str) -> str:
    s = (name or "").strip().lower()
    s = s.replace("《", "").replace("》", "")
    for suf in ("条例", "办法", "规定", "规范", "准则", "细则"):
        if s.endswith(suf):
            s = s[: -len(suf)]
            break
    s = "".join([c for c in s if c.isalnum() or c in ("_", "-", " ")])
    s = " ".join(s.split())
    return s

def _doc_level_rerank(hits: List[Any], score_mode: str) -> List[Dict[str, Any]]:
    by_src: Dict[str, List[Any]] = {}
    for h in hits:
        s = _hit_entity_source(h) or "unknown"
        by_src.setdefault(s, []).append(h)
    ranked = []
    for s, items in by_src.items():
        scs = [_hit_score(x) for x in items]
        score = min(scs) if score_mode == "distance" else max(scs)
        count = len(items)
        agg = (score if score_mode != "distance" else (1.0 / (1e-9 + score))) + 0.05 * count
        ranked.append({"source": s, "score": float(agg)})
    ranked.sort(key=lambda x: x["score"], reverse=True)
    return ranked

def _dedupe_similar_sources(ranked_sources: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    seen = set()
    out = []
    for item in ranked_sources:
        key = _normalize_source_name(item.get("source") or "")
        if key in seen:
            continue
        seen.add(key)
        out.append(item)
    return out

def _extract_regulation_clauses(text: str) -> List[str]:
    t = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    lines = [ln.strip() for ln in t.split("\n") if ln and ln.strip()]
    keys = ["不得", "应当", "禁止", "处罚", "罚款", "责令", "违法", "违反", "处理", "追责", "问责"]
    clauses = []
    for ln in lines:
        low = ln.lower()
        hit = False
        for k in keys:
            if (k in ln) or (k in low):
                hit = True
                break
        if hit:
            clauses.append(ln)
    return clauses[:200]

def _build_hits_from_clauses(source: str, clauses: List[str]) -> List[Dict[str, Any]]:
    out = []
    for c in clauses:
        ent = {"source": source, "text": c, "metadata": {"section": ""}}
        out.append({"entity": ent, "score": 1.0})
    return out

def _normalize_text_for_dedupe(s: str) -> str:
    return "".join((s or "").lower().split())

def _estimate_tokens(s: str) -> int:
    return max(1, len((s or "")) // 4)

def _prepare_structured_items(filename: str, text: str, chunk_size: int, overlap: int, document_ir: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
    if document_ir and (document_ir.get("elements") or []):
        return _document_ir_to_structured_items(document_ir, chunk_size, overlap)
    ext = os.path.splitext(filename or "")[1].lower()
    if ext == ".xlsx":
        return _chunk_units_to_items(_build_plain_chunk_units(text), chunk_size=chunk_size, overlap=overlap)
    return split_text_with_sections(filename, text, chunk_size, overlap)


def _build_sources(final_docs: List[Any], query: str, score_mode: str) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    for i, d in enumerate(final_docs, start=1):
        excerpt = _build_excerpt(_hit_display_text(d), query, 200)
        md = _hit_metadata(d)
        section = (md.get("section") or "").strip()
        items.append({
            "ref": i,
            "source": _hit_entity_source(d) or "unknown",
            "score": _hit_score(d),
            "section": section,
            "chunk_range": _hit_chunk_range(d),
            "text": excerpt
        })
    return items

def _dynamic_thresholds(qtype: str, has_filename_hint: bool) -> Dict[str, float]:
    score_unit = _rrf(0, int(getattr(config, "RRF_K", 60)))
    if has_filename_hint:
        return {"max_distance": 0.99, "min_score": 0.0}
    if qtype in ("single_doc_extract", "regulation_execution"):
        return {"max_distance": 0.95, "min_score": 0.1 * score_unit}
    if qtype in ("screening", "regulation_involved_docs", "summary"):
        return {"max_distance": 0.92, "min_score": 0.15 * score_unit}
    return {"max_distance": float(getattr(config, "MAX_RELEVANCE_DISTANCE", 0.8)),
            "min_score": float(getattr(config, "MIN_RELEVANCE_SCORE", 0.25)) * score_unit}

def _is_stats_intent(query: str) -> bool:
    q = (query or "").lower()
    keys = ["数量", "条数", "条目", "分块", "chunk", "记录数", "多少", "几条", "页数", "sheet", "工作表", "统计"]
    return any(k.lower() in q for k in keys)

def _filter_display_sources(
    docs: List[Any],
    score_mode: str,
    qfilters: Dict[str, Any],
    fnames: List[str],
    qtype: str,
    max_sources: int = 3,
    target_sources: Optional[List[str]] = None,
    observations: Optional[Dict[str, Any]] = None,
) -> List[Any]:
    if not docs:
        return []
    items = docs[:]
    effective_targets = list(target_sources or fnames or [])
    if effective_targets:
        sset = set([_normalize_filename_for_match(x) for x in effective_targets])
        items = [d for d in items if _normalize_filename_for_match(_hit_entity_source(d) or "") in sset]
    if qfilters.get("doc_type"):
        items = [d for d in items if (_hit_metadata(d).get("doc_type") or "") == qfilters["doc_type"]]
    if qfilters.get("topic"):
        items = [d for d in items if qfilters["topic"] in (_hit_metadata(d).get("topics") or [])]
    if not items:
        return []
    if not effective_targets and qtype not in {"screening", "regulation_involved_docs", "compare", "arch"}:
        focus_score = float((observations or {}).get("intra_doc_focus_score") or 0.0)
        window = items[: min(len(items), 5)]
        source_counts: Dict[str, int] = {}
        for doc in window:
            src = _normalize_filename_for_match(_hit_entity_source(doc) or "")
            if src:
                source_counts[src] = source_counts.get(src, 0) + 1
        if source_counts:
            dominant_source, dominant_count = max(source_counts.items(), key=lambda item: (item[1], item[0]))
            dominant_share = float(dominant_count) / float(len(window))
            if dominant_share >= 0.5 or focus_score >= 0.72:
                dominant_items = [d for d in items if _normalize_filename_for_match(_hit_entity_source(d) or "") == dominant_source]
                if dominant_items:
                    items = dominant_items
    if score_mode == "distance":
        best = min([_hit_score(d) for d in items] or [1.0])
        margin = float(getattr(config, "DISPLAY_DISTANCE_MARGIN", 0.02))
        items = [d for d in items if _hit_score(d) <= best + margin]
        items.sort(key=lambda d: _hit_score(d))
    else:
        best = max([_hit_score(d) for d in items] or [0.0])
        ratio = float(getattr(config, "DISPLAY_SCORE_RATIO", 0.8))
        cutoff = best * ratio
        items = [d for d in items if _hit_score(d) >= cutoff]
        items.sort(key=lambda d: _hit_score(d), reverse=True)
    return items[:max_sources]

def _tok_terms(text: str) -> List[str]:
    s = (text or "").lower()
    terms: List[str] = []
    buf = []
    for ch in s:
        if "\u4e00" <= ch <= "\u9fff":
            if buf:
                t = "".join(buf)
                terms.extend([t])
                buf = []
            terms.append(ch)
        elif ch.isalnum() or ch in ("_",):
            buf.append(ch)
        else:
            if buf:
                t = "".join(buf)
                if t:
                    terms.append(t)
                buf = []
    if buf:
        t = "".join(buf)
        if t:
            terms.append(t)
    return terms

def _bm25_scores(query: str, documents: List[str], k1: float = 1.5, b: float = 0.75) -> List[float]:
    if not documents:
        return []
    q_terms = _tok_terms(query)
    if not q_terms:
        return [0.0] * len(documents)
    docs_terms = [_tok_terms(t or "") for t in documents]
    N = len(documents)
    doc_lens = [len(t) for t in docs_terms]
    avgdl = (sum(doc_lens) / N) if N else 0.0
    df: Dict[str, int] = {}
    for terms in docs_terms:
        seen = set()
        for t in terms:
            if t in seen:
                continue
            seen.add(t)
            df[t] = df.get(t, 0) + 1
    idf: Dict[str, float] = {}
    for t in set(q_terms):
        dft = df.get(t, 0)
        idf[t] = math.log(1 + (N - dft + 0.5) / (dft + 0.5))
    scores: List[float] = []
    for terms, dl in zip(docs_terms, doc_lens):
        tf: Dict[str, int] = {}
        for t in terms:
            tf[t] = tf.get(t, 0) + 1
        s = 0.0
        for t in q_terms:
            if t not in idf:
                continue
            ft = tf.get(t, 0)
            if ft == 0:
                continue
            denom = ft + k1 * (1 - b + b * (dl / (avgdl or 1.0)))
            s += idf[t] * (ft * (k1 + 1)) / (denom or 1.0)
        scores.append(s)
    return scores

def _minmax_norm(arr: List[float]) -> List[float]:
    if not arr:
        return []
    mn = min(arr)
    mx = max(arr)
    if mx - mn <= 1e-9:
        return [0.0 for _ in arr]
    return [(x - mn) / (mx - mn) for x in arr]

def _passes_relevance_cluster(docs: List[Any], score_mode: str, thr: Dict[str, float], top_n: int = 3) -> bool:
    if not docs:
        return False
    take = docs[:max(1, min(len(docs), top_n))]
    if score_mode == "distance":
        good = [d for d in take if _hit_score(d) <= thr.get("max_distance", 0.8)]
    else:
        good = [d for d in take if _hit_score(d) >= thr.get("min_score", 0.25)]
    return len(good) > 0


def _fts_storage_text(text: str, metadata: Optional[Dict[str, Any]] = None) -> str:
    md = metadata or {}
    raw_text = md.get("raw_text") if isinstance(md, dict) else None
    if raw_text is not None:
        return str(raw_text)
    return text or ""


def _lex_db_add_chunk_sql(source: str, text: str, section: str, metadata: Dict[str, Any], chunk_id: int):
    conn = _lex_db_connect()
    cur = conn.execute(
        "INSERT INTO chunks_meta(source, chunk_id, section, metadata) VALUES (?,?,?,?)",
        (source, int(chunk_id), section or "", json.dumps(metadata or {}, ensure_ascii=False))
    )
    _crash_inject("after_meta_insert")
    rowid = cur.lastrowid
    conn.execute("INSERT INTO chunks_fts(rowid, text) VALUES (?,?)", (rowid, text or ""))
    _crash_inject("after_fts_insert")


def _chunk_row_ids_for_source_version(source: str, doc_version: Optional[int]) -> List[int]:
    conn = _lex_db_connect()
    rows = conn.execute("SELECT id, metadata FROM chunks_meta WHERE source = ?", (source,)).fetchall()
    out: List[int] = []
    for row_id, metadata_raw in rows:
        metadata = {}
        try:
            metadata = json.loads(metadata_raw or "{}")
        except Exception:
            metadata = {}
        version = metadata.get("doc_version")
        try:
            version = int(version) if version is not None else None
        except Exception:
            version = None
        if doc_version is None:
            out.append(int(row_id))
        elif version == int(doc_version):
            out.append(int(row_id))
    return out


def _delete_milvus_source_version(source: str, doc_version: Optional[int]):
    safe = _safe_filename(source)
    if doc_version is None:
        return
    try:
        vector_db = VectorDBService()
        vector_db.connect()
        rows = vector_db.client.query(
            collection_name=vector_db.collection_name,
            filter=f"source == {json.dumps(safe, ensure_ascii=False)}",
            output_fields=["id", "metadata"],
            limit=10_000,
        )
        ids_to_delete: List[int] = []
        for row in rows or []:
            metadata = row.get("metadata") or {}
            try:
                version = int(metadata.get("doc_version")) if metadata.get("doc_version") is not None else None
            except Exception:
                version = None
            if version == int(doc_version):
                pk = _milvus_row_pk(row)
                if pk is not None:
                    ids_to_delete.append(pk)
        if ids_to_delete:
            batch_size = 1000
            for start in range(0, len(ids_to_delete), batch_size):
                vector_db.client.delete(
                    collection_name=vector_db.collection_name,
                    ids=ids_to_delete[start:start + batch_size],
                )
    except Exception as e:
        logger.warning(f"delete_milvus_source_version_failed: source={safe} doc_version={doc_version} err={e}")


def _lex_db_delete_source_version(source: str, doc_version: Optional[int], *, drop_control_plane: bool = False):
    conn = _lex_db_connect()
    had_outer_tx = bool(getattr(conn, "in_transaction", False))
    version_key = int(doc_version) if doc_version is not None else None
    conn.execute("SAVEPOINT delete_source_version")
    try:
        ids = _chunk_row_ids_for_source_version(source, version_key)
        if ids:
            conn.executemany("DELETE FROM chunks_fts WHERE rowid = ?", [(i,) for i in ids])
            conn.executemany("DELETE FROM chunks_meta WHERE id = ?", [(i,) for i in ids])
        if version_key is None:
            conn.execute("DELETE FROM document_ir_meta WHERE source = ?", (source,))
            conn.execute("DELETE FROM document_ir WHERE source = ?", (source,))
        else:
            conn.execute("DELETE FROM document_ir_meta WHERE source = ? AND doc_version = ?", (source, version_key))
            conn.execute("DELETE FROM document_ir WHERE source = ? AND doc_version = ?", (source, version_key))
        if drop_control_plane:
            conn.execute("DELETE FROM documents_fts WHERE filename = ?", (source,))
            conn.execute("DELETE FROM doc_status WHERE source = ?", (source,))
            conn.execute("DELETE FROM documents WHERE source = ?", (source,))
        conn.execute("RELEASE SAVEPOINT delete_source_version")
        _lex_commit_if_needed(conn, had_outer_tx)
    except Exception:
        try:
            conn.execute("ROLLBACK TO SAVEPOINT delete_source_version")
            conn.execute("RELEASE SAVEPOINT delete_source_version")
        except Exception:
            pass
        raise


def _purge_source_for_reindex(source: str, pending_version: Optional[int] = None):
    # Reindex must preserve the currently published active version and control-plane versions.
    safe = _safe_filename(source)
    _crash_inject("before_purge")
    if pending_version is not None:
        _delete_milvus_source_version(safe, pending_version)
        _lex_db_delete_source_version(safe, pending_version, drop_control_plane=False)
    _crash_inject("after_purge")


def _lex_db_delete_source(source: str):
    conn = _lex_db_connect()
    had_outer_tx = bool(getattr(conn, "in_transaction", False))
    _crash_inject("delete_sqlite")
    conn.execute("SAVEPOINT delete_source")
    try:
        _lex_db_delete_source_version(source, None, drop_control_plane=True)
        conn.execute("RELEASE SAVEPOINT delete_source")
        _lex_commit_if_needed(conn, had_outer_tx)
    except Exception:
        try:
            conn.execute("ROLLBACK TO SAVEPOINT delete_source")
            conn.execute("RELEASE SAVEPOINT delete_source")
        except Exception:
            pass
        raise
def _lexical_recall_indexed(query: str, limit: int, source_filter: Optional[str] = None) -> List[Dict[str, Any]]:
    conn = _lex_db_connect()
    q = query.strip()
    if not q:
        return []
    cur = conn.execute("SELECT rowid, text FROM chunks_fts WHERE chunks_fts MATCH ? ORDER BY rank LIMIT ?", (q, int(limit)))
    rows = cur.fetchall()
    out: List[Dict[str, Any]] = []
    for rid, text in rows:
        meta = conn.execute("SELECT source, section, metadata FROM chunks_meta WHERE id = ?", (rid,)).fetchone()
        if not meta:
            continue
        source, section, metadata = meta
        if source_filter and (_normalize_filename_for_match(source or "") != source_filter):
            continue
        md = {}
        try:
            md = json.loads(metadata or "{}")
        except Exception:
            md = {}
        md["section"] = (section or "")
        ent = {"source": source or "unknown", "text": text or "", "metadata": md}
        out.append({"entity": ent, "score": 0.0})
    return out
def _lex_db_set_status(source: str, status: str):
    conn = _lex_db_connect()
    had_outer_tx = bool(getattr(conn, "in_transaction", False))
    conn.execute(
        "INSERT INTO doc_status(source, status, updated_at) VALUES (?,?,?) ON CONFLICT(source) DO UPDATE SET status=excluded.status, updated_at=excluded.updated_at",
        (source, status, datetime.now().isoformat())
    )
    _lex_commit_if_needed(conn, had_outer_tx)
def _lex_db_get_status(source: str) -> Optional[str]:
    conn = _lex_db_connect()
    row = conn.execute("SELECT status FROM doc_status WHERE source = ?", (source,)).fetchone()
    return (row[0] if row else None)
def _doc_get(source: str) -> Dict[str, Any]:
    conn = _lex_db_connect()
    row = conn.execute(
        "SELECT source, status, active_version, pending_version, last_error, updated_at, canonical_title, title_tokens, aliases, filename_stem, doc_type, topic, "
        "source_id, original_filename, content_sha256, mime_type, detected_ext, file_size, page_count, parser_route, parser_backend, parse_status, parse_quality_score, quality_flags, searchable, publish_gate, duplicate_state, duplicate_of, same_title_group, suspicious_file_type "
        "FROM documents WHERE source = ?",
        (source,),
    ).fetchone()
    if not row:
        return {
            "source": source,
            "status": None,
            "active_version": None,
            "pending_version": None,
            "last_error": None,
            "updated_at": None,
            "canonical_title": None,
            "title_tokens": None,
            "aliases": None,
            "filename_stem": None,
            "doc_type": None,
            "topic": None,
            "source_id": None,
            "original_filename": None,
            "content_sha256": None,
            "mime_type": None,
            "detected_ext": None,
            "file_size": None,
            "page_count": None,
            "parser_route": None,
            "parser_backend": None,
            "parse_status": None,
            "parse_quality_score": None,
            "quality_flags": None,
            "searchable": 0,
            "publish_gate": None,
            "duplicate_state": None,
            "duplicate_of": None,
            "same_title_group": None,
            "suspicious_file_type": 0,
        }
    return {
        "source": row[0],
        "status": row[1],
        "active_version": row[2],
        "pending_version": row[3],
        "last_error": row[4],
        "updated_at": row[5],
        "canonical_title": row[6],
        "title_tokens": row[7],
        "aliases": row[8],
        "filename_stem": row[9],
        "doc_type": row[10],
        "topic": row[11],
        "source_id": row[12],
        "original_filename": row[13],
        "content_sha256": row[14],
        "mime_type": row[15],
        "detected_ext": row[16],
        "file_size": row[17],
        "page_count": row[18],
        "parser_route": row[19],
        "parser_backend": row[20],
        "parse_status": row[21],
        "parse_quality_score": row[22],
        "quality_flags": row[23],
        "searchable": row[24],
        "publish_gate": row[25],
        "duplicate_state": row[26],
        "duplicate_of": row[27],
        "same_title_group": row[28],
        "suspicious_file_type": row[29],
    }
def _doc_upsert(source: str, **fields):
    conn = _lex_db_connect()
    had_outer_tx = bool(getattr(conn, "in_transaction", False))
    now = datetime.now().isoformat()
    cur = _doc_get(source)
    status = fields.get("status", cur["status"])
    active_version = fields.get("active_version", cur["active_version"])
    pending_version = fields.get("pending_version", cur["pending_version"])
    last_error = fields.get("last_error", cur["last_error"])
    canonical_title = fields.get("canonical_title", cur["canonical_title"])
    title_tokens = fields.get("title_tokens", cur["title_tokens"])
    aliases = fields.get("aliases", cur["aliases"])
    filename_stem = fields.get("filename_stem", cur["filename_stem"])
    doc_type = fields.get("doc_type", cur["doc_type"])
    topic = fields.get("topic", cur["topic"])
    source_id = fields.get("source_id", cur["source_id"])
    original_filename = fields.get("original_filename", cur["original_filename"])
    content_sha256 = fields.get("content_sha256", cur["content_sha256"])
    mime_type = fields.get("mime_type", cur["mime_type"])
    detected_ext = fields.get("detected_ext", cur["detected_ext"])
    file_size = fields.get("file_size", cur["file_size"])
    page_count = fields.get("page_count", cur["page_count"])
    parser_route = fields.get("parser_route", cur["parser_route"])
    parser_backend = fields.get("parser_backend", cur["parser_backend"])
    parse_status = fields.get("parse_status", cur["parse_status"])
    parse_quality_score = fields.get("parse_quality_score", cur["parse_quality_score"])
    quality_flags = fields.get("quality_flags", cur["quality_flags"])
    searchable = fields.get("searchable", cur["searchable"])
    publish_gate = fields.get("publish_gate", cur["publish_gate"])
    duplicate_state = fields.get("duplicate_state", cur["duplicate_state"])
    duplicate_of = fields.get("duplicate_of", cur["duplicate_of"])
    same_title_group = fields.get("same_title_group", cur["same_title_group"])
    suspicious_file_type = fields.get("suspicious_file_type", cur["suspicious_file_type"])
    conn.execute(
        "INSERT INTO documents(source, status, active_version, pending_version, last_error, updated_at, canonical_title, title_tokens, aliases, filename_stem, doc_type, topic, source_id, original_filename, content_sha256, mime_type, detected_ext, file_size, page_count, parser_route, parser_backend, parse_status, parse_quality_score, quality_flags, searchable, publish_gate, duplicate_state, duplicate_of, same_title_group, suspicious_file_type) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) "
        "ON CONFLICT(source) DO UPDATE SET status=excluded.status, active_version=excluded.active_version, pending_version=excluded.pending_version, last_error=excluded.last_error, updated_at=excluded.updated_at, canonical_title=excluded.canonical_title, title_tokens=excluded.title_tokens, aliases=excluded.aliases, filename_stem=excluded.filename_stem, doc_type=excluded.doc_type, topic=excluded.topic, source_id=excluded.source_id, original_filename=excluded.original_filename, content_sha256=excluded.content_sha256, mime_type=excluded.mime_type, detected_ext=excluded.detected_ext, file_size=excluded.file_size, page_count=excluded.page_count, parser_route=excluded.parser_route, parser_backend=excluded.parser_backend, parse_status=excluded.parse_status, parse_quality_score=excluded.parse_quality_score, quality_flags=excluded.quality_flags, searchable=excluded.searchable, publish_gate=excluded.publish_gate, duplicate_state=excluded.duplicate_state, duplicate_of=excluded.duplicate_of, same_title_group=excluded.same_title_group, suspicious_file_type=excluded.suspicious_file_type",
        (source, status, active_version, pending_version, last_error, now, canonical_title, title_tokens, aliases, filename_stem, doc_type, topic, source_id, original_filename, content_sha256, mime_type, detected_ext, file_size, page_count, parser_route, parser_backend, parse_status, parse_quality_score, quality_flags, searchable, publish_gate, duplicate_state, duplicate_of, same_title_group, suspicious_file_type)
    )
    _lex_commit_if_needed(conn, had_outer_tx)
def _doc_next_version(source: str) -> int:
    cur = _doc_get(source)
    v = cur["active_version"] if cur["active_version"] is not None else 0
    pv = cur["pending_version"] if cur["pending_version"] is not None else v
    return int(max(v, pv)) + 1

def _get_active_version(source: str) -> Optional[int]:
    doc = _doc_get(source)
    try:
        return int(doc.get("active_version")) if doc.get("active_version") is not None else None
    except Exception:
        return None
def _filename_stem(name: str) -> str:
    s = (name or "").strip().replace("\\", "/").split("/")[-1]
    if "." in s:
        s = ".".join(s.split(".")[:-1])
    return s


def _doc_title_profile(source: str) -> Dict[str, str]:
    stem = _filename_stem(source)
    canonical = re.sub(r"(?:[_\-]\d{4}[-_]\d{2}[-_]\d{2}){1,2}$", "", stem).strip("_-")
    canonical = re.sub(r"[_\-]+", " ", canonical).strip()
    if not canonical:
        canonical = stem.replace("_", " ").strip() or stem
    aliases = _expand_title_aliases(stem, stem.replace("_", " ").strip(), canonical)
    return {
        "stem": stem,
        "canonical_title": canonical,
        "title_tokens": " ".join(aliases),
        "aliases": ",".join(aliases[1:]),
    }


def _json_dumps(data: Any) -> str:
    return json.dumps(data or {}, ensure_ascii=False, sort_keys=True)


def _json_loads(value: Any, fallback: Any):
    if value in (None, ""):
        return fallback
    try:
        return json.loads(value)
    except Exception:
        return fallback


def _content_sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content or b"").hexdigest()


def _content_sha256_text(content: str) -> str:
    return hashlib.sha256((content or "").encode("utf-8")).hexdigest()


def _build_source_id(original_filename: str, content_sha256: str) -> str:
    seed = f"{_safe_filename(original_filename)}:{content_sha256}".encode("utf-8")
    digest = hashlib.sha1(seed).hexdigest()[:16]
    return f"doc_{datetime.now().strftime('%Y%m%d')}_{digest}"


def _same_title_group(canonical_title: str) -> str:
    normalized = _normalize_query(canonical_title).lower()
    if not normalized:
        return ""
    return hashlib.sha1(normalized.encode("utf-8")).hexdigest()[:16]


def _canonical_doc_id_for_source(source: str) -> str:
    safe_source = _normalize_filename_for_match(source)
    if not safe_source:
        return ""
    info = _doc_get(safe_source)
    group = str(info.get("same_title_group") or "").strip()
    if group:
        return group
    canonical_title = str(info.get("canonical_title") or _filename_stem(safe_source) or safe_source).strip()
    group = _same_title_group(canonical_title)
    if group:
        return group
    return _normalize_title_probe_text(canonical_title)


def _sources_equivalent(left: str, right: str) -> bool:
    left_source = _normalize_filename_for_match(left)
    right_source = _normalize_filename_for_match(right)
    if not left_source or not right_source:
        return False
    if left_source == right_source:
        return True
    left_id = _canonical_doc_id_for_source(left_source)
    right_id = _canonical_doc_id_for_source(right_source)
    return bool(left_id and right_id and left_id == right_id)


def _collapse_sources_by_canonical(sources: List[str], limit: Optional[int] = None) -> List[str]:
    out: List[str] = []
    seen: set[str] = set()
    for source in sources:
        safe_source = _normalize_filename_for_match(source)
        if not safe_source:
            continue
        canonical_id = _canonical_doc_id_for_source(safe_source) or f"source:{safe_source}"
        if canonical_id in seen:
            continue
        seen.add(canonical_id)
        out.append(safe_source)
        if limit and len(out) >= max(1, int(limit)):
            break
    return out


def _title_probe_variants(text: str) -> List[str]:
    base = _normalize_query((text or "").replace("相关", " "))
    if not base:
        return []
    variants: List[str] = []
    for candidate in [base, _strip_section_question_tail(base)]:
        value = _normalize_query(candidate)
        if len(value) >= 2 and value not in variants:
            variants.append(value)
        stripped = re.sub(r"(条例|办法|规定|规则|细则|通知|通告|决定|议事规则|实施办法|管理条例|管理办法)$", "", value)
        stripped_core = _normalize_query(_strip_leading_region_prefix(stripped))
        stripped = stripped_core or _normalize_query(stripped)
        if len(stripped) >= 2 and stripped not in variants:
            variants.append(stripped)
    return variants


def _normalized_title_candidate_sources(text: str, limit: int = 5) -> List[str]:
    out: List[str] = []
    probe_limit = max(int(limit) * 2, 6)
    for probe in _title_probe_variants(text):
        for entry in _rank_title_source_matches(probe, limit=probe_limit, include_topic_like=True):
            source = _normalize_filename_for_match(entry.get("source") or "")
            if source and source not in out:
                out.append(source)
        if not out:
            probe_norm = _normalize_title_probe_text(probe)
            if len(probe_norm) >= 2:
                conn = _lex_db_connect()
                rows = conn.execute("SELECT source FROM documents").fetchall()
                ranked: List[Tuple[float, str]] = []
                for row in rows:
                    source = _normalize_filename_for_match((row[0] if row else "") or "")
                    if not source or not _visible_document_exists(source):
                        continue
                    best = 0.0
                    for entity in _canonical_source_core_entities(source):
                        entity_norm = _normalize_title_probe_text(entity)
                        if not entity_norm:
                            continue
                        if probe_norm == entity_norm:
                            best = max(best, 1.0)
                        elif probe_norm in entity_norm:
                            best = max(best, 0.84)
                    if best > 0:
                        ranked.append((best, source))
                ranked.sort(key=lambda item: (-item[0], item[1]))
                for _, source in ranked[:probe_limit]:
                    if source not in out:
                        out.append(source)
        if len(out) >= probe_limit:
            break
    return _collapse_sources_by_canonical(out, limit=limit)


def _normalized_query_title_candidate_sources(query: str, limit: int = 5) -> List[str]:
    probes = _extract_explicit_regulation_mentions(query) or [_strip_query_intent_phrases(query) or query]
    out: List[str] = []
    for probe in probes:
        for source in _normalized_title_candidate_sources(probe, limit=limit):
            if source and source not in out:
                out.append(source)
        if len(out) >= max(1, int(limit)):
            break
    return _collapse_sources_by_canonical(out, limit=limit)


@lru_cache(maxsize=2048)
def _embed_text_sync_cached(text: str) -> Tuple[float, ...]:
    payload = _normalize_query(text)
    if not payload:
        return tuple()
    try:
        import requests

        response = requests.post(
            f"{config.EMBEDDING_URL}/embed",
            json={"texts": [payload], "normalize": True, "batch_size": 1},
            timeout=20,
        )
        response.raise_for_status()
        embeddings = response.json().get("embeddings") or []
        if not embeddings:
            return tuple()
        return tuple(float(value) for value in (embeddings[0] or []))
    except Exception:
        return tuple()


def _normalized_embedding_cosine(left: Tuple[float, ...], right: Tuple[float, ...]) -> float:
    if not left or not right or len(left) != len(right):
        return 0.0
    return float(sum(lv * rv for lv, rv in zip(left, right)))


@lru_cache(maxsize=1)
def _dense_title_probe_entries() -> Tuple[Tuple[str, str, str], ...]:
    conn = _lex_db_connect()
    rows = conn.execute("SELECT source, canonical_title, aliases, filename_stem FROM documents").fetchall()
    out: List[Tuple[str, str, str]] = []
    for source, canonical_title, aliases, filename_stem in rows:
        safe_source = _normalize_filename_for_match(source or "")
        if not safe_source or not _visible_document_exists(safe_source):
            continue
        display_title = (canonical_title or "").strip() or (filename_stem or "").strip() or safe_source
        parts = [display_title, aliases or "", filename_stem or ""] + _canonical_source_core_entities(safe_source)
        probe_text = "\n".join(part.strip() for part in parts if str(part or "").strip())
        if probe_text:
            out.append((safe_source, display_title, probe_text))
    return tuple(out)


def _dense_title_source_matches(text: str, limit: int = 5) -> List[Dict[str, Any]]:
    if not bool(getattr(config, "ENABLE_DENSE_TITLE_FALLBACK", True)):
        return []
    probe = _normalize_reference_text(text) or _normalize_query(text)
    if len(probe) < 2:
        return []
    strip_suffixes = [
        "海洋环境和渔业资源保护条例", "管理条例", "管理办法", "实施办法", "实施细则",
        "保护条例", "条例", "规定", "办法", "规则", "保护", "管理", "要求", "内容",
    ]
    core_fragments: List[str] = []
    for fragment in re.findall(r"[\u4e00-\u9fff]{2,}", _normalize_query(text)):
        cleaned = fragment
        for suffix in strip_suffixes:
            if cleaned.endswith(suffix) and len(cleaned) - len(suffix) >= 2:
                cleaned = cleaned[: -len(suffix)]
                break
        cleaned = _normalize_query(cleaned)
        if len(cleaned) >= 2 and cleaned not in core_fragments:
            core_fragments.append(cleaned)
    max_chars = max(40, int(getattr(config, "DENSE_TITLE_PROBE_MAX_CHARS", 160)))
    query_embedding = _embed_text_sync_cached(probe[:max_chars])
    ranked_map: Dict[str, Dict[str, Any]] = {}
    compact_probe = _normalize_reference_text(text)
    if query_embedding:
        for source, display_title, probe_text in _dense_title_probe_entries():
            title_embedding = _embed_text_sync_cached(probe_text[:max_chars])
            if not title_embedding:
                continue
            score = _normalized_embedding_cosine(query_embedding, title_embedding)
            compact_title = _normalize_reference_text(display_title)
            if compact_probe and compact_title and (compact_probe in compact_title or compact_title in compact_probe):
                score = max(score, 0.86)
            if core_fragments and any(fragment in _normalize_query(display_title) for fragment in core_fragments):
                score = max(score, 0.88)
            if score <= 0.0:
                continue
            ranked_map[source] = {
                "source": source,
                "title": display_title,
                "score": score,
            }
    for entry in _build_doc_recall_plan(text, limit=max(limit * 3, 8)):
        source = _normalize_filename_for_match((entry or {}).get("source") or "")
        if not source:
            continue
        display_title = _source_display_title(source)
        if core_fragments and not any(fragment in _normalize_query(display_title) for fragment in core_fragments):
            continue
        prior = float((entry or {}).get("prior") or 0.0)
        if prior <= 0.0:
            continue
        score = 0.82 + min(prior, 0.12)
        current = ranked_map.get(source)
        if current is None or score > float(current.get("score") or 0.0):
            ranked_map[source] = {
                "source": source,
                "title": display_title,
                "score": score,
            }
    ranked = list(ranked_map.values())
    ranked.sort(key=lambda item: (-float(item.get("score") or 0.0), item.get("source") or ""))
    return ranked[: max(1, int(limit))]


def _doc_profile_get(source: str, doc_version: int) -> Dict[str, Any]:
    conn = _lex_db_connect()
    row = conn.execute(
        "SELECT source, source_id, doc_version, original_filename, canonical_title, region, doc_type, publish_date, effective_date, doc_version_label, parse_quality_score, quality_flags, source_resolution_fields, parser_route, parser_backend, mime_type, detected_ext, file_size, page_count, content_sha256, created_at, updated_at FROM document_profiles WHERE source = ? AND doc_version = ?",
        (source, int(doc_version)),
    ).fetchone()
    if not row:
        return {}
    return {
        "source": row[0],
        "source_id": row[1],
        "doc_version": row[2],
        "original_filename": row[3],
        "canonical_title": row[4],
        "region": row[5],
        "doc_type": row[6],
        "publish_date": row[7],
        "effective_date": row[8],
        "doc_version_label": row[9],
        "parse_quality_score": row[10],
        "quality_flags": _json_loads(row[11], []),
        "source_resolution_fields": _json_loads(row[12], {}),
        "parser_route": row[13],
        "parser_backend": row[14],
        "mime_type": row[15],
        "detected_ext": row[16],
        "file_size": row[17],
        "page_count": row[18],
        "content_sha256": row[19],
        "created_at": row[20],
        "updated_at": row[21],
    }


def _doc_profile_upsert(source: str, doc_version: int, **fields):
    conn = _lex_db_connect()
    had_outer_tx = bool(getattr(conn, "in_transaction", False))
    current = _doc_profile_get(source, int(doc_version))
    now = datetime.now().isoformat()
    payload = {
        "source": source,
        "source_id": fields.get("source_id", current.get("source_id")),
        "doc_version": int(doc_version),
        "original_filename": fields.get("original_filename", current.get("original_filename")),
        "canonical_title": fields.get("canonical_title", current.get("canonical_title")),
        "region": fields.get("region", current.get("region")),
        "doc_type": fields.get("doc_type", current.get("doc_type")),
        "publish_date": fields.get("publish_date", current.get("publish_date")),
        "effective_date": fields.get("effective_date", current.get("effective_date")),
        "doc_version_label": fields.get("doc_version_label", current.get("doc_version_label")),
        "parse_quality_score": fields.get("parse_quality_score", current.get("parse_quality_score")),
        "quality_flags": _json_dumps(fields.get("quality_flags", current.get("quality_flags", []))),
        "source_resolution_fields": _json_dumps(fields.get("source_resolution_fields", current.get("source_resolution_fields", {}))),
        "parser_route": fields.get("parser_route", current.get("parser_route")),
        "parser_backend": fields.get("parser_backend", current.get("parser_backend")),
        "mime_type": fields.get("mime_type", current.get("mime_type")),
        "detected_ext": fields.get("detected_ext", current.get("detected_ext")),
        "file_size": fields.get("file_size", current.get("file_size")),
        "page_count": fields.get("page_count", current.get("page_count")),
        "content_sha256": fields.get("content_sha256", current.get("content_sha256")),
        "created_at": current.get("created_at") or now,
        "updated_at": now,
    }
    conn.execute(
        "INSERT INTO document_profiles(source, source_id, doc_version, original_filename, canonical_title, region, doc_type, publish_date, effective_date, doc_version_label, parse_quality_score, quality_flags, source_resolution_fields, parser_route, parser_backend, mime_type, detected_ext, file_size, page_count, content_sha256, created_at, updated_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) "
        "ON CONFLICT(source, doc_version) DO UPDATE SET source_id=excluded.source_id, original_filename=excluded.original_filename, canonical_title=excluded.canonical_title, region=excluded.region, doc_type=excluded.doc_type, publish_date=excluded.publish_date, effective_date=excluded.effective_date, doc_version_label=excluded.doc_version_label, parse_quality_score=excluded.parse_quality_score, quality_flags=excluded.quality_flags, source_resolution_fields=excluded.source_resolution_fields, parser_route=excluded.parser_route, parser_backend=excluded.parser_backend, mime_type=excluded.mime_type, detected_ext=excluded.detected_ext, file_size=excluded.file_size, page_count=excluded.page_count, content_sha256=excluded.content_sha256, updated_at=excluded.updated_at",
        (payload["source"], payload["source_id"], payload["doc_version"], payload["original_filename"], payload["canonical_title"], payload["region"], payload["doc_type"], payload["publish_date"], payload["effective_date"], payload["doc_version_label"], payload["parse_quality_score"], payload["quality_flags"], payload["source_resolution_fields"], payload["parser_route"], payload["parser_backend"], payload["mime_type"], payload["detected_ext"], payload["file_size"], payload["page_count"], payload["content_sha256"], payload["created_at"], payload["updated_at"]),
    )
    _lex_commit_if_needed(conn, had_outer_tx)


def _replace_doc_aliases(source: str, doc_version: int, aliases: List[Dict[str, Any]]):
    conn = _lex_db_connect()
    had_outer_tx = bool(getattr(conn, "in_transaction", False))
    conn.execute("DELETE FROM document_aliases WHERE source = ? AND doc_version = ?", (source, int(doc_version)))
    for alias in aliases or []:
        text = _normalize_query(str((alias or {}).get("alias") or ""))
        if not text:
            continue
        conn.execute(
            "INSERT OR REPLACE INTO document_aliases(source, doc_version, alias, alias_type, weight) VALUES (?,?,?,?,?)",
            (source, int(doc_version), text, str((alias or {}).get("alias_type") or "auto"), float((alias or {}).get("weight") or 1.0)),
        )
    _lex_commit_if_needed(conn, had_outer_tx)


def _replace_doc_sections(source: str, doc_version: int, sections: List[Dict[str, Any]]):
    conn = _lex_db_connect()
    had_outer_tx = bool(getattr(conn, "in_transaction", False))
    conn.execute("DELETE FROM document_sections WHERE source = ? AND doc_version = ?", (source, int(doc_version)))
    for idx, section in enumerate(sections or []):
        title = _normalize_query(str((section or {}).get("section_title") or ""))
        if not title:
            continue
        section_key = str((section or {}).get("section_key") or f"section_{idx}")
        conn.execute(
            "INSERT OR REPLACE INTO document_sections(source, doc_version, section_key, section_title, section_level, chunk_start, chunk_end, section_path) VALUES (?,?,?,?,?,?,?,?)",
            (source, int(doc_version), section_key, title, int((section or {}).get("section_level") or 1), (section or {}).get("chunk_start"), (section or {}).get("chunk_end"), _json_dumps((section or {}).get("section_path") or [])),
        )
    _lex_commit_if_needed(conn, had_outer_tx)


def _replace_doc_topics(source: str, doc_version: int, topics: List[Dict[str, Any]]):
    conn = _lex_db_connect()
    had_outer_tx = bool(getattr(conn, "in_transaction", False))
    conn.execute("DELETE FROM document_topics WHERE source = ? AND doc_version = ?", (source, int(doc_version)))
    for topic in topics or []:
        text = _normalize_query(str((topic or {}).get("topic") or ""))
        if not text:
            continue
        conn.execute(
            "INSERT OR REPLACE INTO document_topics(source, doc_version, topic, topic_type, weight) VALUES (?,?,?,?,?)",
            (source, int(doc_version), text, str((topic or {}).get("topic_type") or "topic_term"), float((topic or {}).get("weight") or 1.0)),
        )
    _lex_commit_if_needed(conn, had_outer_tx)


def _canonical_extension(ext: str) -> str:
    raw = (ext or "").strip().lower()
    return _EXTENSION_EQUIVALENTS.get(raw, raw)


def _extension_matches_detected(extension: str, detected_ext: str) -> bool:
    ext = _canonical_extension(extension)
    detected = _canonical_extension(detected_ext)
    if not ext or not detected:
        return True
    if ext == detected:
        return True
    if ext in _TEXT_LIKE_EXTENSIONS and detected in _TEXT_LIKE_EXTENSIONS:
        return True
    return False


def _probe_mime_allowed(detected_ext: str, mime_type: str) -> bool:
    allowed = _ALLOWED_MIME_BY_EXTENSION.get(_canonical_extension(detected_ext), set())
    mime = (mime_type or "").strip().lower()
    if not allowed or not mime:
        return True
    if mime in allowed:
        return True
    if _canonical_extension(detected_ext) in _TEXT_LIKE_EXTENSIONS and mime.startswith("text/"):
        return True
    return False


def _build_text_upload_probe(filename: str, content: str) -> Dict[str, Any]:
    safe_name = _safe_filename(filename)
    extension = os.path.splitext(safe_name)[1].lower()
    detected_ext = extension if extension in _TEXT_LIKE_EXTENSIONS else ".txt"
    return {
        "filename": safe_name,
        "extension": extension,
        "detected_ext": detected_ext,
        "mime_type": _sniff_mime_type(safe_name, (content or "").encode("utf-8"), detected_ext),
        "signature": "text",
        "file_size": len((content or "").encode("utf-8")),
        "page_count": 1,
        "route": "plain_text",
        "parser_backend": "direct_text",
        "degraded": False,
    }


def _validate_upload_probe(filename: str, probe: Dict[str, Any], is_text_upload: bool = False):
    safe_name = _safe_filename(filename)
    extension = os.path.splitext(safe_name)[1].lower()
    detected_ext = (probe.get("detected_ext") or extension or "").lower()
    file_size = int(probe.get("file_size") or 0)
    if file_size <= 0:
        raise HTTPException(status_code=400, detail="empty_file: 文件内容为空")
    if file_size > int(config.MAX_FILE_SIZE_MB) * 1024 * 1024:
        raise HTTPException(status_code=400, detail=f"file_too_large: 文件超过 {config.MAX_FILE_SIZE_MB}MB 限制")
    if detected_ext not in SUPPORTED_FILE_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"unsupported_file_type: 不支持的文件格式 {detected_ext or extension or '<none>'}")

    suspicious_reasons: List[str] = []
    if extension and not _extension_matches_detected(extension, detected_ext):
        suspicious_reasons.append("ext_magic_conflict")
    if not _probe_mime_allowed(detected_ext, str(probe.get("mime_type") or "")):
        suspicious_reasons.append("mime_mismatch")
    if str(probe.get("signature") or "") == "binary" and detected_ext in _TEXT_LIKE_EXTENSIONS:
        suspicious_reasons.append("binary_payload_for_text_type")
    if suspicious_reasons:
        raise HTTPException(status_code=400, detail=f"suspicious_file_type: {', '.join(suspicious_reasons)}")

    if detected_ext == ".pdf" and probe.get("encrypted"):
        raise HTTPException(status_code=400, detail="encrypted_file: PDF 已加密，无法解析")
    if detected_ext == ".pdf" and int(probe.get("page_count") or 0) > int(config.MAX_PDF_PAGES):
        raise HTTPException(status_code=400, detail=f"pdf_page_limit_exceeded: PDF 页数超过 {config.MAX_PDF_PAGES} 页限制")
    if detected_ext == ".xlsx":
        if int(probe.get("sheet_count") or 0) > int(config.MAX_XLSX_SHEETS):
            raise HTTPException(status_code=400, detail=f"xlsx_sheet_limit_exceeded: 工作表数量超过 {config.MAX_XLSX_SHEETS}")
        if int(probe.get("max_rows") or 0) > int(config.MAX_XLSX_ROWS):
            raise HTTPException(status_code=400, detail=f"xlsx_row_limit_exceeded: 行数超过 {config.MAX_XLSX_ROWS}")
        if int(probe.get("max_cols") or 0) > int(config.MAX_XLSX_COLS):
            raise HTTPException(status_code=400, detail=f"xlsx_col_limit_exceeded: 列数超过 {config.MAX_XLSX_COLS}")
    if detected_ext in _IMAGE_EXTENSIONS:
        image_pixels = int(probe.get("image_pixels") or 0)
        if image_pixels and image_pixels > int(config.MAX_IMAGE_PIXELS):
            raise HTTPException(status_code=400, detail=f"image_pixel_limit_exceeded: 图片像素超过 {config.MAX_IMAGE_PIXELS}")
    if is_text_upload and detected_ext not in SUPPORTED_FILE_EXTENSIONS:
        raise HTTPException(status_code=400, detail="unsupported_text_upload: 文本上传文件名后缀不在支持列表内")


def _extract_dates_from_text(text: str) -> List[str]:
    normalized = (text or "").replace("年", "-").replace("月", "-").replace("日", "")
    candidates: List[str] = []
    for year, month, day in re.findall(r"(20\d{2}|19\d{2})[-_/\.](\d{1,2})[-_/\.](\d{1,2})", normalized):
        candidates.append(f"{int(year):04d}-{int(month):02d}-{int(day):02d}")
    return _dedupe_keep_order(candidates)


def _infer_document_dates(filename: str, text: str) -> Dict[str, Optional[str]]:
    candidates = _extract_dates_from_text(f"{filename}\n{text}")
    effective_date = None
    publish_date = None
    for match in re.finditer(r"((?:20|19)\d{2}[年\-_/\.]{1}\d{1,2}[月\-_/\.]{1}\d{1,2}日?)", text or ""):
        snippet = (text or "")[max(0, match.start() - 8): match.end() + 8]
        normalized = _extract_dates_from_text(match.group(1))
        if not normalized:
            continue
        value = normalized[0]
        if ("施行" in snippet or "实施" in snippet) and not effective_date:
            effective_date = value
        if any(token in snippet for token in ("发布", "公布", "印发")) and not publish_date:
            publish_date = value
    if not publish_date and candidates:
        publish_date = candidates[0]
    if not effective_date and len(candidates) >= 2:
        effective_date = candidates[1]
    if not effective_date:
        effective_date = publish_date
    version_label = f"{effective_date}施行版" if effective_date else None
    return {
        "publish_date": publish_date,
        "effective_date": effective_date,
        "doc_version_label": version_label,
        "date_terms": [value for value in (publish_date, effective_date) if value],
    }


def _document_primary_title(filename: str, document_ir: Optional[Dict[str, Any]]) -> str:
    for element in (document_ir or {}).get("elements") or []:
        element_type = (element.get("element_type") or "").strip().lower()
        text = _normalize_query((element.get("text_raw") or ""))
        if element_type == "title" and len(text) >= 4:
            return text
    for element in (document_ir or {}).get("elements") or []:
        element_type = (element.get("element_type") or "").strip().lower()
        text = _normalize_query((element.get("text_raw") or ""))
        if element_type in {"heading", "sheet"} and len(text) >= 4:
            return text
    return _doc_title_profile(filename).get("canonical_title") or _filename_stem(filename)


def _document_section_titles(document_ir: Optional[Dict[str, Any]]) -> List[str]:
    titles: List[str] = []
    for element in (document_ir or {}).get("elements") or []:
        element_type = (element.get("element_type") or "").strip().lower()
        raw_text = _normalize_query(element.get("text_raw") or "")
        if element_type in {"title", "heading", "sheet"} and raw_text:
            titles.append(raw_text)
        for part in _normalize_section_path(element.get("section_path")):
            normalized = _normalize_query(part)
            if normalized:
                titles.append(normalized)
    return _dedupe_keep_order(titles)


def _document_topic_terms(text: str, section_titles: List[str]) -> List[str]:
    out = list(_infer_topics(text))
    for title in section_titles:
        compact = _normalize_query(title)
        if not compact or len(compact) <= 2:
            continue
        out.append(compact)
        for suffix in ("法律责任", "处罚", "罚则", "附则", "总则", "登记", "程序", "要求"):
            if suffix in compact:
                out.append(suffix)
    return _dedupe_keep_order(out)


def _assess_document_quality(document_ir: Dict[str, Any], probe: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    plain_text = _document_ir_plain_text(document_ir, normalized=False)
    normalized_text = re.sub(r"\s+", "", plain_text or "")
    section_titles = _document_section_titles(document_ir)
    flags: List[str] = []
    if not normalized_text:
        flags.append("parse_empty")
    if len(normalized_text) < int(config.MIN_PARSE_TEXT_CHARS):
        flags.append("parse_low_quality")
    if (probe or {}).get("garbled_text_majority"):
        flags.append("garbled_text_majority")
    if (probe or {}).get("degraded"):
        flags.append("parser_degraded")
    score = 0.0
    score += min(len(normalized_text) / 1200.0, 1.0) * 0.60
    score += min(len(section_titles) / 8.0, 1.0) * 0.20
    if (probe or {}).get("page_count"):
        score += min(int((probe or {}).get("page_count") or 0) / 20.0, 1.0) * 0.10
    if (probe or {}).get("route") in {"pdf_ocr_layout", "image_ocr_layout"} and len(normalized_text) >= int(config.MIN_PARSE_TEXT_CHARS):
        score += 0.10
    if "parse_empty" in flags:
        score = 0.0
    elif "parse_low_quality" in flags:
        score = min(score, 0.30)
    status = "parsed"
    if "parse_empty" in flags:
        status = "parse_empty"
    elif "parse_low_quality" in flags or score < float(config.MIN_PARSE_QUALITY_SCORE):
        status = "parse_low_quality"
    return {
        "status": status,
        "score": round(float(score), 4),
        "flags": _dedupe_keep_order(flags),
        "text_chars": len(normalized_text),
        "section_count": len(section_titles),
    }


def _build_document_profile(source: str, original_filename: str, source_id: str, content_sha256: str, text: str, document_ir: Dict[str, Any], probe: Dict[str, Any], quality: Dict[str, Any], metadata: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    primary_title = _document_primary_title(original_filename or source, document_ir)
    title_profile = _doc_title_profile(primary_title or source)
    canonical_title = title_profile.get("canonical_title") or primary_title or _filename_stem(source)
    region = _extract_region_token(canonical_title)
    section_titles = _document_section_titles(document_ir)
    topic_terms = _document_topic_terms(text, section_titles)
    date_info = _infer_document_dates(original_filename or source, text)
    doc_type = (metadata or {}).get("doc_type") or _classify_doc_type(original_filename or source, text)
    aliases = _expand_title_aliases(original_filename or source, canonical_title, primary_title)
    if date_info.get("effective_date"):
        aliases.append(f"{canonical_title}{date_info['effective_date']}")
    source_resolution_fields = {
        "canonical_title": canonical_title,
        "title_aliases": _dedupe_keep_order(aliases),
        "region_terms": _dedupe_keep_order([term for term in [region, _strip_region_admin_tokens(region)] if term]),
        "doc_type_terms": _dedupe_keep_order([doc_type, canonical_title[-2:] if len(canonical_title) >= 2 else ""]),
        "section_titles": section_titles[:50],
        "topic_terms": topic_terms[:80],
        "date_terms": date_info.get("date_terms") or [],
    }
    return {
        "source": source,
        "source_id": source_id,
        "original_filename": original_filename,
        "canonical_title": canonical_title,
        "title_aliases": source_resolution_fields["title_aliases"],
        "region": region or None,
        "doc_type": doc_type,
        "publish_date": date_info.get("publish_date"),
        "effective_date": date_info.get("effective_date"),
        "doc_version_label": date_info.get("doc_version_label"),
        "section_titles": section_titles,
        "topic_terms": topic_terms,
        "source_resolution_fields": source_resolution_fields,
        "quality_flags": quality.get("flags") or [],
        "parse_quality_score": quality.get("score"),
        "parser_route": probe.get("route"),
        "parser_backend": probe.get("parser_backend"),
        "mime_type": probe.get("mime_type"),
        "detected_ext": probe.get("detected_ext"),
        "file_size": probe.get("file_size"),
        "page_count": probe.get("page_count"),
        "content_sha256": content_sha256,
    }


def _document_section_rows(document_ir: Dict[str, Any], fallback_title: str = "") -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    seen: set[str] = set()
    for idx, element in enumerate(document_ir.get("elements") or []):
        section_path = _normalize_section_path(element.get("section_path"))
        if not section_path:
            text = _normalize_query(element.get("text_raw") or "")
            if (element.get("element_type") or "").strip().lower() in {"title", "heading", "sheet"} and text:
                section_path = [text]
        if not section_path:
            continue
        key = " > ".join(section_path)
        if key in seen:
            continue
        seen.add(key)
        rows.append({
            "section_key": key,
            "section_title": section_path[-1],
            "section_level": len(section_path),
            "chunk_start": idx,
            "chunk_end": idx,
            "section_path": section_path,
        })
    if rows:
        return rows
    fallback = _normalize_query(fallback_title)
    if fallback:
        rows.append({
            "section_key": fallback,
            "section_title": fallback,
            "section_level": 1,
            "chunk_start": 0,
            "chunk_end": 0,
            "section_path": [fallback],
        })
    return rows


def _profile_alias_rows(profile: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows = [{"alias": alias, "alias_type": "title_alias", "weight": 1.0} for alias in profile.get("title_aliases") or []]
    rows.extend({"alias": title, "alias_type": "section_title", "weight": 0.65} for title in (profile.get("section_titles") or [])[:24])
    return rows


def _profile_topic_rows(profile: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows = [{"topic": topic, "topic_type": "topic_term", "weight": 1.0} for topic in profile.get("topic_terms") or []]
    if profile.get("region"):
        rows.append({"topic": profile["region"], "topic_type": "region_term", "weight": 0.9})
    return rows


def _persist_document_profile(source: str, doc_version: int, profile: Dict[str, Any]):
    profile_payload = dict(profile or {})
    profile_payload.pop("source", None)
    profile_payload.pop("doc_version", None)
    _doc_profile_upsert(source, doc_version, **profile_payload)
    _replace_doc_aliases(source, doc_version, _profile_alias_rows(profile))
    _replace_doc_sections(
        source,
        doc_version,
        _document_section_rows(
            _load_document_ir(source, doc_version) or {},
            fallback_title=str(profile.get("canonical_title") or source),
        ),
    )
    _replace_doc_topics(source, doc_version, _profile_topic_rows(profile))


def _find_existing_by_hash(content_sha256: str) -> Optional[Dict[str, Any]]:
    if not content_sha256:
        return None
    conn = _lex_db_connect()
    row = conn.execute(
        "SELECT source FROM documents WHERE content_sha256 = ? ORDER BY updated_at DESC LIMIT 1",
        (content_sha256,),
    ).fetchone()
    if not row:
        return None
    return _doc_get(row[0])


def _find_same_title_candidates(canonical_title: str, exclude_source: str = "") -> List[str]:
    group = _same_title_group(canonical_title)
    if not group:
        return []
    conn = _lex_db_connect()
    rows = conn.execute(
        "SELECT source FROM documents WHERE same_title_group = ? ORDER BY updated_at DESC LIMIT 20",
        (group,),
    ).fetchall()
    out: List[str] = []
    for (source,) in rows:
        safe_source = _safe_filename(source)
        if safe_source and safe_source != _safe_filename(exclude_source):
            out.append(safe_source)
    return _dedupe_keep_order(out)


def _detect_duplicate_upload(source: str, content_sha256: str, canonical_title: str) -> Dict[str, Any]:
    current = _doc_get(source)
    if current.get("content_sha256") and current.get("content_sha256") == content_sha256:
        return {
            "duplicate_state": "no_change",
            "duplicate_of": source,
            "same_title_candidates": _find_same_title_candidates(canonical_title, exclude_source=source),
        }
    existing_hash = _find_existing_by_hash(content_sha256)
    if existing_hash and _safe_filename(existing_hash.get("source") or "") != _safe_filename(source):
        return {
            "duplicate_state": "already_exists",
            "duplicate_of": existing_hash.get("source"),
            "same_title_candidates": _find_same_title_candidates(canonical_title, exclude_source=source),
        }
    same_title_candidates = _find_same_title_candidates(canonical_title, exclude_source=source)
    return {
        "duplicate_state": "same_title_candidate" if same_title_candidates else "new_upload",
        "duplicate_of": same_title_candidates[0] if same_title_candidates else None,
        "same_title_candidates": same_title_candidates,
    }


def _source_version_stats(source: str, doc_version: int) -> Dict[str, int]:
    conn = _lex_db_connect()
    rows = conn.execute("SELECT id, metadata FROM chunks_meta WHERE source = ?", (_safe_filename(source),)).fetchall()
    sqlite_chunks = 0
    fts_chunks = 0
    row_ids: List[int] = []
    for row_id, metadata in rows:
        md = _json_loads(metadata, {})
        version = md.get("doc_version")
        try:
            version = int(version) if version is not None else None
        except Exception:
            version = None
        if version is None or version == int(doc_version):
            sqlite_chunks += 1
            row_ids.append(int(row_id))
    for row_id in row_ids:
        if conn.execute("SELECT 1 FROM chunks_fts WHERE rowid = ?", (row_id,)).fetchone():
            fts_chunks += 1
    return {"sqlite_chunks": sqlite_chunks, "fts_chunks": fts_chunks}


def _milvus_version_count(source: str, doc_version: int) -> int:
    try:
        vector_db = VectorDBService()
        vector_db.connect()
        rows = vector_db.client.query(
            collection_name=vector_db.collection_name,
            filter=f"source == {json.dumps(_safe_filename(source), ensure_ascii=False)}",
            output_fields=["metadata"],
            limit=10000,
        )
        count = 0
        for row in rows or []:
            metadata = row.get("metadata") or {}
            try:
                if int(metadata.get("doc_version")) == int(doc_version):
                    count += 1
            except Exception:
                continue
        return count
    except Exception:
        return 0


def _build_publish_gate(source: str, doc_version: int) -> Dict[str, Any]:
    doc = _doc_get(source)
    stats = _source_version_stats(source, doc_version)
    milvus_count = _milvus_version_count(source, doc_version)
    profile = _doc_profile_get(source, doc_version)
    conn = _lex_db_connect()
    section_count_row = conn.execute(
        "SELECT COUNT(*) FROM document_sections WHERE source = ? AND doc_version = ?",
        (_safe_filename(source), int(doc_version)),
    ).fetchone()
    section_count = int(section_count_row[0] or 0) if section_count_row else 0
    parse_quality_score = float(profile.get("parse_quality_score") or doc.get("parse_quality_score") or 0.0)
    gate = {
        "sqlite_chunks_ok": stats["sqlite_chunks"] > 0,
        "fts_chunks_ok": stats["sqlite_chunks"] > 0 and stats["sqlite_chunks"] == stats["fts_chunks"],
        "milvus_vectors_ok": stats["sqlite_chunks"] > 0 and milvus_count == stats["sqlite_chunks"],
        "visibility_ok": _verify_version_visible(source, int(doc_version)),
        "profile_ok": bool(profile),
        "section_index_ok": section_count > 0 or stats["sqlite_chunks"] <= 1,
        "parse_quality_ok": parse_quality_score >= float(config.MIN_PARSE_QUALITY_SCORE),
    }
    gate["ready"] = all(gate.values())
    gate["counts"] = {
        "sqlite_chunks": stats["sqlite_chunks"],
        "fts_chunks": stats["fts_chunks"],
        "milvus_vectors": milvus_count,
        "sections": section_count,
    }
    return gate


def _build_upload_response(task_id: Optional[str], source: str, task_status: str, document_status: str, searchable: bool, duplicate_state: Optional[str] = None, duplicate_of: Optional[str] = None, same_title_candidates: Optional[List[str]] = None) -> Dict[str, Any]:
    doc = _doc_get(source)
    return {
        "task_id": task_id,
        "filename": source,
        "task_status": task_status,
        "document_status": document_status,
        "searchable": bool(searchable),
        "active_version": doc.get("active_version"),
        "pending_version": doc.get("pending_version"),
        "duplicate_state": duplicate_state,
        "duplicate_of": duplicate_of,
        "same_title_candidates": list(same_title_candidates or []),
    }


def _doc_searchable_flag(source: str) -> int:
    doc = _doc_get(source)
    if doc.get("status") in {"deleting", "pending_delete", "delete_failed"}:
        return 0
    return 1 if (doc.get("active_version") is not None or bool(doc.get("searchable"))) else 0


def _docfts_upsert(source: str, title: Optional[str] = None, aliases: Optional[str] = None, doc_type: Optional[str] = None, topic: Optional[str] = None):
    conn = _lex_db_connect()
    had_outer_tx = bool(getattr(conn, "in_transaction", False))
    stem = _filename_stem(source)
    conn.execute("DELETE FROM documents_fts WHERE filename = ?", (source,))
    conn.execute("INSERT INTO documents_fts(filename, title, aliases, doc_type, topic, filename_stem) VALUES (?,?,?,?,?,?)", (source, title or stem, aliases or "", doc_type or "", topic or "", stem))
    _lex_commit_if_needed(conn, had_outer_tx)
def _doc_recall_indexed(query: str, limit: int) -> List[str]:
    plan = _build_doc_recall_plan(query, int(limit))
    if plan:
        return [entry["source"] for entry in plan if entry.get("source")]
    return _doc_recall_fallback(query, int(limit))

def _verify_version_visible(source: str, version: int) -> bool:
    try:
        vector_db = VectorDBService()
        vector_db.connect()
        resp = vector_db.client.query(
            collection_name=vector_db.collection_name,
            filter=f"source == '{_safe_filename(source)}'",
            output_fields=["text", "source", "metadata"],
            limit=10_000
        )
        for r in resp or []:
            md = r.get("metadata") or {}
            if (md.get("doc_version") == version):
                return True
        return False
    except Exception:
        return False

def _cleanup_old_versions(source: str, keep_version: int):
    try:
        vector_db = VectorDBService()
        vector_db.connect()
        resp = vector_db.client.query(
            collection_name=vector_db.collection_name,
            filter=f"source == '{_safe_filename(source)}'",
            output_fields=["metadata", "id"],
            limit=10_000
        )
        ids_to_delete = []
        for r in resp or []:
            md = r.get("metadata") or {}
            dv = md.get("doc_version")
            if dv is not None and dv != keep_version:
                pk = r.get("id") or r.get("pk") or r.get("primary_key")
                if pk is not None:
                    ids_to_delete.append(pk)
        if ids_to_delete:
            BATCH = 1000
            deleted = 0
            for i in range(0, len(ids_to_delete), BATCH):
                batch_ids = ids_to_delete[i:i+BATCH]
                try:
                    vector_db.client.delete(
                        collection_name=vector_db.collection_name,
                        ids=batch_ids
                    )
                    deleted += len(batch_ids)
                except Exception as e:
                    logger.warning(f"cleanup_delete_batch_failed: source={source} batch_len={len(batch_ids)} err={e}")
            logger.info(f"cleanup_old_versions_done: source={source} keep_version={keep_version} deleted={deleted}")
        else:
            logger.info(f"cleanup_old_versions_skip: source={source} keep_version={keep_version} no_old_ids_found")
    except Exception:
        pass


def _milvus_row_pk(row: Dict[str, Any]) -> Optional[int]:
    for key in ("id", "pk", "primary_key"):
        value = row.get(key)
        if value is not None:
            try:
                return int(value)
            except Exception:
                return None
    return None


def _find_milvus_container_residuals(vector_db: Any, source: str, limit: int = 2000) -> List[Dict[str, Any]]:
    safe = _safe_filename(source)
    if not safe:
        return []
    pattern = f"%{safe}%"
    rows = vector_db.client.query(
        collection_name=vector_db.collection_name,
        filter=f"source != {json.dumps(safe, ensure_ascii=False)} and text like {json.dumps(pattern, ensure_ascii=False)}",
        output_fields=["id", "source", "metadata", "text"],
        limit=max(1, int(limit)),
    )
    out: List[Dict[str, Any]] = []
    seen = set()
    for row in rows or []:
        pk = _milvus_row_pk(row)
        if pk is None or pk in seen:
            continue
        seen.add(pk)
        out.append(row)
    return out


def _delete_milvus_document_object(vector_db: Any, source: str) -> Dict[str, Any]:
    safe = _safe_filename(source)
    residual_rows = _find_milvus_container_residuals(
        vector_db,
        safe,
        limit=int(os.getenv("MILVUS_DELETE_RELATED_LIMIT", "5000")),
    )
    residual_ids = [_milvus_row_pk(row) for row in residual_rows]
    residual_ids = [rid for rid in residual_ids if rid is not None]
    residual_sources = sorted({(row.get("source") or "").strip() for row in residual_rows if (row.get("source") or "").strip()})

    vector_db.client.delete(
        collection_name=vector_db.collection_name,
        filter=f"source == {json.dumps(safe, ensure_ascii=False)}",
    )

    if residual_ids:
        batch_size = 1000
        for start in range(0, len(residual_ids), batch_size):
            batch_ids = residual_ids[start:start + batch_size]
            vector_db.client.delete(
                collection_name=vector_db.collection_name,
                ids=batch_ids,
            )

    return {
        "exact_source": safe,
        "residual_ids_deleted": len(residual_ids),
        "residual_sources": residual_sources,
    }


def _vector_finalize(source: str):
    safe = _safe_filename(source)
    try:
        vector_db = VectorDBService()
        vector_db.connect()
        resp = vector_db.client.query(
            collection_name=vector_db.collection_name,
            filter=f"source == '{safe}'",
            output_fields=["source"],
            limit=1
        )
        if resp:
            _lex_db_set_status(safe, "completed")
        else:
            _lex_db_set_status(safe, "vector_failed")
    except Exception:
        _lex_db_set_status(safe, "vector_failed")

def _get_chunks_for_source(source: str, doc_version: Optional[int] = None) -> List[Dict[str, Any]]:
    target_version = doc_version
    if target_version is None:
        target_version = _get_active_version(source)
        if target_version is None:
            cur = _doc_get(source)
            target_version = cur.get("pending_version")
    if target_version is not None:
        try:
            target_version = int(target_version)
        except Exception:
            target_version = None
    document_ir = _ensure_document_ir(source, target_version or 1)
    if document_ir and (document_ir.get("elements") or []):
        items = _document_ir_to_structured_items(document_ir, config.CHUNK_SIZE, config.OVERLAP)
        out = []
        for item in items:
            out.append({
                "text": item.get("text") or "",
                "raw_text": item.get("raw_text") or item.get("text") or "",
                "section": item.get("section") or "",
                "chunk_id": int(item.get("chunk_id") or 0),
                "metadata": {
                    "chunk_id": int(item.get("chunk_id") or 0),
                    "section": item.get("section") or "",
                    "section_title": item.get("section_title") or item.get("section") or "",
                    "section_node_id": item.get("section_node_id"),
                    "raw_text": item.get("raw_text") or item.get("text") or "",
                    "text_normalized": item.get("normalized_text") or _normalize_ir_text(item.get("raw_text") or item.get("text") or ""),
                    "page_no": item.get("page_no"),
                    "page_span": item.get("page_span") or [],
                    "section_path": item.get("section_path") or [],
                    "parent_section_id": item.get("parent_section_id"),
                    "parent_section_path": item.get("parent_section_path") or [],
                    "parent_section_title": item.get("parent_section_title"),
                    "section_depth": item.get("section_depth"),
                    "semantic_unit_ids": item.get("semantic_unit_ids") or [],
                    "chunk_role": item.get("chunk_role") or "body",
                    "payload": item.get("payload") or {},
                    "element_id": item.get("element_id"),
                    "element_type": item.get("element_type"),
                    "reading_order": item.get("reading_order"),
                    "doc_version": document_ir.get("doc_version") or target_version or (_get_active_version(source) or 1),
                },
            })
        if out:
            return out
    conn = _lex_db_connect()
    rows = conn.execute("SELECT id, section, metadata FROM chunks_meta WHERE source = ? ORDER BY chunk_id ASC", (source,)).fetchall()
    fallback_version = target_version or _get_active_version(source) or 1
    out = []
    for rid, section, metadata in rows:
        trow = conn.execute("SELECT text FROM chunks_fts WHERE rowid = ?", (rid,)).fetchone()
        text = trow[0] if trow else ""
        md = {}
        try:
            md = json.loads(metadata or "{}")
        except Exception:
            md = {}
        md["section"] = section or ""
        row_version = md.get("doc_version")
        try:
            row_version = int(row_version) if row_version is not None else None
        except Exception:
            row_version = None
        if target_version is not None and row_version not in {None, int(target_version)}:
            continue
        if row_version is None:
            md["doc_version"] = fallback_version
        out.append({
            "text": text or "",
            "raw_text": text or "",
            "section": section or "",
            "chunk_id": int(md.get("chunk_id") or 0),
            "metadata": md,
        })
    return out

async def _rebuild_vectors_for_source(source: str) -> bool:
    """补偿：从 SQLite 读取文本并向 Milvus upsert；成功返回 True"""
    try:
        cur = _doc_get(source)
        target_version = cur.get("pending_version") if cur.get("pending_version") is not None else cur.get("active_version")
        chunks = _get_chunks_for_source(source, target_version)
        if not chunks:
            return False
        items = _contextualize_chunk_items(
            source,
            [
                {
                    "chunk_id": c.get("chunk_id", idx),
                    "section": c.get("section") or (c.get("metadata") or {}).get("section") or "",
                    "text": c.get("raw_text") or c.get("text") or "",
                }
                for idx, c in enumerate(chunks)
            ],
        )
        if not items:
            return False
        texts = [c["text"] for c in items]
        embedding_service = EmbeddingService()
        vector_db = VectorDBService()
        embs = await embedding_service.embed_batched(texts, per_request=32, timeout=60, retries=2)
        vector_db.connect()
        _delete_milvus_source_version(source, target_version)
        docs = []
        now = datetime.now().isoformat()
        for c, item, emb in zip(chunks, items, embs):
            vector_metadata = _milvus_safe_metadata({
                **(c["metadata"] or {}),
                "raw_text": item.get("raw_text") or c.get("raw_text") or c.get("text") or "",
            })
            docs.append({
                "embedding": emb,
                "text": item["text"],
                "source": source,
                "metadata": vector_metadata,
                "created_at": now
            })
        vector_db.insert(docs)
        return True
    except Exception as e:
        logger.warning(f"vector_compensation_failed: source={source} err={e}")
        return False

async def _compensation_worker(interval_sec: int = 5):
    """后台补偿：处理 vector_pending/vector_failed"""
    await asyncio.sleep(2)
    while True:
        try:
            conn = _lex_db_connect()
            pend = conn.execute("SELECT source FROM doc_status WHERE status IN ('vector_pending','vector_failed') LIMIT 20").fetchall()
            for (src,) in pend:
                ok = await _rebuild_vectors_for_source(src)
                cur = _doc_get(src)
                v_next = cur.get("pending_version")
                if ok and v_next is not None:
                    publish_gate = _build_publish_gate(src, int(v_next))
                    if publish_gate.get("ready"):
                        _doc_upsert(src, status="completed", active_version=int(v_next), pending_version=None, last_error=None, searchable=1, publish_gate=_json_dumps(publish_gate))
                        _lex_db_set_status(src, "completed")
                        # async cleanup old versions (best-effort)
                        _cleanup_old_versions(src, int(v_next))
                    else:
                        _doc_upsert(src, status="publish_failed", last_error="publish gate not ready", searchable=_doc_searchable_flag(src), publish_gate=_json_dumps(publish_gate))
                        _lex_db_set_status(src, "publish_failed")
                else:
                    _doc_upsert(src, status="vector_failed", last_error="vector upsert or verify failed", searchable=_doc_searchable_flag(src))
                    _lex_db_set_status(src, "vector_failed")
            deletions = _pending_delete_due(limit=10)
            for item in deletions:
                src = _safe_filename(item.get("source") or "")
                if not src:
                    continue
                lock = _get_source_lock(src)
                if not lock.acquire(blocking=False):
                    continue
                try:
                    vector_db = VectorDBService()
                    vector_db.connect()
                    vector_delete_info = _delete_milvus_document_object(vector_db, src)
                    _lex_db_delete_source(src)
                    file_cleanup = _delete_uploaded_artifacts(src) if item.get("delete_files", True) else {"removed": [], "missing": [], "failed": []}
                    if file_cleanup.get("failed"):
                        _reschedule_pending_delete(src, int(item.get("retry_count", 0)) + 1, "artifact_cleanup_failed")
                        continue
                    _complete_pending_delete(src)
                    logger.info(
                        "pending_delete_completed: source=%s residual_ids=%s files_removed=%s",
                        src,
                        vector_delete_info.get("residual_ids_deleted"),
                        len(file_cleanup.get("removed") or []),
                    )
                except Exception as e:
                    _doc_upsert(src, status="pending_delete", last_error=str(e))
                    _lex_db_set_status(src, "pending_delete")
                    _reschedule_pending_delete(src, int(item.get("retry_count", 0)) + 1, str(e))
                finally:
                    try:
                        lock.release()
                    except Exception:
                        pass
        except Exception as e:
            logger.warning(f"compensation_worker_error: {e}")
        await asyncio.sleep(interval_sec)
def _extract_filename_candidates(query: str) -> List[str]:
    q = (query or "")
    pat = r'([A-Za-z0-9_\-\(\)（）\u4e00-\u9fa5\.]+?\.(?:pdf|docx|xlsx|txt|md|markdown|csv|json|log))(?![A-Za-z0-9_\-\(\)（）\u4e00-\u9fa5])'
    return list(dict.fromkeys(re.findall(pat, q, flags=re.IGNORECASE)))

def _normalize_filename_for_match(name: str) -> str:
    s = (name or "").strip().replace("\\", "/")
    s = s.split("/")[-1]
    return s


# ==================== 向量数据库服务 ====================

class VectorDBService:
    """Milvus 向量数据库服务"""
    
    def __init__(self):
        self.collection_name = "rag_documents"
        self.collection = None
    
    def connect(self):
        """连接 Milvus"""
        from pymilvus import MilvusClient, FieldSchema, CollectionSchema, DataType
        
        self.client = MilvusClient(
            uri=f"http://{config.MILVUS_HOST}:{config.MILVUS_PORT}",
            user=config.MILVUS_USER,
            password=config.MILVUS_PASSWORD,
            secure=config.MILVUS_SECURE
        )
        
        # 检查集合是否存在
        if not self.client.has_collection(collection_name=self.collection_name):
            # 创建集合
            fields = [
                FieldSchema(name="id", dtype=DataType.INT64, is_primary=True, auto_id=True),
                FieldSchema(name="embedding", dtype=DataType.FLOAT_VECTOR, dim=1024),
                FieldSchema(name="text", dtype=DataType.VARCHAR, max_length=65535),
                FieldSchema(name="source", dtype=DataType.VARCHAR, max_length=1024),
                FieldSchema(name="metadata", dtype=DataType.JSON),
                FieldSchema(name="created_at", dtype=DataType.VARCHAR, max_length=64),
            ]
            schema = CollectionSchema(fields, "RAG Documents Collection")
            self.client.create_collection(
                collection_name=self.collection_name,
                schema=schema
            )

        if not self.client.list_indexes(collection_name=self.collection_name):
            index_params = MilvusClient.prepare_index_params()
            index_params.add_index(
                field_name="embedding",
                index_name="embedding_idx",
                index_type="HNSW",
                metric_type="COSINE",
                params={"M": 8, "efConstruction": 200}
            )
            self.client.create_index(
                collection_name=self.collection_name,
                index_params=index_params
            )

        self.client.load_collection(self.collection_name)

    def insert(self, documents: List[Dict]):
        """插入文档"""
        if not hasattr(self, 'client'):
            self.connect()

        self.client.insert(
            collection_name=self.collection_name,
            data=documents
        )
        _crash_inject("after_milvus_insert")

    def search(self, query_embedding: List[float], top_k: int, filters: dict = None):
        """搜索文档"""
        if not hasattr(self, 'client'):
            self.connect()

        search_params = {
            "metric_type": "COSINE",
            "params": {"ef": 100}
        }

        results = self.client.search(
            collection_name=self.collection_name,
            data=[query_embedding],
            limit=top_k,
            output_fields=["text", "source", "metadata"],
            search_params=search_params,
            filter=filters if filters else None
        )

        return results[0] if results else []


def _milvus_source_stats() -> Dict[str, Dict[str, Any]]:
    """按 source 聚合 Milvus 中已存在的 chunk 数与最新 created_at（最佳努力）。"""
    try:
        from pymilvus import Collection, connections

        uri = f"http://{config.MILVUS_HOST}:{config.MILVUS_PORT}"
        connections.connect(
            alias="list_documents",
            uri=uri,
            user=config.MILVUS_USER,
            password=config.MILVUS_PASSWORD,
            secure=config.MILVUS_SECURE,
        )
        collection = Collection("rag_documents", using="list_documents")
        collection.load()
        iterator = collection.query_iterator(
            batch_size=2000,
            limit=-1,
            expr="id >= 0",
            output_fields=["source", "created_at", "metadata"],
        )
        out: Dict[str, Dict[str, Any]] = {}
        try:
            while True:
                batch = iterator.next()
                if not batch:
                    break
                for row in batch:
                    src = _normalize_filename_for_match(row.get("source") or "")
                    if not src:
                        continue
                    item = out.setdefault(src, {"chunks_indexed": 0, "created_at": row.get("created_at")})
                    item["chunks_indexed"] += 1
                    created_at = row.get("created_at") or ""
                    if created_at and created_at > (item.get("created_at") or ""):
                        item["created_at"] = created_at
        finally:
            iterator.close()
        return out
    except Exception as e:
        logger.warning(f"milvus_source_stats_failed: {e}")
        return {}


def _normalize_topics(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, str):
        raw = [x.strip() for x in value.split(",")]
        return [x for x in raw if x]
    if isinstance(value, (list, tuple, set)):
        out = []
        for item in value:
            s = str(item or "").strip()
            if s:
                out.append(s)
        return out
    s = str(value).strip()
    return [s] if s else []


def _count_pick_top(counter: Dict[str, int]) -> Optional[str]:
    if not counter:
        return None
    return sorted(counter.items(), key=lambda x: (-x[1], x[0]))[0][0]


def _build_milvus_source_catalog(rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    catalog: Dict[str, Dict[str, Any]] = {}
    for idx, row in enumerate(rows):
        source = _normalize_filename_for_match(row.get("source") or "")
        if not source:
            continue
        metadata = row.get("metadata") or {}
        if not isinstance(metadata, dict):
            metadata = {}
        created_at = row.get("created_at") or ""
        text = row.get("text") or ""
        item = catalog.setdefault(
            source,
            {
                "source": source,
                "created_at": created_at,
                "rows": [],
                "doc_versions": set(),
                "doc_type_counts": {},
                "topic_counts": {},
            },
        )
        if created_at and created_at > (item.get("created_at") or ""):
            item["created_at"] = created_at
        chunk_id = metadata.get("chunk_id")
        try:
            chunk_id = int(chunk_id)
        except Exception:
            chunk_id = idx
        row_item = {
            "text": text,
            "metadata": metadata,
            "created_at": created_at,
            "chunk_id": chunk_id,
        }
        item["rows"].append(row_item)
        doc_version = metadata.get("doc_version")
        try:
            if doc_version is not None:
                item["doc_versions"].add(int(doc_version))
        except Exception:
            pass
        doc_type = str(metadata.get("doc_type") or "").strip()
        if doc_type:
            item["doc_type_counts"][doc_type] = item["doc_type_counts"].get(doc_type, 0) + 1
        for topic in _normalize_topics(metadata.get("topics")):
            item["topic_counts"][topic] = item["topic_counts"].get(topic, 0) + 1

    out: Dict[str, Dict[str, Any]] = {}
    for source, item in catalog.items():
        rows_sorted = sorted(item["rows"], key=lambda x: (x.get("chunk_id", 0), x.get("created_at") or ""))
        topics = [k for k, _ in sorted(item["topic_counts"].items(), key=lambda x: (-x[1], x[0]))]
        versions = sorted(item["doc_versions"])
        active_version = versions[-1] if versions else (1 if rows_sorted else None)
        out[source] = {
            "source": source,
            "created_at": item.get("created_at"),
            "chunks_indexed": len(rows_sorted),
            "rows": rows_sorted,
            "active_version": active_version,
            "doc_type": _count_pick_top(item["doc_type_counts"]),
            "topics": topics,
        }
    return out


def _milvus_source_catalog(sources: Optional[List[str]] = None) -> Dict[str, Dict[str, Any]]:
    """读取 Milvus 全量/指定 source 的文本与元数据，用于正式回填 SQLite。"""
    try:
        from pymilvus import Collection, connections

        uri = f"http://{config.MILVUS_HOST}:{config.MILVUS_PORT}"
        connections.connect(
            alias="reconcile_sqlite",
            uri=uri,
            user=config.MILVUS_USER,
            password=config.MILVUS_PASSWORD,
            secure=config.MILVUS_SECURE,
        )
        collection = Collection("rag_documents", using="reconcile_sqlite")
        collection.load()
        rows: List[Dict[str, Any]] = []
        if sources:
            for source in sources:
                safe_source = _safe_filename(source)
                rows.extend(
                    collection.query(
                        expr=f'source == "{safe_source}"',
                        output_fields=["source", "text", "metadata", "created_at"],
                        limit=20_000,
                    )
                    or []
                )
        else:
            iterator = collection.query_iterator(
                batch_size=1000,
                limit=-1,
                expr="id >= 0",
                output_fields=["source", "text", "metadata", "created_at"],
            )
            try:
                while True:
                    batch = iterator.next()
                    if not batch:
                        break
                    rows.extend(batch)
            finally:
                iterator.close()
        return _build_milvus_source_catalog(rows)
    except Exception as e:
        logger.warning(f"milvus_source_catalog_failed: {e}")
        return {}


def _sqlite_source_set(conn: Optional[sqlite3.Connection] = None) -> set[str]:
    db = conn or _lex_db_connect()
    out: set[str] = set()
    queries = (
        "SELECT source FROM documents",
        "SELECT source FROM doc_status",
        "SELECT source FROM chunks_meta",
        "SELECT source FROM document_ir_meta",
        "SELECT source FROM document_ir",
        "SELECT filename FROM documents_fts",
    )
    for sql in queries:
        try:
            out.update([(r[0] or "") for r in db.execute(sql).fetchall() if (r[0] or "").strip()])
        except Exception:
            pass
    return {_normalize_filename_for_match(x) for x in out if x}


def _reconcile_sqlite_from_catalog(catalog: Dict[str, Dict[str, Any]], prune_sqlite_orphans: bool = False) -> Dict[str, Any]:
    conn = _lex_db_connect()
    sources = sorted([_normalize_filename_for_match(x) for x in catalog.keys() if x])
    before_sources = _sqlite_source_set(conn)
    pruned_sources: List[str] = []
    upserted_sources: List[str] = []
    had_outer_tx = bool(getattr(conn, "in_transaction", False))
    if not had_outer_tx:
        _lex_tx_begin()
    try:
        if prune_sqlite_orphans:
            for source in sorted(before_sources - set(sources)):
                _lex_db_delete_source(source)
                pruned_sources.append(source)

        for source in sources:
            item = catalog[source]
            _lex_db_delete_source(source)
            for row in item.get("rows") or []:
                metadata = row.get("metadata") or {}
                if not isinstance(metadata, dict):
                    metadata = {}
                if metadata.get("doc_version") is None and item.get("active_version") is not None:
                    metadata = {**metadata, "doc_version": int(item["active_version"])}
                _lex_db_add_chunk_sql(
                    source,
                    _fts_storage_text(row.get("text") or "", metadata),
                    metadata.get("section") or "",
                    metadata,
                    int(row.get("chunk_id") or 0),
                )
            title_profile = _doc_title_profile(source)
            topic = ((item.get("topics") or [None])[0])
            _doc_upsert(
                source,
                status="completed",
                active_version=item.get("active_version"),
                pending_version=None,
                last_error=None,
                canonical_title=title_profile["canonical_title"],
                title_tokens=title_profile["title_tokens"],
                aliases=title_profile["aliases"],
                filename_stem=title_profile["stem"],
                doc_type=item.get("doc_type"),
                topic=topic,
            )
            _docfts_upsert(source, title=title_profile["canonical_title"], aliases=title_profile["aliases"], doc_type=item.get("doc_type"), topic=topic)
            _lex_db_set_status(source, "completed")
            upserted_sources.append(source)

        if not had_outer_tx:
            _lex_tx_commit()
        _lex_db_checkpoint("PASSIVE")
    except Exception:
        if not had_outer_tx:
            _lex_tx_rollback()
        raise

    after_sources = _sqlite_source_set(conn)
    return {
        "milvus_source_count": len(sources),
        "sqlite_source_count_before": len(before_sources),
        "sqlite_source_count_after": len(after_sources),
        "upserted_sources": len(upserted_sources),
        "pruned_sources": pruned_sources,
        "healthy": after_sources == set(sources),
    }


def reconcile_sqlite_with_milvus(prune_sqlite_orphans: bool = False, sources: Optional[List[str]] = None) -> Dict[str, Any]:
    catalog = _milvus_source_catalog(sources=sources)
    if not catalog and not sources:
        return {
            "milvus_source_count": 0,
            "sqlite_source_count_before": len(_sqlite_source_set()),
            "sqlite_source_count_after": len(_sqlite_source_set()),
            "upserted_sources": 0,
            "pruned_sources": [],
            "healthy": False,
            "error": "no_milvus_sources_found",
        }
    report = _reconcile_sqlite_from_catalog(catalog, prune_sqlite_orphans=prune_sqlite_orphans)
    report["sources"] = sorted(catalog.keys())
    return report


# ==================== 嵌入服务 ====================

class EmbeddingService:
    """嵌入服务"""
    
    async def embed(self, texts: List[str]) -> List[List[float]]:
        """生成文本嵌入"""
        import requests
        
        response = requests.post(
            f"{config.EMBEDDING_URL}/embed",
            json={
                "texts": texts,
                "normalize": True,
                "batch_size": 32
            },
            timeout=30
        )
        response.raise_for_status()
        
        return response.json()["embeddings"]

    async def embed_batched(self, texts: List[str], per_request: int = 64, timeout: int = 60, retries: int = 2) -> List[List[float]]:
        import requests, time
        out: List[List[float]] = []
        n = len(texts)
        i = 0
        while i < n:
            batch = texts[i: i + per_request]
            attempt = 0
            while True:
                try:
                    resp = requests.post(
                        f"{config.EMBEDDING_URL}/embed",
                        json={"texts": batch, "normalize": True, "batch_size": min(32, per_request)},
                        timeout=timeout
                    )
                    resp.raise_for_status()
                    embs = resp.json().get("embeddings") or []
                    out.extend(embs)
                    break
                except Exception as e:
                    if attempt >= retries:
                        raise e
                    time.sleep(min(2 ** attempt, 4))
                    attempt += 1
            i += per_request
        return out


# ==================== 重排序服务 ====================

class RerankService:
    """重排序服务"""
    
    async def rerank(self, query: str, documents: List[str], top_k: int) -> List[Dict]:
        """重排序文档"""
        import requests
        
        response = requests.post(
            f"{config.RERANK_URL}/rerank",
            json={
                "query": query,
                "documents": documents,
                "top_n": top_k
            },
            timeout=30
        )
        response.raise_for_status()
        
        return response.json()["results"]


# ==================== 查询处理 ====================

class QueryHandler:
    """查询处理器"""
    
    def __init__(self):
        self.vector_db = VectorDBService()
        self.embedding_service = EmbeddingService()
        self.rerank_service = RerankService()
        self._purify_cache: Dict[str, str] = {}

    async def generate_clarification(self, query: str, candidate_sources: List[str], reason: str = "document_target_required") -> str:
        prompt = _build_retrieval_grounded_clarification_prompt(query, candidate_sources, reason=reason)
        if not (config.LLM_CHAT_COMPLETIONS_URL or config.LLM_API_BASE):
            return ""

        def _chat_url_candidates() -> List[str]:
            if config.LLM_CHAT_COMPLETIONS_URL:
                return [config.LLM_CHAT_COMPLETIONS_URL]
            base = (config.LLM_API_BASE or "").rstrip("/")
            candidates = []
            if base:
                candidates.append(f"{base}/chat/completions")
                if not base.endswith("/v1"):
                    candidates.append(f"{base}/v1/chat/completions")
                if base.endswith("/v1"):
                    candidates.append(f"{base[:-3].rstrip('/')}/chat/completions")
            return candidates

        def _extra_body() -> Dict[str, Any]:
            raw = (config.LLM_EXTRA_BODY or "").strip()
            if not raw:
                return {}
            try:
                obj = json.loads(raw)
                return obj if isinstance(obj, dict) else {}
            except Exception:
                return {}

        def _call_llm() -> str:
            import requests

            headers = {"Content-Type": "application/json"}
            if config.LLM_API_KEY:
                headers["Authorization"] = f"Bearer {config.LLM_API_KEY}"
            payload = {
                "model": config.LLM_MODEL,
                "messages": [
                    {"role": "system", "content": "你是一个专业的AI助手"},
                    {"role": "user", "content": prompt},
                ],
                "temperature": min(float(config.LLM_TEMPERATURE), 0.3),
                "top_p": config.LLM_TOP_P,
                "max_tokens": 180,
                "presence_penalty": 0.0,
            }
            extra = _extra_body()
            if extra:
                for k, v in extra.items():
                    if k not in payload:
                        payload[k] = v
            last_exc = None
            for url in _chat_url_candidates():
                try:
                    resp = requests.post(url, headers=headers, json=payload, timeout=config.LLM_TIMEOUT)
                    if resp.status_code == 404:
                        last_exc = HTTPException(status_code=404, detail="LLM endpoint 404")
                        continue
                    resp.raise_for_status()
                    data = resp.json()
                    choices = data.get("choices") or []
                    if choices and isinstance(choices, list):
                        msg = (choices[0] or {}).get("message") or {}
                        content_val = msg.get("content")
                        content = (content_val or "").strip() if isinstance(content_val, str) else ""
                        if not content:
                            reasoning_val = msg.get("reasoning")
                            content = (reasoning_val or "").strip() if isinstance(reasoning_val, str) else ""
                        if not content:
                            text_val = (choices[0] or {}).get("text")
                            content = (text_val or "").strip() if isinstance(text_val, str) else ""
                        if content:
                            return content
                except Exception as exc:
                    last_exc = exc
                    continue
            raise last_exc or HTTPException(status_code=500, detail="LLM clarification failed")

        try:
            return await asyncio.to_thread(_call_llm)
        except Exception as exc:
            logger.error(f"Clarification generation error: {str(exc)}")
            return ""

    async def _build_rule_backed_clarification(self, query: str, reason: str, seed_sources: Optional[List[str]] = None) -> Dict[str, Any]:
        limit = max(1, int(_policy_get("source_resolution.clarification_examples_limit", 3)))
        candidates = _retrieval_backed_clarification_candidates(query, seed_sources=seed_sources, limit=limit)
        template = _build_document_clarification_prompt(candidates)
        if not candidates:
            return {
                "message": template,
                "candidate_sources": [],
                "used_llm": False,
            }
        llm_message = (await self.generate_clarification(query, candidates, reason=reason)).strip()
        message = llm_message if llm_message and not llm_message.startswith("抱歉") else template
        return {
            "message": message,
            "candidate_sources": candidates,
            "used_llm": bool(llm_message),
        }

    async def _run_target_scoped_recall(
        self,
        query: str,
        retrieval_query: str,
        query_embedding: List[float],
        qtype: str,
        qfilters: Dict[str, Any],
        recall_k: int,
        final_n: int,
        pool_n: int,
        enable_rerank: bool,
        target_source: str,
        compare_subquery: Optional[Dict[str, str]] = None,
    ) -> Dict[str, Any]:
        safe_name = _normalize_filename_for_match(target_source)
        effective_query = _normalize_query((compare_subquery or {}).get("raw_text_query") or retrieval_query) or retrieval_query
        section_query = _normalize_query((compare_subquery or {}).get("section_query") or "")
        doc_prior_query = _normalize_query((compare_subquery or {}).get("doc_prior_query") or effective_query) or effective_query
        if _query_has_compare_intent(query):
            cleaned = _strip_compare_noise_terms(effective_query)
            if cleaned:
                effective_query = cleaned
            cleaned_prior = _strip_compare_noise_terms(doc_prior_query)
            if cleaned_prior:
                doc_prior_query = cleaned_prior
        effective_embedding = query_embedding
        if effective_query != retrieval_query:
            effective_embedding = (await self.embedding_service.embed([effective_query]))[0]
        milvus_filter = f"source == {json.dumps(safe_name, ensure_ascii=False)}"
        docs = self.vector_db.search(effective_embedding, top_k=recall_k, filters=milvus_filter)
        visible_dense = _filter_hits_by_source_state(docs)
        docs = visible_dense["hits"]
        dense_source_scores = _dense_source_score_map(docs)

        doc_recall_plan = _build_doc_recall_plan(doc_prior_query, limit=3, source_filter=safe_name)
        lex_items = _collect_lexical_candidates(effective_query, [safe_name], doc_recall_plan)
        if section_query and section_query != effective_query:
            try:
                lex_items.extend(
                    _lexical_recall_indexed(
                        section_query,
                        max(20, min(160, getattr(config, "LEXICAL_RECALL_LIMIT", 1000) // 5)),
                        source_filter=safe_name,
                    )
                )
            except Exception:
                try:
                    lex_items.extend(
                        _lexical_recall_fallback(
                            section_query,
                            max(20, min(160, getattr(config, "LEXICAL_RECALL_LIMIT", 1000) // 5)),
                            source_filter=safe_name,
                        )
                    )
                except Exception:
                    pass
        visible_lex = _filter_hits_by_source_state(lex_items)
        lex_items = visible_lex["hits"]
        docs_all = docs + lex_items
        if not docs_all:
            return {
                "source": safe_name,
                "evidence_query": effective_query,
                "section_query": section_query,
                "doc_prior_query": doc_prior_query,
                "score_mode": "score",
                "docs": [],
                "selected_docs": [],
                "post_filter_docs": [],
                "retrieve_docs": [],
                "early_filtered": visible_dense["dropped"] + visible_lex["dropped"],
                "visibility_filtered": visible_dense["dropped"] + visible_lex["dropped"],
                "dense_source_scores": dense_source_scores,
                "rerank_used": False,
                "doc_recall_plan": doc_recall_plan,
            }

        dense_rank_map: Dict[str, int] = {}
        for i, d in enumerate(docs):
            src = _normalize_filename_for_match(_hit_entity_source(d) or "")
            if src and src not in dense_rank_map:
                dense_rank_map[src] = i
        lex_rank_map: Dict[str, int] = {}
        for i, d in enumerate(lex_items):
            src = _normalize_filename_for_match(_hit_entity_source(d) or "")
            if src and src not in lex_rank_map:
                lex_rank_map[src] = i

        source_count: Dict[str, int] = {}
        for d in docs_all:
            src = _normalize_filename_for_match(_hit_entity_source(d) or "")
            if src:
                source_count[src] = source_count.get(src, 0) + 1

        weak_query = _is_weak_reference_query(query)
        source_signals = _build_source_signal_map(effective_query, lex_items, doc_recall_plan)
        combined = [(0.0, i) for i in range(len(docs_all))]
        fused_source_scores: Dict[str, float] = {}
        for i, d in enumerate(docs_all):
            src = _normalize_filename_for_match(_hit_entity_source(d) or "")
            if src not in fused_source_scores:
                fused_source_scores[src] = _fusion_source_score(
                    src,
                    effective_query,
                    dense_rank_map,
                    lex_rank_map,
                    source_count,
                    source_signals,
                    {safe_name},
                    set(),
                    weak_query,
                )
            combined[i] = (fused_source_scores[src], i)
        combined.sort(
            key=lambda item: (item[0], _source_dense_tiebreak_score(docs_all[item[1]], dense_source_scores)),
            reverse=True,
        )

        seen_keys = set()
        docs_fused = []
        for fused_score, idx in combined:
            hit = docs_all[idx]
            key = (_hit_entity_source(hit) or "unknown", (_hit_entity_text(hit) or "")[:64])
            if key in seen_keys:
                continue
            seen_keys.add(key)
            docs_fused.append(_clone_hit_with_score(hit, fused_score))
            if len(docs_fused) >= recall_k:
                break
        docs = docs_fused

        chunk_rerank_enabled = _should_apply_chunk_rerank(docs[:pool_n], dense_rank_map, lex_rank_map, source_signals, enable_rerank)
        reranked_chunk = await _chunk_level_rerank(self.rerank_service, query, docs[:pool_n], pool_n, chunk_rerank_enabled)
        docs = reranked_chunk["hits"]
        score_mode = reranked_chunk["score_mode"]
        heading_expansion_query = section_query or effective_query
        docs = _expand_heading_hits_to_article_hits(heading_expansion_query, safe_name, docs, limit=max(final_n, 6))

        merged_docs = _merge_and_dedupe_hits(docs, score_mode=score_mode)
        aggregated_docs = _aggregate_doc_sections(merged_docs, score_mode=score_mode)
        context_docs = _docs_for_query_context(qtype, merged_docs, aggregated_docs)
        retrieve_docs = _apply_retrieval_filters(docs, qfilters, [safe_name])
        retrieve_docs = _filter_low_relevance_sources(retrieve_docs, score_mode=score_mode, query=effective_query)
        retrieve_docs = _intra_doc_chunk_rerank(effective_query, retrieve_docs, score_mode=score_mode, qtype=qtype)
        filtered_docs = _apply_retrieval_filters(context_docs, qfilters, [safe_name])
        selected_docs = _filter_low_relevance_sources(filtered_docs, score_mode=score_mode, query=effective_query)
        selected_docs = _intra_doc_chunk_rerank(effective_query, selected_docs, score_mode=score_mode, qtype=qtype)
        post_filter_docs = selected_docs[:]
        selected_docs = selected_docs[: min(len(selected_docs), final_n)]

        return {
            "source": safe_name,
            "evidence_query": effective_query,
            "section_query": section_query,
            "doc_prior_query": doc_prior_query,
            "score_mode": score_mode,
            "docs": docs,
            "selected_docs": selected_docs,
            "post_filter_docs": post_filter_docs,
            "retrieve_docs": retrieve_docs,
            "early_filtered": visible_dense["dropped"] + visible_lex["dropped"],
            "visibility_filtered": visible_dense["dropped"] + visible_lex["dropped"],
            "dense_source_scores": dense_source_scores,
            "rerank_used": bool(reranked_chunk["used"]),
            "doc_recall_plan": doc_recall_plan,
        }

    async def _run_lightweight_recall(
        self,
        query: str,
        top_k: int,
        enable_rerank: bool,
        filename_hints: Optional[List[str]] = None,
        user_id: str = "anonymous",
    ) -> Dict[str, Any]:
        qtype = _classify_question_type(query)
        llm_parse: Dict[str, Any] = {}
        if bool(getattr(config, "ENABLE_LLM_QUERY_PARSE", True)):
            try:
                parsed = await asyncio.to_thread(_llm_parse_query_cached, query, "")
                if isinstance(parsed, dict):
                    llm_parse = parsed
            except Exception:
                llm_parse = {}
        query_explicit_fnames = [
            _normalize_filename_for_match(name or "")
            for name in _extract_filename_candidates(query)
            if _normalize_filename_for_match(name or "")
        ]
        query_explicit_set = set(query_explicit_fnames)
        source_resolution = await asyncio.to_thread(_resolve_query_target_sources, query, list(filename_hints or []), user_id)
        query_route = (source_resolution.get("route") or _classify_query_route(query, list(filename_hints or [])) or "content_qa")
        query_quality = _query_deep_quality_state(query, llm_parse=llm_parse, source_resolution=source_resolution)
        intent_tier = str(query_quality.get("tier") or "")
        if query_quality["reason"]:
            return {
                "query": query,
                "retrieval_query": _normalize_query(str(llm_parse.get("retrieval_query") or "")) or query,
                "retrieval_query_raw": query,
                "dense_query": _normalize_query(str(llm_parse.get("dense_query") or "")) or _normalize_query(str(llm_parse.get("retrieval_query") or "")) or query,
                "llm_parse": llm_parse,
                "is_comparison": bool(llm_parse.get("is_comparison")),
                "question_type": qtype,
                "score_mode": "score",
                "docs": [],
                "selected_docs": [],
                "qfilters": _query_filters(query),
                "recall_k": 0,
                "final_n": 0,
                "rerank_used": False,
                "query_route": query_quality["reason"],
                "weak_query": _is_weak_reference_query(query),
                "early_filtered": [],
                "visibility_filtered": [],
                "dense_source_scores": {},
                "post_filter_docs": [],
                "retrieve_docs": [],
                "source_lock_required": False,
                "resolved_source_lock": False,
                "target_sources": [],
                "source_lock_candidates": list(source_resolution.get("candidates") or []),
                "source_lock_reason": source_resolution.get("reason") or "",
                "clarification": source_resolution.get("clarification") or "",
                "target_text": source_resolution.get("target_text") or "",
                "lock_mode": "none",
                "lock_confidence": 0.0,
                "lock_message_prefix": "",
                "source_lock_kind": "",
                "source_resolution_trace": dict(source_resolution.get("source_resolution_trace") or {}),
                "inherited_from_context": bool(source_resolution.get("inherited_from_context")),
                "compare_subjects": list(source_resolution.get("compare_subjects") or []),
                "compare_doc_like_subjects": list(source_resolution.get("compare_doc_like_subjects") or []),
                "compare_missing_targets": list(source_resolution.get("compare_missing_targets") or []),
                "compare_common_aspects": list(source_resolution.get("compare_common_aspects") or []),
                "compare_topic_pair": list(source_resolution.get("compare_topic_pair") or []),
                "compare_canonical_aspects": list(source_resolution.get("compare_canonical_aspects") or []),
                "compare_expanded_aspects": list(source_resolution.get("compare_expanded_aspects") or []),
                "compare_source_subqueries": dict(source_resolution.get("compare_source_subqueries") or {}),
                "compare_status": source_resolution.get("compare_status") or "not_compare",
                "compare_plan": dict(source_resolution.get("compare_plan") or {}),
                "compare_source_results": [],
                "blocked_reason": query_quality["reason"],
                "query_quality": query_quality["quality"],
                "intent_tier": intent_tier,
            }
        compare_plan = dict(source_resolution.get("compare_plan") or {})
        is_comparison_hint = bool(compare_plan.get("is_compare")) or _query_has_compare_intent(query)
        clarification_limit = max(1, int(_policy_get("source_resolution.clarification_examples_limit", 3)))
        soft_clarification_eligible = (
            intent_tier in {"tier_2", "tier_3"}
            and not bool(source_resolution.get("resolved"))
            and not is_comparison_hint
            and (
                not bool(source_resolution.get("required"))
                or (
                    (source_resolution.get("route") or "") == "weak_title_reference"
                    and not list(source_resolution.get("candidates") or [])
                    and (source_resolution.get("reason") or "") == "document_target_required"
                )
            )
        )
        if soft_clarification_eligible:
            soft_candidates = _retrieval_backed_clarification_candidates(
                query,
                seed_sources=list(source_resolution.get("candidates") or source_resolution.get("sources") or []),
                limit=clarification_limit,
            )
            if soft_candidates:
                return {
                    "query": query,
                    "retrieval_query": _normalize_query(str(llm_parse.get("retrieval_query") or "")) or query,
                    "retrieval_query_raw": query,
                    "dense_query": _normalize_query(str(llm_parse.get("dense_query") or "")) or _normalize_query(str(llm_parse.get("retrieval_query") or "")) or query,
                    "llm_parse": llm_parse,
                    "is_comparison": bool(llm_parse.get("is_comparison")),
                    "question_type": qtype,
                    "score_mode": "score",
                    "docs": [],
                    "selected_docs": [],
                    "qfilters": _query_filters(query),
                    "recall_k": 0,
                    "final_n": 0,
                    "rerank_used": False,
                    "query_route": query_route,
                    "weak_query": _is_weak_reference_query(query),
                    "early_filtered": [],
                    "visibility_filtered": [],
                    "dense_source_scores": {},
                    "post_filter_docs": [],
                    "retrieve_docs": [],
                    "source_lock_required": False,
                    "resolved_source_lock": False,
                    "target_sources": [],
                    "source_lock_candidates": soft_candidates,
                    "source_lock_reason": source_resolution.get("reason") or "not_needed",
                    "clarification": source_resolution.get("clarification") or "",
                    "target_text": source_resolution.get("target_text") or "",
                    "lock_mode": "none",
                    "lock_confidence": 0.0,
                    "lock_message_prefix": "",
                    "source_lock_kind": "",
                    "source_resolution_trace": dict(source_resolution.get("source_resolution_trace") or {}),
                    "inherited_from_context": bool(source_resolution.get("inherited_from_context")),
                    "compare_subjects": list(source_resolution.get("compare_subjects") or []),
                    "compare_doc_like_subjects": list(source_resolution.get("compare_doc_like_subjects") or []),
                    "compare_missing_targets": list(source_resolution.get("compare_missing_targets") or []),
                    "compare_common_aspects": list(source_resolution.get("compare_common_aspects") or []),
                    "compare_topic_pair": list(source_resolution.get("compare_topic_pair") or []),
                    "compare_canonical_aspects": list(source_resolution.get("compare_canonical_aspects") or []),
                    "compare_expanded_aspects": list(source_resolution.get("compare_expanded_aspects") or []),
                    "compare_source_subqueries": dict(source_resolution.get("compare_source_subqueries") or {}),
                    "compare_status": source_resolution.get("compare_status") or "not_compare",
                    "compare_plan": dict(source_resolution.get("compare_plan") or {}),
                    "compare_source_results": [],
                    "blocked_reason": "",
                    "query_quality": query_quality["quality"],
                    "intent_tier": intent_tier,
                    "soft_clarification_required": True,
                    "soft_clarification_reason": "tier2_soft_confirm" if intent_tier == "tier_2" else "tier3_summary_clarification",
                }
        fnames = list(source_resolution.get("sources") or filename_hints or [])
        active_fnames = list(fnames)
        topical_multi_doc_mode = query_route == "multi_doc_query" and bool(active_fnames) and not bool(source_resolution.get("required"))
        retrieval_query = query
        retrieval_query_override = _normalize_query(source_resolution.get("retrieval_query_override") or "")
        if source_resolution.get("required") and not source_resolution.get("resolved"):
            if (source_resolution.get("reason") or "") == "document_ambiguous":
                candidates = [
                    _normalize_filename_for_match(x or "")
                    for x in (source_resolution.get("candidates") or [])
                    if _normalize_filename_for_match(x or "")
                ]
                candidates = list(dict.fromkeys(candidates))
                chosen = candidates[0] if candidates else ""
                if chosen:
                    title = _source_display_title(chosen) or chosen
                    if len(candidates) <= 1:
                        prefix = f"当前仅找到《{title}》，基于该条例为你解答：\n"
                    else:
                        prefix = f"当前匹配到多个可能的文件，先基于《{title}》为你解答（可能不完全准确）：\n"
                    source_resolution = {
                        **dict(source_resolution),
                        "resolved": True,
                        "sources": [chosen],
                        "reason": "document_ambiguous_soft_lock",
                        "lock_mode": "soft_lock",
                        "lock_confidence": 0.62 if len(candidates) <= 1 else 0.55,
                        "lock_message_prefix": prefix,
                        "source_lock_kind": "ambiguous_soft_lock",
                        "source_resolution_trace": {
                            **dict(source_resolution.get("source_resolution_trace") or {}),
                            "original_reason": "document_ambiguous",
                            "ambiguous_soft_lock": True,
                            "ambiguous_candidates": candidates[:5],
                            "chosen_source": chosen,
                        },
                    }
                    fnames = [chosen]
                    active_fnames = [chosen]
                    query_route = (source_resolution.get("route") or query_route or "content_qa")
                    topical_multi_doc_mode = False
                else:
                    pass
            if not source_resolution.get("resolved"):
                fallback_intent_tier = intent_tier or (
                    "tier_2" if _query_has_strong_business_signal(query) or _query_quality_strong_topic_terms(query) else (
                        "tier_3" if _query_has_weak_business_signal(query) else ""
                    )
                )
                if (
                    fallback_intent_tier in {"tier_2", "tier_3"}
                    and not list(source_resolution.get("candidates") or [])
                    and (source_resolution.get("reason") or "") == "document_target_required"
                ):
                    soft_candidates = _retrieval_backed_clarification_candidates(
                        query,
                        seed_sources=list(source_resolution.get("sources") or []),
                        limit=clarification_limit,
                    )
                    if soft_candidates:
                        return {
                            "query": query,
                            "retrieval_query": retrieval_query,
                            "retrieval_query_raw": query,
                            "dense_query": _normalize_query(str(llm_parse.get("dense_query") or "")) or _normalize_query(str(llm_parse.get("retrieval_query") or "")) or query,
                            "llm_parse": llm_parse,
                            "is_comparison": bool(is_comparison_hint),
                            "question_type": qtype,
                            "score_mode": "score",
                            "docs": [],
                            "selected_docs": [],
                            "qfilters": _query_filters(query),
                            "recall_k": 0,
                            "final_n": 0,
                            "rerank_used": False,
                            "query_route": query_route,
                            "weak_query": _is_weak_reference_query(query),
                            "early_filtered": [],
                            "visibility_filtered": [],
                            "dense_source_scores": {},
                            "post_filter_docs": [],
                            "retrieve_docs": [],
                            "source_lock_required": False,
                            "resolved_source_lock": False,
                            "target_sources": [],
                            "source_lock_candidates": soft_candidates,
                            "source_lock_reason": source_resolution.get("reason") or "document_target_required",
                            "clarification": source_resolution.get("clarification") or "",
                            "target_text": source_resolution.get("target_text") or "",
                            "compare_subjects": list(source_resolution.get("compare_subjects") or []),
                            "compare_doc_like_subjects": list(source_resolution.get("compare_doc_like_subjects") or []),
                            "compare_missing_targets": list(source_resolution.get("compare_missing_targets") or []),
                            "compare_common_aspects": list(source_resolution.get("compare_common_aspects") or []),
                            "compare_topic_pair": list(source_resolution.get("compare_topic_pair") or []),
                            "compare_canonical_aspects": list(source_resolution.get("compare_canonical_aspects") or []),
                            "compare_expanded_aspects": list(source_resolution.get("compare_expanded_aspects") or []),
                            "compare_source_subqueries": dict(source_resolution.get("compare_source_subqueries") or {}),
                            "compare_status": source_resolution.get("compare_status") or compare_plan.get("compare_status") or "not_compare",
                            "compare_plan": compare_plan,
                            "compare_source_results": [],
                            "lock_mode": "none",
                            "lock_confidence": 0.0,
                            "lock_message_prefix": "",
                            "source_lock_kind": "",
                            "source_resolution_trace": dict(source_resolution.get("source_resolution_trace") or {}),
                            "inherited_from_context": bool(source_resolution.get("inherited_from_context")),
                            "blocked_reason": "",
                            "query_quality": query_quality["quality"],
                            "intent_tier": fallback_intent_tier,
                            "soft_clarification_required": True,
                            "soft_clarification_reason": "tier2_soft_confirm" if fallback_intent_tier == "tier_2" else "tier3_summary_clarification",
                        }
                return {
                "query": query,
                "retrieval_query": retrieval_query,
                "is_comparison": bool(is_comparison_hint),
                "question_type": qtype,
                "score_mode": "score",
                "docs": [],
                "selected_docs": [],
                "qfilters": _query_filters(query),
                "recall_k": 0,
                "final_n": 0,
                "rerank_used": False,
                "query_route": query_route,
                "weak_query": _is_weak_reference_query(query),
                "early_filtered": [],
                "visibility_filtered": [],
                "dense_source_scores": {},
                "post_filter_docs": [],
                "retrieve_docs": [],
                "source_lock_required": True,
                "resolved_source_lock": False,
                "target_sources": [],
                "source_lock_candidates": list(source_resolution.get("candidates") or []),
                "source_lock_reason": source_resolution.get("reason") or "document_target_required",
                "clarification": source_resolution.get("clarification") or "",
                "target_text": source_resolution.get("target_text") or "",
                "compare_subjects": list(source_resolution.get("compare_subjects") or []),
                "compare_doc_like_subjects": list(source_resolution.get("compare_doc_like_subjects") or []),
                "compare_missing_targets": list(source_resolution.get("compare_missing_targets") or []),
                "compare_common_aspects": list(source_resolution.get("compare_common_aspects") or []),
                "compare_topic_pair": list(source_resolution.get("compare_topic_pair") or []),
                "compare_canonical_aspects": list(source_resolution.get("compare_canonical_aspects") or []),
                "compare_expanded_aspects": list(source_resolution.get("compare_expanded_aspects") or []),
                "compare_source_subqueries": dict(source_resolution.get("compare_source_subqueries") or {}),
                "compare_status": source_resolution.get("compare_status") or compare_plan.get("compare_status") or "not_compare",
                "compare_plan": compare_plan,
                "compare_source_results": [],
                "lock_mode": source_resolution.get("lock_mode") or "none",
                "lock_confidence": float(source_resolution.get("lock_confidence") or 0.0),
                "lock_message_prefix": source_resolution.get("lock_message_prefix") or "",
                "source_lock_kind": source_resolution.get("source_lock_kind") or "",
                "source_resolution_trace": dict(source_resolution.get("source_resolution_trace") or {}),
                "inherited_from_context": bool(source_resolution.get("inherited_from_context")),
                "query_quality": query_quality["quality"],
                "intent_tier": fallback_intent_tier,
                }
        if retrieval_query_override:
            retrieval_query = retrieval_query_override
        elif source_resolution.get("strip_title_mentions") and active_fnames:
            if source_resolution.get("reason") == "explicit_filename_unique":
                stripped_query = _strip_filename_mentions(query, active_fnames)
            else:
                stripped_query = _strip_source_title_mentions(query, active_fnames)
            if stripped_query:
                retrieval_query = stripped_query
        elif fnames and query_explicit_set:
            mentioned_fnames = [
                _normalize_filename_for_match(name or "")
                for name in fnames
                if _normalize_filename_for_match(name or "") in query_explicit_set
            ]
            if mentioned_fnames:
                stripped_query = _strip_filename_mentions(query, mentioned_fnames)
                if stripped_query:
                    retrieval_query = stripped_query
        elif not active_fnames:
            title_sources = _extract_title_source_candidates(query)
            if title_sources:
                active_fnames = title_sources
                stripped_query = _strip_source_title_mentions(query, title_sources)
                if stripped_query:
                    retrieval_query = stripped_query

        if is_comparison_hint and source_resolution.get("compare_missing_targets"):
            stripped_query = _strip_raw_text_mentions(retrieval_query, list(source_resolution.get("compare_missing_targets") or []))
            if stripped_query:
                retrieval_query = stripped_query

        if query_route == "single_doc_compare" and len(active_fnames) == 1:
            compare_subqueries = dict(source_resolution.get("compare_source_subqueries") or compare_plan.get("source_subqueries") or {})
            single_compare_query = _normalize_query((compare_subqueries.get(active_fnames[0]) or {}).get("raw_text_query") or "")
            if single_compare_query:
                retrieval_query = single_compare_query

        is_comparison = False
        if getattr(config, "ENABLE_COMPARE_INTENT_TAG", True):
            is_comparison = bool(compare_plan.get("is_compare")) or _query_has_compare_intent(query)
        if is_comparison:
            cleaned = _strip_compare_noise_terms(retrieval_query)
            if cleaned:
                retrieval_query = cleaned

        retrieval_query_raw = retrieval_query
        dense_query = retrieval_query
        qfilters = _query_filters(query)
        if llm_parse:
            rq = _normalize_query(str(llm_parse.get("retrieval_query") or ""))
            dq = _normalize_query(str(llm_parse.get("dense_query") or "")) or rq
            if rq and not retrieval_query_override:
                retrieval_query = rq
            if dq and not retrieval_query_override:
                dense_query = dq
            anchors = llm_parse.get("anchors")
            if isinstance(anchors, list) and anchors:
                qfilters["_llm_anchor_override"] = list(anchors)[: int(getattr(config, "QUERY_PARSE_MAX_ANCHORS", 1))]
            aspects = llm_parse.get("aspects")
            if isinstance(aspects, list) and aspects:
                qfilters["_llm_aspects_override"] = list(aspects)[: int(getattr(config, "QUERY_PARSE_MAX_ASPECTS", 4))]
            section_targets = llm_parse.get("section_targets")
            if isinstance(section_targets, list) and section_targets:
                qfilters["_llm_section_targets_override"] = list(section_targets)[: int(getattr(config, "QUERY_PARSE_MAX_SECTION_TARGETS", 4))]

        if len(_normalize_query(retrieval_query)) < max(2, int(getattr(config, "MIN_QUERY_CHARS", 2))):
            retrieval_query = retrieval_query_raw

        if is_comparison:
            cleaned_dense = _strip_compare_noise_terms(dense_query)
            if cleaned_dense:
                dense_query = cleaned_dense

        embeddings = await self.embedding_service.embed([dense_query])
        query_embedding = embeddings[0]

        requested_k = int(top_k or 10)
        recall_k = min(max(requested_k * 2, 20), min(config.TOP_K, int(getattr(config, "RETRIEVAL_CANDIDATE_K", config.RECALL_TOP_K))))
        final_n = min(max(config.FINAL_CONTEXT_N, 3), max(3, int(getattr(config, "FINAL_CONTEXT_N_MAX", 10))))
        pool_n = min(max(max(config.RERANK_KEEP_N, config.CHUNK_RERANK_KEEP_N), requested_k * 2), recall_k)
        if enable_rerank and bool(getattr(config, "ENABLE_CHUNK_RERANK", False)) and bool(getattr(config, "ENABLE_RERANK", True)):
            target_pool = int(getattr(config, "CHUNK_RERANK_POOL_N", 60))
            recall_k = min(max(recall_k, target_pool), min(int(config.TOP_K), int(getattr(config, "RETRIEVAL_CANDIDATE_K", recall_k))))
            pool_n = min(max(pool_n, min(recall_k, target_pool)), recall_k)

        if len(active_fnames) == 1:
            recall_k = min(max(recall_k, int(getattr(config, "LOCKED_DOC_RECALL_K", 60))), int(config.TOP_K))
            pool_n = min(max(pool_n, min(recall_k, requested_k * 3)), recall_k)

        if query_route == "multi_doc_compare" and len(active_fnames) >= 2:
            compare_sources = [
                _normalize_filename_for_match(name)
                for name in active_fnames
                if _normalize_filename_for_match(name)
            ]
            compare_sources = list(dict.fromkeys(compare_sources))
            compare_subqueries = dict(source_resolution.get("compare_source_subqueries") or compare_plan.get("source_subqueries") or {})
            compare_source_results = await asyncio.gather(*[
                self._run_target_scoped_recall(
                    query=query,
                    retrieval_query=retrieval_query,
                    query_embedding=query_embedding,
                    qtype=qtype,
                    qfilters=qfilters,
                    recall_k=recall_k,
                    final_n=final_n,
                    pool_n=pool_n,
                    enable_rerank=enable_rerank,
                    target_source=source,
                    compare_subquery=compare_subqueries.get(source) or None,
                )
                for source in compare_sources
            ])
            compare_final_n = max(final_n * max(1, len(compare_sources)), final_n)
            return {
                "query": query,
                "retrieval_query": retrieval_query,
                "retrieval_query_raw": retrieval_query_raw,
                "dense_query": dense_query,
                "llm_parse": llm_parse,
                "is_comparison": bool(is_comparison),
                "question_type": qtype,
                "score_mode": next((item.get("score_mode") for item in compare_source_results if item.get("score_mode")), "score"),
                "docs": _merge_compare_source_doc_groups(compare_source_results, per_source_limit=max(2, requested_k)),
                "selected_docs": _merge_compare_source_doc_groups(compare_source_results, per_source_limit=max(2, final_n)),
                "qfilters": qfilters,
                "recall_k": recall_k,
                "final_n": compare_final_n,
                "rerank_used": any(bool(item.get("rerank_used")) for item in compare_source_results),
                "query_route": query_route,
                "weak_query": _is_weak_reference_query(query),
                "early_filtered": [entry for item in compare_source_results for entry in (item.get("early_filtered") or [])],
                "visibility_filtered": [entry for item in compare_source_results for entry in (item.get("visibility_filtered") or [])],
                "dense_source_scores": {
                    key: value
                    for item in compare_source_results
                    for key, value in (item.get("dense_source_scores") or {}).items()
                },
                "post_filter_docs": _merge_compare_source_doc_groups(compare_source_results, per_source_limit=max(2, final_n)),
                "retrieve_docs": _merge_compare_source_doc_groups(compare_source_results, per_source_limit=max(2, requested_k)),
                "source_lock_required": False,
                "resolved_source_lock": True,
                "target_sources": compare_sources,
                "source_lock_candidates": compare_sources,
                "source_lock_reason": source_resolution.get("reason") or "",
                "clarification": source_resolution.get("clarification") or "",
                "target_text": source_resolution.get("target_text") or "",
                "compare_subjects": list(source_resolution.get("compare_subjects") or []),
                "compare_doc_like_subjects": list(source_resolution.get("compare_doc_like_subjects") or []),
                "compare_missing_targets": list(source_resolution.get("compare_missing_targets") or []),
                "compare_common_aspects": list(source_resolution.get("compare_common_aspects") or []),
                "compare_topic_pair": list(source_resolution.get("compare_topic_pair") or []),
                "compare_canonical_aspects": list(source_resolution.get("compare_canonical_aspects") or []),
                "compare_expanded_aspects": list(source_resolution.get("compare_expanded_aspects") or []),
                "compare_source_subqueries": compare_subqueries,
                "compare_status": source_resolution.get("compare_status") or compare_plan.get("compare_status") or "plan_ready",
                "compare_plan": compare_plan,
                "compare_source_results": compare_source_results,
                "lock_mode": source_resolution.get("lock_mode") or "hard_lock",
                "lock_confidence": float(source_resolution.get("lock_confidence") or 1.0),
                "lock_message_prefix": source_resolution.get("lock_message_prefix") or "",
                "source_lock_kind": source_resolution.get("source_lock_kind") or "",
                "source_resolution_trace": dict(source_resolution.get("source_resolution_trace") or {}),
                "inherited_from_context": bool(source_resolution.get("inherited_from_context")),
            }

        milvus_filter = None
        safe_names: List[str] = []
        if active_fnames:
            safe_names = [_normalize_filename_for_match(x) for x in active_fnames]
            if len(safe_names) == 1:
                milvus_filter = f"source == {json.dumps(safe_names[0], ensure_ascii=False)}"

        docs = self.vector_db.search(query_embedding, top_k=recall_k, filters=milvus_filter)
        visible_dense = _filter_hits_by_source_state(docs)
        docs = visible_dense["hits"]
        dense_source_scores = _dense_source_score_map(docs)

        lex_items = _collect_lexical_candidates(retrieval_query, safe_names, [])
        visible_lex = _filter_hits_by_source_state(lex_items)
        lex_items = visible_lex["hits"]
        docs_all = docs + lex_items

        if not docs_all:
            return {
                "query": query,
                "retrieval_query": retrieval_query,
                "retrieval_query_raw": retrieval_query_raw,
                "dense_query": dense_query,
                "llm_parse": llm_parse,
                "is_comparison": bool(is_comparison),
                "evidence_query": retrieval_query,
                "question_type": qtype,
                "score_mode": "score",
                "docs": [],
                "selected_docs": [],
                "qfilters": qfilters,
                "recall_k": recall_k,
                "final_n": final_n,
                "rerank_used": False,
                "query_route": query_route,
                "weak_query": _is_weak_reference_query(query),
                "early_filtered": visible_dense["dropped"] + visible_lex["dropped"],
                "visibility_filtered": visible_dense["dropped"] + visible_lex["dropped"],
                "dense_source_scores": dense_source_scores,
                "post_filter_docs": [],
                "retrieve_docs": [],
                "source_lock_required": bool(source_resolution.get("required")),
                "resolved_source_lock": bool(source_resolution.get("resolved")),
                "target_sources": list(active_fnames),
                "source_lock_candidates": list(source_resolution.get("candidates") or []),
                "source_lock_reason": source_resolution.get("reason") or "",
                "clarification": source_resolution.get("clarification") or "",
                "target_text": source_resolution.get("target_text") or "",
                "lock_mode": source_resolution.get("lock_mode") or "none",
                "lock_confidence": float(source_resolution.get("lock_confidence") or 0.0),
                "lock_message_prefix": source_resolution.get("lock_message_prefix") or "",
                "source_lock_kind": source_resolution.get("source_lock_kind") or "",
                "source_resolution_trace": dict(source_resolution.get("source_resolution_trace") or {}),
                "inherited_from_context": bool(source_resolution.get("inherited_from_context")),
            }

        dense_rank_map: Dict[str, int] = {}
        for i, d in enumerate(docs):
            src = _normalize_filename_for_match(_hit_entity_source(d) or "")
            if src and src not in dense_rank_map:
                dense_rank_map[src] = i
        lex_rank_map: Dict[str, int] = {}
        for i, d in enumerate(lex_items):
            src = _normalize_filename_for_match(_hit_entity_source(d) or "")
            if src and src not in lex_rank_map:
                lex_rank_map[src] = i

        source_count: Dict[str, int] = {}
        for d in docs_all:
            src = _normalize_filename_for_match(_hit_entity_source(d) or "")
            if src:
                source_count[src] = source_count.get(src, 0) + 1

        weak_query = _is_weak_reference_query(query)
        fname_set = {_normalize_filename_for_match(x) for x in active_fnames if x}
        source_signals = _build_source_signal_map(retrieval_query, lex_items, [])
        combined = [(0.0, i) for i in range(len(docs_all))]
        fused_source_scores: Dict[str, float] = {}
        for i, d in enumerate(docs_all):
            src = _normalize_filename_for_match(_hit_entity_source(d) or "")
            if src not in fused_source_scores:
                fused_source_scores[src] = _fusion_source_score(
                    src,
                    retrieval_query,
                    dense_rank_map,
                    lex_rank_map,
                    source_count,
                    source_signals,
                    fname_set,
                    set(),
                    weak_query,
                )
            combined[i] = (fused_source_scores[src], i)
        combined.sort(
            key=lambda item: (item[0], _source_dense_tiebreak_score(docs_all[item[1]], dense_source_scores)),
            reverse=True,
        )

        seen_keys = set()
        docs_fused = []
        for fused_score, idx in combined:
            hit = docs_all[idx]
            key = (_hit_entity_source(hit) or "unknown", (_hit_entity_text(hit) or "")[:64])
            if key in seen_keys:
                continue
            seen_keys.add(key)
            docs_fused.append(_clone_hit_with_score(hit, fused_score))
            if len(docs_fused) >= recall_k:
                break
        docs = docs_fused

        chunk_rerank_enabled = _should_apply_chunk_rerank(docs[:pool_n], dense_rank_map, lex_rank_map, source_signals, enable_rerank)
        reranked_chunk = await _chunk_level_rerank(self.rerank_service, retrieval_query, docs[:pool_n], pool_n, chunk_rerank_enabled)
        docs = reranked_chunk["hits"]
        score_mode = reranked_chunk["score_mode"]

        merged_docs = _merge_and_dedupe_hits(docs, score_mode=score_mode)
        aggregated_docs = _aggregate_doc_sections(merged_docs, score_mode=score_mode)
        context_docs = _docs_for_query_context(qtype, merged_docs, aggregated_docs)
        retrieve_docs = _apply_retrieval_filters(docs, qfilters, active_fnames)
        retrieve_docs = _filter_low_relevance_sources(retrieve_docs, score_mode=score_mode, query=retrieval_query)
        retrieve_docs = _intra_doc_chunk_rerank(retrieval_query, retrieve_docs, score_mode=score_mode, qtype=qtype)
        filtered_docs = _apply_retrieval_filters(context_docs, qfilters, active_fnames)
        selected_docs = _filter_low_relevance_sources(filtered_docs, score_mode=score_mode, query=retrieval_query)
        selected_docs = _intra_doc_chunk_rerank(retrieval_query, selected_docs, score_mode=score_mode, qtype=qtype)
        post_filter_docs = selected_docs[:]
        selected_docs = selected_docs[: min(len(selected_docs), final_n)]

        return {
            "query": query,
            "retrieval_query": retrieval_query,
            "retrieval_query_raw": retrieval_query_raw,
            "dense_query": dense_query,
            "llm_parse": llm_parse,
            "is_comparison": bool(is_comparison),
            "evidence_query": retrieval_query,
            "question_type": qtype,
            "score_mode": score_mode,
            "docs": docs,
            "selected_docs": selected_docs,
            "qfilters": qfilters,
            "recall_k": recall_k,
            "final_n": final_n,
            "rerank_used": bool(reranked_chunk["used"]),
            "query_route": query_route,
            "weak_query": weak_query,
            "early_filtered": visible_dense["dropped"] + visible_lex["dropped"],
            "visibility_filtered": visible_dense["dropped"] + visible_lex["dropped"],
            "dense_source_scores": dense_source_scores,
            "post_filter_docs": post_filter_docs,
            "retrieve_docs": retrieve_docs,
            "source_lock_required": bool(source_resolution.get("required")),
            "resolved_source_lock": bool(source_resolution.get("resolved") or (bool(active_fnames) and not topical_multi_doc_mode)),
            "target_sources": list(active_fnames),
            "source_lock_candidates": list(source_resolution.get("candidates") or []),
            "source_lock_reason": source_resolution.get("reason") or "",
            "clarification": source_resolution.get("clarification") or "",
            "target_text": source_resolution.get("target_text") or "",
            "lock_mode": source_resolution.get("lock_mode") or ("hard_lock" if active_fnames else "none"),
            "lock_confidence": float(source_resolution.get("lock_confidence") or (1.0 if active_fnames else 0.0)),
            "lock_message_prefix": source_resolution.get("lock_message_prefix") or "",
            "source_lock_kind": source_resolution.get("source_lock_kind") or "",
            "source_resolution_trace": dict(source_resolution.get("source_resolution_trace") or {}),
            "inherited_from_context": bool(source_resolution.get("inherited_from_context")),
            "compare_subjects": list(source_resolution.get("compare_subjects") or []),
            "compare_doc_like_subjects": list(source_resolution.get("compare_doc_like_subjects") or []),
            "compare_missing_targets": list(source_resolution.get("compare_missing_targets") or []),
                "compare_common_aspects": list(source_resolution.get("compare_common_aspects") or []),
                "compare_topic_pair": list(source_resolution.get("compare_topic_pair") or []),
                "compare_canonical_aspects": list(source_resolution.get("compare_canonical_aspects") or []),
                "compare_expanded_aspects": list(source_resolution.get("compare_expanded_aspects") or []),
                "compare_source_subqueries": dict(source_resolution.get("compare_source_subqueries") or compare_plan.get("source_subqueries") or {}),
                "compare_status": source_resolution.get("compare_status") or compare_plan.get("compare_status") or "not_compare",
                "compare_plan": compare_plan,
            "compare_common_aspects": list(source_resolution.get("compare_common_aspects") or []),
            "compare_topic_pair": list(source_resolution.get("compare_topic_pair") or []),
            "compare_canonical_aspects": list(source_resolution.get("compare_canonical_aspects") or []),
            "compare_expanded_aspects": list(source_resolution.get("compare_expanded_aspects") or []),
            "compare_source_subqueries": dict(source_resolution.get("compare_source_subqueries") or compare_plan.get("source_subqueries") or {}),
            "compare_status": source_resolution.get("compare_status") or compare_plan.get("compare_status") or "not_compare",
            "compare_plan": compare_plan,
            "compare_source_results": [],
            "intent_tier": intent_tier,
        }

    async def retrieve(self, query: str, user_id: str = "anonymous", top_k: int = 10, enable_rerank: bool = True):
        """仅检索，不调用 LLM"""
        try:
            query = _normalize_query(query)
            fnames = _extract_filename_candidates(query)
            if len(query) < config.MIN_QUERY_CHARS:
                return {
                    "documents": [],
                    "sources": [],
                    "metadata": _build_control_plane_metadata(
                        query=query,
                        user_id=user_id,
                        query_route="query_too_short",
                        internal_route="query_too_short",
                        final_channel="blocked",
                        blocked="query_too_short",
                        query_quality="invalid",
                    ),
                }
            if len(query) > config.MAX_QUERY_CHARS:
                return {
                    "documents": [],
                    "sources": [],
                    "metadata": _build_control_plane_metadata(
                        query=query,
                        user_id=user_id,
                        query_route="query_too_long",
                        internal_route="query_too_long",
                        final_channel="blocked",
                        blocked="query_too_long",
                        query_quality="invalid",
                    ),
                }
            blocked = _blocked_reason(query)
            if blocked:
                return {
                    "documents": [],
                    "sources": [],
                    "metadata": _build_control_plane_metadata(
                        query=query,
                        user_id=user_id,
                        query_route=blocked,
                        internal_route=blocked,
                        final_channel="blocked",
                        blocked=blocked,
                        query_quality="invalid",
                    ),
                }
            query_quality = _query_static_quality_state(query)
            if query_quality["reason"]:
                return {
                    "documents": [],
                    "sources": [],
                    "metadata": _build_control_plane_metadata(
                        query=query,
                        user_id=user_id,
                        query_route=query_quality["reason"],
                        internal_route=query_quality["reason"],
                        final_channel="blocked",
                        blocked=query_quality["reason"],
                        query_quality=query_quality["quality"],
                    ),
                }
            recall = await self._run_lightweight_recall(query, top_k=top_k, enable_rerank=enable_rerank, filename_hints=fnames, user_id=user_id)
            if recall.get("blocked_reason"):
                return {
                    "documents": [],
                    "sources": [],
                    "metadata": _build_control_plane_metadata(
                        query=query,
                        user_id=user_id,
                        query_route=recall.get("blocked_reason") or "low_information_query",
                        internal_route=recall.get("blocked_reason") or "low_information_query",
                        final_channel="blocked",
                        blocked=recall.get("blocked_reason") or "low_information_query",
                        query_quality=recall.get("query_quality") or "low_information",
                        recall=recall,
                    ),
                }
            if recall.get("soft_clarification_required"):
                clarification = await self._build_rule_backed_clarification(
                    query,
                    reason=str(recall.get("soft_clarification_reason") or recall.get("intent_tier") or "document_clarification"),
                    seed_sources=list(recall.get("source_lock_candidates") or []),
                )
                return {
                    "documents": [],
                    "sources": [],
                    "metadata": _build_control_plane_metadata(
                        query=query,
                        user_id=user_id,
                        query_route="document_clarification",
                        internal_route=recall.get("soft_clarification_reason") or recall.get("query_route") or "content_qa",
                        final_channel="document_clarification",
                        recall=recall,
                        refusal_reason=recall.get("soft_clarification_reason") or "document_clarification",
                        question_type=recall.get("question_type") or _classify_question_type(query),
                        extra={
                            "refused": recall.get("soft_clarification_reason") or "document_clarification",
                            "clarification": clarification.get("message") or "",
                            "candidate_sources": list(clarification.get("candidate_sources") or recall.get("source_lock_candidates") or []),
                            "clarification_used_llm": bool(clarification.get("used_llm")),
                            "intent_tier": recall.get("intent_tier") or "",
                        },
                    ),
                }
            if recall.get("source_lock_required") and not recall.get("resolved_source_lock"):
                source_lock_reason = recall.get("source_lock_reason") or "document_target_required"
                if source_lock_reason in {"compare_target_not_found", "compare_targets_not_found"}:
                    return {
                        "documents": [],
                        "sources": [],
                        "metadata": _build_control_plane_metadata(
                            query=query,
                            user_id=user_id,
                            query_route=source_lock_reason,
                            internal_route=recall.get("query_route") or source_lock_reason,
                            final_channel="document_not_found",
                            recall=recall,
                            refusal_reason=source_lock_reason,
                            question_type=recall.get("question_type") or _classify_question_type(query),
                            extra={
                                "refused": source_lock_reason,
                                "target_text": recall.get("target_text") or "",
                                "message": _build_compare_target_not_found_prompt(
                                    list(recall.get("compare_missing_targets") or []),
                                    list(recall.get("source_lock_candidates") or recall.get("target_sources") or []),
                                ),
                            },
                        ),
                    }
                if source_lock_reason == "document_not_found":
                    return {
                        "documents": [],
                        "sources": [],
                        "metadata": _build_control_plane_metadata(
                            query=query,
                            user_id=user_id,
                            query_route="document_not_found",
                            internal_route=recall.get("query_route") or "explicit_regulation_reference",
                            final_channel="document_not_found",
                            recall=recall,
                            refusal_reason="document_not_found",
                            question_type=recall.get("question_type") or _classify_question_type(query),
                            extra={
                                "refused": "document_not_found",
                                "target_text": recall.get("target_text") or "",
                                "message": _build_document_not_found_prompt(recall.get("target_text") or query),
                            },
                        ),
                    }
                return {
                    "documents": [],
                    "sources": [],
                    "metadata": _build_control_plane_metadata(
                        query=query,
                        user_id=user_id,
                        query_route="document_ambiguous" if source_lock_reason == "document_ambiguous" else "document_clarification",
                        internal_route=recall.get("query_route") or "weak_title_reference",
                        final_channel="document_ambiguous" if source_lock_reason == "document_ambiguous" else "document_clarification",
                        recall=recall,
                        refusal_reason=source_lock_reason,
                        question_type=recall.get("question_type") or _classify_question_type(query),
                        extra={
                            "refused": source_lock_reason,
                            "clarification": recall.get("clarification") or _build_document_clarification_prompt(recall.get("source_lock_candidates") or []),
                            "candidate_sources": list(recall.get("source_lock_candidates") or []),
                            "target_text": recall.get("target_text") or "",
                        },
                    ),
                }
            resolved_targets = [_normalize_filename_for_match(x) for x in (recall.get("target_sources") or fnames) if _normalize_filename_for_match(x)]
            if recall.get("query_route") == "multi_doc_compare" and recall.get("compare_source_results"):
                compare_retrieve_groups = []
                for item in recall.get("compare_source_results") or []:
                    compare_retrieve_groups.append({
                        "source": item.get("source") or "",
                        "evidence_query": item.get("evidence_query") or "",
                        "docs": _select_retrieve_output_docs(
                            item.get("retrieve_docs") or item.get("post_filter_docs") or item.get("selected_docs") or [],
                            top_k=top_k,
                            default_n=recall["final_n"],
                        ),
                    })
                retrieve_docs = _merge_compare_source_doc_groups(compare_retrieve_groups, per_source_limit=max(2, top_k))
                observations = await _compare_evidence_observations_async(query, compare_retrieve_groups, qfilters=recall["qfilters"])
            else:
                retrieve_docs = _select_retrieve_output_docs(recall.get("retrieve_docs") or recall.get("post_filter_docs") or recall["selected_docs"], top_k=top_k, default_n=recall["final_n"])
                observations = _evidence_observations(
                    recall.get("evidence_query") or recall.get("retrieval_query") or query,
                    retrieve_docs,
                    qfilters=recall["qfilters"],
                    candidate_docs=recall.get("retrieve_docs") or recall.get("post_filter_docs") or recall["selected_docs"],
                    target_sources=resolved_targets,
                    source_lock_resolved=bool(recall.get("resolved_source_lock")),
                    source_lock_reason=str(recall.get("source_lock_reason") or ""),
                    is_comparison=bool(recall.get("is_comparison")),
                    compare_missing_targets=list(recall.get("compare_missing_targets") or []),
                )
            refusal_reason = None
            if observations["answer_scope"] not in {"full", "guarded_full"} and not bool(observations.get("compare_degraded")):
                refusal_reason = observations["evidence_coverage_reason"]
            if refusal_reason:
                return {
                    "documents": [],
                    "sources": [],
                    "metadata": _build_control_plane_metadata(
                        query=query,
                        user_id=user_id,
                        query_route="evidence_insufficient",
                        internal_route=recall.get("query_route") or "content_qa",
                        final_channel="refusal",
                        recall=recall,
                        refusal_reason=refusal_reason,
                        question_type=recall["question_type"],
                        docs_returned=len(retrieve_docs),
                        extra={
                            "refused": refusal_reason,
                            "message": _build_compare_evidence_failure_prompt(observations.get("compare_source_statuses") or []) if recall.get("query_route") == "multi_doc_compare" else "",
                            "compare_status": observations.get("compare_status") or recall.get("compare_status") or "",
                            "visibility_enforced": True,
                            "visibility_filtered": recall["visibility_filtered"],
                            **observations,
                        },
                    ),
                }

            display_docs = _filter_display_sources(
                retrieve_docs,
                recall["score_mode"],
                recall["qfilters"],
                resolved_targets,
                recall["question_type"],
                max_sources=3,
                target_sources=resolved_targets,
                observations=observations,
            )
            sources = _build_sources(display_docs if display_docs else retrieve_docs[:3], query, score_mode=recall["score_mode"])

            documents = []
            for d in retrieve_docs:
                documents.append({
                    "source": _hit_entity_source(d) or "unknown",
                    "score": _hit_score(d),
                    "text": _build_excerpt(_hit_display_text(d), query, 500),
                    "metadata": _hit_metadata(d),
                    "chunk_range": _hit_chunk_range(d),
                })

            return {
                "documents": documents,
                "sources": sources,
                "metadata": _build_control_plane_metadata(
                    query=query,
                    user_id=user_id,
                    query_route="light_rag",
                    internal_route=recall.get("query_route") or "content_qa",
                    final_channel="light_rag",
                    recall=recall,
                    question_type=recall["question_type"],
                    docs_returned=len(retrieve_docs),
                    extra={
                        "compare_status": observations.get("compare_status") or recall.get("compare_status") or "",
                        "docs_recalled": recall["recall_k"],
                        "docs_rerank_kept": len(recall["docs"]),
                        "docs_final": len(retrieve_docs),
                        "rerank_used": recall["rerank_used"],
                        "weak_query_expansion": recall["weak_query"],
                        "early_filtered": recall["early_filtered"],
                        "visibility_enforced": True,
                        "visibility_filtered": recall["visibility_filtered"],
                        **observations,
                    },
                )
            }
        except Exception as e:
            logger.error(f"Retrieve processing error: {str(e)}")
            raise HTTPException(status_code=500, detail=str(e))
    
    async def process(
        self,
        query: str,
        user_id: str = "anonymous",
        top_k: int = 10,
        enable_rerank: bool = True,
        forced_fnames: Optional[List[str]] = None,
    ):
        """处理查询"""
        try:
            query = _normalize_query(query)
            print({
                "event": "query.extracted",
                "query": query,
                "query_len": len(query or ""),
            }, flush=True)
            qtype = _classify_question_type(query)
            original_fnames = list(forced_fnames or _extract_filename_candidates(query))

            def _decorate_answer(answer: str, answer_mode: str) -> str:
                if answer_mode == "rag_related_doc":
                    return "未在当前可见知识库中命中目标文档，以下回答基于其他文档的相似证据，可能不等同于目标文档原文。\n" + answer
                return answer

            def _update_current_locked_document_state(recall: Dict[str, Any], resolved_targets: List[str], observations: Optional[Dict[str, Any]] = None):
                normalized_targets = [_normalize_filename_for_match(item) for item in (resolved_targets or []) if _normalize_filename_for_match(item)]
                if len(normalized_targets) != 1:
                    if recall.get("source_lock_required") and not recall.get("resolved_source_lock"):
                        _clear_current_locked_document(user_id)
                    return
                if recall.get("query_route") in {"multi_doc_compare", "single_doc_compare", "compare_clarification"}:
                    return
                lock_mode = str(recall.get("lock_mode") or "hard_lock")
                reliable = bool(recall.get("resolved_source_lock"))
                if observations is not None:
                    reliable = reliable and observations.get("answer_scope") in {"full", "guarded_full"}
                if not reliable:
                    return
                _set_current_locked_document(
                    user_id,
                    normalized_targets[0],
                    reason=str(recall.get("source_lock_reason") or recall.get("query_route") or ""),
                    reliable=True,
                    lock_mode=lock_mode,
                )

            if len(query) < config.MIN_QUERY_CHARS:
                return {
                    "answer": "请提供更具体的问题描述。",
                    "sources": [],
                    "metadata": _build_control_plane_metadata(
                        query=query,
                        user_id=user_id,
                        query_route="query_too_short",
                        internal_route="query_too_short",
                        final_channel="blocked",
                        blocked="query_too_short",
                        query_quality="invalid",
                        answer_mode="refusal",
                    ),
                }
            if len(query) > config.MAX_QUERY_CHARS:
                return {
                    "answer": f"问题过长（>{config.MAX_QUERY_CHARS} 字符），请精简后再试。",
                    "sources": [],
                    "metadata": _build_control_plane_metadata(
                        query=query,
                        user_id=user_id,
                        query_route="query_too_long",
                        internal_route="query_too_long",
                        final_channel="blocked",
                        blocked="query_too_long",
                        query_quality="invalid",
                        answer_mode="refusal",
                    ),
                }
            blocked = _blocked_reason(query)
            if blocked:
                return {
                    "answer": "抱歉，无法处理该请求。",
                    "sources": [],
                    "metadata": _build_control_plane_metadata(
                        query=query,
                        user_id=user_id,
                        query_route=blocked,
                        internal_route=blocked,
                        final_channel="blocked",
                        blocked=blocked,
                        query_quality="invalid",
                        answer_mode="refusal",
                    ),
                }
            query_quality = _query_static_quality_state(query)
            if query_quality["reason"]:
                return {
                    "answer": _invalid_query_message(query_quality["reason"]),
                    "sources": [],
                    "metadata": _build_control_plane_metadata(
                        query=query,
                        user_id=user_id,
                        query_route=query_quality["reason"],
                        internal_route=query_quality["reason"],
                        final_channel="blocked",
                        blocked=query_quality["reason"],
                        query_quality=query_quality["quality"],
                        answer_mode="refusal",
                    ),
                }
            try:
                recall = await self._run_lightweight_recall(query, top_k=top_k, enable_rerank=enable_rerank, filename_hints=original_fnames, user_id=user_id)
            except Exception as e:
                logger.warning(f"milvus_jitter: search_exception type={type(e).__name__} msg={str(e)}")
                return {
                    "answer": "检索服务暂时不可用，请稍后重试。",
                    "sources": [],
                    "metadata": _build_control_plane_metadata(
                        query=query,
                        user_id=user_id,
                        query_route="retrieval_error",
                        internal_route="retrieval_error",
                        final_channel="refusal",
                        refusal_reason="retrieval_error",
                        docs_returned=0,
                        question_type=qtype,
                        answer_mode="refusal",
                        extra={"refused": "retrieval_error", "error": str(e)},
                    ),
                }
            if recall.get("blocked_reason"):
                clarification = await self._build_rule_backed_clarification(
                    query,
                    reason=str(recall.get("blocked_reason") or "low_information_query"),
                    seed_sources=list(recall.get("source_lock_candidates") or recall.get("target_sources") or []),
                )
                if clarification.get("candidate_sources"):
                    return {
                        "answer": clarification.get("message") or _invalid_query_message(recall.get("blocked_reason") or "low_information_query"),
                        "sources": [],
                        "metadata": _build_control_plane_metadata(
                            query=query,
                            user_id=user_id,
                            query_route="document_clarification",
                            internal_route=recall.get("blocked_reason") or "low_information_query",
                            final_channel="document_clarification",
                            refusal_reason=recall.get("blocked_reason") or "low_information_query",
                            query_quality=recall.get("query_quality") or "low_information",
                            answer_mode="clarification",
                            docs_returned=0,
                            question_type=qtype,
                            recall=recall,
                            extra={
                                "refused": recall.get("blocked_reason") or "low_information_query",
                                "candidate_sources": list(clarification.get("candidate_sources") or []),
                                "clarification": clarification.get("message") or "",
                                "clarification_used_llm": bool(clarification.get("used_llm")),
                            },
                        ),
                    }
                return {
                    "answer": _invalid_query_message(recall.get("blocked_reason") or "low_information_query"),
                    "sources": [],
                    "metadata": _build_control_plane_metadata(
                        query=query,
                        user_id=user_id,
                        query_route=recall.get("blocked_reason") or "low_information_query",
                        internal_route=recall.get("blocked_reason") or "low_information_query",
                        final_channel="blocked",
                        blocked=recall.get("blocked_reason") or "low_information_query",
                        query_quality=recall.get("query_quality") or "low_information",
                        answer_mode="refusal",
                        recall=recall,
                    ),
                }
            if recall.get("soft_clarification_required"):
                clarification = await self._build_rule_backed_clarification(
                    query,
                    reason=str(recall.get("soft_clarification_reason") or recall.get("intent_tier") or "document_clarification"),
                    seed_sources=list(recall.get("source_lock_candidates") or []),
                )
                return {
                    "answer": clarification.get("message") or _build_document_clarification_prompt(recall.get("source_lock_candidates") or []),
                    "sources": [],
                    "metadata": _build_control_plane_metadata(
                        query=query,
                        user_id=user_id,
                        query_route="document_clarification",
                        internal_route=recall.get("soft_clarification_reason") or recall.get("query_route") or "content_qa",
                        final_channel="document_clarification",
                        refusal_reason=recall.get("soft_clarification_reason") or "document_clarification",
                        query_quality=recall.get("query_quality") or "valid",
                        answer_mode="clarification",
                        docs_returned=0,
                        question_type=qtype,
                        recall=recall,
                        extra={
                            "refused": recall.get("soft_clarification_reason") or "document_clarification",
                            "candidate_sources": list(clarification.get("candidate_sources") or recall.get("source_lock_candidates") or []),
                            "clarification": clarification.get("message") or "",
                            "clarification_used_llm": bool(clarification.get("used_llm")),
                            "intent_tier": recall.get("intent_tier") or "",
                        },
                    ),
                }
            resolved_targets = [_normalize_filename_for_match(x) for x in (recall.get("target_sources") or original_fnames) if _normalize_filename_for_match(x)]

            if recall.get("source_lock_required") and not recall.get("resolved_source_lock"):
                _clear_current_locked_document(user_id)
                source_lock_reason = recall.get("source_lock_reason") or "document_target_required"
                if source_lock_reason in {"compare_target_not_found", "compare_targets_not_found"}:
                    return {
                        "answer": _build_compare_target_not_found_prompt(
                            list(recall.get("compare_missing_targets") or []),
                            list(recall.get("source_lock_candidates") or recall.get("target_sources") or []),
                        ),
                        "sources": [],
                        "metadata": _build_control_plane_metadata(
                            query=query,
                            user_id=user_id,
                            query_route=source_lock_reason,
                            internal_route=recall.get("query_route") or source_lock_reason,
                            final_channel="document_not_found",
                            recall=recall,
                            refusal_reason=source_lock_reason,
                            docs_returned=0,
                            question_type=qtype,
                            answer_mode="refusal",
                            extra={
                                "refused": source_lock_reason,
                                "target_text": recall.get("target_text") or "",
                            },
                        ),
                    }
                if source_lock_reason == "document_not_found":
                    return {
                        "answer": _build_document_not_found_prompt(recall.get("target_text") or query),
                        "sources": [],
                        "metadata": _build_control_plane_metadata(
                            query=query,
                            user_id=user_id,
                            query_route="document_not_found",
                            internal_route=recall.get("query_route") or "explicit_regulation_reference",
                            final_channel="document_not_found",
                            recall=recall,
                            refusal_reason="document_not_found",
                            docs_returned=0,
                            question_type=qtype,
                            answer_mode="refusal",
                            extra={"refused": "document_not_found", "target_text": recall.get("target_text") or ""},
                        ),
                    }
                clarification = await self._build_rule_backed_clarification(
                    query,
                    reason=str(source_lock_reason or "document_target_required"),
                    seed_sources=list(recall.get("source_lock_candidates") or []),
                )
                return {
                    "answer": clarification.get("message") or recall.get("clarification") or _build_document_clarification_prompt(recall.get("source_lock_candidates") or []),
                    "sources": [],
                    "metadata": _build_control_plane_metadata(
                        query=query,
                        user_id=user_id,
                        query_route="document_ambiguous" if source_lock_reason == "document_ambiguous" else "document_clarification",
                        internal_route=recall.get("query_route") or "weak_title_reference",
                        final_channel="document_ambiguous" if source_lock_reason == "document_ambiguous" else "document_clarification",
                        recall=recall,
                        refusal_reason=source_lock_reason,
                        docs_returned=0,
                        question_type=qtype,
                        answer_mode="clarification",
                        extra={
                            "refused": source_lock_reason,
                            "candidate_sources": list(clarification.get("candidate_sources") or recall.get("source_lock_candidates") or []),
                            "clarification": clarification.get("message") or recall.get("clarification") or "",
                            "clarification_used_llm": bool(clarification.get("used_llm")),
                            "target_text": recall.get("target_text") or "",
                        },
                    ),
                }

            process_input_docs = recall.get("post_filter_docs") or recall["selected_docs"]
            if bool(recall.get("resolved_source_lock")) and len(resolved_targets) == 1 and recall.get("query_route") not in {"multi_doc_compare", "single_doc_compare"}:
                process_input_docs = recall.get("retrieve_docs") or process_input_docs
            process_docs = _select_process_output_docs(
                query,
                process_input_docs,
                recall["score_mode"],
                recall["qfilters"],
                recall["final_n"],
            )
            if recall.get("query_route") == "multi_doc_compare" and recall.get("compare_source_results"):
                compare_process_groups = []
                for item in recall.get("compare_source_results") or []:
                    compare_process_groups.append({
                        "source": item.get("source") or "",
                        "evidence_query": item.get("evidence_query") or "",
                        "docs": _select_process_output_docs(
                            query,
                            item.get("post_filter_docs") or item.get("selected_docs") or [],
                            item.get("score_mode") or recall["score_mode"],
                            recall["qfilters"],
                            max(2, recall["final_n"]),
                        ),
                    })
                process_docs = _merge_compare_source_doc_groups(compare_process_groups, per_source_limit=max(2, recall["final_n"]))
                observations = await _compare_evidence_observations_async(query, compare_process_groups, qfilters=recall["qfilters"])
            else:
                observations = _evidence_observations(
                    recall.get("evidence_query") or recall.get("retrieval_query") or query,
                    process_docs,
                    qfilters=recall["qfilters"],
                    candidate_docs=process_input_docs,
                    target_sources=resolved_targets,
                    source_lock_resolved=bool(recall.get("resolved_source_lock")),
                    source_lock_reason=str(recall.get("source_lock_reason") or ""),
                    is_comparison=bool(recall.get("is_comparison")),
                    compare_missing_targets=list(recall.get("compare_missing_targets") or []),
                )
            refusal_reason = None
            if (
                observations["answer_scope"] not in {"full", "guarded_full"}
                and not bool(observations.get("compare_degraded"))
            ):
                refusal_reason = observations["evidence_coverage_reason"]
            if refusal_reason:
                return {
                    "answer": _build_compare_evidence_failure_prompt(observations.get("compare_source_statuses") or []) if recall.get("query_route") == "multi_doc_compare" else "未检索到相关证据。",
                    "sources": [],
                    "metadata": _build_control_plane_metadata(
                        query=query,
                        user_id=user_id,
                        query_route="evidence_insufficient",
                        internal_route=recall.get("query_route") or "content_qa",
                        final_channel="refusal",
                        recall=recall,
                        refusal_reason=refusal_reason,
                        docs_returned=len(process_docs),
                        question_type=qtype,
                        answer_mode="refusal",
                        extra={
                            "refused": refusal_reason,
                            "compare_status": observations.get("compare_status") or recall.get("compare_status") or "",
                            "visibility_enforced": True,
                            "visibility_filtered": recall["visibility_filtered"],
                            "llm_query_parse_enabled": bool(getattr(config, "ENABLE_LLM_QUERY_PARSE", True)),
                            "llm_parse": dict(recall.get("llm_parse") or {}),
                            **observations,
                        },
                    ),
                }
            _update_current_locked_document_state(recall, resolved_targets, observations)
            answer_mode = _answer_mode_for_sources(resolved_targets, process_docs)
            if bool(observations.get("compare_degraded")) and recall.get("query_route") == "single_doc_compare":
                qtype = "compare_degraded"
                answer_mode = "compare_degraded"
            elif qtype in {"compare", "compare_degraded"} and observations.get("compare_status") == "compare_asymmetric":
                answer_mode = "compare_asymmetric"
            compare_refs: List[Dict[str, Any]] = []
            if qtype in {"compare", "compare_degraded"} and recall.get("query_route") == "multi_doc_compare" and compare_process_groups:
                evidence, compare_refs = _format_compare_evidence(
                    compare_process_groups,
                    query,
                    score_mode=recall["score_mode"],
                    compare_plan=recall.get("compare_plan"),
                    compare_source_statuses=observations.get("compare_source_statuses") or [],
                )
            elif qtype in {"compare", "compare_degraded"} and recall.get("query_route") == "single_doc_compare":
                evidence, compare_refs = _format_single_doc_compare_evidence(
                    process_docs,
                    query,
                    score_mode=recall["score_mode"],
                    compare_plan=recall.get("compare_plan"),
                )
            else:
                evidence = _format_evidence(process_docs, query, score_mode=recall["score_mode"])
            limits = _answer_limits(qtype)
            logger.info(
                f"DEBUG ANSWER INPUT -> qtype={qtype} answer_mode={answer_mode} "
                f"selected_docs={len(process_docs)} evidence_chars={len(evidence)}"
            )
            answer = await self.generate_answer(
                query,
                evidence,
                qtype=qtype,
                max_tokens=limits["max_tokens"],
                answer_mode=answer_mode,
                compare_missing_targets=list(recall.get("compare_missing_targets") or []),
                compare_source_status_hints=_compare_source_status_prompt_lines(observations.get("compare_source_statuses") or []),
            )
            answer_preview = re.sub(r"\s+", " ", answer or "").strip()
            if len(answer_preview) > 160:
                answer_preview = answer_preview[:160].rstrip() + "..."
            logger.info(
                f"DEBUG ANSWER RAW -> qtype={qtype} answer_mode={answer_mode} "
                f"has_citation={'[' in (answer or '')} raw_answer={answer_preview}"
            )
            compare_target_count = len([
                source for source in (resolved_targets or [])
                if _normalize_filename_for_match(source or "")
            ])
            compare_answer_refs = compare_refs if len(compare_refs) >= 2 else _fallback_compare_refs_from_docs(process_docs)
            refusal_answer = "未在知识库中找到足够相关的证据来回答该问题。"
            if (
                (answer or "").strip() == refusal_answer
                and process_docs
                and str(observations.get("answer_scope") or "") in {"full", "guarded_full"}
            ):
                if qtype in {"compare", "compare_degraded"} and compare_target_count > 1:
                    answer = _build_multi_doc_compare_grounded_answer(
                        compare_answer_refs,
                        recall.get("compare_plan"),
                    )
                elif qtype in {"compare", "compare_degraded"}:
                    answer = _build_single_doc_compare_grounded_answer(compare_refs, recall.get("compare_plan"))
                else:
                    answer = _build_related_doc_grounded_answer(process_docs)
            if answer_mode == "rag_related_doc" and evidence and "[" not in answer:
                answer = _build_related_doc_grounded_answer(process_docs)
            elif qtype in {"compare", "compare_degraded"} and evidence and "[" not in answer and compare_target_count > 1:
                answer = _build_multi_doc_compare_grounded_answer(compare_answer_refs, recall.get("compare_plan"))
            elif qtype in {"compare", "compare_degraded"} and evidence and "[" not in answer:
                answer = _build_single_doc_compare_grounded_answer(compare_refs, recall.get("compare_plan"))
            elif config.REQUIRE_EVIDENCE and evidence and "[" not in answer:
                print({
                    "event": "answer.force_refuse_no_citation",
                    "query": query,
                    "qtype": qtype,
                    "answer_mode": answer_mode,
                    "selected_docs": len(process_docs),
                    "evidence_chars": len(evidence),
                    "require_evidence": config.REQUIRE_EVIDENCE,
                    "raw_answer_preview": answer_preview,
                }, flush=True)
                if qtype in {"compare", "compare_degraded"} and compare_target_count > 1:
                    answer = _build_multi_doc_compare_grounded_answer(
                        compare_answer_refs,
                        recall.get("compare_plan"),
                    )
                elif qtype in {"compare", "compare_degraded"}:
                    answer = _build_single_doc_compare_grounded_answer(compare_refs, recall.get("compare_plan"))
                elif answer_mode in {"target_hit", "rag_related_doc"}:
                    answer = _build_related_doc_grounded_answer(process_docs)
                else:
                    answer = "未在知识库中找到足够相关的证据来回答该问题。"
            answer = _append_answer_scope_semantics(answer, observations)
            answer = _decorate_answer(answer, answer_mode)
            lock_prefix = str(recall.get("lock_message_prefix") or "")
            if lock_prefix and not answer.startswith(lock_prefix):
                answer = lock_prefix + answer
            
            display_docs = _filter_display_sources(
                process_docs,
                recall["score_mode"],
                recall["qfilters"],
                resolved_targets,
                qtype,
                max_sources=3,
                target_sources=resolved_targets,
                observations=observations,
            )
            if not display_docs and process_docs:
                logger.info("obs: display_layer_cleared=1")
            sources = _build_sources(display_docs if display_docs else process_docs[:3], query, score_mode=recall["score_mode"])
            logger.info(f"obs: context_chunk_count={len(process_docs)} final_sources_count={len(display_docs) if display_docs else min(len(process_docs), 3)}")
            return {
                "answer": answer,
                "sources": sources,
                "metadata": _build_control_plane_metadata(
                    query=query,
                    user_id=user_id,
                    query_route="light_rag",
                    internal_route=recall.get("query_route") or "content_qa",
                    final_channel="light_rag",
                    recall={**recall, "target_sources": resolved_targets},
                    docs_returned=len(process_docs),
                    question_type=qtype,
                    answer_mode=answer_mode,
                    extra={
                        "lock_mode": recall.get("lock_mode") or "",
                        "lock_confidence": float(recall.get("lock_confidence") or 0.0),
                        "lock_message_prefix": recall.get("lock_message_prefix") or "",
                        "source_lock_kind": recall.get("source_lock_kind") or "",
                        "source_resolution_trace": dict(recall.get("source_resolution_trace") or {}),
                        "inherited_from_context": bool(recall.get("inherited_from_context")),
                        "compare_status": observations.get("compare_status") or recall.get("compare_status") or "",
                        "docs_recalled": recall["recall_k"],
                        "docs_rerank_kept": len(recall["docs"]),
                        "docs_final": len(process_docs),
                        "rerank_used": recall["rerank_used"],
                        "weak_query_expansion": recall["weak_query"],
                        "early_filtered": recall["early_filtered"],
                        "visibility_enforced": True,
                        "visibility_filtered": recall["visibility_filtered"],
                        "llm_query_parse_enabled": bool(getattr(config, "ENABLE_LLM_QUERY_PARSE", True)),
                        "llm_parse": dict(recall.get("llm_parse") or {}),
                        **observations,
                    },
                )
            }
        except Exception as e:
            logger.error(f"Query processing error: {str(e)}")
            raise HTTPException(status_code=500, detail=str(e))
    
    async def generate_answer(
        self,
        query: str,
        context: str,
        qtype: str = "other",
        max_tokens: Optional[int] = None,
        answer_mode: str = "target_hit",
        compare_missing_targets: Optional[List[str]] = None,
        compare_source_status_hints: str = "",
    ) -> str:
        """生成答案（调用 LLM Chat Completions）"""
        prompt = _build_answer_prompt(
            query=query,
            evidence=context,
            qtype=qtype,
            answer_mode=answer_mode,
            compare_missing_targets=compare_missing_targets,
            compare_source_status_hints=compare_source_status_hints,
        )

        def _chat_url_candidates() -> List[str]:
            if config.LLM_CHAT_COMPLETIONS_URL:
                return [config.LLM_CHAT_COMPLETIONS_URL]

            base = (config.LLM_API_BASE or "").rstrip("/")
            candidates = []
            if base:
                candidates.append(f"{base}/chat/completions")
                if not base.endswith("/v1"):
                    candidates.append(f"{base}/v1/chat/completions")
                if base.endswith("/v1"):
                    candidates.append(f"{base[:-3].rstrip('/')}/chat/completions")
            return candidates

        def _extra_body() -> Dict[str, Any]:
            raw = (config.LLM_EXTRA_BODY or "").strip()
            if not raw:
                return {}
            try:
                obj = json.loads(raw)
                return obj if isinstance(obj, dict) else {}
            except Exception:
                return {}

        def _call_llm() -> str:
            import requests

            headers = {"Content-Type": "application/json"}
            if config.LLM_API_KEY:
                headers["Authorization"] = f"Bearer {config.LLM_API_KEY}"

            payload = {
                "model": config.LLM_MODEL,
                "messages": [
                    {"role": "system", "content": "你是一个专业的AI助手"},
                    {"role": "user", "content": prompt}
                ],
                "temperature": config.LLM_TEMPERATURE,
                "top_p": config.LLM_TOP_P,
                "max_tokens": int(max_tokens or config.LLM_MAX_TOKENS),
                "presence_penalty": config.LLM_PRESENCE_PENALTY
            }

            extra = _extra_body()
            if extra:
                for k, v in extra.items():
                    if k not in payload:
                        payload[k] = v

            last_exc = None
            for url in _chat_url_candidates():
                try:
                    resp = requests.post(url, headers=headers, json=payload, timeout=config.LLM_TIMEOUT)
                    if resp.status_code == 404:
                        last_exc = HTTPException(status_code=404, detail="LLM endpoint 404")
                        continue
                    resp.raise_for_status()
                    data = resp.json()
                    choices = data.get("choices") or []
                    if choices and isinstance(choices, list):
                        msg = (choices[0] or {}).get("message") or {}
                        content_val = msg.get("content")
                        content = (content_val or "").strip() if isinstance(content_val, str) else ""
                        if not content:
                            reasoning_val = msg.get("reasoning")
                            content = (reasoning_val or "").strip() if isinstance(reasoning_val, str) else ""
                        if not content:
                            text_val = (choices[0] or {}).get("text")
                            content = (text_val or "").strip() if isinstance(text_val, str) else ""
                        if content:
                            return content
                    raise HTTPException(
                        status_code=500,
                        detail="LLM response missing choices.message.content (hint: disable thinking via chat_template_kwargs.enable_thinking=false)"
                    )
                except Exception as e:
                    last_exc = e
                    continue

            raise last_exc or HTTPException(status_code=500, detail="LLM call failed")

        try:
            return await asyncio.to_thread(_call_llm)
        except Exception as e:
            logger.error(f"LLM generation error: {str(e)}")
            return "抱歉，生成服务当前不可用。"


def split_text(text: str, chunk_size: int, overlap: int) -> List[str]:
    """简单分块"""
    text = (text or "").strip()
    if not text:
        return []

    if len(text) <= chunk_size:
        return [text]

    chunks = []
    step = max(1, chunk_size - overlap)
    for i in range(0, len(text), step):
        chunk = text[i:i + chunk_size].strip()
        if chunk:
            chunks.append(chunk)
        if i + chunk_size >= len(text):
            break
    return chunks


def decode_text_bytes(content: bytes) -> str:
    """多编码尝试"""
    for encoding in ("utf-8", "utf-8-sig", "gbk", "gb2312", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="ignore")


def _merge_parser_metadata(metadata: Optional[Dict[str, Any]], probe: Dict[str, Any]) -> Dict[str, Any]:
    merged = dict(metadata or {})
    merged["parser_probe"] = probe
    merged["parser_route"] = probe.get("route")
    merged["parser_backend"] = probe.get("parser_backend")
    return merged


def _parse_json_document_ir(filename: str, content: bytes, metadata: Optional[Dict[str, Any]], doc_version: Optional[int], parser_name: str) -> Dict[str, Any]:
    text = decode_text_bytes(content)
    try:
        obj = json.loads(text)
    except Exception:
        return _build_document_ir_from_text(filename, text, metadata=metadata, parser_name="json-text", doc_version=doc_version)
    document_ir = _new_document_ir(filename, metadata=metadata, doc_version=doc_version, parser_name=parser_name)
    pretty = json.dumps(obj, ensure_ascii=False, indent=2)
    if isinstance(obj, dict):
        for key, value in obj.items():
            if isinstance(value, (str, int, float, bool)) or value is None:
                _append_ir_element(document_ir, element_type="key_value", text_raw=f"{key}: {value}", json_payload={"key": key, "value": value}, page_no=1, parser_name=parser_name)
    _append_ir_element(document_ir, element_type="code_block", text_raw=pretty, text_normalized=_normalize_ir_text(pretty), json_payload=obj if isinstance(obj, (dict, list)) else None, page_no=1, parser_name=parser_name)
    return document_ir


def _parse_csv_document_ir(filename: str, content: bytes, metadata: Optional[Dict[str, Any]], doc_version: Optional[int], parser_name: str) -> Dict[str, Any]:
    text = decode_text_bytes(content)
    rows = [row for row in csv.reader(io.StringIO(text))]
    headers = [str(v).strip() for v in (rows[0] if rows else [])]
    data_rows = rows[1:] if len(rows) > 1 else []
    line_items = []
    for row in data_rows:
        parts = []
        for idx, value in enumerate(row):
            val = str(value).strip()
            if not val:
                continue
            header = headers[idx] if idx < len(headers) and headers[idx] else f"列{idx + 1}"
            parts.append(f"{header}: {val}")
        if parts:
            line_items.append(" | ".join(parts))
    table_text = "\n".join(line_items) or "\n".join([", ".join([str(v) for v in row]) for row in rows])
    document_ir = _new_document_ir(filename, metadata=metadata, doc_version=doc_version, parser_name=parser_name)
    _append_ir_element(
        document_ir,
        element_type="table",
        text_raw=table_text,
        text_normalized=_normalize_ir_text(table_text),
        markdown=table_text,
        json_payload={
            "table_json": {"headers": headers, "rows": rows},
            "table_text": table_text,
        },
        page_no=1,
        parser_name=parser_name,
    )
    return document_ir


def _bbox_area_ratio(bbox: Optional[Dict[str, Any]], page_width: Optional[float], page_height: Optional[float]) -> float:
    if not bbox or not page_width or not page_height or page_width <= 0 or page_height <= 0:
        return 0.0
    try:
        width = max(0.0, float(bbox.get("x1") or 0.0) - float(bbox.get("x0") or 0.0))
        height = max(0.0, float(bbox.get("y1") or 0.0) - float(bbox.get("y0") or 0.0))
        area = width * height
        page_area = float(page_width) * float(page_height)
        if page_area <= 0:
            return 0.0
        return max(0.0, min(1.0, area / page_area))
    except Exception:
        return 0.0


def _pdf_page_quality_profile(entries: List[Dict[str, Any]], image_blocks: List[Dict[str, Any]], page_width: Optional[float], page_height: Optional[float]) -> Dict[str, Any]:
    texts = [str((entry or {}).get("text") or "").strip() for entry in entries or [] if str((entry or {}).get("text") or "").strip()]
    text = "\n".join(texts).strip()
    text_chars = len(re.sub(r"\s+", "", text))
    garbled = _looks_like_cid_garbled_text(text)
    image_area_ratio = 0.0
    for block in image_blocks or []:
        image_area_ratio += _bbox_area_ratio(_normalize_bbox(block.get("bbox")), page_width, page_height)
    image_area_ratio = max(0.0, min(1.0, image_area_ratio))
    return {
        "text_chars": text_chars,
        "garbled": bool(garbled),
        "image_area_ratio": round(float(image_area_ratio), 4),
        "entry_count": len(entries or []),
        "image_block_count": len(image_blocks or []),
    }


def _pdf_page_should_use_ocr(page_profile: Dict[str, Any]) -> bool:
    if not page_profile:
        return False
    text_chars = int(page_profile.get("text_chars") or 0)
    image_area_ratio = float(page_profile.get("image_area_ratio") or 0.0)
    if text_chars <= max(20, int(config.PDF_OCR_MAX_TEXT_CHARS_PER_PAGE)):
        return True
    if bool(page_profile.get("garbled")) and text_chars <= max(120, int(config.PDF_OCR_MAX_TEXT_CHARS_PER_PAGE) * 2):
        return True
    if image_area_ratio >= 0.55 and text_chars <= max(180, int(config.PDF_OCR_MAX_TEXT_CHARS_PER_PAGE) * 2):
        return True
    return False


def _ocr_entries_from_payload(payload: Dict[str, Any], page_width: Optional[float], page_height: Optional[float]) -> List[Dict[str, Any]]:
    raw_lines = _extract_ocr_lines(payload)
    if not raw_lines:
        raw_lines = [{"text": text, "line_index": idx} for idx, text in enumerate(_extract_ocr_texts(payload)) if (text or "").strip()]
    entries: List[Dict[str, Any]] = []
    for line in raw_lines:
        text = str(line.get("text") or "").strip()
        if not text:
            continue
        bbox = _normalize_bbox(line.get("bbox"))
        layout = _bbox_layout_metrics(bbox, page_width=page_width, page_height=page_height)
        font_size = _safe_float(line.get("font_size"))
        if font_size is not None:
            layout["font_size"] = font_size
        entries.append({
            "text": text,
            "bbox": bbox,
            "layout": layout,
            "confidence": _safe_float(line.get("confidence")),
            "line_index": int(line.get("line_index") if line.get("line_index") is not None else len(entries)),
            "meta": dict(line.get("meta") or {}),
        })
    return entries


def _ocr_compensate_pdf_page(page: Any, page_no: int, probe: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    parser_backend = _choose_ocr_backend(_detect_parser_capabilities())
    if parser_backend != "external_http_ocr":
        return None
    try:
        import fitz  # type: ignore

        temp_dir = tempfile.mkdtemp(prefix=f"ocr_pdf_page_{page_no:04d}_", dir=_resolve_ocr_temp_dir())
        try:
            os.chmod(temp_dir, 0o755)
        except Exception:
            pass
        image_path = os.path.join(temp_dir, f"page_{page_no:04d}.png")
        pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
        pixmap.save(image_path)
        try:
            os.chmod(image_path, 0o644)
        except Exception:
            pass
        payload = _call_external_ocr(image_path)
        page_width = _safe_float(payload.get("meta", {}).get("page_width") or payload.get("meta", {}).get("width") or getattr(getattr(page, "rect", None), "width", None))
        page_height = _safe_float(payload.get("meta", {}).get("page_height") or payload.get("meta", {}).get("height") or getattr(getattr(page, "rect", None), "height", None))
        entries = _ocr_entries_from_payload(payload, page_width=page_width, page_height=page_height)
        if not entries:
            return None
        return {
            "entries": entries,
            "meta": payload.get("meta") or {},
            "image_path": image_path,
            "page_width": page_width,
            "page_height": page_height,
            "parser_backend": parser_backend,
            "probe": probe,
        }
    except Exception as e:
        logger.warning("pdf_page_ocr_compensation_failed: source=%s page=%s err=%s", _safe_filename(probe.get("filename") or ""), page_no, e)
        return None


def _parse_pdf_fast_document_ir(filename: str, content: bytes, metadata: Optional[Dict[str, Any]], doc_version: Optional[int], backend: str) -> Dict[str, Any]:
    if backend in {"pymupdf", "pymupdf4llm"} and _module_available("fitz"):
        try:
            import fitz  # type: ignore

            document = fitz.open(stream=content, filetype="pdf")
            document_ir = _new_document_ir(filename, metadata=metadata, doc_version=doc_version, parser_name="pymupdf")
            section_stack: List[str] = []
            toc_mode = False
            main_body_started = False
            semantic_truncated = False
            compensation_roots: set[str] = set()
            for page_no, page in enumerate(document, start=1):
                page_width = _safe_float(getattr(getattr(page, "rect", None), "width", None))
                page_height = _safe_float(getattr(getattr(page, "rect", None), "height", None))
                blocks = page.get_text("dict").get("blocks", [])
                entries: List[Dict[str, Any]] = []
                image_blocks: List[Dict[str, Any]] = []
                for block in blocks:
                    if block.get("type") == 1:
                        image_blocks.append(block)
                        continue
                    lines = []
                    span_sizes: List[float] = []
                    bold_hits = 0
                    span_count = 0
                    for line in block.get("lines", []) or []:
                        spans = [str(span.get("text") or "") for span in line.get("spans", []) or []]
                        line_text = "".join(spans).strip()
                        if line_text and not _is_pdf_noise_text(line_text, page_no=page_no):
                            lines.append(line_text)
                        for span in line.get("spans", []) or []:
                            size = _safe_float(span.get("size"))
                            if size is not None and size > 0:
                                span_sizes.append(size)
                            font_name = str(span.get("font") or "").lower()
                            flags = int(span.get("flags") or 0)
                            span_count += 1
                            if "bold" in font_name or (flags & 16):
                                bold_hits += 1
                    text = "\n".join(lines).strip()
                    if not text or _is_pdf_noise_text(text, page_no=page_no):
                        continue
                    bbox = _normalize_bbox(block.get("bbox"))
                    layout = _bbox_layout_metrics(bbox, page_width=page_width, page_height=page_height)
                    if span_sizes:
                        layout["font_size_max"] = max(span_sizes)
                        layout["font_size_median"] = _median_number(span_sizes)
                        layout["font_size"] = max(span_sizes)
                    layout["line_count"] = len(lines)
                    if span_count:
                        layout["bold_ratio"] = bold_hits / max(1, span_count)
                        layout["is_bold"] = layout["bold_ratio"] >= 0.5
                    entries.append({"text": text, "bbox": bbox, "layout": layout})
                page_quality = _pdf_page_quality_profile(entries, image_blocks, page_width=page_width, page_height=page_height)
                page_route = "digital"
                ocr_meta_summary: Dict[str, Any] = {}
                if _pdf_page_should_use_ocr(page_quality):
                    ocr_page = _ocr_compensate_pdf_page(page, page_no, {
                        "filename": filename,
                        "page_quality": page_quality,
                        "route": "pdf_digital_fast",
                    })
                    if ocr_page and ocr_page.get("entries"):
                        entries = list(ocr_page.get("entries") or [])
                        page_width = ocr_page.get("page_width") or page_width
                        page_height = ocr_page.get("page_height") or page_height
                        page_route = "ocr_compensation"
                        ocr_meta_summary = _summarize_ocr_meta(ocr_page.get("meta") or {})
                        if ocr_page.get("image_path"):
                            compensation_roots.add(os.path.dirname(ocr_page["image_path"]))
                page_profile = _build_visual_page_profile(entries, page_width=page_width, page_height=page_height)
                for index, entry in enumerate(entries):
                    text = entry["text"]
                    normalized_text = _normalize_heading_title(text)
                    bbox = entry.get("bbox")
                    layout = dict(entry.get("layout") or {})
                    next_text = ""
                    for candidate in entries[index + 1:]:
                        candidate_text = (candidate.get("text") or "").strip()
                        if candidate_text:
                            next_text = candidate_text
                            break
                    payload = {
                        "layout": layout,
                        "pdf_page_route": page_route,
                        "pdf_page_quality": page_quality,
                    }
                    if ocr_meta_summary:
                        payload["ocr_meta"] = ocr_meta_summary
                        payload["ocr_role"] = "page_compensation"
                    if _looks_like_toc_title(normalized_text):
                        _append_ir_element(document_ir, element_type="heading", text_raw=normalized_text, page_no=page_no, section_path=["toc"], bbox=bbox, json_payload={**payload, "pdf_role": "toc_title"}, parser_name="pymupdf")
                        toc_mode = True
                        continue
                    if toc_mode:
                        if _should_exit_visual_toc(normalized_text, next_text, layout, page_profile):
                            toc_mode = False
                        else:
                            toc_type = "heading" if _pdf_is_toc_entry_text(normalized_text) else "paragraph"
                            _append_ir_element(document_ir, element_type=toc_type, text_raw=normalized_text, page_no=page_no, section_path=["toc"], bbox=bbox, json_payload={**payload, "pdf_role": "toc_entry"}, parser_name="pymupdf")
                            continue
                    truncation_label = _law_semantic_truncation_label(normalized_text)
                    if truncation_label and main_body_started:
                        semantic_truncated = True
                        section_stack = []
                        _append_ir_element(document_ir, element_type="heading", text_raw=truncation_label, page_no=page_no, section_path=["appendix"], bbox=bbox, json_payload={**payload, "pdf_role": "appendix_heading", "appendix_label": truncation_label}, parser_name="pymupdf")
                        continue
                    if _is_visual_title_candidate(normalized_text, layout, page_profile, page_no, bool(document_ir.get("elements"))):
                        _append_ir_element(document_ir, element_type="title", text_raw=normalized_text, page_no=page_no, bbox=bbox, json_payload={**payload, "pdf_role": "title"}, parser_name="pymupdf")
                        continue
                    heading_level = None if semantic_truncated else _infer_visual_heading_level(normalized_text, layout, page_profile, next_text)
                    if heading_level is not None:
                        section_stack = section_stack[:max(0, heading_level - 2)] + [normalized_text]
                        main_body_started = True
                        _append_ir_element(document_ir, element_type="heading", text_raw=normalized_text, page_no=page_no, section_path=section_stack[:-1], bbox=bbox, json_payload={**payload, "pdf_role": "heading", "heading_level": heading_level}, parser_name="pymupdf")
                        continue
                    element_type = "list_item" if re.match(r"^([\-*•]|\d+[\.)])\s+", normalized_text) else "paragraph"
                    target_section_path = ["appendix"] if semantic_truncated else section_stack
                    if _clause_heading_label(normalized_text):
                        main_body_started = True
                    _append_ir_element(document_ir, element_type=element_type, text_raw=normalized_text, page_no=page_no, section_path=target_section_path, bbox=bbox, json_payload=payload, parser_name="pymupdf")
                if page_no < document.page_count:
                    _append_ir_element(document_ir, element_type="page_break", page_no=page_no, parser_name="pymupdf")
            document.close()
            for root in compensation_roots:
                _remove_temp_path(root)
            if document_ir.get("elements"):
                return document_ir
        except Exception:
            pass
    try:
        from pypdf import PdfReader
    except ImportError:
        raise HTTPException(status_code=500, detail="缺少 pypdf 依赖，无法解析 PDF")
    reader = PdfReader(io.BytesIO(content))
    if getattr(reader, "is_encrypted", False):
        raise HTTPException(status_code=400, detail="PDF 为加密文档，未提供解密密码，无法解析")
    document_ir = _new_document_ir(filename, metadata=metadata, doc_version=doc_version, parser_name="pypdf")
    for page_no, page in enumerate(reader.pages, start=1):
        _append_text_block_to_ir(document_ir, page.extract_text() or "", page_no=page_no, parser_name="pypdf")
        if page_no < len(reader.pages):
            _append_ir_element(document_ir, element_type="page_break", page_no=page_no, parser_name="pypdf")
    return document_ir


def _parse_pdf_ocr_document_ir(filename: str, content: bytes, metadata: Optional[Dict[str, Any]], doc_version: Optional[int], probe: Dict[str, Any]) -> Dict[str, Any]:
    parser_name = probe.get("parser_backend") or "ocr-route-degraded"
    document_ir = _new_document_ir(filename, metadata=metadata, doc_version=doc_version, parser_name=parser_name)
    if parser_name != "external_http_ocr":
        notice = "扫描 PDF / 图片文档已命中 OCR + layout 路由，但当前运行环境未接入可执行的 OCR/layout backend，暂未提取正文。"
        _append_ir_element(document_ir, element_type="figure", text_raw=notice, page_no=1, parser_name=parser_name, json_payload={"probe": probe, "degraded": True}, ocr_used=False)
        return document_ir
    rendered_pages: List[Dict[str, Any]] = []
    try:
        rendered_pages = _render_pdf_pages_for_ocr(content)
        page_results: List[Dict[str, Any]] = []
        for rendered in rendered_pages:
            payload = _call_external_ocr(rendered["image_path"])
            page_results.append({
                "page_no": rendered["page_no"],
                "image_path": rendered.get("image_path"),
                "lines": _extract_ocr_lines(payload),
                "texts": _extract_ocr_texts(payload),
                "meta": payload.get("meta") or {},
            })
        return _build_ocr_document_ir(filename, metadata, doc_version, parser_name, probe, page_results, empty_notice="OCR 服务已调用，但扫描 PDF 未识别到可索引正文。")
    except Exception as e:
        logger.warning("pdf_ocr_failed: source=%s err=%s", _safe_filename(filename), e)
        notice = "扫描 PDF 已命中 OCR 路由，但 OCR 服务暂不可用或未识别到可用正文。"
        _append_ir_element(document_ir, element_type="figure", text_raw=notice, page_no=1, parser_name=parser_name, json_payload={"probe": probe, "degraded": True, "error": str(e)}, ocr_used=False)
        return document_ir
    finally:
        cleanup_roots = {os.path.dirname(page.get("image_path") or "") for page in rendered_pages if page.get("image_path")}
        for root in cleanup_roots:
            _remove_temp_path(root)


def _parse_image_document_ir(filename: str, content: bytes, metadata: Optional[Dict[str, Any]], doc_version: Optional[int], probe: Dict[str, Any]) -> Dict[str, Any]:
    parser_name = probe.get("parser_backend") or "image-route-degraded"
    document_ir = _new_document_ir(filename, metadata=metadata, doc_version=doc_version, parser_name=parser_name)
    if parser_name != "external_http_ocr":
        _append_ir_element(
            document_ir,
            element_type="figure",
            text_raw="图片文档已路由到 OCR + layout 路径，但当前环境未安装 OCR backend。",
            page_no=1,
            parser_name=parser_name,
            json_payload={"probe": probe, "degraded": True},
            ocr_used=False,
        )
        return document_ir
    suffix = os.path.splitext(filename)[1] or ".img"
    temp_path = _write_temp_binary(content, suffix=suffix)
    try:
        payload = _call_external_ocr(temp_path)
        return _build_ocr_document_ir(
            filename,
            metadata,
            doc_version,
            parser_name,
            probe,
            [{"page_no": 1, "lines": _extract_ocr_lines(payload), "texts": _extract_ocr_texts(payload), "meta": payload.get("meta") or {}}],
            empty_notice="OCR 服务已调用，但图片中未识别到可索引正文。",
        )
    except Exception as e:
        logger.warning("image_ocr_failed: source=%s err=%s", _safe_filename(filename), e)
        _append_ir_element(
            document_ir,
            element_type="figure",
            text_raw="图片文档已路由到 OCR + layout 路径，但 OCR 服务暂不可用或未识别到可用正文。",
            page_no=1,
            parser_name=parser_name,
            json_payload={"probe": probe, "degraded": True, "error": str(e)},
            ocr_used=False,
        )
        return document_ir
    finally:
        _remove_temp_path(temp_path)


def _write_temp_binary(content: bytes, suffix: str) -> str:
    fd, path = tempfile.mkstemp(prefix="ocr_", suffix=suffix, dir=_resolve_ocr_temp_dir())
    try:
        with os.fdopen(fd, "wb") as f:
            f.write(content)
        try:
            os.chmod(path, 0o644)
        except Exception:
            pass
    except Exception:
        try:
            os.close(fd)
        except Exception:
            pass
        _remove_temp_path(path)
        raise
    return path


def _remove_temp_path(path: str):
    if not path:
        return
    try:
        if os.path.isdir(path):
            shutil.rmtree(path)
        else:
            os.remove(path)
    except FileNotFoundError:
        pass
    except Exception:
        pass


def _call_external_ocr(image_path: str) -> Dict[str, Any]:
    import requests

    service_url = (config.OCR_SERVICE_URL or "").strip()
    if not service_url:
        raise RuntimeError("ocr_service_url_not_configured")
    host_visible_path = _host_visible_ocr_path(image_path)
    response = requests.post(
        service_url,
        headers={"Content-Type": "application/json"},
        json={
            "image_path": host_visible_path,
            "mode": config.OCR_MODE,
            "lang": config.OCR_LANG,
        },
        timeout=config.OCR_TIMEOUT,
    )
    response.raise_for_status()
    payload = response.json()
    if not isinstance(payload, dict):
        raise RuntimeError("ocr_response_invalid")
    return payload


def _extract_ocr_texts(payload: Dict[str, Any]) -> List[str]:
    return [line.get("text") or "" for line in _extract_ocr_lines(payload) if (line.get("text") or "").strip()]


def _render_pdf_pages_for_ocr(content: bytes) -> List[Dict[str, Any]]:
    if not _module_available("fitz"):
        raise RuntimeError("pymupdf_not_available_for_pdf_ocr")
    import fitz  # type: ignore

    temp_dir = tempfile.mkdtemp(prefix="ocr_pdf_", dir=_resolve_ocr_temp_dir())
    try:
        os.chmod(temp_dir, 0o755)
    except Exception:
        pass
    pages: List[Dict[str, Any]] = []
    try:
        document = fitz.open(stream=content, filetype="pdf")
        for page_no, page in enumerate(document, start=1):
            pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            image_path = os.path.join(temp_dir, f"page_{page_no:04d}.png")
            pixmap.save(image_path)
            try:
                os.chmod(image_path, 0o644)
            except Exception:
                pass
            pages.append({"page_no": page_no, "image_path": image_path})
        document.close()
        if not pages:
            raise RuntimeError("pdf_has_no_renderable_pages")
        return pages
    except Exception:
        _remove_temp_path(temp_dir)
        raise


def _build_ocr_document_ir(
    filename: str,
    metadata: Optional[Dict[str, Any]],
    doc_version: Optional[int],
    parser_name: str,
    probe: Dict[str, Any],
    page_results: List[Dict[str, Any]],
    empty_notice: str,
) -> Dict[str, Any]:
    document_ir = _new_document_ir(filename, metadata=metadata, doc_version=doc_version, parser_name=parser_name)
    has_text = False
    section_stack: List[str] = []
    toc_mode = False
    main_body_started = False
    semantic_truncated = False
    for index, page in enumerate(page_results):
        page_no = int(page.get("page_no") or (index + 1))
        meta = page.get("meta") or {}
        meta_summary = _summarize_ocr_meta(meta)
        page_width = _safe_float(meta.get("page_width") or meta.get("width"))
        page_height = _safe_float(meta.get("page_height") or meta.get("height"))
        raw_lines = list(page.get("lines") or [])
        if not raw_lines:
            raw_lines = [{"text": text, "line_index": line_index} for line_index, text in enumerate(page.get("texts") or []) if (text or "").strip()]
        entries: List[Dict[str, Any]] = []
        for line in raw_lines:
            text = str(line.get("text") or "").strip()
            if not text:
                continue
            bbox = _normalize_bbox(line.get("bbox"))
            layout = _bbox_layout_metrics(bbox, page_width=page_width, page_height=page_height)
            font_size = _safe_float(line.get("font_size"))
            if font_size is not None:
                layout["font_size"] = font_size
            entries.append({
                "text": text,
                "bbox": bbox,
                "layout": layout,
                "confidence": _safe_float(line.get("confidence")),
                "line_index": int(line.get("line_index") if line.get("line_index") is not None else len(entries)),
                "meta": dict(line.get("meta") or {}),
            })
        page_profile = _build_visual_page_profile(entries, page_width=page_width, page_height=page_height)
        for line_index, entry in enumerate(entries):
            text = entry["text"]
            normalized_text = _normalize_heading_title(text)
            has_text = True
            bbox = entry.get("bbox")
            layout = dict(entry.get("layout") or {})
            next_text = ""
            for candidate in entries[line_index + 1:]:
                candidate_text = (candidate.get("text") or "").strip()
                if candidate_text:
                    next_text = candidate_text
                    break
            payload = {
                "probe": probe,
                "ocr_meta": meta_summary,
                "ocr_line_index": entry.get("line_index", line_index),
                "layout": layout,
            }
            if entry.get("meta"):
                payload["ocr_line_meta"] = _summarize_ocr_meta(entry.get("meta"))
            if _looks_like_toc_title(normalized_text):
                _append_ir_element(
                    document_ir,
                    element_type="heading",
                    text_raw=normalized_text,
                    page_no=page_no,
                    section_path=["toc"],
                    bbox=bbox,
                    parser_name=parser_name,
                    json_payload={**payload, "ocr_role": "toc_title"},
                    ocr_used=True,
                    ocr_confidence=entry.get("confidence"),
                )
                toc_mode = True
                continue
            if toc_mode:
                if _should_exit_visual_toc(normalized_text, next_text, layout, page_profile):
                    toc_mode = False
                else:
                    _append_ir_element(
                        document_ir,
                        element_type="paragraph",
                        text_raw=normalized_text,
                        page_no=page_no,
                        section_path=["toc"],
                        bbox=bbox,
                        parser_name=parser_name,
                        json_payload={**payload, "ocr_role": "toc_entry"},
                        ocr_used=True,
                        ocr_confidence=entry.get("confidence"),
                    )
                    continue
            truncation_label = _law_semantic_truncation_label(normalized_text)
            if truncation_label and main_body_started:
                semantic_truncated = True
                section_stack = []
                _append_ir_element(
                    document_ir,
                    element_type="heading",
                    text_raw=truncation_label,
                    page_no=page_no,
                    section_path=["appendix"],
                    bbox=bbox,
                    parser_name=parser_name,
                    json_payload={**payload, "ocr_role": "appendix_heading", "appendix_label": truncation_label},
                    ocr_used=True,
                    ocr_confidence=entry.get("confidence"),
                )
                continue
            heading_level = None if semantic_truncated else _infer_visual_heading_level(normalized_text, layout, page_profile, next_text)
            if heading_level is not None:
                section_stack = section_stack[:max(0, heading_level - 2)] + [normalized_text]
                main_body_started = True
                _append_ir_element(
                    document_ir,
                    element_type="heading",
                    text_raw=normalized_text,
                    page_no=page_no,
                    section_path=section_stack[:-1],
                    bbox=bbox,
                    parser_name=parser_name,
                    json_payload={**payload, "ocr_role": "heading", "heading_level": heading_level},
                    ocr_used=True,
                    ocr_confidence=entry.get("confidence"),
                )
                continue
            target_section_path = ["appendix"] if semantic_truncated else section_stack
            if _clause_heading_label(normalized_text):
                main_body_started = True
            _append_ir_element(
                document_ir,
                element_type="paragraph",
                text_raw=normalized_text,
                page_no=page_no,
                section_path=target_section_path,
                bbox=bbox,
                parser_name=parser_name,
                json_payload=payload,
                ocr_used=True,
                ocr_confidence=entry.get("confidence"),
            )
        if index < len(page_results) - 1:
            _append_ir_element(document_ir, element_type="page_break", page_no=page_no, parser_name=parser_name, json_payload={"probe": probe}, ocr_used=False)
    if not has_text:
        _append_ir_element(
            document_ir,
            element_type="figure",
            text_raw=empty_notice,
            page_no=1,
            parser_name=parser_name,
            json_payload={"probe": probe, "degraded": True, "ocr_pages": len(page_results)},
            ocr_used=False,
        )
    cleanup_roots = {os.path.dirname(page.get("image_path") or "") for page in page_results if page.get("image_path")}
    for root in cleanup_roots:
        _remove_temp_path(root)
    return document_ir


def _resolve_ocr_temp_dir() -> str:
    candidate = (config.OCR_SHARED_CONTAINER_DIR or "").strip() or UPLOAD_DIR
    try:
        os.makedirs(candidate, exist_ok=True)
        try:
            os.chmod(candidate, 0o755)
        except Exception:
            pass
        return candidate
    except Exception:
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        return UPLOAD_DIR


def _host_visible_ocr_path(local_path: str) -> str:
    local = (local_path or "").strip()
    container_dir = (config.OCR_SHARED_CONTAINER_DIR or "").strip()
    host_dir = (config.OCR_SHARED_HOST_DIR or "").strip()
    if not local or not container_dir or not host_dir:
        return local
    try:
        rel = os.path.relpath(local, container_dir)
    except Exception:
        return local
    if rel.startswith(".."):
        return local
    return os.path.normpath(os.path.join(host_dir, rel))


def _iter_docx_block_items(parent):
    from docx.document import Document as DocxDocument
    from docx.table import Table, _Cell
    from docx.text.paragraph import Paragraph

    if isinstance(parent, DocxDocument):
        parent_elm = parent.element.body
    elif isinstance(parent, _Cell):
        parent_elm = parent._tc
    else:
        parent_elm = getattr(parent, "_element", None)
    if parent_elm is None:
        return
    for child in parent_elm.iterchildren():
        local_name = child.tag.rsplit("}", 1)[-1]
        if local_name == "p":
            yield Paragraph(child, parent)
        elif local_name == "tbl":
            yield Table(child, parent)


def _docx_paragraph_has_page_break(paragraph: Any) -> bool:
    for run in getattr(paragraph, "runs", []) or []:
        if run._element.findall(f".//{{{DOCX_NS}}}br"):
            return True
    return False


def _docx_extract_figures_from_paragraph(paragraph: Any) -> List[Dict[str, Any]]:
    figures: List[Dict[str, Any]] = []
    for doc_pr in paragraph._element.findall(f".//{{{WP_NS}}}docPr"):
        figures.append({
            "name": doc_pr.get("name") or "image",
            "descr": doc_pr.get("descr") or "",
        })
    return figures


def _docx_style_heading_level(style_name: str) -> Optional[int]:
    name = (style_name or "").strip().lower()
    if not name:
        return None
    if "heading" in name or "标题" in name:
        level_match = re.search(r"(\d+)", name)
        return int(level_match.group(1)) if level_match else 1
    return None


def _docx_is_centered_paragraph(paragraph: Any) -> bool:
    try:
        alignment = getattr(getattr(paragraph, "paragraph_format", None), "alignment", None)
        return alignment is not None and int(alignment) == 1
    except Exception:
        return False


def _docx_has_bold_signal(paragraph: Any) -> bool:
    runs = [run for run in (getattr(paragraph, "runs", None) or []) if (run.text or "").strip()]
    if not runs:
        return False
    bold_runs = 0
    for run in runs:
        if bool(getattr(run, "bold", False)):
            bold_runs += 1
    return bold_runs > 0 and bold_runs >= max(1, len(runs) // 2)


def _docx_is_toc_title(text: str) -> bool:
    compact = re.sub(r"\s+", "", text or "")
    return compact in {"目录", "目次"}


def _docx_is_toc_entry_text(text: str) -> bool:
    s = (text or "").strip()
    if not s:
        return False
    if _docx_is_toc_title(s):
        return True
    if re.search(r"[.．·•…]{2,}\s*\d{1,3}$", s):
        return True
    if re.search(r"\d{1,3}$", s) and len(s) <= 40 and not _clause_heading_label(s):
        return True
    chapter = _chapter_heading_title(s)
    if chapter:
        return True
    if len(s) <= 20 and not _clause_heading_label(s) and not re.search(r"[。；：，,.]", s):
        return True
    return False


def _docx_heading_level_from_text(text: str) -> Optional[int]:
    s = (text or "").strip()
    if not s:
        return None
    if re.match(r"^第[一二三四五六七八九十百千0-9]+编(?:\s|$)", s):
        return 1
    if re.match(r"^第[一二三四五六七八九十百千0-9]+章(?:\s|$)", s):
        return 2
    if re.match(r"^第[一二三四五六七八九十百千0-9]+节(?:\s|$)", s):
        return 3
    if len(s) <= 20 and not _clause_heading_label(s) and not re.search(r"[。；：，,.]", s):
        return 2
    return None


def _docx_is_plain_heading_candidate(paragraph: Any, text: str, next_text: str = "") -> bool:
    s = (text or "").strip()
    if not s or _docx_is_toc_title(s) or _clause_heading_label(s):
        return False
    if re.match(r"^([\-*•]|\d+[\.)])\s+", s):
        return False
    if len(s) > 30 or re.search(r"[。；：，,.]", s):
        return False
    if _chapter_heading_title(s):
        return True
    next_value = (next_text or "").strip()
    if next_value and _clause_heading_label(next_value):
        return _docx_is_centered_paragraph(paragraph) or _docx_has_bold_signal(paragraph) or len(s) <= 10
    return (_docx_is_centered_paragraph(paragraph) or _docx_has_bold_signal(paragraph)) and len(s) <= 20


def _docx_should_exit_toc(paragraph: Any, text: str, next_text: str = "") -> bool:
    s = (text or "").strip()
    if not s:
        return False
    if _clause_heading_label(s):
        return True
    if _chapter_heading_title(s):
        next_value = (next_text or "").strip()
        return bool(next_value and (_clause_heading_label(next_value) or not _docx_is_toc_entry_text(next_value)))
    if _docx_is_plain_heading_candidate(paragraph, s, next_text):
        return True
    return len(s) > 40 and not _docx_is_toc_entry_text(s)


def _docx_heading_level(paragraph: Any, text: str, style_name: str, next_text: str = "") -> Optional[int]:
    style_level = _docx_style_heading_level(style_name)
    if style_level is not None:
        return style_level
    if _docx_is_plain_heading_candidate(paragraph, text, next_text):
        return _docx_heading_level_from_text(text)
    return None


def _docx_table_rows(table: Any) -> List[List[str]]:
    rows: List[List[str]] = []
    for row in getattr(table, "rows", []) or []:
        rows.append([((cell.text or "").strip()) for cell in getattr(row, "cells", []) or []])
    return rows


def _docx_extract_notes(content: bytes, part_name: str, note_tag: str) -> List[Dict[str, Any]]:
    notes: List[Dict[str, Any]] = []
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            if part_name not in zf.namelist():
                return notes
            root = ElementTree.fromstring(zf.read(part_name))
            for note in root.findall(f".//{{{DOCX_NS}}}{note_tag}"):
                note_type = note.get(f"{{{DOCX_NS}}}type")
                if note_type:
                    continue
                text = "".join([(node.text or "") for node in note.findall(f".//{{{DOCX_NS}}}t")]).strip()
                if text:
                    notes.append({"id": note.get(f"{{{DOCX_NS}}}id"), "text": text})
    except Exception:
        return []
    return notes


def _docx_extract_comments(content: bytes) -> List[Dict[str, Any]]:
    comments: List[Dict[str, Any]] = []
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            if "word/comments.xml" not in zf.namelist():
                return comments
            root = ElementTree.fromstring(zf.read("word/comments.xml"))
            for comment in root.findall(f".//{{{DOCX_NS}}}comment"):
                text = "".join([(node.text or "") for node in comment.findall(f".//{{{DOCX_NS}}}t")]).strip()
                if text:
                    comments.append({
                        "id": comment.get(f"{{{DOCX_NS}}}id"),
                        "author": comment.get(f"{{{DOCX_NS}}}author"),
                        "text": text,
                    })
    except Exception:
        return []
    return comments


def _parse_docx_document_ir(filename: str, content: bytes, metadata: Optional[Dict[str, Any]], doc_version: Optional[int], parser_name: str) -> Dict[str, Any]:
    try:
        from docx import Document
        from docx.table import Table
        from docx.text.paragraph import Paragraph
    except ImportError:
        raise HTTPException(status_code=500, detail="缺少 python-docx 依赖，无法解析 DOCX")
    document = Document(io.BytesIO(content))
    document_ir = _new_document_ir(filename, metadata=metadata, doc_version=doc_version, parser_name=parser_name)
    section_stack: List[str] = []

    def append_table(table: Any, page_no: int, scope_path: Optional[List[str]] = None):
        active_path = list(scope_path or section_stack)
        rows = _docx_table_rows(table)
        markdown = "\n".join([" | ".join(row) for row in rows if any(cell for cell in row)])
        _append_ir_element(document_ir, element_type="table", text_raw=markdown, text_normalized=_normalize_ir_text(markdown), markdown=markdown, json_payload={"table_json": {"rows": rows}, "table_text": markdown}, page_no=page_no, section_path=active_path, parser_name=parser_name)

    def append_paragraph(paragraph: Any, page_no: int, scope_path: Optional[List[str]] = None, *, next_text: str = "", toc_mode: bool = False) -> bool:
        nonlocal section_stack
        text = (paragraph.text or "").strip()
        style_name = ((paragraph.style.name or "") if paragraph.style else "").strip().lower()
        active_path = list(scope_path or section_stack)
        in_special_scope = bool(scope_path)
        figures = _docx_extract_figures_from_paragraph(paragraph)
        for figure in figures:
            _append_ir_element(document_ir, element_type="figure", text_raw=figure.get("descr") or figure.get("name") or "image", page_no=page_no, section_path=active_path, parser_name=parser_name, json_payload=figure)
        if text:
            if not in_special_scope and _docx_is_toc_title(text):
                _append_ir_element(document_ir, element_type="heading", text_raw=text, page_no=page_no, section_path=["toc"], parser_name=parser_name, json_payload={"docx_role": "toc_title"})
                return True
            if toc_mode and not in_special_scope:
                if _docx_should_exit_toc(paragraph, text, next_text):
                    toc_mode = False
                    active_path = list(section_stack)
                else:
                    _append_ir_element(document_ir, element_type="paragraph", text_raw=text, page_no=page_no, section_path=["toc"], parser_name=parser_name, json_payload={"docx_role": "toc_entry"})
                    if _docx_paragraph_has_page_break(paragraph):
                        _append_ir_element(document_ir, element_type="page_break", page_no=page_no, section_path=["toc"], parser_name=parser_name)
                    return True
            heading_level = None if in_special_scope else _docx_heading_level(paragraph, text, style_name, next_text)
            if heading_level is not None:
                level = heading_level
                if level <= 1 and not document_ir.get("elements"):
                    _append_ir_element(document_ir, element_type="title", text_raw=text, page_no=page_no, section_path=active_path, parser_name=parser_name)
                else:
                    section_stack = section_stack[:max(0, level - 2)] + [text]
                    _append_ir_element(document_ir, element_type="heading", text_raw=text, page_no=page_no, section_path=section_stack[:-1], parser_name=parser_name)
            elif style_name.startswith("list") or re.match(r"^([\-*•]|\d+[\.)])\s+", text):
                _append_ir_element(document_ir, element_type="list_item", text_raw=text, page_no=page_no, section_path=active_path, parser_name=parser_name)
            elif "caption" in style_name:
                _append_ir_element(document_ir, element_type="caption", text_raw=text, page_no=page_no, section_path=active_path, parser_name=parser_name)
            else:
                _append_text_block_to_ir(document_ir, text, page_no=page_no, base_section_path=active_path, parser_name=parser_name)
        if _docx_paragraph_has_page_break(paragraph):
            _append_ir_element(document_ir, element_type="page_break", page_no=page_no, section_path=active_path, parser_name=parser_name)
        return toc_mode

    body_blocks = list(_iter_docx_block_items(document) or [])
    toc_mode = False
    for index, block in enumerate(body_blocks):
        if block is None:
            continue
        next_text = ""
        for candidate in body_blocks[index + 1:]:
            if candidate is None or candidate.__class__.__name__ != "Paragraph":
                continue
            next_text = ((candidate.text or "").strip())
            if next_text:
                break
        if block.__class__.__name__ == "Paragraph":
            toc_mode = append_paragraph(block, page_no=1, next_text=next_text, toc_mode=toc_mode)
        elif block.__class__.__name__ == "Table":
            append_table(block, page_no=1)

    for section_idx, section in enumerate(document.sections, start=1):
        for region_name, region in (("header", section.header), ("footer", section.footer)):
            region_path = [f"{region_name}_{section_idx}"]
            for block in _iter_docx_block_items(region):
                if block is None:
                    continue
                if block.__class__.__name__ == "Paragraph":
                    append_paragraph(block, page_no=section_idx, scope_path=region_path)
                elif block.__class__.__name__ == "Table":
                    append_table(block, page_no=section_idx, scope_path=region_path)

    for footnote in _docx_extract_notes(content, "word/footnotes.xml", "footnote"):
        _append_ir_element(document_ir, element_type="key_value", text_raw=footnote["text"], page_no=1, section_path=["footnotes"], parser_name=parser_name, json_payload=footnote)
    for endnote in _docx_extract_notes(content, "word/endnotes.xml", "endnote"):
        _append_ir_element(document_ir, element_type="key_value", text_raw=endnote["text"], page_no=1, section_path=["endnotes"], parser_name=parser_name, json_payload=endnote)
    for comment in _docx_extract_comments(content):
        _append_ir_element(document_ir, element_type="key_value", text_raw=comment["text"], page_no=1, section_path=["comments"], parser_name=parser_name, json_payload=comment)
    return document_ir


def _normalize_sheet_cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text or None
    return value


def _sheet_range_has_values(sheet: Any, min_row: int, min_col: int, max_row: int, max_col: int) -> bool:
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if _normalize_sheet_cell_value(cell.value) is not None:
                return True
    return False


def _build_xlsx_table_payload(workbook_name: str, sheet: Any, value_sheet: Any, ref: str, table_name: Optional[str] = None) -> Dict[str, Any]:
    from openpyxl.utils import get_column_letter, range_boundaries

    min_col, min_row, max_col, max_row = range_boundaries(ref)
    headers = []
    for col_idx in range(min_col, max_col + 1):
        header_value = _normalize_sheet_cell_value(value_sheet.cell(min_row, col_idx).value)
        headers.append(str(header_value) if header_value is not None else get_column_letter(col_idx))
    merged_ranges = []
    for merged in getattr(sheet.merged_cells, "ranges", []) or []:
        m_min_col, m_min_row, m_max_col, m_max_row = range_boundaries(str(merged))
        if not (m_max_row < min_row or m_min_row > max_row or m_max_col < min_col or m_min_col > max_col):
            merged_ranges.append(str(merged))
    rows_json = []
    formulas = []
    row_texts = []
    for row_idx in range(min_row + 1, max_row + 1):
        row_hidden = bool(getattr(sheet.row_dimensions.get(row_idx), "hidden", False))
        row_cells = []
        text_parts = []
        for col_idx in range(min_col, max_col + 1):
            column_letter = get_column_letter(col_idx)
            coord = f"{column_letter}{row_idx}"
            formula_value = sheet[coord].value
            display_value = _normalize_sheet_cell_value(value_sheet[coord].value)
            hidden = row_hidden or bool(getattr(sheet.column_dimensions.get(column_letter), "hidden", False))
            header = headers[col_idx - min_col] if (col_idx - min_col) < len(headers) else column_letter
            row_cells.append({
                "coord": coord,
                "row": row_idx,
                "column": column_letter,
                "header": header,
                "display_value": display_value,
                "formula": formula_value if isinstance(formula_value, str) and formula_value.startswith("=") else None,
                "hidden": hidden,
            })
            if isinstance(formula_value, str) and formula_value.startswith("="):
                formulas.append({"coord": coord, "formula": formula_value, "display_value": display_value})
            if display_value is not None:
                text_parts.append(f"{header}({coord}): {display_value}")
        rows_json.append({"row_index": row_idx, "hidden": row_hidden, "cells": row_cells})
        if text_parts:
            row_texts.append(" | ".join(text_parts))
    table_text = f"Sheet: {sheet.title}\nRange: {ref}\n" + "\n".join(row_texts)
    table_json = {
        "workbook": workbook_name,
        "sheet": sheet.title,
        "table_name": table_name or f"{sheet.title}!{ref}",
        "range": ref,
        "header_row": min_row,
        "headers": headers,
        "row_count": max(0, max_row - min_row),
        "column_count": max(0, max_col - min_col + 1),
        "merged_cells": merged_ranges,
        "hidden_policy": "preserve_with_visibility_flags",
        "formula_policy": "preserve_formula_and_display",
        "empty_value_strategy": "null",
        "repeated_value_strategy": "preserve_source",
        "rows": rows_json,
        "formulas": formulas,
    }
    return {"table_text": table_text.strip(), "table_json": table_json}


def _parse_xlsx_document_ir(filename: str, content: bytes, metadata: Optional[Dict[str, Any]], doc_version: Optional[int], parser_name: str) -> Dict[str, Any]:
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="缺少 openpyxl 依赖，无法解析 XLSX")
    workbook_formula = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    workbook_value = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    document_ir = _new_document_ir(filename, metadata=metadata, doc_version=doc_version, parser_name=parser_name)
    for sheet_idx, (sheet, value_sheet) in enumerate(zip(workbook_formula.worksheets, workbook_value.worksheets), start=1):
        section_path = [sheet.title]
        _append_ir_element(document_ir, element_type="sheet", text_raw=sheet.title, page_no=sheet_idx, section_path=section_path, parser_name=parser_name, json_payload={"sheet_state": sheet.sheet_state})
        seen_refs = set()
        for table_name, table_obj in (getattr(sheet, "tables", {}) or {}).items():
            ref = getattr(table_obj, "ref", None) or str(table_obj)
            if not ref:
                continue
            seen_refs.add(ref)
            payload = _build_xlsx_table_payload(filename, sheet, value_sheet, ref, table_name=table_name)
            _append_ir_element(document_ir, element_type="table", text_raw=payload["table_text"], text_normalized=_normalize_ir_text(payload["table_text"]), markdown=payload["table_text"], json_payload=payload, page_no=sheet_idx, section_path=section_path, parser_name=parser_name)
        try:
            fallback_ref = sheet.calculate_dimension()
        except Exception:
            fallback_ref = None
        if fallback_ref and fallback_ref not in seen_refs:
            from openpyxl.utils import range_boundaries

            min_col, min_row, max_col, max_row = range_boundaries(fallback_ref)
            if _sheet_range_has_values(value_sheet, min_row, min_col, max_row, max_col):
                payload = _build_xlsx_table_payload(filename, sheet, value_sheet, fallback_ref)
                _append_ir_element(document_ir, element_type="table", text_raw=payload["table_text"], text_normalized=_normalize_ir_text(payload["table_text"]), markdown=payload["table_text"], json_payload=payload, page_no=sheet_idx, section_path=section_path, parser_name=parser_name)
    workbook_formula.close()
    workbook_value.close()
    return document_ir


def extract_document_ir_from_file(filename: str, content: bytes, metadata: Optional[Dict[str, Any]] = None, doc_version: Optional[int] = None) -> Dict[str, Any]:
    """从探测与路由后的解析路径提取统一 Document IR"""
    probe = _probe_file_for_parser(filename, content)
    detected_ext = (probe.get("detected_ext") or probe.get("extension") or "").lower()
    effective_metadata = _merge_parser_metadata(metadata, probe)

    if detected_ext not in SUPPORTED_FILE_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的文件格式: {detected_ext or os.path.splitext(filename)[1].lower()}。支持格式: {', '.join(sorted(SUPPORTED_FILE_EXTENSIONS))}"
        )

    if detected_ext in {".txt", ".md", ".markdown", ".log"}:
        return _build_document_ir_from_text(filename, decode_text_bytes(content), metadata=effective_metadata, parser_name=probe.get("parser_backend") or detected_ext.lstrip("."), doc_version=doc_version)
    if detected_ext == ".json":
        return _parse_json_document_ir(filename, content, effective_metadata, doc_version, parser_name=probe.get("parser_backend") or "json")
    if detected_ext == ".csv":
        return _parse_csv_document_ir(filename, content, effective_metadata, doc_version, parser_name=probe.get("parser_backend") or "csv")
    if detected_ext == ".pdf":
        if probe.get("route") == "pdf_ocr_layout":
            return _parse_pdf_ocr_document_ir(filename, content, effective_metadata, doc_version, probe)
        return _parse_pdf_fast_document_ir(filename, content, effective_metadata, doc_version, backend=probe.get("parser_backend") or "pypdf")
    if detected_ext == ".docx":
        return _parse_docx_document_ir(filename, content, effective_metadata, doc_version, parser_name="python-docx")
    if detected_ext == ".xlsx":
        return _parse_xlsx_document_ir(filename, content, effective_metadata, doc_version, parser_name="openpyxl")
    if detected_ext in {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".gif"}:
        return _parse_image_document_ir(filename, content, effective_metadata, doc_version, probe)
    raise HTTPException(status_code=400, detail="无法处理该文件格式")


def extract_text_from_file(filename: str, content: bytes) -> str:
    return _document_ir_plain_text(extract_document_ir_from_file(filename, content), normalized=False)


async def index_document(filename: str, text: str, metadata: Optional[Dict[str, Any]] = None, document_ir: Optional[Dict[str, Any]] = None) -> int:
    """文档分块、向量化并写入 Milvus"""
    if document_ir and (document_ir.get("elements") or []):
        _store_document_ir(filename, document_ir)
    chunk_items = _contextualize_chunk_items(
        filename,
        _prepare_structured_items(filename, text, config.CHUNK_SIZE, config.OVERLAP, document_ir=document_ir),
    )
    if not chunk_items:
        raise HTTPException(status_code=400, detail="文档内容为空，无法索引")

    embedding_service = EmbeddingService()
    vector_db = VectorDBService()
    embeddings = await embedding_service.embed_batched([c["text"] for c in chunk_items], per_request=64, timeout=60, retries=2)

    now = datetime.now().isoformat()
    source_text = _document_ir_plain_text(document_ir, normalized=False) if document_ir else text
    doc_type = _classify_doc_type(filename, source_text)
    topics = _infer_topics(source_text)
    base_metadata = {**(metadata or {}), "doc_type": doc_type, "topics": topics}
    docs = []
    for item, embedding in zip(chunk_items, embeddings):
        chunk = item["text"]
        raw_text = item.get("raw_text") or chunk
        section = (item.get("section") or "").strip()
        section_id = item.get("section_id")
        chunk_id = int(item.get("chunk_id", 0))
        vector_payload = _summarize_chunk_payload(item.get("payload") or {})
        vector_metadata = _milvus_safe_metadata({
            **base_metadata,
            "chunk_id": chunk_id,
            "chunk_count": len(chunk_items),
            "section": section,
            "section_id": section_id,
            "section_title": item.get("section_title") or section,
            "section_node_id": item.get("section_node_id"),
            "clause_label": item.get("clause_label") or "",
            "article_no": item.get("article_no") or item.get("clause_label") or "",
            "unit_kind": item.get("unit_kind") or "paragraph",
            "raw_text": raw_text,
            "text_normalized": item.get("normalized_text") or _normalize_ir_text(raw_text),
            "fts_text": item.get("fts_text") or raw_text,
            "page_no": item.get("page_no"),
            "page_span": item.get("page_span") or [],
            "section_path": item.get("section_path") or [],
            "parent_section_id": item.get("parent_section_id"),
            "parent_section_path": item.get("parent_section_path") or [],
            "parent_section_title": item.get("parent_section_title"),
            "section_depth": item.get("section_depth"),
            "semantic_unit_ids": item.get("semantic_unit_ids") or [],
            "chunk_role": item.get("chunk_role") or "body",
            "payload": vector_payload,
            "element_id": item.get("element_id"),
            "element_type": item.get("element_type"),
            "reading_order": item.get("reading_order"),
            "prev_chunk_id": item.get("prev_chunk_id"),
            "next_chunk_id": item.get("next_chunk_id"),
        })
        docs.append({
            "embedding": embedding,
            "text": chunk,
            "source": filename,
            "metadata": vector_metadata,
            "created_at": now
        })

    vector_db.insert(docs)
    for item in docs:
        _lex_db_add_chunk_sql(
            filename,
            _fts_storage_text(item["text"], item.get("metadata") or {}),
            item.get("metadata", {}).get("section") or "",
            item.get("metadata") or {},
            int(item.get("metadata", {}).get("chunk_id", 0)),
        )
    return len(docs)

async def index_document_incremental(task_id: str, filename: str, text: str, metadata: Optional[Dict[str, Any]] = None, document_ir: Optional[Dict[str, Any]] = None) -> int:
    if document_ir and (document_ir.get("elements") or []):
        _store_document_ir(filename, document_ir)
    items = _contextualize_chunk_items(
        filename,
        _prepare_structured_items(filename, text, config.CHUNK_SIZE, config.OVERLAP, document_ir=document_ir),
    )
    if not items:
        raise HTTPException(status_code=400, detail="文档内容为空，无法索引")

    embedding_service = EmbeddingService()
    vector_db = VectorDBService()
    now = datetime.now().isoformat()
    source_text = _document_ir_plain_text(document_ir, normalized=False) if document_ir else text
    doc_type = _classify_doc_type(filename, source_text)
    topics = _infer_topics(source_text)
    base_metadata = {**(metadata or {}), "doc_type": doc_type, "topics": topics}
    total = len(items)
    done = 0
    batch: List[Dict[str, Any]] = []
    batch_tokens = 0
    max_batch_tokens = 8000
    max_batch_items = 64
    chunk_id = 0

    def _build_doc_entry(batch_item: Dict[str, Any], embedding_value: Any) -> Dict[str, Any]:
        nonlocal chunk_id
        vector_payload = _summarize_chunk_payload(batch_item.get("payload") or {})
        vector_metadata = _milvus_safe_metadata({
            **base_metadata,
            "chunk_id": chunk_id,
            "chunk_count": total,
            "section": batch_item["section"],
            "section_id": batch_item.get("section_id"),
            "section_title": batch_item.get("section_title") or batch_item.get("section") or "",
            "section_node_id": batch_item.get("section_node_id"),
            "clause_label": batch_item.get("clause_label") or "",
            "article_no": batch_item.get("article_no") or batch_item.get("clause_label") or "",
            "unit_kind": batch_item.get("unit_kind") or "paragraph",
            "raw_text": batch_item.get("raw_text") or batch_item["text"],
            "text_normalized": batch_item.get("normalized_text") or _normalize_ir_text(batch_item.get("raw_text") or batch_item["text"]),
            "fts_text": batch_item.get("fts_text") or batch_item.get("raw_text") or batch_item["text"],
            "page_no": batch_item.get("page_no"),
            "page_span": batch_item.get("page_span") or [],
            "section_path": batch_item.get("section_path") or [],
            "parent_section_id": batch_item.get("parent_section_id"),
            "parent_section_path": batch_item.get("parent_section_path") or [],
            "parent_section_title": batch_item.get("parent_section_title"),
            "section_depth": batch_item.get("section_depth"),
            "semantic_unit_ids": batch_item.get("semantic_unit_ids") or [],
            "chunk_role": batch_item.get("chunk_role") or "body",
            "payload": vector_payload,
            "element_id": batch_item.get("element_id"),
            "element_type": batch_item.get("element_type"),
            "reading_order": batch_item.get("reading_order"),
            "prev_chunk_id": batch_item.get("prev_chunk_id"),
            "next_chunk_id": batch_item.get("next_chunk_id"),
            "doc_version": _doc_get(filename).get("pending_version") or _doc_get(filename).get("active_version") or 1,
            "rebuild_seq": now,
        })
        entry = {
            "embedding": embedding_value,
            "text": batch_item["text"],
            "source": filename,
            "metadata": vector_metadata,
            "created_at": now,
        }
        chunk_id += 1
        return entry

    async def _flush_batch(savepoint_name: str):
        nonlocal batch, batch_tokens, done
        if not batch:
            return
        _lex_tx_savepoint(savepoint_name)
        texts = [item["text"] for item in batch]
        docs: List[Dict[str, Any]] = []
        if not config.TEST_LEX_ONLY:
            embs = await embedding_service.embed_batched(texts, per_request=32, timeout=60, retries=2)
            for batch_item, emb in zip(batch, embs):
                docs.append(_build_doc_entry(batch_item, emb))
            vector_db.insert(docs)
        else:
            for batch_item in batch:
                docs.append(_build_doc_entry(batch_item, None))
        done += len(batch)
        TASKS[task_id]["status"] = "indexing"
        TASKS[task_id]["stage"] = "embedding_partial"
        TASKS[task_id]["chunks_indexed"] = done
        _task_log(task_id, "embedding_batch_done", {"done": done, "total": total})
        for item in docs:
            _lex_db_add_chunk_sql(
                filename,
                _fts_storage_text(item["text"], item.get("metadata") or {}),
                item.get("metadata", {}).get("section") or "",
                item.get("metadata") or {},
                int(item.get("metadata", {}).get("chunk_id", 0)),
            )
        _lex_tx_release(savepoint_name)
        batch = []
        batch_tokens = 0

    for item in items:
        text_chunk = item.get("text") or ""
        raw_text = item.get("raw_text") or text_chunk
        token_count = _estimate_tokens(text_chunk)
        if batch and (batch_tokens + token_count > max_batch_tokens or len(batch) >= max_batch_items):
            await _flush_batch("batch_write")
        batch.append({
            "section": (item.get("section") or "").strip(),
            "text": text_chunk,
            "raw_text": raw_text,
            "normalized_text": item.get("normalized_text") or _normalize_ir_text(raw_text),
            "fts_text": item.get("fts_text") or raw_text,
            "page_no": item.get("page_no"),
            "section_path": item.get("section_path") or [],
            "element_id": item.get("element_id"),
            "element_type": item.get("element_type"),
            "reading_order": item.get("reading_order"),
        })
        batch_tokens += token_count

    if batch:
        await _flush_batch("batch_write_last")
    return done


# ==================== API 端点 ====================

# 优先处理 API 路由，然后再处理静态文件
@app.get("/")
async def root():
    # 如果是浏览器访问首页，重定向到 index.html
    return FileResponse(os.path.join(WEB_DIR, "index.html"))

@app.get("/favicon.ico")
async def favicon():
    return Response(status_code=204)


@app.get("/health")
async def health_check():
    """健康检查"""
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "services": {
            "vector_db": "connected",
            "embedding": "connected",
            "rerank": "connected"
        }
    }

@app.on_event("startup")
async def _startup():
    _bootstrap_runtime_state_from_legacy_or_milvus()
    _load_tasks()
    try:
        report = reconcile_sqlite_with_milvus(prune_sqlite_orphans=True)
        logger.info(
            "startup_reconcile: healthy=%s milvus=%s sqlite_after=%s upserted=%s pruned=%s",
            report.get("healthy"),
            report.get("milvus_source_count"),
            report.get("sqlite_source_count_after"),
            report.get("upserted_sources"),
            len(report.get("pruned_sources") or []),
        )
    except Exception as e:
        logger.warning(f"startup_reconcile_failed: {e}")
    asyncio.create_task(_compensation_worker(interval_sec=5))


@app.post("/query", response_model=QueryResponse)
async def query(query_req: QueryRequest):
    """处理查询"""
    try:
        payload = query_req.model_dump()
    except Exception:
        try:
            payload = query_req.dict()
        except Exception:
            payload = str(query_req)

    print({
        "event": "query.req",
        "payload": payload,
        "query_req_type": type(query_req).__name__,
    }, flush=True)
    handler = QueryHandler()
    pending = _get_pending_clarification(query_req.user_id)
    if pending:
        candidates = list((pending or {}).get("candidates") or [])
        picked_idx = _parse_pending_candidate_selection(query_req.query, len(candidates))
        if picked_idx is not None:
            locked_source = _normalize_filename_for_match(candidates[picked_idx] if picked_idx < len(candidates) else "")
            base_query = _normalize_query((pending or {}).get("query") or "")
            if locked_source and base_query:
                result = await handler.process(
                    query=base_query,
                    user_id=query_req.user_id,
                    top_k=query_req.top_k,
                    enable_rerank=bool(query_req.enable_rerank),
                    forced_fnames=[locked_source],
                )
                _clear_pending_clarification(query_req.user_id)
                meta = dict(result.get("metadata") or {})
                meta["clarification_resolved"] = True
                meta["clarification_selected_source"] = locked_source
                meta["clarification_original_query"] = base_query
                result["metadata"] = meta
                return QueryResponse(**result)

    result = await handler.process(
        query=query_req.query,
        user_id=query_req.user_id,
        top_k=query_req.top_k,
        enable_rerank=bool(query_req.enable_rerank)
    )
    meta = dict(result.get("metadata") or {})
    candidates = list(meta.get("candidate_sources") or [])
    if meta.get("answer_mode") == "clarification" and candidates:
        _set_pending_clarification(
            user_id=query_req.user_id,
            query=meta.get("query") or query_req.query,
            candidates=candidates,
            reason=(meta.get("refused") or meta.get("blocked") or "document_clarification"),
        )
    return QueryResponse(**result)


@app.post("/retrieve")
async def retrieve(query_req: QueryRequest):
    """仅检索，不调用 LLM"""
    handler = QueryHandler()
    result = await handler.retrieve(
        query=query_req.query,
        user_id=query_req.user_id,
        top_k=query_req.top_k,
        enable_rerank=bool(query_req.enable_rerank)
    )
    return result


@app.post("/documents")
async def upload_document(doc_req: DocumentRequest):
    safe_name = _safe_filename(doc_req.filename)
    probe = _build_text_upload_probe(safe_name, doc_req.content)
    _validate_upload_probe(safe_name, probe, is_text_upload=True)
    title_profile = _doc_title_profile(safe_name)
    content_sha256 = _content_sha256_text(doc_req.content)
    duplicate_info = _detect_duplicate_upload(safe_name, content_sha256, title_profile["canonical_title"])
    if duplicate_info.get("duplicate_state") in {"no_change", "already_exists"}:
        return _build_upload_response(
            task_id=None,
            source=safe_name,
            task_status="completed",
            document_status=str(duplicate_info.get("duplicate_state") or "no_change"),
            searchable=bool(_doc_get(duplicate_info.get("duplicate_of") or safe_name).get("searchable")),
            duplicate_state=duplicate_info.get("duplicate_state"),
            duplicate_of=duplicate_info.get("duplicate_of"),
            same_title_candidates=duplicate_info.get("same_title_candidates") or [],
        )
    task_id = _new_task_id()
    TASKS[task_id] = {
        "status": "accepted",
        "stage": "validating",
        "filename": safe_name,
        "created_at": datetime.now().isoformat(),
        "payload": {"text": doc_req.content, "metadata": doc_req.metadata},
        "document_status": "accepted",
    }
    _task_log(task_id, "validating", {"filename": safe_name})
    _lex_db_set_status(safe_name, "accepted")
    dt = (doc_req.metadata or {}).get("doc_type")
    topic = (doc_req.metadata or {}).get("topic")
    source_id = _build_source_id(doc_req.filename, content_sha256)
    _doc_upsert(safe_name, status="accepted", canonical_title=title_profile["canonical_title"], title_tokens=title_profile["title_tokens"], aliases=title_profile["aliases"], filename_stem=title_profile["stem"], doc_type=dt, topic=topic, source_id=source_id, original_filename=doc_req.filename, content_sha256=content_sha256, mime_type=probe.get("mime_type"), detected_ext=probe.get("detected_ext"), file_size=probe.get("file_size"), page_count=probe.get("page_count"), parser_route=probe.get("route"), parser_backend=probe.get("parser_backend"), parse_status="accepted", searchable=0, duplicate_state=duplicate_info.get("duplicate_state"), duplicate_of=duplicate_info.get("duplicate_of"), same_title_group=_same_title_group(title_profile["canonical_title"]), suspicious_file_type=0)
    _docfts_upsert(safe_name, title=title_profile["canonical_title"], aliases=title_profile["aliases"], doc_type=dt, topic=topic)
    async def _run():
        try:
            lock = _get_source_lock(safe_name)
            if not lock.acquire(timeout=30):
                raise HTTPException(status_code=429, detail="同一文档正在处理，请稍后重试")
            _lex_tx_begin()
            _lex_db_set_status(safe_name, "reindexing")
            v_next = _doc_next_version(safe_name)
            _doc_upsert(safe_name, status="reindexing", pending_version=v_next, parse_status="parsing", searchable=_doc_searchable_flag(safe_name))
            TASKS[task_id]["status"] = "indexing"
            TASKS[task_id]["stage"] = "parsing"
            _task_log(task_id, "parsing")
            document_ir = _build_document_ir_from_text(
                safe_name,
                doc_req.content,
                metadata=doc_req.metadata,
                parser_name="direct_text",
                doc_version=v_next,
            )
            text = _document_ir_plain_text(document_ir, normalized=False)
            quality = _assess_document_quality(document_ir, probe)
            if quality["status"] == "parse_empty":
                raise HTTPException(status_code=400, detail="parse_empty: 解析后正文为空")
            if quality["status"] == "parse_low_quality":
                raise HTTPException(status_code=400, detail="parse_low_quality: 解析质量不足，未进入发布")
            TASKS[task_id]["status"] = "indexing"
            TASKS[task_id]["stage"] = "profile_building"
            _task_log(task_id, "profile_building", {"quality": quality})
            profile = _build_document_profile(safe_name, doc_req.filename, source_id, content_sha256, text, document_ir, probe, quality, metadata=doc_req.metadata)
            _doc_upsert(safe_name, status="reindexing", pending_version=v_next, parse_status=quality["status"], parse_quality_score=quality["score"], quality_flags=_json_dumps(quality["flags"]), canonical_title=profile["canonical_title"], title_tokens=" ".join(profile.get("title_aliases") or []), aliases=",".join((profile.get("title_aliases") or [])[1:]), filename_stem=_filename_stem(safe_name), doc_type=profile.get("doc_type"), topic=",".join((profile.get("topic_terms") or [])[:8]), source_id=source_id, original_filename=doc_req.filename, content_sha256=content_sha256, mime_type=probe.get("mime_type"), detected_ext=probe.get("detected_ext"), file_size=probe.get("file_size"), page_count=probe.get("page_count"), parser_route=probe.get("route"), parser_backend=probe.get("parser_backend"), searchable=_doc_searchable_flag(safe_name), duplicate_state=duplicate_info.get("duplicate_state"), duplicate_of=duplicate_info.get("duplicate_of"), same_title_group=_same_title_group(profile["canonical_title"]))
            TASKS[task_id]["status"] = "indexing"
            TASKS[task_id]["stage"] = "embedding"
            _task_log(task_id, "embedding")
            # Reindex is versioned: only purge remnants of the target pending version.
            _purge_source_for_reindex(safe_name, v_next)
            total_done = await index_document_incremental(
                task_id=task_id,
                filename=safe_name,
                text=text,
                metadata=doc_req.metadata,
                document_ir=document_ir,
            )
            _persist_document_profile(safe_name, v_next, profile)
            _crash_inject("before_commit")
            _lex_db_set_status(safe_name, "vector_pending")
            _doc_upsert(safe_name, status="vector_pending", pending_version=v_next, last_error=None, parse_status=quality["status"], searchable=_doc_searchable_flag(safe_name))
            _lex_tx_commit()
            _lex_db_checkpoint("PASSIVE")
            TASKS[task_id]["status"] = "completed"
            TASKS[task_id]["stage"] = "publish_pending"
            TASKS[task_id]["chunks_indexed"] = total_done
            TASKS[task_id]["document_status"] = "vector_pending"
            _task_log(task_id, "publish_pending", {"chunks_indexed": total_done, "pending_version": v_next})
        except asyncio.CancelledError:
            _lex_tx_rollback()
            TASKS[task_id]["status"] = "failed"
            TASKS[task_id]["stage"] = "cancelled"
            TASKS[task_id]["error"] = "cancelled_for_delete"
            _doc_upsert(safe_name, status="pending_delete", last_error="cancelled_for_delete", searchable=0)
            _task_log(task_id, "cancelled", {"reason": "delete_requested"})
            raise
        except HTTPException as e:
            _lex_tx_rollback()
            detail = str(e.detail or "upload_failed")
            status_code = detail.split(":", 1)[0]
            TASKS[task_id]["status"] = "failed"
            TASKS[task_id]["stage"] = status_code
            TASKS[task_id]["error"] = detail
            _doc_upsert(safe_name, status=status_code, last_error=detail, parse_status=status_code, searchable=_doc_searchable_flag(safe_name))
            _task_log(task_id, "failed", {"error": detail})
        except Exception as e:
            _lex_tx_rollback()
            TASKS[task_id]["status"] = "failed"
            TASKS[task_id]["stage"] = "failed"
            TASKS[task_id]["error"] = str(e)
            _doc_upsert(safe_name, status="vector_failed", last_error=str(e), searchable=_doc_searchable_flag(safe_name))
            _task_log(task_id, "failed", {"error": str(e)})
        finally:
            try:
                lock.release()
            except Exception:
                pass
    _register_source_async_task(safe_name, asyncio.create_task(_run()))
    return _build_upload_response(task_id, safe_name, "accepted", "accepted", False, duplicate_state=duplicate_info.get("duplicate_state"), duplicate_of=duplicate_info.get("duplicate_of"), same_title_candidates=duplicate_info.get("same_title_candidates") or [])


@app.post("/documents/upload")
async def upload_document_file(file: UploadFile = File(...)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名不能为空")
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="文件内容为空")
    safe_name = _safe_filename(file.filename)
    probe = _probe_file_for_parser(safe_name, raw)
    _validate_upload_probe(safe_name, probe, is_text_upload=False)
    title_profile = _doc_title_profile(safe_name)
    content_sha256 = _content_sha256_bytes(raw)
    duplicate_info = _detect_duplicate_upload(safe_name, content_sha256, title_profile["canonical_title"])
    if duplicate_info.get("duplicate_state") in {"no_change", "already_exists"}:
        return _build_upload_response(
            task_id=None,
            source=safe_name,
            task_status="completed",
            document_status=str(duplicate_info.get("duplicate_state") or "no_change"),
            searchable=bool(_doc_get(duplicate_info.get("duplicate_of") or safe_name).get("searchable")),
            duplicate_state=duplicate_info.get("duplicate_state"),
            duplicate_of=duplicate_info.get("duplicate_of"),
            same_title_candidates=duplicate_info.get("same_title_candidates") or [],
        )
    task_id = _new_task_id()
    path = os.path.join(UPLOAD_DIR, f"{task_id}__{safe_name}")
    with open(path, "wb") as f:
        f.write(raw)
    TASKS[task_id] = {
        "status": "accepted",
        "stage": "validating",
        "filename": safe_name,
        "path": path,
        "created_at": datetime.now().isoformat(),
        "document_status": "accepted",
    }
    _task_log(task_id, "validating", {"filename": safe_name, "probe": probe})
    _lex_db_set_status(safe_name, "accepted")
    source_id = _build_source_id(file.filename, content_sha256)
    _doc_upsert(safe_name, status="accepted", canonical_title=title_profile["canonical_title"], title_tokens=title_profile["title_tokens"], aliases=title_profile["aliases"], filename_stem=title_profile["stem"], doc_type=None, topic=None, source_id=source_id, original_filename=file.filename, content_sha256=content_sha256, mime_type=probe.get("mime_type"), detected_ext=probe.get("detected_ext"), file_size=probe.get("file_size"), page_count=probe.get("page_count"), parser_route=probe.get("route"), parser_backend=probe.get("parser_backend"), parse_status="accepted", searchable=0, duplicate_state=duplicate_info.get("duplicate_state"), duplicate_of=duplicate_info.get("duplicate_of"), same_title_group=_same_title_group(title_profile["canonical_title"]), suspicious_file_type=0)
    _docfts_upsert(safe_name, title=title_profile["canonical_title"], aliases=title_profile["aliases"], doc_type=None, topic=None)
    async def _run():
        try:
            lock = _get_source_lock(safe_name)
            if not lock.acquire(timeout=30):
                raise HTTPException(status_code=429, detail="同一文档正在处理，请稍后重试")
            _lex_tx_begin()
            _lex_db_set_status(safe_name, "reindexing")
            v_next = _doc_next_version(safe_name)
            _doc_upsert(safe_name, status="reindexing", pending_version=v_next, parse_status="parsing", searchable=_doc_searchable_flag(safe_name))
            TASKS[task_id]["status"] = "indexing"
            TASKS[task_id]["stage"] = "parsing"
            _task_log(task_id, "parsing")
            document_ir = extract_document_ir_from_file(
                safe_name,
                raw,
                metadata={"file_type": os.path.splitext(safe_name)[1].lstrip(".").lower(), "file_size": len(raw)},
                doc_version=v_next,
            )
            text = _document_ir_plain_text(document_ir, normalized=False)
            quality = _assess_document_quality(document_ir, probe)
            if quality["status"] == "parse_empty":
                raise HTTPException(status_code=400, detail="parse_empty: 解析后正文为空")
            if quality["status"] == "parse_low_quality":
                raise HTTPException(status_code=400, detail="parse_low_quality: 解析质量不足，未进入发布")
            TASKS[task_id]["stage"] = "profile_building"
            _task_log(task_id, "profile_building", {"quality": quality})
            profile = _build_document_profile(safe_name, file.filename, source_id, content_sha256, text, document_ir, probe, quality, metadata={"file_type": os.path.splitext(safe_name)[1].lstrip(".").lower(), "file_size": len(raw)})
            _doc_upsert(safe_name, status="reindexing", pending_version=v_next, parse_status=quality["status"], parse_quality_score=quality["score"], quality_flags=_json_dumps(quality["flags"]), canonical_title=profile["canonical_title"], title_tokens=" ".join(profile.get("title_aliases") or []), aliases=",".join((profile.get("title_aliases") or [])[1:]), filename_stem=_filename_stem(safe_name), doc_type=profile.get("doc_type"), topic=",".join((profile.get("topic_terms") or [])[:8]), source_id=source_id, original_filename=file.filename, content_sha256=content_sha256, mime_type=probe.get("mime_type"), detected_ext=probe.get("detected_ext"), file_size=probe.get("file_size"), page_count=probe.get("page_count"), parser_route=probe.get("route"), parser_backend=probe.get("parser_backend"), searchable=_doc_searchable_flag(safe_name), duplicate_state=duplicate_info.get("duplicate_state"), duplicate_of=duplicate_info.get("duplicate_of"), same_title_group=_same_title_group(profile["canonical_title"]))
            TASKS[task_id]["status"] = "indexing"
            TASKS[task_id]["stage"] = "embedding"
            _task_log(task_id, "embedding")
            # Reindex is versioned: only purge remnants of the target pending version.
            _purge_source_for_reindex(safe_name, v_next)
            total_done = await index_document_incremental(
                task_id=task_id,
                filename=safe_name,
                text=text,
                metadata={"file_type": os.path.splitext(safe_name)[1].lstrip(".").lower(), "file_size": len(raw)},
                document_ir=document_ir,
            )
            _persist_document_profile(safe_name, v_next, profile)
            _crash_inject("before_commit")
            _lex_db_set_status(safe_name, "vector_pending")
            _doc_upsert(safe_name, status="vector_pending", pending_version=v_next, last_error=None, parse_status=quality["status"], searchable=_doc_searchable_flag(safe_name))
            _lex_tx_commit()
            _lex_db_checkpoint("PASSIVE")
            TASKS[task_id]["status"] = "completed"
            TASKS[task_id]["stage"] = "publish_pending"
            TASKS[task_id]["chunks_indexed"] = total_done
            TASKS[task_id]["document_status"] = "vector_pending"
            _task_log(task_id, "publish_pending", {"chunks_indexed": total_done, "pending_version": v_next})
        except asyncio.CancelledError:
            _lex_tx_rollback()
            TASKS[task_id]["status"] = "failed"
            TASKS[task_id]["stage"] = "cancelled"
            TASKS[task_id]["error"] = "cancelled_for_delete"
            _doc_upsert(safe_name, status="pending_delete", last_error="cancelled_for_delete", searchable=0)
            _task_log(task_id, "cancelled", {"reason": "delete_requested"})
            raise
        except HTTPException as e:
            _lex_tx_rollback()
            detail = str(e.detail or "upload_failed")
            status_code = detail.split(":", 1)[0]
            TASKS[task_id]["status"] = "failed"
            TASKS[task_id]["stage"] = status_code
            TASKS[task_id]["error"] = detail
            _doc_upsert(safe_name, status=status_code, last_error=detail, parse_status=status_code, searchable=_doc_searchable_flag(safe_name))
            _task_log(task_id, "failed", {"error": detail})
        except Exception as e:
            _lex_tx_rollback()
            TASKS[task_id]["status"] = "failed"
            TASKS[task_id]["stage"] = "failed"
            TASKS[task_id]["error"] = str(e)
            _doc_upsert(safe_name, status="vector_failed", last_error=str(e), searchable=_doc_searchable_flag(safe_name))
            _task_log(task_id, "failed", {"error": str(e)})
        finally:
            try:
                lock.release()
            except Exception:
                pass
    _register_source_async_task(safe_name, asyncio.create_task(_run()))
    return _build_upload_response(task_id, safe_name, "accepted", "accepted", False, duplicate_state=duplicate_info.get("duplicate_state"), duplicate_of=duplicate_info.get("duplicate_of"), same_title_candidates=duplicate_info.get("same_title_candidates") or [])


@app.get("/documents")
async def list_documents():
    """列出文档"""
    conn = _lex_db_connect()
    rows = conn.execute(
        "SELECT source, status, active_version, pending_version, last_error, updated_at, canonical_title, same_title_group FROM documents"
    ).fetchall()
    docs_map: Dict[str, Dict[str, Any]] = {}
    for source, status, active_version, pending_version, last_error, updated_at, canonical_title, same_title_group in rows:
        item = {
            "filename": source,
            "created_at": updated_at,
            "status": status or "not_found",
            "document_status": status or "not_found",
            "task_status": None,
            "task_id": None,
            "chunks_indexed": None,
            "error": last_error,
            "searchable": bool(_doc_searchable_flag(source)),
            "doc_type": None,
            "topics": [],
            "canonical_title": canonical_title or _filename_stem(source),
            "canonical_doc_id": same_title_group or _same_title_group(canonical_title or _filename_stem(source)),
        }
        docs_map[source] = item

    for tid, t in TASKS.items():
        fname = t.get("filename")
        if not fname:
            continue
        created_at = t.get("created_at") or ""
        status = _public_task_status(t.get("status"))
        task_item = {
            "filename": fname,
            "created_at": created_at,
            "status": status,
            "document_status": (docs_map.get(fname) or {}).get("document_status") or status,
            "task_status": status,
            "task_id": tid,
            "chunks_indexed": t.get("chunks_indexed") if status == "completed" else None,
            "error": t.get("error") if status == "failed" else None,
            "searchable": bool(_doc_searchable_flag(fname)),
            "doc_type": None,
            "topics": [],
            "canonical_title": (_doc_get(fname).get("canonical_title") or _filename_stem(fname)),
            "canonical_doc_id": _canonical_doc_id_for_source(fname),
        }
        existing = docs_map.get(fname)
        if not existing:
            if status not in ("accepted", "indexing"):
                continue
            docs_map[fname] = task_item
            continue
        # 如果 Milvus 已有索引，则保持 completed；否则用任务状态
        if (existing.get("status") or "") not in ("completed", "vector_pending"):
            docs_map[fname] = task_item
        else:
            # 统一选择最新的 created_at
            newer = (created_at or "") > (existing.get("created_at") or "")
            if newer:
                existing["created_at"] = created_at
            existing["error"] = None
    milvus_stats = _milvus_source_stats()
    for source, stats in milvus_stats.items():
        existing = docs_map.get(source)
        if existing is None:
            docs_map[source] = {
                "filename": source,
                "created_at": stats.get("created_at"),
                "status": "completed",
                "document_status": "completed",
                "task_status": None,
                "task_id": None,
                "chunks_indexed": stats.get("chunks_indexed"),
                "error": None,
                "searchable": True,
                "doc_type": None,
                "topics": [],
                "canonical_title": (_doc_get(source).get("canonical_title") or _filename_stem(source)),
                "canonical_doc_id": _canonical_doc_id_for_source(source),
            }
            continue
        existing["chunks_indexed"] = stats.get("chunks_indexed")
        if stats.get("created_at") and (stats.get("created_at") > (existing.get("created_at") or "")):
            existing["created_at"] = stats.get("created_at")
        # 如果 Milvus 中已经有可见 chunk，而控制面没有 completed 状态，则按可查看状态展示
        if int(stats.get("chunks_indexed") or 0) > 0 and (existing.get("status") or "") not in ("completed", "vector_pending"):
            existing["status"] = "completed"
            existing["document_status"] = "completed"
            existing["error"] = None
            existing["searchable"] = True

    documents = sorted(docs_map.values(), key=lambda x: x.get("created_at") or "", reverse=True)
    return {"documents": documents}

@app.delete("/documents/{filename}")
async def delete_document(filename: str):
    safe_name = _safe_filename(filename)
    delete_task_id = _new_task_id()
    TASKS[delete_task_id] = {
        "op": "delete",
        "status": "indexing",
        "stage": "deleting_milvus",
        "filename": safe_name,
        "created_at": datetime.now().isoformat(),
    }
    _task_log(delete_task_id, "delete_started", {"filename": safe_name})
    _lex_db_set_status(safe_name, "deleting")
    _doc_upsert(safe_name, status="deleting", last_error=None)

    cancelled_tasks = await _cancel_source_async_tasks(safe_name)
    if cancelled_tasks:
        _task_log(delete_task_id, "cancelled_source_tasks", {"count": cancelled_tasks})

    lock = _get_source_lock(safe_name)
    if not lock.acquire(timeout=30):
        TASKS[delete_task_id]["status"] = "failed"
        TASKS[delete_task_id]["stage"] = "deleting_locked"
        TASKS[delete_task_id]["error"] = "同一文档正在处理，请稍后重试"
        _task_log(delete_task_id, "failed", {"error": TASKS[delete_task_id]["error"]})
        raise HTTPException(status_code=429, detail="同一文档正在处理，请稍后重试")

    try:
        try:
            vector_db = VectorDBService()
            vector_db.connect()
            _crash_inject("delete_milvus")
            vector_delete_info = _delete_milvus_document_object(vector_db, safe_name)
        except Exception as e:
            msg = str(e)
            lower = msg.lower()
            unavailable = ("503" in lower) or ("unavailable" in lower) or ("connection" in lower) or ("timeout" in lower)
            TASKS[delete_task_id]["status"] = "accepted"
            TASKS[delete_task_id]["stage"] = "queued_for_compensation"
            TASKS[delete_task_id]["error"] = msg
            _lex_db_set_status(safe_name, "pending_delete")
            _doc_upsert(safe_name, status="pending_delete", last_error=msg)
            _enqueue_pending_delete(safe_name, last_error=msg, delete_files=True)
            _task_log(delete_task_id, "queued_for_compensation", {"stage": "deleting_milvus", "error": msg})
            return JSONResponse(
                status_code=202 if unavailable else 202,
                content={
                    "filename": safe_name,
                    "status": "pending_delete",
                    "task_id": delete_task_id,
                    "stage": "queued_for_compensation",
                    "error": msg,
                    "retryable": True,
                    "queued": True,
                }
            )

        TASKS[delete_task_id]["status"] = "indexing"
        TASKS[delete_task_id]["stage"] = "deleting_lexical"
        _task_log(delete_task_id, "deleting_lexical", vector_delete_info)
        try:
            _lex_db_delete_source(safe_name)
        except Exception as e:
            msg = str(e)
            TASKS[delete_task_id]["status"] = "accepted"
            TASKS[delete_task_id]["stage"] = "queued_for_compensation"
            TASKS[delete_task_id]["error"] = msg
            _lex_db_set_status(safe_name, "pending_delete")
            _doc_upsert(safe_name, status="pending_delete", last_error=msg)
            _enqueue_pending_delete(safe_name, last_error=msg, delete_files=True)
            _task_log(delete_task_id, "queued_for_compensation", {"stage": "deleting_lexical", "error": msg})
            return JSONResponse(
                status_code=202,
                content={
                    "filename": safe_name,
                    "status": "pending_delete",
                    "task_id": delete_task_id,
                    "stage": "queued_for_compensation",
                    "error": msg,
                    "retryable": True,
                    "queued": True,
                }
            )

        file_cleanup = _delete_uploaded_artifacts(safe_name)
        if file_cleanup.get("failed"):
            _enqueue_pending_delete(safe_name, last_error="artifact_cleanup_failed", delete_files=True)

        TASKS[delete_task_id]["status"] = "completed"
        TASKS[delete_task_id]["stage"] = "completed"
        _task_log(delete_task_id, "completed", {"action": "delete", "files_removed": len(file_cleanup.get("removed") or [])})

        removed = []
        for tid in list(TASKS.keys()):
            t = TASKS.get(tid) or {}
            if t.get("filename") == safe_name:
                removed.append(tid)
                del TASKS[tid]
        _save_tasks()
        return {
            "filename": safe_name,
            "status": "completed",
            "task_id": delete_task_id,
            "tasks_removed": removed,
            "vector_cleanup": vector_delete_info,
            "file_cleanup": file_cleanup,
        }
    finally:
        try:
            lock.release()
        except Exception:
            pass
@app.get("/tasks")
async def list_tasks():
    items = []
    for tid, t in TASKS.items():
        item = {"task_id": tid, **t}
        item["task_status"] = _public_task_status(t.get("status"))
        item["status"] = item["task_status"]
        if t.get("filename"):
            doc = _doc_get(t["filename"])
            item["document_status"] = doc.get("status")
            item["searchable"] = bool(_doc_searchable_flag(t["filename"]))
        items.append(item)
    items.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return {"tasks": items}

@app.get("/tasks/{task_id}")
async def get_task(task_id: str):
    t = TASKS.get(task_id)
    if not t:
        raise HTTPException(status_code=404, detail="task not found")
    item = {"task_id": task_id, **t}
    item["task_status"] = _public_task_status(t.get("status"))
    item["status"] = item["task_status"]
    if t.get("filename"):
        doc = _doc_get(t["filename"])
        item["document_status"] = doc.get("status")
        item["searchable"] = bool(_doc_searchable_flag(t["filename"]))
    return item

@app.post("/documents/{task_id}/retry")
async def retry_task(task_id: str):
    t = TASKS.get(task_id)
    if not t:
        raise HTTPException(status_code=404, detail="任务不存在")
    if (t.get("op") or "index") != "index":
        raise HTTPException(status_code=400, detail="该任务类型不支持该重试方式，请重新发起删除")
    filename = t.get("filename") or "unknown"
    _task_log(task_id, "retry", {"filename": filename})
    TASKS[task_id]["status"] = "accepted"
    TASKS[task_id]["stage"] = "accepted"
    TASKS[task_id]["error"] = None
    TASKS[task_id]["chunks_indexed"] = None
    _save_tasks()

    async def _run():
        lock = _get_source_lock(filename)
        try:
            if not lock.acquire(timeout=30):
                raise HTTPException(status_code=429, detail="同一文档正在处理，请稍后重试")
            TASKS[task_id]["status"] = "indexing"
            TASKS[task_id]["stage"] = "parsing"
            _task_log(task_id, "parsing")
            if t.get("path"):
                # 文件任务重试
                path = t.get("path")
                if not path or not os.path.exists(path):
                    raise HTTPException(status_code=404, detail="原文件缺失，无法重试")
                with open(path, "rb") as f:
                    raw = f.read()
                document_ir = extract_document_ir_from_file(
                    filename,
                    raw,
                    metadata={"file_type": os.path.splitext(filename)[1].lstrip(".").lower(), "file_size": len(raw)},
                    doc_version=_doc_get(filename).get("active_version") or 1,
                )
                text = _document_ir_plain_text(document_ir, normalized=False)
                TASKS[task_id]["status"] = "indexing"
                TASKS[task_id]["stage"] = "embedding"
                _task_log(task_id, "embedding")
                chunks = await index_document(
                    filename=filename,
                    text=text,
                    metadata={"file_type": os.path.splitext(filename)[1].lstrip(".").lower(), "file_size": len(raw)},
                    document_ir=document_ir,
                )
            else:
                # 文本任务重试
                payload = t.get("payload") or {}
                text = payload.get("text") or ""
                metadata = payload.get("metadata")
                document_ir = _build_document_ir_from_text(
                    filename,
                    text,
                    metadata=metadata,
                    parser_name="direct_text",
                    doc_version=_doc_get(filename).get("active_version") or 1,
                )
                TASKS[task_id]["status"] = "indexing"
                TASKS[task_id]["stage"] = "embedding"
                _task_log(task_id, "embedding")
                chunks = await index_document(
                    filename=filename,
                    text=text,
                    metadata=metadata,
                    document_ir=document_ir,
                )
            TASKS[task_id]["status"] = "completed"
            TASKS[task_id]["stage"] = "completed"
            TASKS[task_id]["chunks_indexed"] = chunks
            _task_log(task_id, "completed", {"chunks_indexed": chunks})
        except asyncio.CancelledError:
            TASKS[task_id]["status"] = "failed"
            TASKS[task_id]["stage"] = "cancelled"
            TASKS[task_id]["error"] = "cancelled_for_delete"
            _doc_upsert(filename, status="pending_delete", last_error="cancelled_for_delete")
            _task_log(task_id, "cancelled", {"reason": "delete_requested"})
            raise
        except Exception as e:
            TASKS[task_id]["status"] = "failed"
            TASKS[task_id]["stage"] = "failed"
            TASKS[task_id]["error"] = str(e)
            _task_log(task_id, "failed", {"error": str(e)})
        finally:
            try:
                lock.release()
            except Exception:
                pass
    _register_source_async_task(filename, asyncio.create_task(_run()))
    return {"task_id": task_id, "status": "accepted", "filename": filename}

@app.get("/documents/{filename}")
async def get_document_detail(filename: str):
    """查看文档详情（按 source 聚合已索引分块）"""
    safe_name = _safe_filename(filename)

    # 取该文件最新任务，辅助判定处理中/失败语义
    latest_task_id = None
    latest_task = None
    latest_created_at = ""
    for tid, t in TASKS.items():
        if (t.get("filename") or "") != safe_name:
            continue
        created_at = t.get("created_at") or ""
        if created_at >= latest_created_at:
            latest_created_at = created_at
            latest_task_id = tid
            latest_task = t

    try:
        vector_db = None
        try:
            vector_db = VectorDBService()
            vector_db.connect()
        except Exception:
            # Milvus 不可用时，返回控制面状态
            doc = _doc_get(safe_name)
            doc_status = _lex_db_get_status(safe_name) or (doc.get("status") or "not_found")
            code = 200
            if doc_status in {"accepted", "indexing", "reindexing", "vector_pending"}:
                code = 202
            elif doc_status in {"failed", "vector_failed", "delete_failed"}:
                code = 409
            elif (doc.get("status") is None) and (doc_status == "not_found"):
                raise HTTPException(status_code=404, detail="文档不存在")
            return JSONResponse(
                status_code=code,
                content={
                    "filename": safe_name,
                    "status": doc_status,
                    "document_status": doc_status,
                    "task_status": _public_task_status((latest_task or {}).get("status")),
                    "searchable": bool(_doc_searchable_flag(safe_name)),
                    "task_id": latest_task_id,
                    "stage": (latest_task or {}).get("stage") or "",
                    "active_version": doc.get("active_version"),
                    "pending_version": doc.get("pending_version"),
                    "last_error": doc.get("last_error"),
                    "chunks": [],
                    "chunk_count": 0,
                },
            )

        safe_filename = json.dumps(safe_name, ensure_ascii=False)
        response = vector_db.client.query(
            collection_name=vector_db.collection_name,
            filter=f"source == {safe_filename}",
            output_fields=["text", "source", "metadata", "created_at"],
            limit=5000
        )

        # 仅展示 active_version 的内容（发布控制）
        av = _get_active_version(safe_name)
        if av is not None:
            response = [r for r in (response or []) if ((r.get("metadata") or {}).get("doc_version") == av)]

        if not response:
            doc_status = _lex_db_get_status(safe_name) or "not_found"
            task_status = _public_task_status((latest_task or {}).get("status"))
            task_stage = (latest_task or {}).get("stage") or ""
            processing_statuses = {"accepted", "indexing", "reindexing", "vector_pending"}

            if task_status in processing_statuses:
                return JSONResponse(
                    status_code=202,
                    content={
                        "filename": safe_name,
                        "status": doc_status,
                        "task_id": latest_task_id,
                        "stage": task_stage or task_status,
                        "active_version": _get_active_version(safe_name),
                        "pending_version": _doc_get(safe_name).get("pending_version"),
                        "last_error": _doc_get(safe_name).get("last_error"),
                        "chunks": [],
                        "chunk_count": 0,
                    },
                )

            if task_status == "failed":
                return JSONResponse(
                    status_code=409,
                    content={
                        "filename": safe_name,
                        "status": doc_status or "vector_failed",
                        "task_id": latest_task_id,
                        "error": (latest_task or {}).get("error") or "索引失败",
                        "active_version": _get_active_version(safe_name),
                        "pending_version": _doc_get(safe_name).get("pending_version"),
                        "last_error": _doc_get(safe_name).get("last_error"),
                        "chunks": [],
                        "chunk_count": 0,
                    },
                )

            raise HTTPException(status_code=404, detail="文档不存在")

        def _chunk_id(item: Dict[str, Any]) -> int:
            metadata = item.get("metadata") or {}
            try:
                return int(metadata.get("chunk_id", 0))
            except Exception:
                return 0

        chunks = sorted(response, key=_chunk_id)
        document_ir = _ensure_document_ir(safe_name, av)
        full_text = _document_ir_plain_text(document_ir, normalized=False) if document_ir else "\n\n".join([((c.get("metadata") or {}).get("raw_text") or c.get("text") or "") for c in chunks]).strip()

        return {
            "filename": safe_name,
            "created_at": chunks[0].get("created_at"),
            "chunk_count": len(chunks),
            "status": _lex_db_get_status(safe_name) or "completed",
            "content": full_text,
            "ir_available": bool(document_ir),
            "document_metadata": (document_ir or {}).get("metadata") or {},
            "elements": [
                {
                    "element_id": element.get("element_id"),
                    "page_no": element.get("page_no"),
                    "section_path": element.get("section_path") or [],
                    "element_type": element.get("element_type"),
                    "reading_order": element.get("reading_order"),
                    "text_raw": element.get("text_raw") or "",
                    "text_normalized": element.get("text_normalized") or "",
                    "ocr_used": bool(element.get("ocr_used")),
                    "ocr_confidence": element.get("ocr_confidence"),
                    "parser_name": element.get("parser_name"),
                    "parser_version": element.get("parser_version"),
                }
                for element in (document_ir or {}).get("elements") or []
            ],
            "chunks": [
                {
                    "chunk_id": _chunk_id(c),
                    "text": ((c.get("metadata") or {}).get("raw_text") or c.get("text") or ""),
                    "metadata": c.get("metadata", {}),
                }
                for c in chunks
            ],
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"get_document_detail_error: filename={safe_name} err={e}")
        # 最后兜底：返回控制面状态，以避免 500
        doc = _doc_get(safe_name)
        doc_status = _lex_db_get_status(safe_name) or (doc.get("status") or "not_found")
        if (doc.get("status") is None) and (doc_status == "not_found"):
            raise HTTPException(status_code=404, detail="文档不存在")
        return JSONResponse(
            status_code=200,
            content={
                "filename": safe_name,
                "status": doc_status,
                "task_id": latest_task_id,
                "stage": (latest_task or {}).get("stage") or "",
                "active_version": doc.get("active_version"),
                "pending_version": doc.get("pending_version"),
                "last_error": doc.get("last_error"),
                "chunks": [],
                "chunk_count": 0,
            },
        )


# 挂载静态文件（前端）
if os.path.exists(WEB_DIR):
    app.mount("/", StaticFiles(directory=WEB_DIR, html=True), name="static")
else:
    logger.warning(f"Web directory not found: {WEB_DIR}")


if __name__ == "__main__":
    import uvicorn
    logger.info("Starting RAG Application...")
    uvicorn.run(app, host="0.0.0.0", port=8080)
